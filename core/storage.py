"""
MindForge v5.5.3 存储引擎
支持四层记忆架构：感官记忆 → 短期记忆 → 长期记忆 → 永久记忆
"""

import sqlite3
import json
import math
import uuid
import os
import time
import logging
import threading
import tempfile
from pathlib import Path
from collections import OrderedDict
from dataclasses import dataclass, field, asdict
from typing import List, Optional, Dict, Any, Tuple
from datetime import datetime, timezone

# v5.5.5: psutil 为可选依赖，不可用时降级到 ctypes / 手工检测
try:
    import psutil
    _HAS_PSUTIL = True
except ImportError:
    psutil = None
    _HAS_PSUTIL = False

logger = logging.getLogger(__name__)

from .types import (
    PrivacyLevel, Importance, MemoryType, MemoryLayer,
    DramaGenre, DramaStatus, DramaSeries, DramaScene,
    DramaCharacter, DramaLine,
)
from .encryption import EncryptionEngine, EncryptedBlob, SecurityError


# ===== 路径安全校验（v5.2.9 新增：存储层统一防护）=====

# v5.3.5 安全加固：检测 Windows 短文件名（8.3）绕过尝试
def _is_suspicious_windows_path(comp: str) -> bool:
    """检测 Windows 短文件名绕过模式（如 PROGRA~1、FILE~1.TXT）"""
    if not comp or len(comp) == 0:
        return False
    # v5.4.5 修复 #11：豁免 Unix 根路径 '/'，否则 Linux/Mac 上所有导出功能不可用
    if comp == '/':
        return False
    import re as _re
    # v5.3.7 修复：豁免 Windows 盘符根（如 C:\、D:），之前误报导致所有导出功能失效
    if len(comp) <= 3 and _re.match(r'^[A-Za-z]:\\?$', comp):
        return False
    # 匹配 短名模式：基础名 + ~N + 可选扩展名
    if _re.match(r'^[^~]{1,6}~\d(\..{1,3})?$', comp, _re.IGNORECASE):
        return True
    # v5.5.1 fix: colon only dangerous on Windows
    import sys as _sys
    _dangerous = ('..', '/', '\\', '\x00', ':') if _sys.platform == 'win32' else ('..', '/', '\\', '\x00')
    if any(s in comp for s in _dangerous):
        return True
    return False


def _safe_path(path_str, must_exist=False, allow_symlinks=False,
               max_size=None, allowed_exts=None, max_len=4096):
    """校验文件路径安全性，防止路径遍历攻击"""
    if not path_str or not isinstance(path_str, str):
        raise ValueError("路径不能为空")
    if len(path_str) > max_len:
        raise ValueError(f"路径过长（上限 {max_len} 字符）")
    # v5.3.5 安全：过滤 Unicode 双向和控制字符
    import unicodedata
    for ch in path_str:
        cat = unicodedata.category(ch)
        # 过滤双向控制字符（RLO/LRO 等）和 NUL，防止显示欺骗
        if cat in ('Cf', 'Cc') and ch not in '\n\r\t':
            raise ValueError("路径中包含非法控制字符")
    # v5.3.5 安全：逐组件检测 Windows 短文件名绕过
    target = Path(path_str)
    if not target.is_absolute():
        target = Path.cwd() / target
    for comp in target.parts:
        if comp and _is_suspicious_windows_path(comp):
            raise ValueError(f"路径组件不安全: {comp}")

    try:
        resolved = target.resolve()
    except (OSError, RuntimeError) as e:
        raise ValueError(f"路径解析失败: {e}")

    if not allow_symlinks:
        check_path = resolved
        while check_path != check_path.parent:
            if check_path.is_symlink():
                raise ValueError(f"不允许操作符号链接: {check_path}")
            check_path = check_path.parent

    if allowed_exts is not None:
        ext = resolved.suffix.lower()
        if ext not in allowed_exts:
            raise ValueError(
                f"不支持的文件类型: {ext}（允许: {', '.join(sorted(allowed_exts))}）"
            )

    if must_exist and not resolved.exists():
        raise FileNotFoundError(f"文件不存在: {resolved}")

    if max_size is not None and resolved.exists() and resolved.is_file():
        size = resolved.stat().st_size
        if size > max_size:
            raise ValueError(f"文件过大: {size} 字节（上限 {max_size}）")

    return resolved


# v5.3.5 安全加固：JSON 反序列化深度限制，防止深度攻击
def _safe_json_loads(data: str, max_depth: int = 32, max_size: int = 10_000_000):
    """安全加载 JSON，限制嵌套深度和总大小"""
    if not isinstance(data, str):
        raise ValueError("JSON 数据类型错误")
    if len(data) > max_size:
        raise ValueError(f"JSON 数据过大（{len(data)} > {max_size} 字节）")

    import re as _re

    def _check_depth(s: str) -> int:
        """用括号匹配粗检查深度（在 json.loads 之前快速失败）"""
        cur = 0
        mx = 0
        in_str = False
        esc = False
        for ch in s:
            if in_str:
                if esc:
                    esc = False
                elif ch == '\\':
                    esc = True
                elif ch == '"':
                    in_str = False
                continue
            if ch == '"':
                in_str = True
            elif ch in ('{', '['):
                cur += 1
                if cur > mx:
                    mx = cur
                    if mx > max_depth:
                        return mx
            elif ch in ('}', ']'):
                cur = max(0, cur - 1)
        return mx

    if _check_depth(data) > max_depth:
        raise ValueError(f"JSON 嵌套过深（上限 {max_depth} 层）")
    try:
        return json.loads(data)
    except json.JSONDecodeError:
        raise ValueError("JSON 解析失败")


# v5.3.5 安全加固：限制 SQL 查询返回行数，防止大数据量 DoS
def _limited_fetch(cursor, limit: int = 10000):
    """带行数上限的 fetch 辅助函数，防止数据量过大"""
    rows = cursor.fetchmany(limit + 1)
    if len(rows) > limit:
        raise ValueError(f"查询结果超过行数上限 {limit}")
    return rows


# v5.3.7 安全加固：Unicode 控制字符过滤，防止双向字符（RLO/LRO）显示欺骗
def _filter_unicode_ctrl(s: str) -> str:
    """过滤 Unicode Cf/Cc 类控制字符（保留 \\n\\r\\t），防止路径/ID 显示欺骗"""
    if not isinstance(s, str) or not s:
        return s
    import unicodedata
    return ''.join(
        ch for ch in s
        if unicodedata.category(ch) not in ('Cf', 'Cc') or ch in '\n\r\t'
    )


# v5.4.1 修复：内容长度上限提升为模块级常量，统一作用于
# add_memory / update_memory / batch_add，堵住此前仅 add_memory
# 校验导致的超长内容绕过（DoS）问题。
MAX_CONTENT_LEN = 50000


def _validate_content_len(content: Optional[str]) -> None:
    """v5.4.1 安全加固：统一内容长度校验，超限直接拒绝（不做静默截断）"""
    if content and isinstance(content, str) and len(content) > MAX_CONTENT_LEN:
        raise ValueError(f"content exceeds {MAX_CONTENT_LEN} chars (got {len(content)})")


# v5.3.3 安全加固：LIKE 通配符转义，防止 % 和 _ 被解释为 SQL LIKE 通配符
def _escape_like(value: str) -> str:
    """转义 LIKE 模式中的通配符 % 和 _，防止通配符注入"""
    if not isinstance(value, str):
        return ""
    return value.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")


# v5.3.3 安全加固 + v5.4.7 M-5 修复：增强 HTML/XSS 消毒
_XSS_RE = __import__("re").compile(
    r'<[^>]*>|javascript:|vbscript:|data:text/html|on(?:error|load|click|mouseover|focus|blur|submit|change|input|keydown|keyup|keypress|dblclick|mousedown|mouseup|mousemove|mouseout|mouseenter|mouseleave|contextmenu|wheel|drag|drop|copy|cut|paste|abort|canplay|ended|pause|play|playing|progress|ratechange|seeked|seeking|stalled|suspend|timeupdate|volumechange|waiting|animationstart|animationend|animationiteration|transitionend|toggle|resize|scroll|storage|message|online|offline|popstate|hashchange|beforeunload|pagehide|pageshow|unload)\s*=|<script|</script|<iframe|</iframe|<object|<embed|<svg|<math|<form|<input|<button|<textarea|<select|<option|<applet|<meta|<link|<base',
    __import__("re").IGNORECASE
)

def _sanitize_html(value: str, max_len: int = 10000) -> str:
    """清洗 HTML 内容，防止存储型 XSS

    移除 HTML 标签和危险的事件处理器属性。
    v5.4.7 M-5 修复：扩展事件处理器列表，增加更多危险标签和协议。
    """
    if not isinstance(value, str):
        return ""
    if len(value) > max_len:
        value = value[:max_len]
    # 移除 HTML 标签和危险内容
    cleaned = _XSS_RE.sub("", value)
    return cleaned


# v5.3.3 安全加固：敏感操作频率限制器
class _RateLimiter:
    """简单的内存频率限制器，防止暴力攻击"""

    def __init__(self):
        self._windows: Dict[str, List[float]] = {}
        # v5.3.5 安全：记录最后清理时间，防止内存泄漏
        self._last_purge: float = time.time()
        self._purge_interval: int = 3600  # 每小时清理一次

    def check(self, key: str, max_calls: int = 10, window_seconds: int = 60) -> bool:
        """检查是否超过频率限制

        Args:
            key: 限制键（如 agent_id + operation）
            max_calls: 窗口内最大调用次数
            window_seconds: 时间窗口（秒）

        Returns:
            True=允许，False=超限
        """
        now = time.time()
        # v5.3.5 安全：定期清理全部过期条目，防止内存泄漏
        self._maybe_purge(now)

        if key not in self._windows:
            self._windows[key] = []

        # 清理过期记录
        self._windows[key] = [t for t in self._windows[key] if now - t < window_seconds]

        # v5.3.5 安全：该键窗口清空后直接移除，减少字典体积
        if not self._windows[key]:
            del self._windows[key]
            self._windows[key] = []

        if len(self._windows[key]) >= max_calls:
            return False

        self._windows[key].append(now)
        return True

    def _maybe_purge(self, now: float):
        """定期清理全表过期条目"""
        if now - self._last_purge < self._purge_interval:
            return
        self._last_purge = now
        max_age = max(self._purge_interval * 2, 86400)
        dead_keys = [k for k, ts in self._windows.items()
                     if not ts or now - ts[-1] > max_age]
        for k in dead_keys:
            del self._windows[k]


_rate_limiter = _RateLimiter()


@dataclass
class MemoryEntry:
    """记忆条目"""
    id: str
    content: str = ""
    category: str = "general"
    tags: List[str] = field(default_factory=list)
    privacy: PrivacyLevel = PrivacyLevel.INTERNAL
    importance: Importance = Importance.MEDIUM
    memory_type: MemoryType = MemoryType.TEXT
    layer: MemoryLayer = MemoryLayer.SHORT_TERM
    source_session: str = ""
    source_agent: str = ""
    created_at: float = 0.0
    updated_at: float = 0.0
    last_accessed_at: float = 0.0
    access_count: int = 0
    consolidation_count: int = 0
    forgetting_score: float = 0.0
    strength: float = 1.0
    starred: bool = False
    pinned: bool = False
    expires_at: float = 0.0  # v5.5.2: TTL expiration timestamp (0 = never expires)
    metadata: Dict[str, Any] = field(default_factory=dict)
    encrypted: bool = False
    ciphertext: Optional[bytes] = None
    nonce: Optional[bytes] = None
    salt: Optional[bytes] = None

    @property
    def preview(self) -> str:
        if self.encrypted and not self.content:
            return "[已加密]"
        return self.content[:100]

    def to_dict(self) -> dict:
        return {
            "id": self.id,
            "content": self.content,
            "category": self.category,
            "tags": self.tags,
            "privacy": self.privacy.value,
            "importance": self.importance.value,
            "memory_type": self.memory_type.value,
            "layer": self.layer.value,
            "source_session": self.source_session,
            "source_agent": self.source_agent,
            "created_at": self.created_at,
            "updated_at": self.updated_at,
            "last_accessed_at": self.last_accessed_at,
            "access_count": self.access_count,
            "consolidation_count": self.consolidation_count,
            "forgetting_score": self.forgetting_score,
            "strength": self.strength,
            "starred": self.starred,
            "pinned": self.pinned,
            "expires_at": self.expires_at,
            "metadata": self.metadata,
        }


@dataclass
class AuditRecord:
    """审计记录"""
    id: str
    action: str
    memory_id: str
    actor: str
    session_id: str
    privacy_level: str
    timestamp: float
    details: Dict[str, Any] = field(default_factory=dict)


class HardwareProfiler:
    """硬件性能分析器（v5.5.5 新增）

    自动检测 CPU、内存、磁盘性能，给出记忆检索的推荐配置。
    """

    @staticmethod
    def detect() -> Dict[str, Any]:
        """检测硬件性能，返回性能画像

        返回：
        {
            "cpu_cores": int,
            "memory_gb": float,
            "disk_type": "ssd" | "hdd" | "unknown",
            "performance_level": "low" | "medium" | "high",
            "recommended_limit": int,  # 推荐检索数量
            "recommended_depth": int,  # 推荐检索深度
            "cache_size_mb": int,  # 推荐缓存大小
        }
        """
        # --- CPU 检测 ---
        cpu_cores = os.cpu_count() or 1

        # --- 内存检测 ---
        memory_gb = HardwareProfiler._detect_memory_gb()

        # --- 磁盘检测 ---
        disk_type = HardwareProfiler._detect_disk_type()

        # --- 性能分级 ---
        performance_level = HardwareProfiler._classify_performance(cpu_cores, memory_gb)

        # --- 推荐配置 ---
        rec = HardwareProfiler._recommend_config(performance_level)

        return {
            "cpu_cores": cpu_cores,
            "memory_gb": memory_gb,
            "disk_type": disk_type,
            "performance_level": performance_level,
            "recommended_limit": rec["limit"],
            "recommended_depth": rec["depth"],
            "cache_size_mb": rec["cache_size_mb"],
        }

    @staticmethod
    def _detect_memory_gb() -> float:
        """检测系统总内存（GB），优先使用 psutil，降级到 ctypes"""
        if _HAS_PSUTIL:
            try:
                mem = psutil.virtual_memory()
                return round(mem.total / (1024 ** 3), 2)
            except Exception:
                pass

        # Fallback: Windows API via ctypes
        try:
            import ctypes
            kernel32 = ctypes.windll.kernel32

            class MEMORYSTATUSEX(ctypes.Structure):
                _fields_ = [
                    ("dwLength", ctypes.c_ulong),
                    ("dwMemoryLoad", ctypes.c_ulong),
                    ("ullTotalPhys", ctypes.c_ulonglong),
                    ("ullAvailPhys", ctypes.c_ulonglong),
                    ("ullTotalPageFile", ctypes.c_ulonglong),
                    ("ullAvailPageFile", ctypes.c_ulonglong),
                    ("ullTotalVirtual", ctypes.c_ulonglong),
                    ("ullAvailVirtual", ctypes.c_ulonglong),
                    ("ullAvailExtendedVirtual", ctypes.c_ulonglong),
                ]

            ms = MEMORYSTATUSEX()
            ms.dwLength = ctypes.sizeof(MEMORYSTATUSEX)
            if kernel32.GlobalMemoryStatusEx(ctypes.byref(ms)):
                return round(ms.ullTotalPhys / (1024 ** 3), 2)
        except Exception:
            pass

        # 终极 fallback：假设 4GB
        return 4.0

    @staticmethod
    def _detect_disk_type() -> str:
        """检测磁盘类型（SSD/HDD），优先使用 psutil，降级到写入速度测试"""
        if _HAS_PSUTIL:
            try:
                # psutil 在部分平台可直接获取磁盘类型
                for disk in psutil.disk_partitions():
                    try:
                        usage = psutil.disk_usage(disk.mountpoint)
                        # 无法直接从 psutil 获取 SSD/HDD，跳过
                    except Exception:
                        continue
                # psutil 不直接提供 SSD/HDD 判断，使用写入速度测试
            except Exception:
                pass

        # Fallback: 通过小文件写入速度粗略判断
        try:
            test_data = os.urandom(5 * 1024 * 1024)  # 5MB
            with tempfile.NamedTemporaryFile(delete=False, suffix=".tmp") as f:
                tmp_path = f.name
                start = time.time()
                f.write(test_data)
                f.flush()
                os.fsync(f.fileno())
                elapsed = time.time() - start
            try:
                os.unlink(tmp_path)
            except OSError:
                pass

            if elapsed > 0:
                speed_mbps = (5 * 1024 * 1024 / elapsed) / (1024 * 1024)  # MB/s
                if speed_mbps > 50:
                    return "ssd"
                else:
                    return "hdd"
        except Exception:
            pass

        return "unknown"

    @staticmethod
    def _classify_performance(cpu_cores: int, memory_gb: float) -> str:
        """根据 CPU 和内存分级性能

        - LOW: CPU <= 2 核 或 内存 < 4GB
        - MEDIUM: CPU 2-4 核 且 内存 4-8GB
        - HIGH: CPU > 4 核 且 内存 > 8GB
        """
        if cpu_cores <= 2 or memory_gb < 4:
            return "low"
        elif cpu_cores > 4 and memory_gb > 8:
            return "high"
        else:
            return "medium"

    @staticmethod
    def _recommend_config(performance_level: str) -> Dict[str, Any]:
        """根据性能级别给出推荐配置"""
        configs = {
            "low": {"limit": 10, "depth": 1, "cache_size_mb": 16},
            "medium": {"limit": 20, "depth": 2, "cache_size_mb": 64},
            "high": {"limit": 50, "depth": 3, "cache_size_mb": 256},
        }
        return configs.get(performance_level, configs["medium"])


class MemoryCache:
    """记忆缓存（v5.5.5 新增）

    LRU 缓存高频访问的记忆摘要，减少磁盘 IO 和重复计算。
    """

    def __init__(self, max_size: int = 1000):
        self.max_size = max_size
        self._cache: OrderedDict[str, Dict[str, Any]] = OrderedDict()
        self.hits = 0
        self.misses = 0

    def get(self, memory_id: str) -> Optional[Dict[str, Any]]:
        """获取缓存，命中时移到末尾（最近使用）"""
        if memory_id in self._cache:
            self._cache.move_to_end(memory_id)
            self.hits += 1
            return self._cache[memory_id]
        self.misses += 1
        return None

    def set(self, memory_id: str, data: Dict[str, Any]) -> None:
        """设置缓存，如果满了淘汰最久未使用的"""
        if memory_id in self._cache:
            self._cache.move_to_end(memory_id)
            self._cache[memory_id] = data
        else:
            if len(self._cache) >= self.max_size:
                # 淘汰最久未使用的（队首）
                self._cache.popitem(last=False)
            self._cache[memory_id] = data

    def invalidate(self, memory_id: str) -> None:
        """使单条缓存失效"""
        if memory_id in self._cache:
            del self._cache[memory_id]

    def clear(self) -> None:
        """清空所有缓存"""
        self._cache.clear()
        self.hits = 0
        self.misses = 0

    def stats(self) -> Dict[str, Any]:
        """缓存统计"""
        total = self.hits + self.misses
        hit_rate = round(self.hits / total, 4) if total > 0 else 0.0
        return {
            "size": len(self._cache),
            "max_size": self.max_size,
            "hits": self.hits,
            "misses": self.misses,
            "hit_rate": hit_rate,
        }


class StorageEngine:
    """存储引擎"""

    def __init__(self, db_path: str = "./data/memory.db",
                 encryption: Optional[EncryptionEngine] = None,
                 encrypted: bool = True):
        self.db_path = Path(db_path)
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self.encryption = encryption
        self.encrypted = encrypted and encryption is not None
        # v5.4.6 线程安全：SQLite 连接不能跨线程复用。
        # 使用 threading.local 为每个线程维护独立连接（REST API 场景必需）。
        self._conn_local = threading.local()
        # v5.5.5 硬件自适应：检测硬件性能并初始化记忆缓存
        self._hardware_profile = HardwareProfiler.detect()
        cache_size = self._hardware_profile["recommended_limit"] * 50  # 按推荐 limit 估算缓存条目
        self._memory_cache = MemoryCache(max_size=cache_size)
        self._init_db()

    def _get_conn(self) -> sqlite3.Connection:
        conn = getattr(self._conn_local, "conn", None)
        if conn is None:
            conn = sqlite3.connect(str(self.db_path), timeout=10.0)
            conn.row_factory = sqlite3.Row
            conn.execute("PRAGMA journal_mode=WAL")
            conn.execute("PRAGMA foreign_keys=ON")
            conn.execute("PRAGMA busy_timeout=5000")  # v5.4.7 修复 L-4：并发写入时等待 5 秒
            self._conn_local.conn = conn
        return conn

    def _init_db(self):
        conn = self._get_conn()
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS memories (
                id TEXT PRIMARY KEY,
                content TEXT,
                ciphertext BLOB,
                nonce BLOB,
                salt BLOB,
                category TEXT DEFAULT 'general',
                tags TEXT DEFAULT '[]',
                privacy TEXT DEFAULT 'INTERNAL',
                importance TEXT DEFAULT 'MEDIUM',
                memory_type TEXT DEFAULT 'text',
                layer TEXT DEFAULT 'short_term',
                source_session TEXT DEFAULT '',
                source_agent TEXT DEFAULT '',
                created_at REAL,
                updated_at REAL,
                last_accessed_at REAL,
                access_count INTEGER DEFAULT 0,
                consolidation_count INTEGER DEFAULT 0,
                forgetting_score REAL DEFAULT 0.0,
                strength REAL DEFAULT 1.0,
                starred INTEGER DEFAULT 0,
                pinned INTEGER DEFAULT 0,
                metadata TEXT DEFAULT '{}',
                encrypted INTEGER DEFAULT 0
            );

            CREATE INDEX IF NOT EXISTS idx_category ON memories(category);
            CREATE INDEX IF NOT EXISTS idx_privacy ON memories(privacy);
            CREATE INDEX IF NOT EXISTS idx_layer ON memories(layer);
            CREATE INDEX IF NOT EXISTS idx_importance ON memories(importance);
            CREATE INDEX IF NOT EXISTS idx_created_at ON memories(created_at);
            CREATE INDEX IF NOT EXISTS idx_memory_type ON memories(memory_type);
            CREATE INDEX IF NOT EXISTS idx_starred ON memories(starred);
            CREATE INDEX IF NOT EXISTS idx_pinned ON memories(pinned);

            -- 记忆关联（v5.2.5 新增）
            CREATE TABLE IF NOT EXISTS memory_links (
                id TEXT PRIMARY KEY,
                source_id TEXT NOT NULL,
                target_id TEXT NOT NULL,
                link_type TEXT DEFAULT 'related',
                note TEXT DEFAULT '',
                created_at REAL,
                UNIQUE(source_id, target_id)
            );
            CREATE INDEX IF NOT EXISTS idx_link_source ON memory_links(source_id);
            CREATE INDEX IF NOT EXISTS idx_link_target ON memory_links(target_id);
            CREATE INDEX IF NOT EXISTS idx_link_type ON memory_links(link_type);

            -- 记忆版本历史（v5.2.7 新增）
            CREATE TABLE IF NOT EXISTS memory_versions (
                id TEXT PRIMARY KEY,
                memory_id TEXT NOT NULL,
                version_number INTEGER NOT NULL,
                content TEXT NOT NULL,
                category TEXT,
                tags TEXT,
                importance TEXT,
                actor TEXT DEFAULT '',
                changed_at REAL,
                FOREIGN KEY (memory_id) REFERENCES memories(id),
                UNIQUE(memory_id, version_number)
            );
            CREATE INDEX IF NOT EXISTS idx_versions_memory_id ON memory_versions(memory_id);
            CREATE INDEX IF NOT EXISTS idx_versions_changed_at ON memory_versions(changed_at);

            CREATE VIRTUAL TABLE IF NOT EXISTS memory_fts USING fts5(
                content, category, tags,
                content='',
                tokenize='trigram'
            );

            CREATE TABLE IF NOT EXISTS audit_log (
                id TEXT PRIMARY KEY,
                action TEXT,
                memory_id TEXT,
                actor TEXT,
                session_id TEXT,
                privacy_level TEXT,
                timestamp REAL,
                details TEXT DEFAULT '{}'
            );

            CREATE INDEX IF NOT EXISTS idx_audit_memory ON audit_log(memory_id);
            CREATE INDEX IF NOT EXISTS idx_audit_timestamp ON audit_log(timestamp);

            CREATE TABLE IF NOT EXISTS knowledge_graph (
                id TEXT PRIMARY KEY,
                entity TEXT,
                entity_type TEXT,
                description TEXT,
                metadata TEXT DEFAULT '{}',
                created_at REAL
            );

            CREATE TABLE IF NOT EXISTS graph_relations (
                id TEXT PRIMARY KEY,
                from_entity TEXT,
                to_entity TEXT,
                relation_type TEXT,
                weight REAL DEFAULT 1.0,
                memory_ids TEXT DEFAULT '[]',
                metadata TEXT DEFAULT '{}',
                created_at REAL
            );

            CREATE INDEX IF NOT EXISTS idx_kg_entity ON knowledge_graph(entity);
            CREATE INDEX IF NOT EXISTS idx_relation_from ON graph_relations(from_entity);
            CREATE INDEX IF NOT EXISTS idx_relation_to ON graph_relations(to_entity);

            -- AI 短剧记忆模块（v5.2.1 新增）
            CREATE TABLE IF NOT EXISTS drama_series (
                id TEXT PRIMARY KEY,
                title TEXT NOT NULL,
                genre TEXT DEFAULT 'other',
                total_episodes INTEGER DEFAULT 0,
                current_episode INTEGER DEFAULT 0,
                status TEXT DEFAULT 'planned',
                platform TEXT DEFAULT '',
                rating REAL DEFAULT 0.0,
                description TEXT DEFAULT '',
                tags TEXT DEFAULT '[]',
                cover_url TEXT DEFAULT '',
                metadata TEXT DEFAULT '{}',
                created_at REAL,
                updated_at REAL,
                last_watched_at REAL DEFAULT 0.0
            );
            CREATE INDEX IF NOT EXISTS idx_drama_genre ON drama_series(genre);
            CREATE INDEX IF NOT EXISTS idx_drama_status ON drama_series(status);
            CREATE INDEX IF NOT EXISTS idx_drama_rating ON drama_series(rating);

            CREATE TABLE IF NOT EXISTS drama_scenes (
                id TEXT PRIMARY KEY,
                drama_id TEXT NOT NULL,
                episode INTEGER DEFAULT 0,
                scene_number INTEGER DEFAULT 0,
                title TEXT DEFAULT '',
                content TEXT DEFAULT '',
                location TEXT DEFAULT '',
                time_of_day TEXT DEFAULT '',
                tags TEXT DEFAULT '[]',
                metadata TEXT DEFAULT '{}',
                created_at REAL
            );
            CREATE INDEX IF NOT EXISTS idx_scene_drama ON drama_scenes(drama_id);
            CREATE INDEX IF NOT EXISTS idx_scene_episode ON drama_scenes(episode);

            CREATE TABLE IF NOT EXISTS drama_characters (
                id TEXT PRIMARY KEY,
                drama_id TEXT NOT NULL,
                name TEXT NOT NULL,
                role TEXT DEFAULT 'supporting',
                actor TEXT DEFAULT '',
                description TEXT DEFAULT '',
                personality TEXT DEFAULT '',
                avatar_url TEXT DEFAULT '',
                tags TEXT DEFAULT '[]',
                metadata TEXT DEFAULT '{}',
                created_at REAL
            );
            CREATE INDEX IF NOT EXISTS idx_char_drama ON drama_characters(drama_id);
            CREATE INDEX IF NOT EXISTS idx_char_name ON drama_characters(name);

            CREATE TABLE IF NOT EXISTS drama_lines (
                id TEXT PRIMARY KEY,
                drama_id TEXT NOT NULL,
                scene_id TEXT DEFAULT '',
                character_id TEXT DEFAULT '',
                character_name TEXT DEFAULT '',
                line_text TEXT NOT NULL,
                context TEXT DEFAULT '',
                episode INTEGER DEFAULT 0,
                timestamp TEXT DEFAULT '',
                is_classic INTEGER DEFAULT 0,
                memory_id TEXT DEFAULT '',
                tags TEXT DEFAULT '[]',
                metadata TEXT DEFAULT '{}',
                created_at REAL
            );
            CREATE INDEX IF NOT EXISTS idx_line_drama ON drama_lines(drama_id);
            CREATE INDEX IF NOT EXISTS idx_line_character ON drama_lines(character_id);
            CREATE INDEX IF NOT EXISTS idx_line_classic ON drama_lines(is_classic);

            -- 记忆笔记/批注（v5.2.4 新增）
            CREATE TABLE IF NOT EXISTS memory_notes (
                id TEXT PRIMARY KEY,
                memory_id TEXT NOT NULL,
                content TEXT NOT NULL,
                author TEXT DEFAULT '',
                tags TEXT DEFAULT '[]',
                created_at REAL,
                updated_at REAL
            );
            CREATE INDEX IF NOT EXISTS idx_note_memory ON memory_notes(memory_id);
            CREATE INDEX IF NOT EXISTS idx_note_author ON memory_notes(author);

            -- 记忆模板（v5.2.4 新增）
            CREATE TABLE IF NOT EXISTS memory_templates (
                id TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                content_template TEXT NOT NULL,
                category TEXT DEFAULT 'general',
                tags TEXT DEFAULT '[]',
                importance TEXT DEFAULT 'MEDIUM',
                layer TEXT DEFAULT 'short_term',
                description TEXT DEFAULT '',
                use_count INTEGER DEFAULT 0,
                created_at REAL,
                updated_at REAL
            );
            CREATE INDEX IF NOT EXISTS idx_template_name ON memory_templates(name);
            CREATE INDEX IF NOT EXISTS idx_template_category ON memory_templates(category);

            -- 复习计划（v5.2.4 新增）
            CREATE TABLE IF NOT EXISTS review_schedules (
                id TEXT PRIMARY KEY,
                memory_id TEXT NOT NULL,
                scheduled_at REAL NOT NULL,
                interval_days REAL DEFAULT 1.0,
                review_count INTEGER DEFAULT 0,
                last_reviewed_at REAL DEFAULT 0.0,
                status TEXT DEFAULT 'pending',
                created_at REAL
            );
            CREATE INDEX IF NOT EXISTS idx_schedule_memory ON review_schedules(memory_id);
            CREATE INDEX IF NOT EXISTS idx_schedule_status ON review_schedules(status);
            CREATE INDEX IF NOT EXISTS idx_schedule_due ON review_schedules(scheduled_at);

            -- 记忆嵌入向量（v5.4.5 新增：向量检索）
            CREATE TABLE IF NOT EXISTS memory_embeddings (
                memory_id TEXT PRIMARY KEY,
                embedding BLOB NOT NULL,
                model_name TEXT DEFAULT '',
                dimension INTEGER DEFAULT 0,
                created_at REAL,
                updated_at REAL
            );
            CREATE INDEX IF NOT EXISTS idx_embedding_model ON memory_embeddings(model_name);

            -- 归档记忆（v5.4.6 新增：auto-archive 机制）
            CREATE TABLE IF NOT EXISTS archived_memories (
                id TEXT PRIMARY KEY,
                original_id TEXT NOT NULL,
                content TEXT DEFAULT '',
                category TEXT DEFAULT 'general',
                tags TEXT DEFAULT '[]',
                privacy TEXT DEFAULT 'INTERNAL',
                importance TEXT DEFAULT 'MEDIUM',
                memory_type TEXT DEFAULT 'text',
                layer TEXT DEFAULT 'short_term',
                source_session TEXT DEFAULT '',
                source_agent TEXT DEFAULT '',
                metadata TEXT DEFAULT '{}',
                original_created_at REAL DEFAULT 0,
                original_updated_at REAL DEFAULT 0,
                archived_at REAL DEFAULT 0,
                archived_reason TEXT DEFAULT 'expired'
            );
            CREATE INDEX IF NOT EXISTS idx_archived_layer ON archived_memories(layer);
            CREATE INDEX IF NOT EXISTS idx_archived_category ON archived_memories(category);
            CREATE INDEX IF NOT EXISTS idx_archived_at ON archived_memories(archived_at);
        """)
        conn.commit()

        # v5.5.2 migration: add expires_at column for TTL support
        try:
            conn.execute("ALTER TABLE memories ADD COLUMN expires_at REAL DEFAULT 0")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_expires_at ON memories(expires_at)")
            conn.commit()
        except sqlite3.OperationalError:
            pass  # column already exists

    @staticmethod
    def _strip_control(text: Optional[str]) -> str:
        """v5.4.0 安全加固：过滤控制字符（保留 \\t\\n\\r，过滤 \\x00-\\x1f 和 \\x7f）。
        同时做 str 类型强制，非字符串转空串。"""
        if text is None:
            return ""
        if not isinstance(text, str):
            text = str(text)
        out_chars = []
        for c in text:
            cp = ord(c)
            if cp >= 0x20 and cp != 0x7F:
                out_chars.append(c)
            elif c in "\n\r\t":
                out_chars.append(c)
        return "".join(out_chars)

    @staticmethod
    def _sanitize_tags(tags: Optional[List[Any]],
                       max_tag_len: int = 64,
                       max_tags: int = 64) -> List[str]:
        """v5.4.0 安全加固：tags 列表清洗——类型强制 + 控制字符过滤 + 长度限制 + 去重。"""
        if tags is None:
            return []
        # 只接受 list/tuple/set/frozenset；其他（str、dict 等）直接返回空防误转
        if isinstance(tags, str):
            return []
        if not isinstance(tags, (list, tuple, set, frozenset)):
            return []
        tags_list = list(tags)
        seen: set[str] = set()
        cleaned: list[str] = []
        for t in tags_list:
            if t is None:
                continue
            s = StorageEngine._strip_control(str(t)).strip()
            s = _XSS_RE.sub("", s)  # v5.4.8 安全修复：tags HTML 消毒
            if not s:
                continue
            if len(s) > max_tag_len:
                s = s[:max_tag_len]
            if s in seen:
                continue
            seen.add(s)
            cleaned.append(s)
            if len(cleaned) >= max_tags:
                break
        return cleaned

    @staticmethod
    def _sanitize_metadata(metadata: Optional[Dict[str, Any]],
                           max_depth: int = 5,
                           max_string_len: int = 2000) -> Dict[str, Any]:
        """v5.4.0 安全加固：metadata 递归清洗——控制字符过滤 + 深度限制 + 字符串长度限制。"""
        if metadata is None:
            return {}
        if not isinstance(metadata, dict):
            return {}

        def _sanitize(v, depth):
            if depth > max_depth:
                return None
            if v is None:
                return None
            if isinstance(v, bool):
                return v
            if isinstance(v, int):
                return v if abs(v) <= 2**63 else int(v // 2)
            if isinstance(v, float):
                if math.isnan(v) or math.isinf(v):
                    return None  # NaN/Inf JSON 不可序列化，剔除
                return v
            if isinstance(v, str):
                s = StorageEngine._strip_control(v)
                s = _XSS_RE.sub("", s)  # v5.4.8 安全修复：metadata HTML 消毒
                if len(s) > max_string_len:
                    s = s[:max_string_len]
                # 全空白字符串返回空串而非 None（用户可能存空格占位）
                return s
            if isinstance(v, (list, tuple, set, frozenset)):
                out = []
                for item in list(v)[:256]:
                    cleaned = _sanitize(item, depth + 1)
                    if cleaned is not None:
                        out.append(cleaned)
                return out
            if isinstance(v, dict):
                out = {}
                for k, val in list(v.items())[:256]:
                    raw_key = StorageEngine._strip_control(str(k))
                    cleaned_key = raw_key.strip()[:128]
                    if not cleaned_key:
                        continue
                    cleaned_val = _sanitize(val, depth + 1)
                    if cleaned_val is not None:
                        out[cleaned_key] = cleaned_val
                return out
            return None

        result = _sanitize(metadata, 0)
        return result if isinstance(result, dict) else {}

    @staticmethod
    def _downgrade_enum(value: Any, enum_cls: type, default: Any) -> Any:
        """v5.4.0 安全加固：枚举值降级——非法值自动降级到默认值。
        同时支持大小写（.value 匹配 / 原始大小写 / .name 匹配）。

        v5.5.4 fix: 兼容双重导入路径导致的枚举类不一致问题。
        当 isinstance() 因不同模块路径加载同名枚举而失败时，
        通过 .value / .name 属性匹配来识别。
        """
        if isinstance(value, enum_cls):
            return value
        # 兼容双重导入：isinstance 失败但 value 有枚举属性时，用 .value 匹配
        if hasattr(value, 'value') and hasattr(value, 'name'):
            try:
                return enum_cls(value.value)
            except (ValueError, KeyError):
                pass
        if value is None:
            return default
        sval = str(value).strip()
        if not sval:
            return default
        # 处理 "Importance.LOW" → "LOW" 的情况（双重导入 str() 结果）
        if '.' in sval and not sval.startswith('.'):
            parts = sval.split('.')
            if len(parts) == 2:
                sval = parts[-1]
        # 1) value 精确匹配（大小写不敏感）
        for m in enum_cls:
            if m.value.lower() == sval.lower():
                return m
        # 2) name 精确匹配（大小写不敏感）
        try:
            return enum_cls[sval.upper()]
        except KeyError:
            pass
        try:
            return enum_cls[sval]
        except KeyError:
            pass
        return default

    def add_memory(self,
                   content: str,
                   category: str = "general",
                   tags: Optional[List[str]] = None,
                   privacy: PrivacyLevel = PrivacyLevel.INTERNAL,
                   importance: Importance = Importance.MEDIUM,
                   memory_type: MemoryType = MemoryType.TEXT,
                   layer: MemoryLayer = MemoryLayer.SHORT_TERM,
                   source_session: str = "",
                   source_agent: str = "",
                   starred: bool = False,
                   pinned: bool = False,
                   expires_at: float = 0.0,
                   metadata: Optional[Dict[str, Any]] = None) -> MemoryEntry:
        """添加记忆

        v5.5.2 新增 expires_at 参数（TTL 过期时间戳，0=永不过期）。
        v5.5.6 新增 pinned 参数（置顶标记）。
        """
        # v5.4.7 修复 L-8：拒绝 None 或空内容
        if content is None:
            raise ValueError("content cannot be None")
        if isinstance(content, str) and not content.strip():
            raise ValueError("content cannot be empty or whitespace-only")
        # v5.3.9 安全加固：内容长度上限防 DoS（超限直接拒绝，不做静默截断）
        # v5.4.1 重构：统一走 _validate_content_len（模块级常量）
        _validate_content_len(content)
        if category and isinstance(category, str) and len(category) > 128:
            category = category[:128]
        if source_agent and isinstance(source_agent, str) and len(source_agent) > 128:
            source_agent = source_agent[:128]
        if source_session and isinstance(source_session, str) and len(source_session) > 128:
            source_session = source_session[:128]

        # v5.4.0 安全加固：控制字符过滤（所有字符串字段 + tags + metadata 递归）
        content = self._strip_control(content)
        content = _sanitize_html(content)  # v5.4.8 P0 修复：XSS 消毒
        category = self._strip_control(category)
        source_session = self._strip_control(source_session)
        source_agent = self._strip_control(source_agent)
        clean_tags = self._sanitize_tags(tags)
        clean_metadata = self._sanitize_metadata(metadata)

        # v5.4.0 安全加固：枚举值降级（非法值自动降级到默认值）
        importance = self._downgrade_enum(importance, Importance, Importance.MEDIUM)
        privacy = self._downgrade_enum(privacy, PrivacyLevel, PrivacyLevel.INTERNAL)
        layer = self._downgrade_enum(layer, MemoryLayer, MemoryLayer.SHORT_TERM)
        memory_type = self._downgrade_enum(memory_type, MemoryType, MemoryType.TEXT)

        # starred 类型强制
        starred = bool(starred)
        pinned = bool(pinned)

        now = time.time()
        entry_id = str(uuid.uuid4())

        ciphertext = None
        nonce = None
        salt = None
        stored_content = content

        if self.encrypted and self.encryption:
            blob = self.encryption.encrypt(content)
            ciphertext = blob.ciphertext
            nonce = blob.nonce
            salt = blob.salt
            stored_content = ""

        entry = MemoryEntry(
            id=entry_id,
            content=stored_content,
            category=category,
            tags=clean_tags,
            privacy=privacy,
            importance=importance,
            memory_type=memory_type,
            layer=layer,
            source_session=source_session,
            source_agent=source_agent,
            created_at=now,
            updated_at=now,
            last_accessed_at=now,
            metadata=clean_metadata,
            starred=starred,
            pinned=pinned,
            expires_at=float(expires_at) if expires_at else 0.0,
            encrypted=self.encrypted,
            ciphertext=ciphertext,
            nonce=nonce,
            salt=salt,
        )

        conn = self._get_conn()
        conn.execute("""
            INSERT INTO memories (
                id, content, ciphertext, nonce, salt, category, tags,
                privacy, importance, memory_type, layer,
                source_session, source_agent, created_at, updated_at,
                last_accessed_at, access_count, consolidation_count,
                forgetting_score, strength, starred, pinned, metadata, encrypted, expires_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            entry.id, entry.content, entry.ciphertext, entry.nonce, entry.salt,
            entry.category, json.dumps(entry.tags, ensure_ascii=False),
            entry.privacy.value, entry.importance.value, entry.memory_type.value,
            entry.layer.value, entry.source_session, entry.source_agent,
            entry.created_at, entry.updated_at, entry.last_accessed_at,
            entry.access_count, entry.consolidation_count,
            entry.forgetting_score, entry.strength, int(entry.starred),
            int(entry.pinned),
            json.dumps(entry.metadata, ensure_ascii=False), int(entry.encrypted),
            entry.expires_at
        ))

        if not self.encrypted:
            conn.execute("""
                INSERT INTO memory_fts (rowid, content, category, tags)
                VALUES ((SELECT rowid FROM memories WHERE id = ?), ?, ?, ?)
            """, (entry.id, content, category, json.dumps(clean_tags, ensure_ascii=False)))

        conn.commit()
        self._add_audit("add", entry.id, source_agent, source_session, privacy.value)

        # v5.4.5: 生成嵌入向量（失败不影响记忆写入）
        try:
            self._store_embedding(entry.id, content)
        except Exception:
            pass

        return entry

    def get_memory(self, memory_id: str,
                   actor: str = "", session_id: str = "") -> Optional[MemoryEntry]:
        """获取记忆

        v5.5.2: 自动过期检查——若 expires_at > 0 且已过期，自动移入回收站并返回 None。
        """
        conn = self._get_conn()
        # v5.5.5 fix: 排除回收站记忆
        row = conn.execute("SELECT * FROM memories WHERE id = ? AND category != 'trash'", (memory_id,)).fetchone()
        if not row:
            return None

        entry = self._row_to_entry(row)

        # v5.5.2: auto-expire check
        # v5.5.3 fix: 使用 UPDATE WHERE 防止并发竞态条件；静默失败避免重复写入
        if entry.expires_at and entry.expires_at > 0 and entry.expires_at <= time.time():
            try:
                now = time.time()
                result = conn.execute(
                    "UPDATE memories SET category = 'trash', updated_at = ? WHERE id = ? AND category != 'trash'",
                    (now, memory_id)
                )
                if result.rowcount > 0:
                    conn.commit()
                    self._add_audit("forget", memory_id, actor, session_id,
                                     entry.privacy.value, details={"reason": "ttl_expired", "expires_at": entry.expires_at})
                else:
                    # 已被其他线程/进程处理，无需重复操作
                    pass
            except Exception as e:
                logger.error("auto-expire failed for %s: %s", memory_id, e)
                # 过期失败时仍返回原始记忆，避免数据不一致
                self._update_access(entry, actor, session_id)
                return entry
            return None

        self._update_access(entry, actor, session_id)
        return entry

    def decrypt_content(self, entry: MemoryEntry) -> str:
        """解密记忆内容"""
        if not entry.encrypted or not self.encryption:
            return entry.content

        if entry.ciphertext and entry.nonce and entry.salt:
            blob = EncryptedBlob(
                ciphertext=entry.ciphertext,
                nonce=entry.nonce,
                salt=entry.salt,
            )
            return self.encryption.decrypt(blob)
        return entry.content

    def get_indexable_documents(self, limit: int = 100000) -> Dict[str, str]:
        """返回可索引的 {memory_id: content} 映射（v5.2.8 新增）

        用于 IndexEngine 在新进程启动时水合 TF-IDF 内存索引，
        修复 CLI 跨进程搜索不到历史记忆的问题。
        跳过回收站与加密条目（密文无法直接索引）。

        Args:
            limit: 最大加载条数（安全上限）

        Returns:
            {memory_id: content} 字典
        """
        conn = self._get_conn()
        rows = conn.execute(
            "SELECT id, content FROM memories"
            " WHERE category != 'trash' AND encrypted = 0"
            " ORDER BY created_at DESC LIMIT ?",
            (limit,),
        ).fetchall()
        return {row["id"]: (row["content"] or "") for row in rows}

    def list_memories(self,
                      category: Optional[str] = None,
                      layer: Optional[MemoryLayer] = None,
                      privacy: Optional[PrivacyLevel] = None,
                      starred: Optional[bool] = None,
                      pinned: Optional[bool] = None,
                      created_after: Optional[float] = None,
                      created_before: Optional[float] = None,
                      limit: int = 50,
                      offset: int = 0,
                      sort_by: str = "created_at",
                      sort_order: str = "desc") -> List[MemoryEntry]:
        """列出记忆（v5.2.5 新增 pinned 筛选和置顶优先排序）"""
        conn = self._get_conn()
        # v5.5.5 fix: 默认排除回收站记忆，但显式查询 trash 时除外
        if category == "trash":
            query = "SELECT * FROM memories WHERE 1=1"
        else:
            query = "SELECT * FROM memories WHERE category != 'trash'"
        params = []

        if category:
            query += " AND category = ?"
            params.append(category)
        if layer:
            query += " AND layer = ?"
            params.append(layer.value)
        if privacy:
            query += " AND privacy = ?"
            params.append(privacy.value)
        if starred is not None:
            query += " AND starred = ?"
            params.append(1 if starred else 0)
        if pinned is not None:
            query += " AND pinned = ?"
            params.append(1 if pinned else 0)
        if created_after is not None:
            query += " AND created_at >= ?"
            params.append(created_after)
        if created_before is not None:
            query += " AND created_at <= ?"
            params.append(created_before)

        valid_sort_columns = ["created_at", "updated_at", "last_accessed_at", "access_count", "strength", "forgetting_score"]
        if sort_by not in valid_sort_columns:
            sort_by = "created_at"
        sort_order = "DESC" if sort_order.lower() == "desc" else "ASC"

        # v5.2.5: 置顶记忆优先展示
        query += f" ORDER BY pinned DESC, {sort_by} {sort_order} LIMIT ? OFFSET ?"
        params.extend([limit, offset])

        rows = conn.execute(query, params).fetchall()
        return [self._row_to_entry(row) for row in rows]

    def update_memory(self,
                      entry_id: str,
                      content: Optional[str] = None,
                      category: Optional[str] = None,
                      tags: Optional[List[str]] = None,
                      privacy: Optional[PrivacyLevel] = None,
                      importance: Optional[Importance] = None,
                      layer: Optional[MemoryLayer] = None,
                      starred: Optional[bool] = None,
                      pinned: Optional[bool] = None,
                      metadata: Optional[Dict[str, Any]] = None,
                      actor: str = "",
                      session_id: str = "") -> bool:
        """更新记忆

        v5.0.6 修复：更新 content/category/tags 时同步刷新 FTS 索引，
        避免搜索返回旧内容（与 v5.0.5 的 delete FTS 修复对应）。

        contentless FTS5 表的 'delete' 命令需要索引中**当前的值**（旧值）才能
        正确从倒排索引中移除 token。因此必须：先读旧值 → 用旧值 delete FTS
        → 更新 memories 表 → 用新值 insert FTS。
        """
        conn = self._get_conn()
        now = time.time()

        # v5.4.0 安全加固：长度限制
        MAX_CONTENT_LEN = 50000
        if content is not None:
            if isinstance(content, str) and len(content) > MAX_CONTENT_LEN:
                raise ValueError(f"content exceeds {MAX_CONTENT_LEN} chars (got {len(content)})")
            content = self._strip_control(content)
            content = _sanitize_html(content)  # v5.4.8 安全修复：XSS 消毒
        if category is not None:
            if isinstance(category, str) and len(category) > 128:
                category = category[:128]
            category = self._strip_control(category)

        # v5.4.0 安全加固：控制字符过滤 + 枚举降级
        clean_tags = self._sanitize_tags(tags) if tags is not None else None
        clean_metadata = self._sanitize_metadata(metadata) if metadata is not None else None
        privacy_v = self._downgrade_enum(privacy, PrivacyLevel, None) if privacy is not None else None
        importance_v = self._downgrade_enum(importance, Importance, None) if importance is not None else None
        layer_v = self._downgrade_enum(layer, MemoryLayer, None) if layer is not None else None
        actor = self._strip_control(actor)[:128]
        session_id = self._strip_control(session_id)[:128]

        updates = []
        params = []
        fts_dirty = False  # 是否需要刷新 FTS

        if content is not None:
            # v5.4.1 修复：update 路径此前绕过内容长度校验，可注入超长内容（DoS）
            _validate_content_len(content)
            if self.encrypted and self.encryption:
                blob = self.encryption.encrypt(content)
                updates.append("ciphertext = ?")
                updates.append("nonce = ?")
                updates.append("salt = ?")
                params.extend([blob.ciphertext, blob.nonce, blob.salt])
                updates.append("content = ?")
                params.append("")
            else:
                updates.append("content = ?")
                params.append(content)
                fts_dirty = True

        if category is not None:
            updates.append("category = ?")
            params.append(category)
            fts_dirty = True
        if clean_tags is not None:
            updates.append("tags = ?")
            params.append(json.dumps(clean_tags, ensure_ascii=False))
            fts_dirty = True
        if privacy_v is not None:
            updates.append("privacy = ?")
            params.append(privacy_v.value)
        if importance_v is not None:
            updates.append("importance = ?")
            params.append(importance_v.value)
        if layer_v is not None:
            updates.append("layer = ?")
            params.append(layer_v.value)
        if starred is not None:
            updates.append("starred = ?")
            params.append(1 if bool(starred) else 0)
        if pinned is not None:
            updates.append("pinned = ?")
            params.append(1 if bool(pinned) else 0)
        if clean_metadata is not None:
            updates.append("metadata = ?")
            params.append(json.dumps(clean_metadata, ensure_ascii=False))

        if not updates:
            return False

        # FTS 刷新步骤 1：更新前先读取旧值，用旧值删除 FTS 索引条目
        old_fts_row = None
        if fts_dirty and not self.encrypted:
            old_fts_row = conn.execute(
                "SELECT rowid, content, category, tags FROM memories WHERE id = ?",
                (entry_id,)
            ).fetchone()
            if old_fts_row:
                try:
                    conn.execute(
                        "INSERT INTO memory_fts(memory_fts, rowid, content, category, tags) "
                        "VALUES('delete', ?, ?, ?, ?)",
                        (old_fts_row[0], old_fts_row[1] or "", old_fts_row[2] or "", old_fts_row[3] or "[]")
                    )
                except sqlite3.OperationalError:
                    pass

        # v5.2.7: 更新前读取旧内容，用于保存历史版本
        old_entry = None
        if content is not None:
            old_row = conn.execute(
                "SELECT content, category, tags, importance FROM memories WHERE id = ?",
                (entry_id,)
            ).fetchone()
            if old_row:
                old_entry = {
                    "content": old_row[0] or "",
                    "category": old_row[1] or "",
                    "tags": old_row[2],
                    "importance": old_row[3] or "",
                }

        # 更新 memories 表
        updates.append("updated_at = ?")
        params.append(now)
        params.append(entry_id)
        cursor = conn.execute(f"UPDATE memories SET {', '.join(updates)} WHERE id = ?", params)
        if cursor.rowcount == 0:
            return False  # v5.4.7 修复 C-1：更新不存在的 ID 时返回 False

        # FTS 刷新步骤 2：用新值重新插入 FTS 索引条目
        if old_fts_row is not None:
            new_row = conn.execute(
                "SELECT rowid, content, category, tags FROM memories WHERE id = ?",
                (entry_id,)
            ).fetchone()
            if new_row:
                try:
                    conn.execute(
                        "INSERT INTO memory_fts (rowid, content, category, tags) VALUES (?, ?, ?, ?)",
                        (new_row[0], new_row[1] or "", new_row[2] or "", new_row[3] or "[]")
                    )
                except sqlite3.OperationalError:
                    pass

        # v5.2.7: 保存历史版本（仅当 content 变更时）
        if content is not None and old_entry:
            try:
                old_tags = old_entry.get("tags", "")
                if isinstance(old_tags, str) and old_tags.startswith("["):
                    try:
                        old_tags = json.loads(old_tags)
                    except Exception:
                        pass
                self.save_version(
                    memory_id=entry_id,
                    content=old_entry.get("content", ""),
                    category=old_entry.get("category", ""),
                    tags=old_tags,
                    importance=old_entry.get("importance", ""),
                    actor=actor,
                )
            except Exception:
                pass  # 版本保存失败不影响主流程

        conn.commit()

        self._add_audit("update", entry_id, actor, session_id,
                        privacy_v.value if privacy_v else "")

        # v5.4.5: 内容变更时重新生成嵌入向量（失败不影响更新）
        if content is not None:
            try:
                self._store_embedding(entry_id, content)
            except Exception:
                pass

        return True

    def adjust_importance(self, entry_id: str, delta: float,
                          actor: str = "", session_id: str = "") -> bool:
        """调整记忆重要性（增量方式，v5.5.6 新增）

        用于冲突衰减等场景，将重要性按 delta 增量调整（正数增加，负数降低）。
        自动限制在 Importance 枚举范围内。

        Args:
            entry_id: 记忆 ID
            delta: 重要性增量（如 -0.15 表示降低 0.15）
            actor: 操作者
            session_id: 会话 ID

        Returns:
            是否成功
        """
        if not entry_id or delta == 0:
            return False
        conn = self._get_conn()
        now = time.time()
        try:
            row = conn.execute(
                "SELECT importance, privacy FROM memories WHERE id = ? AND category != 'trash'",
                (entry_id,)
            ).fetchone()
            if not row:
                return False

            current_imp = self._downgrade_enum(row[0], Importance, Importance.MEDIUM)
            imp_order = list(Importance)
            try:
                current_idx = imp_order.index(current_imp)
            except ValueError:
                current_idx = len(imp_order) // 2

            new_idx = current_idx + int(round(delta * len(imp_order)))
            new_idx = max(0, min(len(imp_order) - 1, new_idx))
            new_imp = imp_order[new_idx]

            if new_imp == current_imp:
                return False

            conn.execute(
                "UPDATE memories SET importance = ?, updated_at = ? WHERE id = ?",
                (new_imp.value, now, entry_id)
            )
            conn.commit()
            self._add_audit(
                "update", entry_id, actor, session_id,
                row[1] if row[1] else "",
                details={"importance_delta": delta, "new_importance": new_imp.value}
            )
            return True
        except Exception:
            conn.rollback()
            return False

    def append_tags(self, entry_id: str, tags: List[str],
                    actor: str = "", session_id: str = "") -> bool:
        """追加标签到单条记忆（v5.5.6 新增）

        与 batch_add_tags 类似，但只操作单条记忆，用于冲突衰减等场景。

        Args:
            entry_id: 记忆 ID
            tags: 要追加的标签列表
            actor: 操作者
            session_id: 会话 ID

        Returns:
            是否成功（标签有变化返回 True）
        """
        if not entry_id or not tags:
            return False
        conn = self._get_conn()
        now = time.time()
        try:
            row = conn.execute(
                "SELECT tags, privacy FROM memories WHERE id = ? AND category != 'trash'",
                (entry_id,)
            ).fetchone()
            if not row:
                return False

            existing_tags = self._safe_json_loads(row[0], [])
            new_tags = list(set(existing_tags + list(tags)))
            if len(new_tags) == len(existing_tags):
                return False

            conn.execute(
                "UPDATE memories SET tags = ?, updated_at = ? WHERE id = ?",
                (json.dumps(new_tags, ensure_ascii=False), now, entry_id)
            )
            self._refresh_fts(conn, entry_id)
            conn.commit()
            self._add_audit(
                "update", entry_id, actor, session_id,
                row[1] if row[1] else "",
                details={"added_tags": tags, "result_tags": new_tags}
            )
            return True
        except Exception:
            conn.rollback()
            return False

    def _refresh_fts(self, conn: sqlite3.Connection, entry_id: str):
        """刷新单条记忆的 FTS 索引（contentless FTS5：先 delete 旧条目再 insert 新条目）

        v5.0.6 新增：辅助方法，读取当前 memories 表中的值来刷新 FTS。
        注意：此方法用 memories 表的**当前值**做 delete 和 insert，仅适用于
        memories 表尚未被更新的场景（如索引修复）。update_memory 中的 FTS
        同步已内联实现（需在更新前读旧值），不调用此方法。
        """
        row = conn.execute(
            "SELECT rowid, content, category, tags FROM memories WHERE id = ?",
            (entry_id,)
        ).fetchone()
        if not row:
            return
        try:
            conn.execute(
                "INSERT INTO memory_fts(memory_fts, rowid, content, category, tags) "
                "VALUES('delete', ?, ?, ?, ?)",
                (row[0], row[1] or "", row[2] or "", row[3] or "[]")
            )
        except sqlite3.Error:
            pass
        try:
            conn.execute(
                "INSERT INTO memory_fts (rowid, content, category, tags) VALUES (?, ?, ?, ?)",
                (row[0], row[1] or "", row[2] or "", row[3] or "[]")
            )
        except sqlite3.Error:
            pass

    def rebuild_fts(self) -> dict:
        """重建 FTS 全文索引（v5.0.6 新增）

        清空并重新构建 memory_fts 表，消除孤立记录，确保索引与 memories 表一致。
        配合 health_check 发现的 fts_orphans 问题使用。

        注意：contentless FTS5 表（content=''）不支持 DELETE FROM，
        必须用 DROP + CREATE 重建表结构来清空。

        Returns:
            {
                "rebuilt": True,
                "indexed": int,      # 重建索引的条目数
                "duration_ms": float,
            }
        """
        import time as _time
        start = _time.time()
        conn = self._get_conn()

        # contentless FTS5 表不能用 DELETE FROM，用 DROP + CREATE 重建
        conn.execute("DROP TABLE IF EXISTS memory_fts")
        conn.execute("""
            CREATE VIRTUAL TABLE memory_fts USING fts5(
                content, category, tags,
                content='',
                tokenize='trigram'
            )
        """)

        # 重新索引所有非加密记忆
        rows = conn.execute(
            "SELECT rowid, content, category, tags FROM memories WHERE encrypted = 0"
        ).fetchall()
        indexed = 0
        for row in rows:
            try:
                conn.execute(
                    "INSERT INTO memory_fts (rowid, content, category, tags) VALUES (?, ?, ?, ?)",
                    (row[0], row[1] or "", row[2] or "", row[3] or "[]")
                )
                indexed += 1
            except sqlite3.OperationalError:
                pass

        conn.commit()
        elapsed = (_time.time() - start) * 1000
        return {
            "rebuilt": True,
            "indexed": indexed,
            "duration_ms": round(elapsed, 2),
        }

    def purge_trash(self,
                    actor: str = "system",
                    session_id: str = "") -> int:
        """清空回收站，永久删除所有 category='trash' 的记忆（v5.0.6 新增）

        软删除（delete_memory(hard_delete=False)）会把 category 改为 'trash'，
        本方法将这些记录彻底删除，并同步清理 FTS 索引。

        Args:
            actor: 操作者（审计日志用）
            session_id: 会话 ID

        Returns:
            永久删除的记忆数量
        """
        conn = self._get_conn()
        rows = conn.execute(
            "SELECT id, rowid, content, category, tags, privacy FROM memories WHERE category = 'trash'"
        ).fetchall()

        if not rows:
            return 0

        ids = [row[0] for row in rows]
        placeholders = ",".join(["?"] * len(ids))

        conn.execute(f"DELETE FROM memories WHERE id IN ({placeholders})", ids)

        # 同步清理 FTS（contentless FTS5 用 'delete' 命令）
        for row in rows:
            try:
                conn.execute(
                    "INSERT INTO memory_fts(memory_fts, rowid, content, category, tags) "
                    "VALUES('delete', ?, ?, ?, ?)",
                    (row[1], row[2] or "", row[3] or "", row[4] or "[]")
                )
            except sqlite3.OperationalError:
                pass

        conn.commit()
        # v5.5.5 fix: 审计日志记录真实隐私级别
        for row in rows:
            self._add_audit("purge", row[0], actor, session_id,
                             row[5] if row[5] else "")

        return len(ids)

    def delete_memory(self, entry_id: str,
                      actor: str = "", session_id: str = "",
                      hard_delete: bool = False) -> bool:
        """删除记忆

        v5.0.5 修复：硬删除时同步清理 FTS 索引，避免搜索时返回已删除的记忆。
        注意：contentless FTS5 表（content=''）不支持标准 DELETE，
        必须用 'delete' 特殊命令。
        """
        # v5.4.0 安全加固：actor/session_id 控制字符过滤 + 长度限制
        actor = self._strip_control(actor)[:128]
        session_id = self._strip_control(session_id)[:128]
        hard_delete = bool(hard_delete)

        conn = self._get_conn()

        # v5.4.7 修复 C-1：先检查记忆是否存在
        # v5.5.5 fix: 同时获取 privacy 用于审计日志
        exists_row = conn.execute(
            "SELECT privacy FROM memories WHERE id = ?", (entry_id,)
        ).fetchone()
        if not exists_row:
            return False
        privacy_val = exists_row[0] if exists_row[0] else ""

        if hard_delete:
            # 先取 rowid 和 FTS 字段用于清理
            row = conn.execute(
                "SELECT rowid, content, category, tags FROM memories WHERE id = ?",
                (entry_id,)
            ).fetchone()
            # v5.4.1 修复：先清理带外键的从表，再删主表，避免 FOREIGN KEY 约束失败
            # v5.5.5 fix: 补充缺失的关联表清理（memory_links, memory_notes, memory_embeddings, memory_templates）
            conn.execute("DELETE FROM memory_versions WHERE memory_id = ?", (entry_id,))
            conn.execute("DELETE FROM review_schedules WHERE memory_id = ?", (entry_id,))
            conn.execute("DELETE FROM memory_links WHERE source_id = ? OR target_id = ?", (entry_id, entry_id))
            conn.execute("DELETE FROM memory_notes WHERE memory_id = ?", (entry_id,))
            conn.execute("DELETE FROM memory_embeddings WHERE memory_id = ?", (entry_id,))
            try:
                conn.execute("DELETE FROM memory_templates WHERE memory_id = ?", (entry_id,))
            except sqlite3.OperationalError:
                pass
            try:
                conn.execute("DELETE FROM kg_edges WHERE source = ? OR target = ?", (entry_id, entry_id))
            except sqlite3.OperationalError:
                pass
            conn.execute("DELETE FROM memories WHERE id = ?", (entry_id,))
            if row:
                try:
                    # contentless FTS5 必须用 'delete' 命令
                    conn.execute(
                        "INSERT INTO memory_fts(memory_fts, rowid, content, category, tags) "
                        "VALUES('delete', ?, ?, ?, ?)",
                        (row[0], row[1] or "", row[2] or "", row[3] or "[]")
                    )
                except sqlite3.OperationalError:
                    pass
        else:
            now = time.time()
            # v5.1.1 修复：软删除时保存原分类到 metadata，便于恢复
            row = conn.execute(
                "SELECT category, metadata FROM memories WHERE id = ?",
                (entry_id,)
            ).fetchone()
            if row:
                try:
                    meta = json.loads(row["metadata"]) if row["metadata"] else {}
                except (json.JSONDecodeError, TypeError):
                    meta = {}
                meta["_original_category"] = row["category"]
                conn.execute(
                    "UPDATE memories SET category = 'trash', updated_at = ?, metadata = ? WHERE id = ?",
                    (now, json.dumps(meta, ensure_ascii=False), entry_id)
                )
            else:
                conn.execute("UPDATE memories SET category = 'trash', updated_at = ? WHERE id = ?",
                             (now, entry_id))

        conn.commit()
        # v5.5.5 fix: 审计日志记录真实隐私级别
        self._add_audit("delete", entry_id, actor, session_id, privacy_val)
        return True

    def batch_delete(self,
                     category: Optional[str] = None,
                     layer: Optional[MemoryLayer] = None,
                     starred: Optional[bool] = None,
                     created_after: Optional[float] = None,
                     created_before: Optional[float] = None,
                     hard_delete: bool = False,
                     actor: str = "",
                     session_id: str = "") -> int:
        """批量删除记忆，返回删除数量

        v5.0.5 修复：硬删除时同步清理 FTS 索引（用 'delete' 特殊命令）。
        """
        # v5.4.0 安全加固
        category = self._strip_control(category)[:128] if category else None
        layer = self._downgrade_enum(layer, MemoryLayer, None) if layer is not None else None
        hard_delete = bool(hard_delete)
        actor = self._strip_control(actor)[:128]
        session_id = self._strip_control(session_id)[:128]

        conn = self._get_conn()
        # v5.5.5 fix: 添加 privacy 用于审计日志；排除回收站记忆
        query = "SELECT id, rowid, content, category, tags, metadata, privacy FROM memories WHERE category != 'trash'"
        params = []

        if category:
            query += " AND category = ?"
            params.append(category)
        if layer:
            query += " AND layer = ?"
            params.append(layer.value)
        if starred is not None:
            query += " AND starred = ?"
            params.append(1 if bool(starred) else 0)
        if created_after is not None:
            try:
                _ca = float(created_after)
                if math.isfinite(_ca):
                    query += " AND created_at >= ?"
                    params.append(_ca)
            except (TypeError, ValueError):
                pass
        if created_before is not None:
            try:
                _cb = float(created_before)
                if math.isfinite(_cb):
                    query += " AND created_at <= ?"
                    params.append(_cb)
            except (TypeError, ValueError):
                pass

        rows = conn.execute(query, params).fetchall()

        if not rows:
            return 0

        ids = [row[0] for row in rows]
        now = time.time()

        if hard_delete:
            placeholders = ",".join(["?"] * len(ids))
            # v5.4.1 修复：先清从表外键再删主表
            # v5.5.5 fix: 补充缺失的关联表清理
            conn.execute(f"DELETE FROM memory_versions WHERE memory_id IN ({placeholders})", ids)
            conn.execute(f"DELETE FROM review_schedules WHERE memory_id IN ({placeholders})", ids)
            conn.execute(f"DELETE FROM memory_links WHERE source_id IN ({placeholders}) OR target_id IN ({placeholders})", ids + ids)
            conn.execute(f"DELETE FROM memory_notes WHERE memory_id IN ({placeholders})", ids)
            conn.execute(f"DELETE FROM memory_embeddings WHERE memory_id IN ({placeholders})", ids)
            try:
                conn.execute(f"DELETE FROM memory_templates WHERE memory_id IN ({placeholders})", ids)
            except sqlite3.OperationalError:
                pass
            try:
                conn.execute(
                    f"DELETE FROM kg_edges WHERE source IN ({placeholders}) OR target IN ({placeholders})",
                    ids + ids
                )
            except sqlite3.OperationalError:
                pass
            conn.execute(f"DELETE FROM memories WHERE id IN ({placeholders})", ids)
            # 同步清理 FTS（contentless FTS5 用 'delete' 命令）
            for row in rows:
                try:
                    conn.execute(
                        "INSERT INTO memory_fts(memory_fts, rowid, content, category, tags) "
                        "VALUES('delete', ?, ?, ?, ?)",
                        (row[1], row[2] or "", row[3] or "", row[4] or "[]")
                    )
                except sqlite3.OperationalError:
                    pass
        else:
            # v5.1.1 修复：批量软删除时也保存原分类到 metadata
            for row in rows:
                try:
                    meta = json.loads(row["metadata"]) if row["metadata"] else {}
                except (json.JSONDecodeError, TypeError):
                    meta = {}
                meta["_original_category"] = row["category"]
                conn.execute(
                    "UPDATE memories SET category = 'trash', updated_at = ?, metadata = ? WHERE id = ?",
                    (now, json.dumps(meta, ensure_ascii=False), row["id"])
                )

        conn.commit()
        # v5.5.5 fix: 审计日志记录真实隐私级别
        for row in rows:
            self._add_audit("delete", row[0], actor, session_id,
                             row[6] if row[6] else "")

        return len(ids)

    def restore_memory(self, entry_id: str,
                       actor: str = "", session_id: str = "") -> bool:
        """从回收站恢复记忆（v5.1.1 新增）

        将 category='trash' 的记忆恢复到软删除前的原分类；
        如果找不到原分类，则恢复到 'default'。
        """
        conn = self._get_conn()
        # v5.5.5 fix: 同时获取 privacy 用于审计日志
        row = conn.execute(
            "SELECT category, metadata, privacy FROM memories WHERE id = ?",
            (entry_id,)
        ).fetchone()

        if not row or row["category"] != "trash":
            return False

        try:
            meta = json.loads(row["metadata"]) if row["metadata"] else {}
        except (json.JSONDecodeError, TypeError):
            meta = {}

        original_category = meta.pop("_original_category", "default")
        now = time.time()

        conn.execute(
            "UPDATE memories SET category = ?, metadata = ?, updated_at = ? WHERE id = ?",
            (original_category, json.dumps(meta, ensure_ascii=False), now, entry_id)
        )
        conn.commit()
        self._add_audit(
            "restore", entry_id, actor, session_id,
            row["privacy"] if row["privacy"] else "",
            details={"message": f"恢复到 {original_category}", "original_category": original_category}
        )
        return True

    def count_memories(self, category: Optional[str] = None,
                       layer: Optional[MemoryLayer] = None) -> int:
        """统计记忆数量

        v5.5.4 修复：排除软删除（category='trash'）的记忆。
        """
        conn = self._get_conn()
        query = "SELECT COUNT(*) FROM memories WHERE category != 'trash'"
        params = []

        if category:
            query += " AND category = ?"
            params.append(category)
        if layer:
            query += " AND layer = ?"
            params.append(layer.value)

        return conn.execute(query, params).fetchone()[0]

    def get_stats(self) -> dict:
        """获取统计信息

        v5.5.4 修复：排除软删除（category='trash'）的记忆，与 get_detailed_stats 保持一致，
        并新增 trash_count 字段。
        """
        conn = self._get_conn()
        total = conn.execute(
            "SELECT COUNT(*) FROM memories WHERE category != 'trash'"
        ).fetchone()[0]

        trash_count = conn.execute(
            "SELECT COUNT(*) FROM memories WHERE category = 'trash'"
        ).fetchone()[0]

        by_privacy = {}
        for row in conn.execute(
            "SELECT privacy, COUNT(*) FROM memories WHERE category != 'trash' GROUP BY privacy"
        ):
            by_privacy[row[0]] = row[1]

        by_layer = {}
        for row in conn.execute(
            "SELECT layer, COUNT(*) FROM memories WHERE category != 'trash' GROUP BY layer"
        ):
            by_layer[row[0]] = row[1]

        by_importance = {}
        for row in conn.execute(
            "SELECT importance, COUNT(*) FROM memories WHERE category != 'trash' GROUP BY importance"
        ):
            by_importance[row[0]] = row[1]

        top_categories = {}
        for row in conn.execute(
            "SELECT category, COUNT(*) as c FROM memories WHERE category != 'trash' "
            "GROUP BY category ORDER BY c DESC LIMIT 10"
        ):
            top_categories[row[0]] = row[1]

        starred_count = conn.execute(
            "SELECT COUNT(*) FROM memories WHERE starred = 1 AND category != 'trash'"
        ).fetchone()[0]

        # v5.5.6 新增：置顶记忆统计
        pinned_count = conn.execute(
            "SELECT COUNT(*) FROM memories WHERE pinned = 1 AND category != 'trash'"
        ).fetchone()[0]

        tag_counts = {}
        for row in conn.execute(
            "SELECT tags FROM memories WHERE category != 'trash'"
        ):
            if row[0]:
                try:
                    tags = json.loads(row[0])
                    for tag in tags:
                        tag_counts[tag] = tag_counts.get(tag, 0) + 1
                except (json.JSONDecodeError, TypeError):
                    pass

        top_tags = dict(sorted(tag_counts.items(), key=lambda x: x[1], reverse=True)[:10])

        db_size = self.db_path.stat().st_size if self.db_path.exists() else 0

        return {
            "total": total,
            "trash_count": trash_count,
            "db_size_bytes": db_size,
            "by_privacy": by_privacy,
            "by_layer": by_layer,
            "by_importance": by_importance,
            "top_categories": top_categories,
            "starred_count": starred_count,
            "pinned_count": pinned_count,
            "top_tags": top_tags,
        }

    def search_by_tag(self, tag: str,
                      category: Optional[str] = None,
                      layer: Optional[MemoryLayer] = None,
                      limit: int = 50,
                      offset: int = 0) -> List[MemoryEntry]:
        """按标签搜索记忆

        v5.0.4 优化：移除双重过滤（SQL LIKE 已足够精确，Python 端再过滤是冗余）。
        仅在边界情况下（tags 字段非合法 JSON）跳过该项。
        """
        conn = self._get_conn()
        # v5.3.3 安全加固：LIKE 通配符转义
        safe_tag = tag[:128] if isinstance(tag, str) else ""
        query = "SELECT * FROM memories WHERE tags LIKE ? ESCAPE '\\' AND category != 'trash'"
        params = [f'%"{_escape_like(safe_tag)}"%']

        if category:
            query += " AND category = ?"
            params.append(category)
        if layer:
            query += " AND layer = ?"
            params.append(layer.value)

        query += " ORDER BY created_at DESC LIMIT ? OFFSET ?"
        params.extend([limit, offset])

        rows = conn.execute(query, params).fetchall()
        results = []
        for row in rows:
            entry = self._row_to_entry(row)
            # 仅校验 JSON 解析结果（防御性），不再做二次标签过滤
            if tag in entry.tags:
                results.append(entry)
        return results

    def deduplicate(self,
                    category: Optional[str] = None,
                    similarity_threshold: float = 0.95,
                    dry_run: bool = True,
                    actor: str = "system",
                    session_id: str = "") -> dict:
        """记忆去重 - 检测并合并高度相似的记忆条目

        v5.0.4 新增功能。

        算法：
        - 同分类下，对 content 做标准化（去空白/小写）后比较
        - 完全相同（相似度=1.0）：保留最早一条，删除其余
        - 高度相似（>= similarity_threshold）：保留 starred 优先 / 重要性更高 / 更新时间更晚的一条
        - dry_run=True 时仅返回报告，不实际删除

        Args:
            category: 限定分类，None 表示全部分类
            similarity_threshold: 相似度阈值，默认 0.95
            dry_run: 试运行模式，只报告不删除
            actor: 操作者（用于审计日志）
            session_id: 会话 ID

        Returns:
            {
                "duplicates_found": int,    # 发现的重复组数
                "would_remove": int,        # 待删除条数（dry_run=True 时）
                "removed": int,             # 实际删除条数（dry_run=False 时）
                "details": [...],           # 详情
            }
        """
        conn = self._get_conn()
        # v5.5.5 fix: 排除回收站记忆
        query = "SELECT * FROM memories WHERE category != 'trash'"
        params = []
        if category:
            query += " AND category = ?"
            params.append(category)
        query += " ORDER BY created_at ASC"
        rows = conn.execute(query, params).fetchall()
        entries = [self._row_to_entry(r) for r in rows]

        def normalize(s: str) -> str:
            return "".join(s.lower().split())

        def similarity(a: str, b: str) -> float:
            na, nb = normalize(a), normalize(b)
            if not na or not nb:
                return 0.0
            if na == nb:
                return 1.0
            # 简单的字符级 Jaccard 相似度
            sa, sb = set(na), set(nb)
            inter = len(sa & sb)
            union = len(sa | sb)
            return inter / union if union else 0.0

        # 按分类分组
        groups: Dict[str, List[MemoryEntry]] = {}
        for e in entries:
            groups.setdefault(e.category, []).append(e)

        duplicates = []
        for cat, items in groups.items():
            n = len(items)
            for i in range(n):
                for j in range(i + 1, n):
                    sim = similarity(items[i].content, items[j].content)
                    if sim >= similarity_threshold:
                        duplicates.append((cat, sim, items[i], items[j]))

        details = []
        removed = 0
        for cat, sim, a, b in duplicates:
            # 选择保留哪一条：starred > importance > 更新时间更晚
            def keep_score(e: MemoryEntry) -> tuple:
                imp_order = {"critical": 4, "high": 3, "medium": 2, "low": 1}
                return (
                    int(e.starred),
                    imp_order.get(e.importance.value, 2),
                    e.updated_at,
                )

            keeper, loser = (a, b) if keep_score(a) >= keep_score(b) else (b, a)
            details.append({
                "category": cat,
                "similarity": round(sim, 4),
                "keeper_id": keeper.id,
                "loser_id": loser.id,
                "keeper_preview": keeper.preview,
                "loser_preview": loser.preview,
            })

            if not dry_run:
                self.delete_memory(loser.id, hard_delete=True, actor=actor, session_id=session_id)
                removed += 1

        return {
            "duplicates_found": len(duplicates),
            "would_remove": len(duplicates) if dry_run else 0,
            "removed": removed,
            "details": details,
        }

    def export_as_markdown(self,
                           output_path: str,
                           category: Optional[str] = None,
                           layer: Optional[MemoryLayer] = None,
                           starred_only: bool = False) -> Path:
        """导出记忆为 Markdown 格式

        v5.0.4 新增功能。

        Args:
            output_path: 输出文件路径
            category: 限定分类
            layer: 限定层级
            starred_only: 仅导出收藏的记忆

        Returns:
            导出文件的 Path 对象
        """
        import html as _html

        entries = self.list_memories(
            category=category,
            layer=layer,
            starred=starred_only if starred_only else None,
            limit=100000,
        )

        # v5.4.7 修复 M-8：使用 _safe_path 校验输出路径
        out = _safe_path(output_path, allowed_exts={".md"})
        out.parent.mkdir(parents=True, exist_ok=True)

        def _fmt_time(ts: float) -> str:
            return datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%Y-%m-%d %H:%M")

        lines = [
            "# MindForge 记忆导出",
            "",
            f"- 导出时间：{datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}",
            f"- 记忆总数：{len(entries)}",
            f"- 筛选条件：分类={_html.escape(str(category)) if category else '全部'}, 层级={layer.value if layer else '全部'}, 仅收藏={'是' if starred_only else '否'}",
            "",
            "---",
            "",
        ]

        # 按分类分组
        groups: Dict[str, List[MemoryEntry]] = {}
        for e in entries:
            groups.setdefault(e.category, []).append(e)

        for cat in sorted(groups.keys()):
            lines.append(f"## 📂 {_html.escape(str(cat))}")
            lines.append("")
            for e in groups[cat]:
                star = "⭐ " if e.starred else ""
                lines.append(f"### {star}{_html.escape(str(e.preview[:60]))}")
                lines.append("")
                lines.append(f"- **ID**: `{_html.escape(str(e.id))}`")
                lines.append(f"- **层级**: {e.layer.value}")
                lines.append(f"- **隐私**: {e.privacy.value}")
                lines.append(f"- **重要性**: {e.importance.value}")
                lines.append(f"- **类型**: {e.memory_type.value}")
                lines.append(f"- **标签**: {', '.join(f'#{_html.escape(str(t))}' for t in e.tags) if e.tags else '无'}")
                lines.append(f"- **创建**: {_fmt_time(e.created_at)}")
                lines.append(f"- **访问**: {e.access_count} 次")
                lines.append("")
                lines.append("**内容**：")
                lines.append("")
                lines.append(_html.escape(str(e.content)))
                lines.append("")
                lines.append("---")
                lines.append("")

        out.write_text("\n".join(lines), encoding="utf-8")
        return out

    def health_check(self) -> dict:
        """数据库健康检查（v5.0.5 新增）

        检查项目：
        - 数据库完整性（PRAGMA integrity_check）
        - 索引完整性（确认所有预期索引存在）
        - FTS 索引同步状态（检测孤立 FTS 记录）
        - 孤立审计日志（指向已不存在的 memory_id）
        - 孤立 FTS 记录（FTS 中的 rowid 在 memories 中已不存在）
        - 加密一致性（是否有标记 encrypted 但缺 ciphertext 的条目）

        Returns:
            {
                "status": "healthy" | "warning" | "critical",
                "integrity_check": str,
                "indexes": {"expected": int, "found": int, "missing": [...]},
                "fts_orphans": int,            # FTS 中有但 memories 中没有的
                "audit_orphans": int,          # 审计日志指向已删除的 memory_id
                "encrypted_inconsistent": int, # 标记加密但缺密文的
                "total_memories": int,
                "db_size_bytes": int,
                "recommendations": [...],
            }
        """
        conn = self._get_conn()

        # 1. 完整性检查
        integrity = conn.execute("PRAGMA integrity_check").fetchone()[0]

        # 2. 索引检查
        expected_indexes = {
            "idx_category", "idx_privacy", "idx_layer", "idx_importance",
            "idx_created_at", "idx_memory_type", "idx_starred",
            "idx_audit_memory", "idx_audit_timestamp",
            "idx_kg_entity", "idx_relation_from", "idx_relation_to",
        }
        actual = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='index' AND name NOT LIKE 'sqlite_%'"
        ).fetchall()
        found_indexes = {row[0] for row in actual}
        missing_indexes = expected_indexes - found_indexes

        # 3. FTS 孤立记录
        fts_orphans = conn.execute(
            "SELECT COUNT(*) FROM memory_fts WHERE rowid NOT IN (SELECT rowid FROM memories)"
        ).fetchone()[0]

        # 4. 审计日志孤立记录
        audit_orphans = conn.execute(
            "SELECT COUNT(*) FROM audit_log WHERE memory_id != '' "
            "AND memory_id NOT IN (SELECT id FROM memories)"
        ).fetchone()[0]

        # 5. 加密一致性
        encrypted_inconsistent = conn.execute(
            "SELECT COUNT(*) FROM memories WHERE encrypted = 1 "
            "AND (ciphertext IS NULL OR nonce IS NULL OR salt IS NULL)"
        ).fetchone()[0]

        # 6. 总数和大小
        total = conn.execute("SELECT COUNT(*) FROM memories").fetchone()[0]
        db_size = self.db_path.stat().st_size if self.db_path.exists() else 0

        # 生成建议
        recommendations = []
        status = "healthy"

        if integrity != "ok":
            status = "critical"
            recommendations.append("数据库完整性检查失败，建议立即从备份恢复")

        if missing_indexes:
            status = "warning" if status == "healthy" else status
            recommendations.append(f"缺失索引：{', '.join(missing_indexes)}，建议重建数据库")

        if fts_orphans > 0:
            status = "warning" if status == "healthy" else status
            recommendations.append(f"发现 {fts_orphans} 条孤立 FTS 记录，建议执行 vacuum")

        if audit_orphans > 0:
            recommendations.append(f"发现 {audit_orphans} 条孤立审计日志（不影响功能，可忽略）")

        if encrypted_inconsistent > 0:
            status = "critical"
            recommendations.append(f"{encrypted_inconsistent} 条加密记忆缺密文，数据可能损坏")

        if not recommendations:
            recommendations.append("一切正常，无需操作")

        return {
            "status": status,
            "integrity_check": integrity,
            "indexes": {
                "expected": len(expected_indexes),
                "found": len(found_indexes & expected_indexes),
                "missing": sorted(missing_indexes),
            },
            "fts_orphans": fts_orphans,
            "audit_orphans": audit_orphans,
            "encrypted_inconsistent": encrypted_inconsistent,
            "total_memories": total,
            "db_size_bytes": db_size,
            "recommendations": recommendations,
        }

    def summarize(self,
                  category: Optional[str] = None,
                  group_by: str = "category") -> dict:
        """生成记忆摘要（v5.0.5 新增）

        Args:
            category: 限定分类，None 表示全部
            group_by: 分组维度，支持 'category' | 'layer' | 'importance' | 'privacy'

        Returns:
            {
                "total": int,
                "grouped": {group_key: {"count": int, "latest": str, "oldest": str, "samples": [...]}},
                "recent_activity": {"last_7d": int, "last_30d": int},
                "top_tags": [...],
            }
        """
        conn = self._get_conn()
        now = time.time()

        valid_groups = {"category", "layer", "importance", "privacy"}
        if group_by not in valid_groups:
            group_by = "category"

        where = " WHERE 1=1"
        params = []
        if category:
            where += " AND category = ?"
            params.append(category)

        total = conn.execute(f"SELECT COUNT(*) FROM memories{where}", params).fetchone()[0]

        # 分组统计
        grouped = {}
        rows = conn.execute(
            f"SELECT {group_by}, COUNT(*), MAX(created_at), MIN(created_at) "
            f"FROM memories{where} GROUP BY {group_by} ORDER BY COUNT(*) DESC",
            params
        ).fetchall()
        for row in rows:
            key = row[0] or "unknown"
            count = row[1]
            latest_ts = row[2]
            oldest_ts = row[3]
            # 取该组前 3 条预览
            samples = conn.execute(
                f"SELECT content FROM memories{where} AND {group_by} = ? "
                f"ORDER BY created_at DESC LIMIT 3",
                params + [row[0]]
            ).fetchall()
            grouped[key] = {
                "count": count,
                "latest": datetime.fromtimestamp(latest_ts, tz=timezone.utc).strftime("%Y-%m-%d") if latest_ts else "",
                "oldest": datetime.fromtimestamp(oldest_ts, tz=timezone.utc).strftime("%Y-%m-%d") if oldest_ts else "",
                "samples": [r[0][:80] for r in samples],
            }

        # 近期活动
        last_7d = conn.execute(
            f"SELECT COUNT(*) FROM memories{where} AND created_at >= ?",
            params + [now - 7 * 86400]
        ).fetchone()[0]
        last_30d = conn.execute(
            f"SELECT COUNT(*) FROM memories{where} AND created_at >= ?",
            params + [now - 30 * 86400]
        ).fetchone()[0]

        # 热门标签
        tag_counts: Dict[str, int] = {}
        for row in conn.execute(f"SELECT tags FROM memories{where}", params):
            if row[0]:
                try:
                    for t in json.loads(row[0]):
                        tag_counts[t] = tag_counts.get(t, 0) + 1
                except (json.JSONDecodeError, TypeError):
                    pass
        top_tags = sorted(tag_counts.items(), key=lambda x: x[1], reverse=True)[:10]

        return {
            "total": total,
            "grouped": grouped,
            "recent_activity": {
                "last_7d": last_7d,
                "last_30d": last_30d,
            },
            "top_tags": top_tags,
        }

    def agent_stats(self, agent_id: Optional[str] = None) -> Dict[str, Any]:
        """Agent 记忆统计（v5.2.2 新增）

        统计按 Agent 来源分组的记忆数据，支持查询特定 Agent。

        Args:
            agent_id: 指定 Agent ID（None 表示统计全部 Agent）

        Returns:
            {
                "total_agents": 总 Agent 数,
                "by_agent": {agent_id: {count, last_active, top_categories}},
                "agent_detail": 指定 Agent 的详情（如果提供了 agent_id）
            }
        """
        conn = self._get_conn()

        if agent_id:
            # 查询特定 Agent
            total = conn.execute(
                "SELECT COUNT(*) FROM memories WHERE source_agent = ?",
                (agent_id,)
            ).fetchone()[0]

            categories = conn.execute(
                "SELECT category, COUNT(*) as cnt FROM memories WHERE source_agent = ? GROUP BY category ORDER BY cnt DESC LIMIT 10",
                (agent_id,)
            ).fetchall()

            last_active = conn.execute(
                "SELECT MAX(created_at) FROM memories WHERE source_agent = ?",
                (agent_id,)
            ).fetchone()[0] or 0

            layers = conn.execute(
                "SELECT layer, COUNT(*) as cnt FROM memories WHERE source_agent = ? GROUP BY layer",
                (agent_id,)
            ).fetchall()

            return {
                "agent_id": agent_id,
                "total_memories": total,
                "last_active": last_active,
                "by_category": {r[0]: r[1] for r in categories},
                "by_layer": {r[0]: r[1] for r in layers},
            }

        # 统计所有 Agent
        agents = conn.execute(
            "SELECT source_agent, COUNT(*) as cnt, MAX(created_at) as last_active "
            "FROM memories WHERE source_agent != '' GROUP BY source_agent ORDER BY cnt DESC"
        ).fetchall()

        by_agent = {}
        for row in agents:
            agent = row[0]
            count = row[1]
            last_active = row[2]

            # 获取每个 Agent 的 top 分类
            top_cats = conn.execute(
                "SELECT category, COUNT(*) as cnt FROM memories WHERE source_agent = ? "
                "GROUP BY category ORDER BY cnt DESC LIMIT 5",
                (agent,)
            ).fetchall()

            by_agent[agent] = {
                "count": count,
                "last_active": last_active,
                "top_categories": [r[0] for r in top_cats[:5]],
            }

        return {
            "total_agents": len(by_agent),
            "by_agent": by_agent,
        }

    def list_by_agent(self,
                      agent_id: str,
                      limit: int = 100,
                      offset: int = 0) -> List[MemoryEntry]:
        """列出特定 Agent 的记忆（v5.2.2 新增）"""
        conn = self._get_conn()
        # v5.5.5 fix: 排除回收站记忆
        rows = conn.execute(
            "SELECT * FROM memories WHERE source_agent = ? AND category != 'trash' ORDER BY created_at DESC LIMIT ? OFFSET ?",
            (agent_id, limit, offset)
        ).fetchall()
        return [self._row_to_entry(r) for r in rows]

    def evolve_memories(self,
                        dry_run: bool = False,
                        actor: str = "system",
                        session_id: str = "evolve") -> Dict[str, Any]:
        """记忆演化 - 基于艾宾浩斯遗忘曲线自动升级记忆层级（v5.2.2 新增）

        规则：
        - 短期记忆创建超过 24 小时且被访问过 → 升级为长期记忆
        - 长期记忆创建超过 7 天且收藏/重要 → 升级为永久记忆
        - 超过 30 天未访问的短期记忆 → 标记为待清理

        Args:
            dry_run: 仅统计不执行
            actor: 操作者
            session_id: 会话 ID

        Returns:
            演化统计结果
        """
        conn = self._get_conn()
        now = time.time()
        day_seconds = 86400

        short_to_long_days = 1
        long_to_perm_days = 7
        stale_days = 30

        # 重新计算
        short_to_long = conn.execute(
            "SELECT COUNT(*) FROM memories "
            "WHERE layer = ? AND category != 'trash' "
            "AND (julianday('now') - julianday(created_at, 'unixepoch')) >= 1 "
            "AND access_count > 0",
            (MemoryLayer.SHORT_TERM.value,)
        ).fetchone()[0]

        long_to_perm = conn.execute(
            "SELECT COUNT(*) FROM memories "
            "WHERE layer = ? AND category != 'trash' "
            "AND (julianday('now') - julianday(created_at, 'unixepoch')) >= 7 "
            "AND (starred = 1 OR importance IN ('HIGH', 'CRITICAL'))",
            (MemoryLayer.LONG_TERM.value,)
        ).fetchone()[0]

        stale_short = conn.execute(
            "SELECT COUNT(*) FROM memories "
            "WHERE layer = ? AND category != 'trash' "
            "AND (julianday('now') - julianday(created_at, 'unixepoch')) >= 30 "
            "AND access_count = 0",
            (MemoryLayer.SHORT_TERM.value,)
        ).fetchone()[0]

        result = {
            "short_to_long": short_to_long,
            "long_to_permanent": long_to_perm,
            "stale_short_term": stale_short,
            "total_evolvable": short_to_long + long_to_perm,
            "executed": not dry_run,
        }

        if dry_run:
            return result

        # 执行升级
        upgraded_short = 0
        rows = conn.execute(
            "SELECT id, privacy FROM memories "
            "WHERE layer = ? AND category != 'trash' "
            "AND (julianday('now') - julianday(created_at, 'unixepoch')) >= 1 "
            "AND access_count > 0 LIMIT 200",
            (MemoryLayer.SHORT_TERM.value,)
        ).fetchall()
        for row in rows:
            conn.execute(
                "UPDATE memories SET layer = ?, updated_at = ? WHERE id = ?",
                (MemoryLayer.LONG_TERM.value, now, row[0])
            )
            self._add_audit(
                "evolve", row[0], actor, session_id,
                row[1] if row[1] else "",
                {"transition": "short_term→long_term"}
            )
            upgraded_short += 1

        upgraded_long = 0
        rows = conn.execute(
            "SELECT id, privacy FROM memories "
            "WHERE layer = ? AND category != 'trash' "
            "AND (julianday('now') - julianday(created_at, 'unixepoch')) >= 7 "
            "AND (starred = 1 OR importance IN ('HIGH', 'CRITICAL')) LIMIT 100",
            (MemoryLayer.LONG_TERM.value,)
        ).fetchall()
        for row in rows:
            conn.execute(
                "UPDATE memories SET layer = ?, updated_at = ? WHERE id = ?",
                (MemoryLayer.PERMANENT.value, now, row[0])
            )
            self._add_audit(
                "evolve", row[0], actor, session_id,
                row[1] if row[1] else "",
                {"transition": "long_term→permanent"}
            )
            upgraded_long += 1

        conn.commit()
        result["upgraded_to_long"] = upgraded_short
        result["upgraded_to_permanent"] = upgraded_long
        return result

    def transfer_agent_memories(self,
                                from_agent: str,
                                to_agent: str,
                                category: Optional[str] = None,
                                actor: str = "system",
                                session_id: str = "transfer") -> Dict[str, Any]:
        """Agent 记忆迁移 - 将一个 Agent 的记忆转移给另一个（v5.2.2 新增）

        Args:
            from_agent: 源 Agent ID
            to_agent: 目标 Agent ID
            category: 可选，仅迁移指定分类
            actor: 操作者
            session_id: 会话 ID

        Returns:
            迁移统计
        """
        conn = self._get_conn()
        now = time.time()

        query = "SELECT id, privacy FROM memories WHERE source_agent = ? AND category != 'trash'"
        params = [from_agent]

        if category:
            query += " AND category = ?"
            params.append(category)

        rows = conn.execute(query, params).fetchall()
        total = len(rows)

        for row in rows:
            conn.execute(
                "UPDATE memories SET source_agent = ?, updated_at = ? WHERE id = ?",
                (to_agent, now, row["id"])
            )
            self._add_audit(
                "agent_transfer", row["id"], actor, session_id,
                row["privacy"] if row["privacy"] else "",
                {"from_agent": from_agent, "to_agent": to_agent}
            )

        conn.commit()
        return {
            "from_agent": from_agent,
            "to_agent": to_agent,
            "transferred": total,
            "category_filter": category,
        }

    def clean_agent_memories(self,
                             agent_id: str,
                             older_than_days: int = 90,
                             max_importance: Optional[str] = None,
                             dry_run: bool = False,
                             actor: str = "system",
                             session_id: str = "clean") -> Dict[str, Any]:
        """清理 Agent 的旧记忆（v5.2.2 新增）

        清理指定 Agent 创建的、超过指定天数、重要度低于等于指定级别的记忆，移入回收站。

        Args:
            agent_id: Agent ID
            older_than_days: 清理超过多少天的记忆
            max_importance: 最高清理的重要级别（LOW/MEDIUM/HIGH/CRITICAL），None 表示清理所有
            dry_run: 仅统计不执行
            actor: 操作者
            session_id: 会话 ID

        Returns:
            清理统计
        """
        conn = self._get_conn()
        now = time.time()

        importance_order = ["LOW", "MEDIUM", "HIGH", "CRITICAL"]
        clean_importances = []

        if max_importance:
            try:
                max_imp = Importance.from_string(max_importance)
                max_idx = importance_order.index(max_imp.value)
                clean_importances = importance_order[:max_idx + 1]
            except (ValueError, KeyError):
                clean_importances = importance_order
        else:
            clean_importances = importance_order

        query = (
            "SELECT id, content, privacy FROM memories "
            "WHERE source_agent = ? AND category != 'trash' "
            "AND (julianday('now') - julianday(created_at, 'unixepoch')) >= ? "
            "AND importance IN ({}) AND starred = 0"
        ).format(",".join(["?"] * len(clean_importances)))
        params = [agent_id, older_than_days] + clean_importances

        rows = conn.execute(query, params).fetchall()
        total = len(rows)

        result = {
            "agent_id": agent_id,
            "older_than_days": older_than_days,
            "max_importance": max_importance,
            "clean_importances": clean_importances,
            "to_clean": total,
            "cleaned": 0,
            "executed": not dry_run,
        }

        if dry_run:
            return result

        for row in rows:
            conn.execute(
                "UPDATE memories SET category = 'trash', updated_at = ? WHERE id = ?",
                (now, row[0])
            )
            self._add_audit("agent_clean", row[0], actor, session_id,
                             row[2] if row[2] else "")

        conn.commit()
        result["cleaned"] = total
        return result

    def quality_score(self, memory_id: str) -> Optional[Dict[str, Any]]:
        """记忆质量评分（v5.2.2 新增）"""
        entry = self.get_memory(memory_id)
        if not entry:
            return None

        import time
        now = time.time()

        scores = {}
        total = 0.0

        # 1. 内容长度评分（0-20分）
        content_len = len(entry.content)
        if 50 <= content_len <= 500:
            scores["content_length"] = 20
        elif content_len < 50:
            scores["content_length"] = max(5, content_len // 10)
        elif content_len <= 1000:
            scores["content_length"] = 15
        else:
            scores["content_length"] = 10
        total += scores["content_length"]

        # 2. 访问频率评分（0-25分）
        access_count = entry.access_count
        if access_count >= 10:
            scores["access_frequency"] = 25
        elif access_count >= 5:
            scores["access_frequency"] = 20
        elif access_count >= 2:
            scores["access_frequency"] = 15
        elif access_count >= 1:
            scores["access_frequency"] = 10
        else:
            scores["access_frequency"] = 5
        total += scores["access_frequency"]

        # 3. 收藏状态（0-15分）
        scores["starred"] = 15 if entry.starred else 0
        total += scores["starred"]

        # 4. 重要性评分（0-20分）
        importance_scores = {
            "CRITICAL": 20,
            "HIGH": 15,
            "MEDIUM": 10,
            "LOW": 5,
        }
        scores["importance"] = importance_scores.get(entry.importance.value, 10)
        total += scores["importance"]

        # 5. 标签丰富度（0-10分）
        tag_count = len(entry.tags)
        if tag_count >= 5:
            scores["tag_richness"] = 10
        elif tag_count >= 3:
            scores["tag_richness"] = 7
        elif tag_count >= 1:
            scores["tag_richness"] = 5
        else:
            scores["tag_richness"] = 0
        total += scores["tag_richness"]

        # 6. 时间衰减（0-10分）
        age_days = (now - entry.created_at) / 86400
        if age_days <= 1:
            scores["freshness"] = 10
        elif age_days <= 7:
            scores["freshness"] = 8
        elif age_days <= 30:
            scores["freshness"] = 6
        elif age_days <= 90:
            scores["freshness"] = 4
        else:
            scores["freshness"] = 2
        total += scores["freshness"]

        # 7. 记忆层级加分（0-5分）
        layer_scores = {
            "permanent": 5,
            "long_term": 3,
            "short_term": 1,
            "sensory": 0,
        }
        scores["layer_bonus"] = layer_scores.get(entry.layer.value, 0)
        total += scores["layer_bonus"]

        return {
            "memory_id": memory_id,
            "total_score": round(total, 1),
            "max_score": 100,
            "percentage": round(total, 1),
            "grade": self._score_to_grade(total),
            "breakdown": scores,
        }

    def _score_to_grade(self, score: float) -> str:
        """分数转等级"""
        if score >= 85:
            return "优秀"
        elif score >= 70:
            return "良好"
        elif score >= 55:
            return "中等"
        elif score >= 40:
            return "及格"
        else:
            return "需改进"

    # ===== Agent 记忆增强（v5.2.9 新增）=====

    def rank_agents(self,
                    by: str = "count",
                    limit: int = 20) -> List[Dict[str, Any]]:
        """Agent 排行榜（v5.2.9 新增）

        Args:
            by: 排序维度（count / last_active / avg_importance / starred）
            limit: 返回数量

        Returns:
            Agent 排名列表
        """
        conn = self._get_conn()
        # v5.2.9 安全加固：by 白名单枚举校验
        _ALLOWED = {"count", "last_active", "avg_importance", "starred"}
        if by not in _ALLOWED:
            by = "count"

        imp_weights = {"CRITICAL": 4, "HIGH": 3, "MEDIUM": 2, "LOW": 1, "TRIVIAL": 0}

        rows = conn.execute(
            "SELECT source_agent, COUNT(*) as cnt, MAX(created_at) as last_active, "
            "SUM(starred) as starred_cnt FROM memories "
            "WHERE source_agent != '' GROUP BY source_agent"
        ).fetchall()

        result = []
        for r in rows:
            agent = r[0]
            count = r[1]
            last_active = r[2] or 0
            starred_cnt = r[3] or 0
            avg_imp = 0.0
            if count > 0:
                imps = conn.execute(
                    "SELECT importance FROM memories WHERE source_agent = ?", (agent,)
                ).fetchall()
                total_imp = sum(imp_weights.get(str(i[0]), 2) for i in imps)
                avg_imp = total_imp / count

            result.append({
                "agent_id": agent,
                "count": count,
                "last_active": last_active,
                "starred_count": starred_cnt,
                "avg_importance": round(avg_imp, 2),
            })

        sort_map = {
            "count": lambda x: -x["count"],
            "last_active": lambda x: -x["last_active"],
            "avg_importance": lambda x: -x["avg_importance"],
            "starred": lambda x: -x["starred_count"],
        }
        result.sort(key=sort_map[by])
        return result[:limit]

    def forget_agent_memories(self,
                              agent_id: str,
                              min_quality_score: int = 30,
                              older_than_days: int = 30,
                              dry_run: bool = False,
                              actor: str = "cli",
                              session_id: str = "forget") -> Dict[str, Any]:
        """遗忘 Agent 低质量旧记忆（v5.2.9 新增）

        Args:
            agent_id: 目标 Agent ID
            min_quality_score: 低于此质量分数的记忆才会被清理
            older_than_days: 只清理超过此天数未更新的记忆
            dry_run: 仅预览
            actor: 操作者
            session_id: 会话 ID

        Returns:
            {evaluated, selected, cleaned}
        """
        conn = self._get_conn()
        # v5.2.9 安全加固：数值边界
        min_quality_score = max(0, min(100, int(min_quality_score)))
        older_than_days = max(0, int(older_than_days))
        if not agent_id or not isinstance(agent_id, str) or len(agent_id) > 128:
            return {"evaluated": 0, "selected": 0, "cleaned": 0,
                    "error": "无效 agent_id"}

        now = time.time()
        cutoff = now - older_than_days * 86400

        mems = conn.execute(
            "SELECT id, source_agent, updated_at, privacy FROM memories "
            "WHERE source_agent = ? AND updated_at < ? AND category != 'trash'",
            (agent_id[:128], cutoff)
        ).fetchall()

        selected_ids = []
        # v5.5.5 fix: 记录 privacy 用于审计日志
        privacy_map = {}
        for r in mems:
            mid = r[0]
            privacy_map[mid] = r[3] if r[3] else ""
            qs = self.quality_score(mid)
            if qs and qs["total_score"] < min_quality_score:
                selected_ids.append(mid)

        if not dry_run and selected_ids:
            placeholders = ",".join("?" * len(selected_ids))
            conn.execute(
                f"UPDATE memories SET category='trash', updated_at=? "
                f"WHERE id IN ({placeholders})",
                [now] + selected_ids,
            )
            for mid in selected_ids:
                self._add_audit("agent_forget", mid, actor, session_id,
                                 privacy_map.get(mid, ""))
            conn.commit()

        return {
            "agent_id": agent_id,
            "evaluated": len(mems),
            "selected": len(selected_ids),
            "cleaned": 0 if dry_run else len(selected_ids),
            "selected_ids": selected_ids[:20],
        }

    # ===== AI 短剧增强（v5.2.9 新增）=====

    def list_lines_by_scene(self,
                            scene_id: str,
                            limit: int = 500,
                            offset: int = 0) -> List[DramaLine]:
        """按场次列出所有台词（v5.2.9 新增）"""
        conn = self._get_conn()
        # 安全：ID 长度限制
        sid = scene_id[:64] if isinstance(scene_id, str) else ""
        limit = max(1, min(10000, int(limit)))
        offset = max(0, int(offset))
        rows = conn.execute(
            "SELECT * FROM drama_lines WHERE scene_id = ? "
            "ORDER BY COALESCE(episode, 0), COALESCE(created_at, 0) LIMIT ? OFFSET ?",
            (sid, limit, offset),
        ).fetchall()
        return [self._row_to_line(r) for r in rows]

    def list_lines_by_character(self,
                                character_id: str,
                                drama_id: Optional[str] = None,
                                limit: int = 500,
                                offset: int = 0) -> List[DramaLine]:
        """按角色列出所有台词（v5.2.9 新增）"""
        conn = self._get_conn()
        cid = character_id[:64] if isinstance(character_id, str) else ""
        limit = max(1, min(10000, int(limit)))
        offset = max(0, int(offset))

        if drama_id:
            did = drama_id[:64] if isinstance(drama_id, str) else ""
            rows = conn.execute(
                "SELECT * FROM drama_lines WHERE character_id = ? AND drama_id = ? "
                "ORDER BY COALESCE(episode, 0), COALESCE(created_at, 0) LIMIT ? OFFSET ?",
                (cid, did, limit, offset),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM drama_lines WHERE character_id = ? "
                "ORDER BY drama_id, COALESCE(episode, 0), COALESCE(created_at, 0) LIMIT ? OFFSET ?",
                (cid, limit, offset),
            ).fetchall()
        return [self._row_to_line(r) for r in rows]

    def top_rated_dramas(self,
                         genre: Optional[str] = None,
                         min_rating: float = 0.0,
                         limit: int = 50) -> List[DramaSeries]:
        """高分短剧排行榜（v5.2.9 新增，别名 drama-stars）"""
        conn = self._get_conn()
        limit = max(1, min(1000, int(limit)))
        min_rating = max(0.0, min(10.0, float(min_rating)))

        sql = "SELECT * FROM drama_series WHERE rating >= ?"
        params: List[Any] = [min_rating]

        # 枚举白名单校验 genre
        _ALLOWED_GENRES = {g.value for g in DramaGenre}
        if genre:
            g = genre.strip().upper() if isinstance(genre, str) else ""
            if g in _ALLOWED_GENRES:
                sql += " AND genre = ?"
                params.append(g)

        sql += " ORDER BY rating DESC, current_episode DESC LIMIT ?"
        params.append(limit)

        rows = conn.execute(sql, params).fetchall()
        return [self._row_to_drama(r) for r in rows]

    def import_dramas_from_json(self,
                                input_path: str,
                                skip_existing: bool = True) -> Dict[str, int]:
        """从 JSON 文件批量导入短剧（v5.2.9 新增）

        导入结构与 export_dramas 相同，包含：dramas[], scenes[], characters[], lines[]。

        Args:
            input_path: JSON 文件路径（已在核心层外部做路径校验更好，此处也再校验一次）
            skip_existing: 跳过已存在的短剧（按 title 匹配）

        Returns:
            {dramas, scenes, characters, lines, skipped, failed}
        """
        # v5.2.9 安全加固：路径二次校验
        safe_path = _safe_path(input_path, must_exist=True,
                               allowed_exts={".json"}, max_size=500 * 1024 * 1024)

        import json
        with open(str(safe_path), "r", encoding="utf-8") as f:
            data = json.load(f)

        stats = {"dramas": 0, "scenes": 0, "characters": 0, "lines": 0,
                 "skipped": 0, "failed": 0}

        conn = self._get_conn()
        dramas_raw = data.get("dramas", [])
        if len(dramas_raw) > 1000:
            dramas_raw = dramas_raw[:1000]

        for drama_data in dramas_raw:
            try:
                title = str(drama_data.get("title", ""))[:512]
                if not title:
                    stats["failed"] += 1
                    continue

                # skip_existing
                if skip_existing:
                    exists = conn.execute(
                        "SELECT id FROM drama_series WHERE title = ?", (title,)
                    ).fetchone()
                    if exists:
                        stats["skipped"] += 1
                        continue

                # 创建短剧
                genre_str = str(drama_data.get("genre", "OTHER"))[:64]
                try:
                    genre_enum = DramaGenre(genre_str)
                except ValueError:
                    genre_enum = DramaGenre.OTHER

                status_str = str(drama_data.get("status", "PLANNED"))[:64]
                try:
                    status_enum = DramaStatus(status_str)
                except ValueError:
                    status_enum = DramaStatus.PLANNED

                new_drama = self.add_drama(
                    title=title,
                    description=str(drama_data.get("description", ""))[:5000],
                    genre=genre_enum,
                    total_episodes=max(0, int(drama_data.get("total_episodes", 0) or 0)),
                    current_episode=max(0, int(drama_data.get("current_episode", 0) or 0)),
                    rating=max(0.0, min(10.0, float(drama_data.get("rating", 0) or 0))),
                    platform=str(drama_data.get("platform", ""))[:128],
                    cover_url=str(drama_data.get("cover_url", ""))[:512],
                    tags=list(drama_data.get("tags", []) or [])[:64],
                    actors=list(drama_data.get("actors", []) or [])[:128],
                    director=str(drama_data.get("director", ""))[:128],
                    status=status_enum,
                )
                stats["dramas"] += 1
                new_did = new_drama.id

                # 导入场次
                for scene_data in drama_data.get("scenes", [])[:1000]:
                    try:
                        s = self.add_scene(
                            drama_id=new_did,
                            title=str(scene_data.get("title", ""))[:512],
                            description=str(scene_data.get("description", ""))[:5000],
                            episode=max(0, int(scene_data.get("episode", 0) or 0)),
                            location=str(scene_data.get("location", ""))[:256],
                            time_of_day=str(scene_data.get("time_of_day", ""))[:64],
                            notes=str(scene_data.get("notes", ""))[:5000],
                        )
                        stats["scenes"] += 1
                    except Exception:
                        stats["failed"] += 1

                # 导入角色
                for char_data in drama_data.get("characters", [])[:500]:
                    try:
                        c = self.add_character(
                            drama_id=new_did,
                            name=str(char_data.get("name", ""))[:128],
                            actor_name=str(char_data.get("actor_name", ""))[:128],
                            description=str(char_data.get("description", ""))[:5000],
                            role_type=str(char_data.get("role_type", "supporting"))[:64],
                            tags=list(char_data.get("tags", []) or [])[:64],
                        )
                        stats["characters"] += 1
                    except Exception:
                        stats["failed"] += 1

                # 导入台词
                for line_data in drama_data.get("lines", [])[:50000]:
                    try:
                        l = self.add_line(
                            drama_id=new_did,
                            scene_id=str(line_data.get("scene_id", ""))[:64],
                            character_id=str(line_data.get("character_id", ""))[:64],
                            line_text=str(line_data.get("line_text", ""))[:5000],
                            character_name=str(line_data.get("character_name", ""))[:128],
                            context=str(line_data.get("context", ""))[:5000],
                            episode=max(0, int(line_data.get("episode", 0) or 0)),
                            timestamp=str(line_data.get("timestamp", ""))[:32],
                            is_classic=bool(line_data.get("is_classic", False)),
                            tags=list(line_data.get("tags", []) or [])[:64],
                        )
                        stats["lines"] += 1
                    except Exception:
                        stats["failed"] += 1

            except Exception:
                stats["failed"] += 1

        return stats

    # ===== Agent 记忆增强（v5.3.0 新增）=====

    def agent_profile(self, agent_id: str) -> Dict[str, Any]:
        """Agent 记忆画像（v5.3.0 新增）

        聚合分析指定 Agent 的记忆全景：总量/质量分布/层级分布/分类分布/
        活跃时间线/知识领域（标签 Top-N）/收藏与置顶统计。

        Args:
            agent_id: Agent ID

        Returns:
            画像字典
        """
        # v5.3.0 安全加固：ID 长度限制
        if not agent_id or not isinstance(agent_id, str) or len(agent_id) > 128:
            return {"error": "无效 agent_id"}

        conn = self._get_conn()
        aid = agent_id[:128]

        total = conn.execute(
            "SELECT COUNT(*) FROM memories WHERE source_agent = ?", (aid,)
        ).fetchone()[0]

        if total == 0:
            return {"agent_id": aid, "total_memories": 0, "message": "该 Agent 暂无记忆"}

        # 层级分布
        layers = conn.execute(
            "SELECT layer, COUNT(*) as cnt FROM memories WHERE source_agent = ? GROUP BY layer",
            (aid,)
        ).fetchall()

        # 分类分布 Top-10
        cats = conn.execute(
            "SELECT category, COUNT(*) as cnt FROM memories WHERE source_agent = ? "
            "GROUP BY category ORDER BY cnt DESC LIMIT 10",
            (aid,)
        ).fetchall()

        # 重要度分布
        imps = conn.execute(
            "SELECT importance, COUNT(*) as cnt FROM memories WHERE source_agent = ? GROUP BY importance",
            (aid,)
        ).fetchall()

        # 收藏/置顶统计
        starred = conn.execute(
            "SELECT COUNT(*) FROM memories WHERE source_agent = ? AND starred = 1", (aid,)
        ).fetchone()[0]
        pinned = conn.execute(
            "SELECT COUNT(*) FROM memories WHERE source_agent = ? AND pinned = 1", (aid,)
        ).fetchone()[0]

        # 标签 Top-10（知识领域）
        tag_rows = conn.execute(
            "SELECT tags FROM memories WHERE source_agent = ? AND tags != ''", (aid,)
        ).fetchall()
        tag_counter: Dict[str, int] = {}
        for r in tag_rows:
            for t in self._safe_json_loads(r[0], []):
                t = str(t)[:64]
                tag_counter[t] = tag_counter.get(t, 0) + 1
        top_tags = sorted(tag_counter.items(), key=lambda x: -x[1])[:10]

        # 活跃时间线（按天聚合，最近 30 天）
        now = time.time()
        cutoff = now - 30 * 86400
        timeline = conn.execute(
            "SELECT created_at FROM memories WHERE source_agent = ? AND created_at >= ?",
            (aid, cutoff)
        ).fetchall()
        daily: Dict[str, int] = {}
        for r in timeline:
            ts = r[0] or 0
            day = time.strftime("%Y-%m-%d", time.localtime(ts))
            daily[day] = daily.get(day, 0) + 1

        # 质量分布（采样最多 200 条，避免大 Agent 性能问题）
        sample_ids = conn.execute(
            "SELECT id FROM memories WHERE source_agent = ? LIMIT 200", (aid,)
        ).fetchall()
        quality_dist = {"优秀": 0, "良好": 0, "中等": 0, "及格": 0, "需改进": 0}
        for r in sample_ids:
            qs = self.quality_score(r[0])
            if qs:
                quality_dist[qs["grade"]] = quality_dist.get(qs["grade"], 0) + 1

        last_active = conn.execute(
            "SELECT MAX(created_at) FROM memories WHERE source_agent = ?", (aid,)
        ).fetchone()[0] or 0
        first_active = conn.execute(
            "SELECT MIN(created_at) FROM memories WHERE source_agent = ?", (aid,)
        ).fetchone()[0] or 0

        return {
            "agent_id": aid,
            "total_memories": total,
            "first_active": first_active,
            "last_active": last_active,
            "by_layer": {r[0]: r[1] for r in layers},
            "by_category": {r[0]: r[1] for r in cats},
            "by_importance": {r[0]: r[1] for r in imps},
            "starred_count": starred,
            "pinned_count": pinned,
            "top_tags": [{"tag": t, "count": c} for t, c in top_tags],
            "activity_timeline_30d": daily,
            "quality_distribution_sample": quality_dist,
            "quality_sample_size": len(sample_ids),
        }

    def merge_agent_memories(self,
                             from_agent: str,
                             to_agent: str,
                             dedup: str = "exact",
                             dry_run: bool = False,
                             actor: str = "cli",
                             session_id: str = "merge") -> Dict[str, Any]:
        """合并两个 Agent 的记忆（v5.3.0 新增）

        将 from_agent 的记忆迁移到 to_agent，支持去重：
        - exact: 内容完全相同则跳过
        - none: 不去重，全部迁移

        Args:
            from_agent: 源 Agent ID
            to_agent: 目标 Agent ID
            dedup: 去重模式（exact / none）
            dry_run: 仅预览
            actor: 操作者
            session_id: 会话 ID

        Returns:
            {evaluated, migrated, skipped_duplicates, failed}
        """
        # v5.3.0 安全加固
        if not from_agent or not isinstance(from_agent, str) or len(from_agent) > 128:
            return {"evaluated": 0, "migrated": 0, "error": "无效 from_agent"}
        if not to_agent or not isinstance(to_agent, str) or len(to_agent) > 128:
            return {"evaluated": 0, "migrated": 0, "error": "无效 to_agent"}
        if from_agent == to_agent:
            return {"evaluated": 0, "migrated": 0, "error": "源和目标 Agent 相同"}

        _ALLOWED_DEDUP = {"exact", "none"}
        if dedup not in _ALLOWED_DEDUP:
            dedup = "exact"

        conn = self._get_conn()
        fa = from_agent[:128]
        ta = to_agent[:128]

        src_rows = conn.execute(
            "SELECT id, content, privacy FROM memories WHERE source_agent = ? AND category != 'trash'", (fa,)
        ).fetchall()

        # 如果去重，预取目标 Agent 已有内容集合
        existing_contents: set = set()
        if dedup == "exact":
            tgt_rows = conn.execute(
                "SELECT content FROM memories WHERE source_agent = ? AND category != 'trash'", (ta,)
            ).fetchall()
            for r in tgt_rows:
                existing_contents.add(r[0])

        now = time.time()
        migrated = 0
        skipped = 0
        failed = 0

        for r in src_rows:
            mid = r[0]
            content = r[1]
            if dedup == "exact" and content in existing_contents:
                skipped += 1
                continue
            if not dry_run:
                try:
                    conn.execute(
                        "UPDATE memories SET source_agent = ?, updated_at = ? WHERE id = ?",
                        (ta, now, mid)
                    )
                    self._add_audit("agent_merge", mid, actor, session_id,
                                    r["privacy"] if r["privacy"] else "",
                                    {"from_agent": fa, "to_agent": ta})
                except Exception:
                    failed += 1
                    continue
            migrated += 1

        if not dry_run and migrated > 0:
            conn.commit()

        return {
            "from_agent": fa,
            "to_agent": ta,
            "dedup": dedup,
            "evaluated": len(src_rows),
            "migrated": migrated,
            "skipped_duplicates": skipped,
            "failed": failed,
        }

    def export_agent_memories(self,
                              agent_id: str,
                              output_path: str,
                              include_audit: bool = False) -> Dict[str, Any]:
        """导出 Agent 全部记忆为独立 JSON 包（v5.3.0 新增）

        Args:
            agent_id: Agent ID
            output_path: 输出 JSON 路径
            include_audit: 是否包含审计日志

        Returns:
            {agent_id, total, file_path}
        """
        # v5.3.0 安全加固
        if not agent_id or not isinstance(agent_id, str) or len(agent_id) > 128:
            return {"error": "无效 agent_id"}

        # 路径校验
        path = _safe_path(output_path, allowed_exts={".json"})

        aid = agent_id[:128]
        entries = self.list_by_agent(agent_id=aid, limit=100000, offset=0)

        export_data: Dict[str, Any] = {
            "version": __version__,
            "export_type": "agent_memories",
            "agent_id": aid,
            "export_time": "",
            "total": len(entries),
            "memories": [],
        }
        from datetime import datetime
        export_data["export_time"] = datetime.now().isoformat()

        for e in entries:
            d = e.to_dict() if hasattr(e, "to_dict") else vars(e)
            export_data["memories"].append(d)

        if include_audit:
            conn = self._get_conn()
            audits = conn.execute(
                "SELECT * FROM audit_log WHERE memory_id IN "
                "(SELECT id FROM memories WHERE source_agent = ?) LIMIT 10000",
                (aid,)
            ).fetchall()
            export_data["audit_logs"] = [dict(r) for r in audits]

        path.parent.mkdir(parents=True, exist_ok=True)
        import json
        with open(str(path), "w", encoding="utf-8") as f:
            json.dump(export_data, f, ensure_ascii=False, indent=2, default=str)

        # v5.3.0 安全加固：权限收紧
        try:
            import stat
            import os
            os.chmod(str(path), stat.S_IRUSR | stat.S_IWUSR | stat.S_IRGRP | stat.S_IROTH)
        except (OSError, ImportError):
            pass

        return {
            "agent_id": aid,
            "total": len(entries),
            "file_path": str(path),
        }

    # ===== AI 短剧增强（v5.3.0 新增）=====

    def drama_detail_stats(self, drama_id: str) -> Optional[Dict[str, Any]]:
        """短剧深度统计（v5.3.0 新增）

        Args:
            drama_id: 短剧 ID

        Returns:
            包含台词数/角色数/场次数/总字数/经典占比等
        """
        conn = self._get_conn()
        did = drama_id[:64] if isinstance(drama_id, str) else ""
        if not did:
            return None

        drama = self.get_drama(did)
        if not drama:
            return None

        scene_count = conn.execute(
            "SELECT COUNT(*) FROM drama_scenes WHERE drama_id = ?", (did,)
        ).fetchone()[0]
        char_count = conn.execute(
            "SELECT COUNT(*) FROM drama_characters WHERE drama_id = ?", (did,)
        ).fetchone()[0]
        line_count = conn.execute(
            "SELECT COUNT(*) FROM drama_lines WHERE drama_id = ?", (did,)
        ).fetchone()[0]
        classic_count = conn.execute(
            "SELECT COUNT(*) FROM drama_lines WHERE drama_id = ? AND is_classic = 1", (did,)
        ).fetchone()[0]

        # 总字数 & 平均台词长度
        text_info = conn.execute(
            "SELECT SUM(LENGTH(line_text)) as total_chars, "
            "AVG(LENGTH(line_text)) as avg_len FROM drama_lines WHERE drama_id = ?",
            (did,)
        ).fetchone()
        total_chars = text_info[0] or 0
        avg_line_len = round(text_info[1] or 0, 1)

        # 每集台词数分布
        ep_dist = conn.execute(
            "SELECT episode, COUNT(*) as cnt FROM drama_lines "
            "WHERE drama_id = ? GROUP BY episode ORDER BY episode", (did,)
        ).fetchall()

        # 台词最多的角色 Top-5
        top_chars = conn.execute(
            "SELECT character_name, COUNT(*) as cnt FROM drama_lines "
            "WHERE drama_id = ? AND character_name != '' "
            "GROUP BY character_name ORDER BY cnt DESC LIMIT 5", (did,)
        ).fetchall()

        classic_ratio = round(classic_count / line_count * 100, 1) if line_count > 0 else 0.0

        return {
            "drama_id": did,
            "title": drama.title,
            "genre": drama.genre.value if hasattr(drama.genre, "value") else str(drama.genre),
            "status": drama.status.value if hasattr(drama.status, "value") else str(drama.status),
            "rating": drama.rating,
            "total_episodes": drama.total_episodes,
            "current_episode": drama.current_episode,
            "scene_count": scene_count,
            "character_count": char_count,
            "line_count": line_count,
            "classic_line_count": classic_count,
            "classic_ratio": classic_ratio,
            "total_text_chars": total_chars,
            "avg_line_length": avg_line_len,
            "episode_distribution": {str(r[0]): r[1] for r in ep_dist},
            "top_characters_by_lines": [
                {"name": r[0], "line_count": r[1]} for r in top_chars
            ],
        }

    def random_lines(self,
                     drama_id: Optional[str] = None,
                     character_id: Optional[str] = None,
                     is_classic: Optional[bool] = None,
                     count: int = 1) -> List[DramaLine]:
        """随机抽取台词（v5.3.0 新增）

        Args:
            drama_id: 限定短剧
            character_id: 限定角色
            is_classic: 仅经典台词
            count: 抽取数量

        Returns:
            随机台词列表
        """
        conn = self._get_conn()
        count = max(1, min(100, int(count)))

        sql = "SELECT * FROM drama_lines WHERE 1=1"
        params: List[Any] = []

        if drama_id:
            sql += " AND drama_id = ?"
            params.append(drama_id[:64])
        if character_id:
            sql += " AND character_id = ?"
            params.append(character_id[:64])
        if is_classic is not None:
            sql += " AND is_classic = ?"
            params.append(1 if is_classic else 0)

        sql += " ORDER BY RANDOM() LIMIT ?"
        params.append(count)

        rows = conn.execute(sql, params).fetchall()
        return [self._row_to_line(r) for r in rows]

    def character_profile(self,
                          character_id: str,
                          drama_id: Optional[str] = None) -> Optional[Dict[str, Any]]:
        """角色画像分析（v5.3.0 新增）

        Args:
            character_id: 角色 ID
            drama_id: 限定短剧（可选）

        Returns:
            角色画像：台词数/经典台词数/出场场次/平均台词长度/台词风格
        """
        conn = self._get_conn()
        cid = character_id[:64] if isinstance(character_id, str) else ""
        if not cid:
            return None

        # 查找角色信息
        char = None
        if drama_id:
            did = drama_id[:64]
            chars = conn.execute(
                "SELECT * FROM drama_characters WHERE id = ? AND drama_id = ?",
                (cid, did)
            ).fetchall()
        else:
            chars = conn.execute(
                "SELECT * FROM drama_characters WHERE id = ?", (cid,)
            ).fetchall()

        if chars:
            char = self._row_to_character(chars[0])
        else:
            # 可能角色 ID 不在 characters 表但台词中有记录
            name_row = conn.execute(
                "SELECT character_name FROM drama_lines WHERE character_id = ? LIMIT 1",
                (cid,)
            ).fetchone()
            if not name_row:
                return None

        # 统计台词
        if drama_id:
            did = drama_id[:64]
            line_sql = "SELECT COUNT(*), SUM(is_classic), SUM(LENGTH(line_text)), " \
                       "COUNT(DISTINCT scene_id) FROM drama_lines " \
                       "WHERE character_id = ? AND drama_id = ?"
            line_params = [cid, did]
        else:
            line_sql = "SELECT COUNT(*), SUM(is_classic), SUM(LENGTH(line_text)), " \
                       "COUNT(DISTINCT scene_id) FROM drama_lines " \
                       "WHERE character_id = ?"
            line_params = [cid]

        info = conn.execute(line_sql, line_params).fetchone()
        total_lines = info[0] or 0
        classic_lines = info[1] or 0
        total_chars = info[2] or 0
        scene_count = info[3] or 0

        avg_line_len = round(total_chars / total_lines, 1) if total_lines > 0 else 0.0

        # 出场短剧列表
        dramas = conn.execute(
            "SELECT DISTINCT drama_id FROM drama_lines WHERE character_id = ?", (cid,)
        ).fetchall()

        # 最长台词（代表性台词）
        longest = conn.execute(
            "SELECT line_text FROM drama_lines WHERE character_id = ? "
            "ORDER BY LENGTH(line_text) DESC LIMIT 1", (cid,)
        ).fetchone()
        longest_line = longest[0] if longest else ""

        # 经典台词数
        classic_ratio = round(classic_lines / total_lines * 100, 1) if total_lines > 0 else 0.0

        name = ""
        if char:
            name = char.name
        elif info:
            name_row = conn.execute(
                "SELECT character_name FROM drama_lines WHERE character_id = ? LIMIT 1", (cid,)
            ).fetchone()
            name = name_row[0] if name_row else ""

        return {
            "character_id": cid,
            "name": name,
            "drama_id": drama_id[:64] if drama_id else None,
            "total_lines": total_lines,
            "classic_lines": classic_lines,
            "classic_ratio": classic_ratio,
            "scene_appearances": scene_count,
            "drama_appearances": len(dramas),
            "drama_ids": [r[0] for r in dramas],
            "avg_line_length": avg_line_len,
            "total_text_chars": total_chars,
            "longest_line": longest_line[:300],
        }

    # ===== v5.3.1 新增 =====

    def search_agent_memories(self,
                              agent_id: str,
                              keyword: str,
                              limit: int = 50,
                              offset: int = 0) -> List[Any]:
        """在指定 Agent 的记忆中搜索关键词（v5.3.1 新增）

        Args:
            agent_id: Agent ID
            keyword: 搜索关键词
            limit: 返回数量上限
            offset: 偏移量

        Returns:
            匹配的记忆列表
        """
        conn = self._get_conn()
        aid = _filter_unicode_ctrl(agent_id[:128]) if isinstance(agent_id, str) else ""
        kw = keyword[:200] if isinstance(keyword, str) else ""
        if not aid or not kw:
            return []

        limit = max(1, min(500, int(limit)))
        offset = max(0, int(offset))

        # v5.3.1 安全加固：参数化查询防 SQL 注入；软删除通过 category='trash' 标记
        # v5.3.3 安全加固：LIKE 通配符转义，防 % 和 _ 注入
        pattern = f"%{_escape_like(kw)}%"
        rows = conn.execute(
            "SELECT * FROM memories WHERE source_agent = ? AND category != 'trash' "
            "AND (content LIKE ? ESCAPE '\\' OR category LIKE ? ESCAPE '\\' OR tags LIKE ? ESCAPE '\\') "
            "ORDER BY importance DESC, created_at DESC LIMIT ? OFFSET ?",
            (aid, pattern, pattern, pattern, limit, offset)
        ).fetchall()
        return [self._row_to_entry(r) for r in rows]

    def compare_agents(self,
                       agent_a: str,
                       agent_b: str) -> Dict[str, Any]:
        """对比两个 Agent 的记忆差异（v5.3.1 新增）

        Args:
            agent_a: Agent A ID
            agent_b: Agent B ID

        Returns:
            对比结果：各自记忆数、共同分类、独有分类、共同标签
        """
        conn = self._get_conn()
        aid_a = agent_a[:128] if isinstance(agent_a, str) else ""
        aid_b = agent_b[:128] if isinstance(agent_b, str) else ""
        if not aid_a or not aid_b:
            return {"error": "Agent ID 不能为空"}

        # 各自记忆总数（软删除通过 category='trash' 标记）
        count_a = conn.execute(
            "SELECT COUNT(*) FROM memories WHERE source_agent = ? AND category != 'trash'",
            (aid_a,)
        ).fetchone()[0]
        count_b = conn.execute(
            "SELECT COUNT(*) FROM memories WHERE source_agent = ? AND category != 'trash'",
            (aid_b,)
        ).fetchone()[0]

        # 各自分类集合
        cats_a = set(r[0] for r in conn.execute(
            "SELECT DISTINCT category FROM memories WHERE source_agent = ? AND category != 'trash' AND category IS NOT NULL",
            (aid_a,)
        ).fetchall())
        cats_b = set(r[0] for r in conn.execute(
            "SELECT DISTINCT category FROM memories WHERE source_agent = ? AND category != 'trash' AND category IS NOT NULL",
            (aid_b,)
        ).fetchall())

        # 各自标签集合
        tags_a = set()
        for r in conn.execute(
            "SELECT DISTINCT tags FROM memories WHERE source_agent = ? AND category != 'trash' AND tags IS NOT NULL",
            (aid_a,)
        ).fetchall():
            if r[0]:
                tags_a.update(t.strip() for t in r[0].split(",") if t.strip())
        tags_b = set()
        for r in conn.execute(
            "SELECT DISTINCT tags FROM memories WHERE source_agent = ? AND category != 'trash' AND tags IS NOT NULL",
            (aid_b,)
        ).fetchall():
            if r[0]:
                tags_b.update(t.strip() for t in r[0].split(",") if t.strip())

        # 平均重要度
        avg_a = conn.execute(
            "SELECT AVG(CASE importance WHEN 'CRITICAL' THEN 4 WHEN 'HIGH' THEN 3 WHEN 'MEDIUM' THEN 2 ELSE 1 END) "
            "FROM memories WHERE source_agent = ? AND category != 'trash'",
            (aid_a,)
        ).fetchone()[0] or 0
        avg_b = conn.execute(
            "SELECT AVG(CASE importance WHEN 'CRITICAL' THEN 4 WHEN 'HIGH' THEN 3 WHEN 'MEDIUM' THEN 2 ELSE 1 END) "
            "FROM memories WHERE source_agent = ? AND category != 'trash'",
            (aid_b,)
        ).fetchone()[0] or 0

        common_cats = sorted(cats_a & cats_b)
        only_a_cats = sorted(cats_a - cats_b)
        only_b_cats = sorted(cats_b - cats_a)
        common_tags = sorted(tags_a & tags_b)

        return {
            "agent_a": aid_a,
            "agent_b": aid_b,
            "count_a": count_a,
            "count_b": count_b,
            "avg_importance_a": round(avg_a, 2),
            "avg_importance_b": round(avg_b, 2),
            "common_categories": common_cats,
            "only_a_categories": only_a_cats,
            "only_b_categories": only_b_cats,
            "common_tags": common_tags[:50],
            "tags_a_count": len(tags_a),
            "tags_b_count": len(tags_b),
        }

    def search_dramas(self,
                      keyword: str,
                      genre: Optional[str] = None,
                      min_rating: float = 0.0,
                      limit: int = 50,
                      offset: int = 0) -> List[Any]:
        """按关键词搜索短剧（v5.3.1 新增）

        Args:
            keyword: 搜索关键词（匹配标题/描述/标签）
            genre: 类型过滤
            min_rating: 最低评分
            limit: 返回数量上限
            offset: 偏移量

        Returns:
            匹配的短剧列表
        """
        conn = self._get_conn()
        kw = keyword[:200] if isinstance(keyword, str) else ""
        if not kw:
            return []

        limit = max(1, min(500, int(limit)))
        offset = max(0, int(offset))
        min_rating = max(0.0, min(10.0, float(min_rating)))

        # v5.3.3 安全加固：LIKE 通配符转义，防 % 和 _ 注入
        pattern = f"%{_escape_like(kw)}%"
        query = ("SELECT * FROM drama_series WHERE "
                 "(title LIKE ? ESCAPE '\\' OR description LIKE ? ESCAPE '\\' OR tags LIKE ? ESCAPE '\\')")
        params = [pattern, pattern, pattern]

        if genre:
            query += " AND genre = ?"
            params.append(genre)
        if min_rating > 0:
            query += " AND rating >= ?"
            params.append(min_rating)

        query += " ORDER BY rating DESC, updated_at DESC LIMIT ? OFFSET ?"
        params.extend([limit, offset])

        rows = conn.execute(query, params).fetchall()
        return [self._row_to_drama(r) for r in rows]

    def character_ranking(self,
                          drama_id: Optional[str] = None,
                          sort_by: str = "lines",
                          limit: int = 20) -> List[Dict[str, Any]]:
        """角色台词排行榜（v5.3.1 新增）

        Args:
            drama_id: 限定短剧（可选，不指定则全局排行）
            sort_by: 排序维度 lines/classic/scenes
            limit: 返回数量上限

        Returns:
            角色排行列表
        """
        conn = self._get_conn()
        limit = max(1, min(100, int(limit)))

        valid_sorts = {"lines": "total_lines", "classic": "classic_lines", "scenes": "scene_count"}
        sort_col = valid_sorts.get(sort_by, "total_lines")

        if drama_id:
            did = drama_id[:64] if isinstance(drama_id, str) else ""
            sql = (
                "SELECT character_id, character_name, "
                "COUNT(*) as total_lines, "
                "SUM(is_classic) as classic_lines, "
                "COUNT(DISTINCT scene_id) as scene_count, "
                "SUM(LENGTH(line_text)) as total_chars "
                "FROM drama_lines WHERE drama_id = ? "
                "GROUP BY character_id, character_name "
                f"ORDER BY {sort_col} DESC LIMIT ?"
            )
            rows = conn.execute(sql, (did, limit)).fetchall()
        else:
            sql = (
                "SELECT character_id, character_name, "
                "COUNT(*) as total_lines, "
                "SUM(is_classic) as classic_lines, "
                "COUNT(DISTINCT scene_id) as scene_count, "
                "SUM(LENGTH(line_text)) as total_chars "
                "FROM drama_lines "
                "GROUP BY character_id, character_name "
                f"ORDER BY {sort_col} DESC LIMIT ?"
            )
            rows = conn.execute(sql, (limit,)).fetchall()

        result = []
        for i, r in enumerate(rows):
            total = r[2] or 0
            classic = r[3] or 0
            scenes = r[4] or 0
            chars = r[5] or 0
            result.append({
                "rank": i + 1,
                "character_id": r[0] or "",
                "name": r[1] or "",
                "total_lines": total,
                "classic_lines": classic,
                "classic_ratio": round(classic / total * 100, 1) if total > 0 else 0.0,
                "scene_count": scenes,
                "total_chars": chars,
                "avg_line_length": round(chars / total, 1) if total > 0 else 0.0,
            })
        return result

    # ===== v5.3.2 新增 =====

    def agent_diff_memories(self,
                            agent_id: str,
                            days_a: int = 7,
                            days_b: int = 1) -> Dict[str, Any]:
        """对比同一 Agent 在不同时间段的记忆差异（v5.3.2 新增）

        Args:
            agent_id: Agent ID
            days_a: 时间段 A 回溯天数（较早，如 7 天前至今）
            days_b: 时间段 B 回溯天数（较近，如 1 天前至今）

        Returns:
            差异报告字典
        """
        conn = self._get_conn()
        aid = _filter_unicode_ctrl(agent_id[:128]) if isinstance(agent_id, str) else ""
        if not aid:
            return {"error": "Agent ID 不能为空"}

        days_a = max(1, min(3650, int(days_a)))
        days_b = max(1, min(3650, int(days_b)))
        if days_b > days_a:
            days_a, days_b = days_b, days_a

        now = __import__("time").time()
        ts_a = now - days_a * 86400
        ts_b = now - days_b * 86400

        # v5.3.2 安全加固：参数化 SQL
        # A 时间段（从 ts_a 到 ts_b 的增量）
        period_a_rows = conn.execute(
            "SELECT id, content, category, importance, created_at FROM memories "
            "WHERE source_agent = ? AND category != 'trash' AND created_at >= ? AND created_at < ?",
            (aid, ts_a, ts_b)
        ).fetchall()

        # B 时间段（从 ts_b 到现在）
        period_b_rows = conn.execute(
            "SELECT id, content, category, importance, created_at FROM memories "
            "WHERE source_agent = ? AND category != 'trash' AND created_at >= ?",
            (aid, ts_b)
        ).fetchall()

        # 分类聚合
        def _agg(rows):
            cats = {}
            imp_sum = {"LOW": 0, "MEDIUM": 0, "HIGH": 0, "CRITICAL": 0}
            for r in rows:
                cat = r[2] or "general"
                cats[cat] = cats.get(cat, 0) + 1
                imp = r[3] or "MEDIUM"
                imp_sum[imp] = imp_sum.get(imp, 0) + 1
            return {"count": len(rows), "by_category": cats, "by_importance": imp_sum}

        period_a = _agg(period_a_rows)
        period_b = _agg(period_b_rows)

        # 分类差集
        cats_a = set(period_a["by_category"].keys())
        cats_b = set(period_b["by_category"].keys())
        new_cats = sorted(cats_b - cats_a)
        dropped_cats = sorted(cats_a - cats_b)

        return {
            "agent_id": aid,
            "period_a": {"days": days_a, **period_a, "time_range": f"[{days_a}天前 ~ {days_b}天前)"},
            "period_b": {"days": days_b, **period_b, "time_range": f"[{days_b}天前 ~ 现在]"},
            "new_categories": new_cats,
            "dropped_categories": dropped_cats,
            "total_diff": period_b["count"] - period_a["count"],
        }

    def agent_purge(self,
                    agent_id: str,
                    actor: str = "system",
                    session_id: str = "",
                    dry_run: bool = True) -> Dict[str, Any]:
        """清空指定 Agent 的全部记忆（v5.3.2 新增，高危操作）

        v5.3.3 安全加固：添加频率限制，防止暴力清空攻击。

        Args:
            agent_id: 目标 Agent ID
            actor: 操作者（审计日志）
            session_id: 会话 ID
            dry_run: True=仅预览，False=实际执行

        Returns:
            清理结果字典
        """
        conn = self._get_conn()
        aid = _filter_unicode_ctrl(agent_id[:128]) if isinstance(agent_id, str) else ""
        if not aid:
            return {"error": "Agent ID 不能为空"}

        # v5.3.3 安全加固：频率限制，实际执行时每分钟最多 3 次
        if not dry_run:
            rate_key = f"purge:{aid}"
            if not _rate_limiter.check(rate_key, max_calls=3, window_seconds=60):
                return {"error": "操作过于频繁，请 60 秒后重试（频率限制）"}

        # 找出所有目标记忆（含软删除也一并清，因为是 purge）
        rows = conn.execute(
            "SELECT id, rowid FROM memories WHERE source_agent = ?",
            (aid,)
        ).fetchall()

        total = len(rows)
        if total == 0:
            return {"agent_id": aid, "total_found": 0, "purged": 0, "dry_run": dry_run}

        if dry_run:
            # 预览各分类数量
            cat_stats = {}
            for r in conn.execute(
                "SELECT category, COUNT(*) FROM memories WHERE source_agent = ? GROUP BY category",
                (aid,)
            ).fetchall():
                cat_stats[r[0] or "general"] = r[1]
            return {
                "agent_id": aid,
                "total_found": total,
                "by_category": cat_stats,
                "purged": 0,
                "dry_run": True,
                "note": "加 --force 参数执行实际删除",
            }

        # 实际删除（v5.3.2 安全：批量 id 列表参数化）
        ids = [r[0] for r in rows]
        rowids = [r[1] for r in rows]
        placeholders = ",".join(["?"] * len(ids))

        # 清理关联表：memory_versions、memory_links、memory_notes（以记忆 id 为外键）
        for table, col in [("memory_versions", "memory_id"),
                           ("memory_links", "source_id"),
                           ("memory_links", "target_id"),
                           ("memory_notes", "memory_id")]:
            try:
                conn.execute(
                    f"DELETE FROM {table} WHERE {col} IN ({placeholders})", ids
                )
            except Exception:
                pass

        # 清理 FTS（contentless FTS5 特殊语法）
        for rid in rowids:
            try:
                conn.execute(
                    "INSERT INTO memory_fts(memory_fts, rowid, content, category, tags) "
                    "VALUES('delete', ?, ?, ?, ?)",
                    (rid, "", "", "[]")
                )
            except Exception:
                pass

        conn.execute(f"DELETE FROM memories WHERE id IN ({placeholders})", ids)

        # v5.4.2 修复：改走 _add_audit，享受 fail-closed 保护（审计失败时拒绝操作）
        self._add_audit(
            action="agent_purge",
            memory_id=aid,
            actor=actor,
            session_id=session_id,
            privacy_level="INTERNAL",
            details={"agent_id": aid, "count": total},
        )

        conn.commit()
        return {"agent_id": aid, "total_found": total, "purged": total, "dry_run": False}

    def drama_update_progress(self,
                               drama_id: str,
                               current_episode: int,
                               status: Optional[str] = None,
                               user_rating: Optional[float] = None,
                               actor: str = "system") -> Dict[str, Any]:
        """更新短剧观看进度（v5.3.2 新增）

        利用 drama_series 表的 metadata 字段存储用户个性化进度，
        无需新增数据库列即可实现该功能。

        Args:
            drama_id: 短剧 ID
            current_episode: 当前看到第几集（≥1）
            status: 观看状态 WATCHING/COMPLETED/DROPPED/PLANNING（可选）
            user_rating: 用户评分 0-10（可选）
            actor: 操作者

        Returns:
            更新结果字典
        """
        conn = self._get_conn()
        did = drama_id[:64] if isinstance(drama_id, str) else ""
        if not did:
            return {"error": "短剧 ID 不能为空"}

        current_episode = max(1, min(10000, int(current_episode)))

        # v5.3.2 安全：status 枚举白名单
        valid_status = {"WATCHING", "COMPLETED", "DROPPED", "PLANNING"}
        if status:
            status = status.upper()
            if status not in valid_status:
                status = None

        if user_rating is not None:
            user_rating = max(0.0, min(10.0, float(user_rating)))

        # 检查短剧是否存在
        row = conn.execute(
            "SELECT id, metadata FROM drama_series WHERE id = ?",
            (did,)
        ).fetchone()
        if not row:
            return {"error": f"短剧不存在: {did}"}

        # 解析现有 metadata（v5.3.2 安全：统一 safe json 加载）
        import json as _json
        try:
            metadata = _json.loads(row[1]) if (row[1] and row[1].strip()) else {}
        except Exception:
            metadata = {}

        # 写入进度字段（前缀 user_progress 避免与官方字段冲突）
        prog = metadata.get("user_progress", {})
        prog["current_episode"] = current_episode
        prog["updated_at"] = __import__("time").time()
        if status:
            prog["status"] = status
        if user_rating is not None:
            prog["user_rating"] = user_rating
        metadata["user_progress"] = prog

        try:
            meta_json = _json.dumps(metadata, ensure_ascii=False)[:10000]
        except Exception:
            meta_json = "{}"

        conn.execute(
            "UPDATE drama_series SET metadata = ?, updated_at = ? WHERE id = ?",
            (meta_json, __import__("time").time(), did)
        )
        conn.commit()

        # 返回完整当前进度
        return {
            "drama_id": did,
            "current_episode": current_episode,
            "status": prog.get("status"),
            "user_rating": prog.get("user_rating"),
            "updated_at": prog.get("updated_at"),
        }

    def drama_recommend_v2(self,
                            genre: Optional[str] = None,
                            min_rating: float = 0.0,
                            mode: str = "unwatched",
                            limit: int = 20) -> List[Dict[str, Any]]:
        """短剧智能推荐 v2（v5.3.2 新增）

        支持按观看状态过滤：优先推荐未观看 / 正在追 / 已弃剧。

        Args:
            genre: 类型过滤（枚举白名单）
            min_rating: 最低评分
            mode: 推荐模式 unwatched/watching/dropped/all
            limit: 返回数量上限

        Returns:
            推荐短剧列表
        """
        conn = self._get_conn()
        import json as _json

        limit = max(1, min(200, int(limit)))
        min_rating = max(0.0, min(10.0, float(min_rating)))

        # v5.3.2 安全：白名单
        valid_genres = {"ROMANCE", "ACTION", "COMEDY", "THRILLER", "SCIFI",
                        "HISTORICAL", "URBAN", "FANTASY", "MYSTERY", "DRAMA"}
        if genre:
            genre = genre.upper()
            if genre not in valid_genres:
                genre = None

        valid_modes = {"unwatched", "watching", "dropped", "all"}
        if mode not in valid_modes:
            mode = "unwatched"

        rows = conn.execute("SELECT * FROM drama_series ORDER BY rating DESC").fetchall()

        candidates = []
        for r in rows:
            d = self._row_to_drama(r)
            if genre and (getattr(d, "genre", "") or "").upper() != genre:
                continue
            rating_val = getattr(d, "rating", 0) or 0
            if rating_val < min_rating:
                continue

            # 从 metadata 解析进度
            status = ""
            cur_ep = 0
            try:
                md = getattr(d, "metadata", {}) or {}
                if isinstance(md, str):
                    md = _json.loads(md) if md.strip() else {}
                prog = md.get("user_progress", {})
                status = prog.get("status", "") or ""
                cur_ep = prog.get("current_episode", 0) or 0
            except Exception:
                pass

            # 模式过滤
            total_eps = getattr(d, "total_episodes", 0) or 0
            if mode == "unwatched":
                if status == "COMPLETED" or (cur_ep > 0 and cur_ep >= total_eps and total_eps > 0):
                    continue
                if status == "WATCHING":  # 追剧中不算未观看
                    continue
                if cur_ep > 0:
                    continue
            elif mode == "watching":
                if status != "WATCHING" and not (0 < cur_ep < (total_eps or 9999)):
                    continue
            elif mode == "dropped":
                if status != "DROPPED":
                    continue

            meta_safe = getattr(d, "metadata", {})
            if isinstance(meta_safe, str):
                try:
                    meta_safe = _json.loads(meta_safe) if meta_safe.strip() else {}
                except Exception:
                    meta_safe = {}
            prog = (meta_safe or {}).get("user_progress", {})
            candidates.append({
                "drama_id": getattr(d, "id", ""),
                "title": getattr(d, "title", ""),
                "genre": getattr(d, "genre", ""),
                "rating": rating_val,
                "status": getattr(d, "status", ""),
                "total_episodes": total_eps,
                "watch_status": prog.get("status", ""),
                "current_episode": prog.get("current_episode", 0),
                "user_rating": prog.get("user_rating"),
            })

        # 按综合分排序：官方评分 * 0.7 + （未看加权 1.3）
        def _score(c):
            s = c["rating"] * 0.7
            if not c["current_episode"]:
                s *= 1.3
            if c["watch_status"] != "DROPPED":
                s += 0.5
            return s

        candidates.sort(key=_score, reverse=True)
        return candidates[:limit]

    # ===== v5.3.3 新增：Agent 记忆时间线 & 热力图 =====

    def agent_timeline(self,
                       agent_id: str,
                       days: int = 30) -> Dict[str, Any]:
        """Agent 记忆时间线分析（v5.3.3 新增）

        按天/小时统计记忆创建趋势，识别 Agent 活跃时段。

        Args:
            agent_id: Agent ID
            days: 回溯天数（1-365）

        Returns:
            时间线分析结果：按天计数、按小时分布、活跃峰、趋势
        """
        conn = self._get_conn()
        aid = _filter_unicode_ctrl(agent_id[:128]) if isinstance(agent_id, str) else ""
        if not aid:
            return {"error": "Agent ID 不能为空"}

        days = max(1, min(365, int(days)))
        now = time.time()
        since = now - days * 86400

        # v5.3.3 安全加固：参数化 SQL
        rows = conn.execute(
            "SELECT created_at FROM memories "
            "WHERE source_agent = ? AND category != 'trash' AND created_at >= ?",
            (aid, since)
        ).fetchall()

        if not rows:
            return {
                "agent_id": aid,
                "days": days,
                "total_memories": 0,
                "by_day": {},
                "by_hour": {},
                "peak_day": None,
                "peak_hour": None,
                "trend": "no_data",
            }

        from datetime import datetime as _dt

        by_day: Dict[str, int] = {}
        by_hour: Dict[int, int] = {h: 0 for h in range(24)}

        for r in rows:
            ts = r[0] or 0
            if ts <= 0:
                continue
            dt = _dt.fromtimestamp(ts)
            day_key = dt.strftime("%Y-%m-%d")
            by_day[day_key] = by_day.get(day_key, 0) + 1
            by_hour[dt.hour] = by_hour.get(dt.hour, 0) + 1

        # 找活跃峰
        peak_day = max(by_day, key=by_day.get) if by_day else None
        peak_hour = max(by_hour, key=by_hour.get) if by_hour else None

        # 趋势分析：后半段 vs 前半段
        sorted_days = sorted(by_day.keys())
        mid = len(sorted_days) // 2
        if mid > 0:
            first_half = sum(by_day[d] for d in sorted_days[:mid])
            second_half = sum(by_day[d] for d in sorted_days[mid:])
            if second_half > first_half * 1.1:
                trend = "rising"
            elif first_half > second_half * 1.1:
                trend = "declining"
            else:
                trend = "stable"
        else:
            trend = "insufficient_data"

        # 活跃时段标签
        active_hours = sorted(by_hour.items(), key=lambda x: x[1], reverse=True)
        top_hours = [h for h, c in active_hours[:3] if c > 0]

        return {
            "agent_id": aid,
            "days": days,
            "total_memories": len(rows),
            "by_day": by_day,
            "by_hour": {str(h): by_hour[h] for h in range(24) if by_hour[h] > 0},
            "peak_day": {"date": peak_day, "count": by_day[peak_day]} if peak_day else None,
            "peak_hour": {"hour": peak_hour, "count": by_hour[peak_hour]} if peak_hour is not None else None,
            "top_active_hours": top_hours,
            "trend": trend,
            "avg_per_day": round(len(rows) / days, 2) if days > 0 else 0,
        }

    def agent_heatmap(self,
                      agent_id: str,
                      days: int = 30) -> Dict[str, Any]:
        """Agent 记忆热力图矩阵（v5.3.3 新增）

        生成 分类 × 重要度 的记忆密度矩阵，可视化 Agent 记忆分布。

        Args:
            agent_id: Agent ID
            days: 回溯天数（1-365）

        Returns:
            热力图矩阵：分类行 × 重要度列的计数矩阵
        """
        conn = self._get_conn()
        aid = _filter_unicode_ctrl(agent_id[:128]) if isinstance(agent_id, str) else ""
        if not aid:
            return {"error": "Agent ID 不能为空"}

        days = max(1, min(365, int(days)))
        now = time.time()
        since = now - days * 86400

        # v5.3.3 安全加固：参数化 SQL
        rows = conn.execute(
            "SELECT category, importance FROM memories "
            "WHERE source_agent = ? AND category != 'trash' AND created_at >= ?",
            (aid, since)
        ).fetchall()

        importance_levels = ["LOW", "MEDIUM", "HIGH", "CRITICAL"]
        matrix: Dict[str, Dict[str, int]] = {}

        for r in rows:
            cat = (r[0] or "general")[:128]
            imp = (r[1] or "MEDIUM").upper()
            if imp not in importance_levels:
                imp = "MEDIUM"
            if cat not in matrix:
                matrix[cat] = {lv: 0 for lv in importance_levels}
            matrix[cat][imp] += 1

        # 计算行/列总计和密度
        row_totals = {cat: sum(vals.values()) for cat, vals in matrix.items()}
        col_totals = {lv: sum(matrix[cat][lv] for cat in matrix) for lv in importance_levels}
        grand_total = sum(row_totals.values())

        # 找密度最高的单元格
        max_cell = None
        max_count = 0
        for cat, vals in matrix.items():
            for imp, cnt in vals.items():
                if cnt > max_count:
                    max_count = cnt
                    max_cell = {"category": cat, "importance": imp, "count": cnt}

        return {
            "agent_id": aid,
            "days": days,
            "total_memories": grand_total,
            "matrix": matrix,
            "row_totals": row_totals,
            "col_totals": col_totals,
            "importance_levels": importance_levels,
            "categories": sorted(matrix.keys()),
            "max_density_cell": max_cell,
        }

    # ===== v5.3.3 新增：AI 短剧追剧统计 & 角色关系网络 =====

    def drama_binge_stats(self,
                          drama_id: Optional[str] = None) -> Dict[str, Any]:
        """追剧统计（v5.3.3 新增）

        统计观看进度记录，包括连续观看天数、最长追剧周期、平均完成时长。

        Args:
            drama_id: 指定短剧（可选，None=全部）

        Returns:
            追剧统计结果
        """
        conn = self._get_conn()
        import json as _json

        # v5.3.3 安全加固：参数化 SQL + ID 长度限制
        if drama_id:
            did = drama_id[:64] if isinstance(drama_id, str) else ""
            rows = conn.execute(
                "SELECT id, title, total_episodes, current_episode, status, "
                "updated_at, last_watched_at, metadata, rating "
                "FROM drama_series WHERE id = ?",
                (did,)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT id, title, total_episodes, current_episode, status, "
                "updated_at, last_watched_at, metadata, rating "
                "FROM drama_series"
            ).fetchall()

        if not rows:
            return {
                "total_dramas": 0,
                "watching": 0,
                "completed": 0,
                "dropped": 0,
                "planned": 0,
                "binge_stats": {},
            }

        total = len(rows)
        watching = 0
        completed = 0
        dropped = 0
        planned = 0
        total_episodes_watched = 0
        total_episodes_planned = 0
        watch_records = []

        for r in rows:
            status = r[4] or "planned"
            cur_ep = r[3] or 0
            total_ep = r[2] or 0
            rating = r[8] or 0.0
            last_watched = r[6] or 0
            updated = r[5] or 0

            if status == "watching":
                watching += 1
            elif status == "completed":
                completed += 1
            elif status == "dropped":
                dropped += 1
            else:
                planned += 1

            total_episodes_watched += cur_ep
            total_episodes_planned += total_ep

            # 从 metadata 解析进度历史
            watch_history = []
            try:
                md_raw = r[7]
                if isinstance(md_raw, str) and md_raw.strip():
                    md = _json.loads(md_raw)
                elif isinstance(md_raw, dict):
                    md = md_raw
                else:
                    md = {}
                watch_history = md.get("user_progress", {}).get("history", [])
            except Exception:
                pass

            watch_records.append({
                "drama_id": r[0],
                "title": r[1],
                "status": status,
                "current_episode": cur_ep,
                "total_episodes": total_ep,
                "rating": rating,
                "last_watched_at": last_watched,
                "updated_at": updated,
                "watch_history_count": len(watch_history),
            })

        # 完成率
        completion_rate = round(
            (total_episodes_watched / total_episodes_planned * 100)
            if total_episodes_planned > 0 else 0.0, 2
        )

        # 最近观看的短剧 Top-5
        recent_watched = sorted(
            [w for w in watch_records if w["last_watched_at"] > 0],
            key=lambda x: x["last_watched_at"], reverse=True
        )[:5]

        # 评分分布
        rated = [w for w in watch_records if w["rating"] > 0]
        avg_rating = round(sum(w["rating"] for w in rated) / len(rated), 2) if rated else 0.0

        return {
            "total_dramas": total,
            "watching": watching,
            "completed": completed,
            "dropped": dropped,
            "planned": planned,
            "total_episodes_watched": total_episodes_watched,
            "total_episodes_planned": total_episodes_planned,
            "completion_rate": completion_rate,
            "average_rating": avg_rating,
            "rated_count": len(rated),
            "recent_watched": recent_watched,
        }

    def character_network(self,
                          drama_id: str) -> Dict[str, Any]:
        """角色关系网络分析（v5.3.3 新增）

        分析短剧中角色间的共同出场频率，构建角色关系网络数据。

        Args:
            drama_id: 短剧 ID

        Returns:
            角色关系网络：节点列表 + 边列表（含共同出场次数）
        """
        conn = self._get_conn()
        did = _filter_unicode_ctrl(drama_id[:64]) if isinstance(drama_id, str) and drama_id else ""
        if not did:
            return {"error": "短剧 ID 不能为空"}

        # 获取所有角色
        char_rows = conn.execute(
            "SELECT id, name, role FROM drama_characters WHERE drama_id = ?",
            (did,)
        ).fetchall()

        if not char_rows:
            return {
                "drama_id": did,
                "nodes": [],
                "edges": [],
                "total_characters": 0,
            }

        characters = {r[0]: {"id": r[0], "name": r[1], "role": r[2]} for r in char_rows}

        # 获取所有场次的台词，按场次统计角色出场
        scene_chars: Dict[str, set] = {}
        line_rows = conn.execute(
            "SELECT scene_id, character_id FROM drama_lines "
            "WHERE drama_id = ? AND character_id != ''",
            (did,)
        ).fetchall()

        for r in line_rows:
            scene_id = r[0] or ""
            char_id = r[1] or ""
            if scene_id and char_id and char_id in characters:
                if scene_id not in scene_chars:
                    scene_chars[scene_id] = set()
                scene_chars[scene_id].add(char_id)

        # 构建共现矩阵
        co_occurrence: Dict[str, Dict[str, int]] = {}
        for scene_id, char_set in scene_chars.items():
            char_list = list(char_set)
            for i in range(len(char_list)):
                for j in range(i + 1, len(char_list)):
                    first_char, second_char = char_list[i], char_list[j]
                    if first_char not in co_occurrence:
                        co_occurrence[first_char] = {}
                    if second_char not in co_occurrence[first_char]:
                        co_occurrence[first_char][second_char] = 0
                    co_occurrence[first_char][second_char] += 1

        # 构建边列表
        edges = []
        seen_pairs = set()
        for primary_char, partners in co_occurrence.items():
            for partner_char, count in partners.items():
                pair_key = tuple(sorted([primary_char, partner_char]))
                if pair_key in seen_pairs:
                    continue
                seen_pairs.add(pair_key)
                edges.append({
                    "source": a,
                    "source_name": characters.get(a, {}).get("name", a),
                    "target": b,
                    "target_name": characters.get(b, {}).get("name", b),
                    "weight": count,
                })

        edges.sort(key=lambda e: e["weight"], reverse=True)

        # 节点：附加出场次数和关联数
        nodes = []
        for char_id, info in characters.items():
            scene_count = sum(1 for s in scene_chars.values() if char_id in s)
            connections = sum(1 for e in edges if e["source"] == char_id or e["target"] == char_id)
            nodes.append({
                "id": char_id,
                "name": info["name"],
                "role": info["role"],
                "scene_count": scene_count,
                "connections": connections,
            })

        nodes.sort(key=lambda n: n["connections"], reverse=True)

        return {
            "drama_id": did,
            "nodes": nodes,
            "edges": edges,
            "total_characters": len(nodes),
            "total_edges": len(edges),
            "total_scenes_analyzed": len(scene_chars),
        }

    # ===== v5.3.4 新增：Agent 记忆情感分析 + 记忆衰减评分 =====

    def agent_sentiment(self,
                        agent_id: str,
                        days: int = 30) -> Dict[str, Any]:
        """Agent 记忆情感分析（v5.3.4 新增）

        基于关键词匹配分析 Agent 记忆的整体情感倾向。

        Args:
            agent_id: Agent ID
            days: 回溯天数（1-365）

        Returns:
            情感分析结果：正面/负面/中性计数、情感分布、主导情感
        """
        conn = self._get_conn()
        aid = _filter_unicode_ctrl(agent_id[:128]) if isinstance(agent_id, str) else ""
        if not aid:
            return {"error": "Agent ID 不能为空"}

        days = max(1, min(365, int(days)))
        now = time.time()
        since = now - days * 86400

        # v5.3.4 安全：参数化 SQL
        rows = conn.execute(
            "SELECT content, importance FROM memories "
            "WHERE source_agent = ? AND category != 'trash' AND created_at >= ?",
            (aid, since)
        ).fetchall()

        if not rows:
            return {
                "agent_id": aid,
                "days": days,
                "total_memories": 0,
                "positive": 0,
                "negative": 0,
                "neutral": 0,
                "dominant_sentiment": "no_data",
            }

        # 情感关键词词典
        positive_words = {
            "好", "棒", "优秀", "成功", "完成", "解决", "开心", "满意", "喜欢",
            "good", "great", "excellent", "success", "happy", "love", "perfect",
            "完成", "突破", "提升", "优化", "改进", "有效", "正确", "赞同",
        }
        negative_words = {
            "坏", "差", "失败", "错误", "问题", "bug", "崩溃", "讨厌", "不满",
            "bad", "fail", "error", "broken", "crash", "hate", "wrong", "issue",
            "缺失", "丢失", "异常", "警告", "危险", "漏洞", "冲突", "阻塞",
        }

        positive_count = 0
        negative_count = 0
        neutral_count = 0
        sentiment_by_imp: Dict[str, Dict[str, int]] = {}

        for r in rows:
            content = (r[0] or "").lower()
            imp = (r[1] or "MEDIUM").upper()

            pos_hits = sum(1 for w in positive_words if w in content)
            neg_hits = sum(1 for w in negative_words if w in content)

            if pos_hits > neg_hits:
                sentiment = "positive"
                positive_count += 1
            elif neg_hits > pos_hits:
                sentiment = "negative"
                negative_count += 1
            else:
                sentiment = "neutral"
                neutral_count += 1

            if imp not in sentiment_by_imp:
                sentiment_by_imp[imp] = {"positive": 0, "negative": 0, "neutral": 0}
            sentiment_by_imp[imp][sentiment] += 1

        total = len(rows)
        if positive_count > negative_count and positive_count > neutral_count:
            dominant = "positive"
        elif negative_count > positive_count and negative_count > neutral_count:
            dominant = "negative"
        else:
            dominant = "neutral"

        return {
            "agent_id": aid,
            "days": days,
            "total_memories": total,
            "positive": positive_count,
            "negative": negative_count,
            "neutral": neutral_count,
            "positive_ratio": round(positive_count / total, 4) if total else 0,
            "negative_ratio": round(negative_count / total, 4) if total else 0,
            "neutral_ratio": round(neutral_count / total, 4) if total else 0,
            "dominant_sentiment": dominant,
            "by_importance": sentiment_by_imp,
        }

    def memory_decay(self,
                     agent_id: str,
                     days: int = 30) -> Dict[str, Any]:
        """记忆衰减评分（v5.3.4 新增）

        基于艾宾浩斯遗忘曲线模型，评估 Agent 记忆的衰减状态。
        衰减评分 = 重要性权重 × recency_factor × retention_rate

        Args:
            agent_id: Agent ID
            days: 回溯天数（1-365）

        Returns:
            衰减分析结果：平均衰减率、高危记忆数、各衰减级别分布
        """
        conn = self._get_conn()
        aid = _filter_unicode_ctrl(agent_id[:128]) if isinstance(agent_id, str) else ""
        if not aid:
            return {"error": "Agent ID 不能为空"}

        days = max(1, min(365, int(days)))
        now = time.time()
        since = now - days * 86400

        # v5.3.4 安全：参数化 SQL
        rows = conn.execute(
            "SELECT id, content, importance, created_at, access_count "
            "FROM memories "
            "WHERE source_agent = ? AND category != 'trash' AND created_at >= ?",
            (aid, since)
        ).fetchall()

        if not rows:
            return {
                "agent_id": aid,
                "days": days,
                "total_memories": 0,
                "avg_retention": 0,
                "critical_decay": 0,
                "decay_distribution": {},
            }

        # 重要性权重
        imp_weights = {"LOW": 0.5, "MEDIUM": 0.7, "HIGH": 0.85, "CRITICAL": 1.0}

        total_retention = 0.0
        decay_levels = {"strong": 0, "stable": 0, "fading": 0, "critical": 0}
        critical_memories = []

        for r in rows:
            mid = r[0]
            content = (r[1] or "")[:80]
            imp = (r[2] or "MEDIUM").upper()
            created = r[3] or now
            access_count = r[4] or 0

            imp_weight = imp_weights.get(imp, 0.7)

            # 艾宾浩斯遗忘曲线：retention = e^(-t/S)
            # S = 稳定性，受重要性和访问次数影响
            days_elapsed = max(0.01, (now - created) / 86400)
            stability = imp_weight * (1 + min(access_count, 10) * 0.1) * 7  # 基础7天
            retention = __import__("math").exp(-days_elapsed / stability)

            total_retention += retention

            if retention >= 0.7:
                decay_levels["strong"] += 1
            elif retention >= 0.4:
                decay_levels["stable"] += 1
            elif retention >= 0.15:
                decay_levels["fading"] += 1
            else:
                decay_levels["critical"] += 1
                if len(critical_memories) < 20:
                    critical_memories.append({
                        "id": mid,
                        "content_preview": content,
                        "importance": imp,
                        "retention": round(retention, 4),
                        "days_elapsed": round(days_elapsed, 1),
                        "access_count": access_count,
                    })

        total = len(rows)
        avg_retention = round(total_retention / total, 4) if total else 0

        return {
            "agent_id": aid,
            "days": days,
            "total_memories": total,
            "avg_retention": avg_retention,
            "critical_decay": decay_levels["critical"],
            "decay_distribution": decay_levels,
            "critical_memories": critical_memories,
        }

    # ===== v5.3.4 新增：AI 短剧对比 + 角色成长弧线 =====

    def drama_compare(self,
                      drama_ids: List[str]) -> Dict[str, Any]:
        """短剧对比分析（v5.3.4 新增）

        对比多部短剧的评分、集数、角色数、经典台词数等维度。

        Args:
            drama_ids: 短剧 ID 列表（最多 5 部）

        Returns:
            对比分析结果
        """
        conn = self._get_conn()
        # v5.3.4 安全：数量限制 + ID 长度截断
        if not drama_ids or not isinstance(drama_ids, list):
            return {"error": "短剧 ID 列表不能为空"}
        ids = [d[:64] for d in drama_ids if isinstance(d, str) and d][:5]
        if not ids:
            return {"error": "无有效短剧 ID"}

        dramas = []
        for did in ids:
            row = conn.execute(
                "SELECT id, title, genre, total_episodes, current_episode, "
                "status, rating, metadata "
                "FROM drama_series WHERE id = ?",
                (did,)
            ).fetchone()
            if not row:
                dramas.append({"id": did, "error": "未找到"})
                continue

            # 统计角色数
            char_count = conn.execute(
                "SELECT COUNT(*) FROM drama_characters WHERE drama_id = ?",
                (did,)
            ).fetchone()[0]

            # 统计台词数和经典台词数
            line_stats = conn.execute(
                "SELECT COUNT(*), SUM(CASE WHEN is_classic = 1 THEN 1 ELSE 0 END) "
                "FROM drama_lines WHERE drama_id = ?",
                (did,)
            ).fetchone()

            dramas.append({
                "id": row[0],
                "title": row[1] or "未命名",
                "genre": row[2] or "未知",
                "total_episodes": row[3] or 0,
                "current_episode": row[4] or 0,
                "status": row[5] or "planned",
                "rating": row[6] or 0.0,
                "character_count": char_count,
                "total_lines": line_stats[0] or 0,
                "classic_lines": line_stats[1] or 0,
            })

        # 计算对比维度
        valid = [d for d in dramas if "error" not in d]
        if len(valid) >= 2:
            best_rating = max(valid, key=lambda x: x["rating"])
            most_episodes = max(valid, key=lambda x: x["total_episodes"])
            most_characters = max(valid, key=lambda x: x["character_count"])
            most_classic = max(valid, key=lambda x: x["classic_lines"])
        else:
            best_rating = most_episodes = most_characters = most_classic = None

        return {
            "dramas": dramas,
            "comparison": {
                "best_rated": best_rating["title"] if best_rating else None,
                "most_episodes": most_episodes["title"] if most_episodes else None,
                "most_characters": most_characters["title"] if most_characters else None,
                "most_classic_lines": most_classic["title"] if most_classic else None,
            },
            "total_compared": len(valid),
        }

    def character_arc(self,
                      drama_id: str,
                      character_id: str) -> Dict[str, Any]:
        """角色成长弧线分析（v5.3.4 新增）

        分析角色在不同场景中的台词量变化，识别角色的成长轨迹。

        Args:
            drama_id: 短剧 ID
            character_id: 角色 ID

        Returns:
            角色成长弧线数据：按场景的台词量变化、活跃峰值、成长阶段
        """
        conn = self._get_conn()
        did = _filter_unicode_ctrl(drama_id[:64]) if isinstance(drama_id, str) and drama_id else ""
        cid = character_id[:64] if isinstance(character_id, str) and character_id else ""
        if not did or not cid:
            return {"error": "短剧 ID 和角色 ID 不能为空"}

        # 获取角色信息
        char_row = conn.execute(
            "SELECT id, name, role FROM drama_characters WHERE id = ? AND drama_id = ?",
            (cid, did)
        ).fetchone()
        if not char_row:
            return {"error": "角色不存在"}

        # 获取该角色在各个场景中的台词数
        scene_rows = conn.execute(
            "SELECT scene_id, COUNT(*) as line_count "
            "FROM drama_lines WHERE drama_id = ? AND character_id = ? "
            "GROUP BY scene_id ORDER BY scene_id",
            (did, cid)
        ).fetchall()

        if not scene_rows:
            return {
                "drama_id": did,
                "character_id": cid,
                "character_name": char_row[1],
                "total_scenes": 0,
                "total_lines": 0,
                "arc_points": [],
                "peak_scene": None,
                "growth_stage": "no_data",
            }

        arc_points = []
        for sr in scene_rows:
            arc_points.append({
                "scene_id": sr[0],
                "line_count": sr[1],
            })

        total_lines = sum(p["line_count"] for p in arc_points)
        total_scenes = len(arc_points)

        # 找活跃峰值场景
        peak = max(arc_points, key=lambda x: x["line_count"])

        # 成长阶段分析：将场景分为前中后三段
        third = max(1, total_scenes // 3)
        early = sum(p["line_count"] for p in arc_points[:third])
        mid = sum(p["line_count"] for p in arc_points[third:third * 2])
        late = sum(p["line_count"] for p in arc_points[third * 2:])

        if late > early * 1.2:
            growth_stage = "rising"  # 后期崛起
        elif early > late * 1.2:
            growth_stage = "falling"  # 前期活跃，后期淡出
        elif mid > early * 1.1 and mid > late * 1.1:
            growth_stage = "peak_middle"  # 中期高峰
        else:
            growth_stage = "stable"  # 稳定出场

        return {
            "drama_id": did,
            "character_id": cid,
            "character_name": char_row[1],
            "character_role": char_row[2],
            "total_scenes": total_scenes,
            "total_lines": total_lines,
            "arc_points": arc_points,
            "peak_scene": {"scene_id": peak["scene_id"], "line_count": peak["line_count"]},
            "growth_stage": growth_stage,
            "stage_distribution": {"early": early, "mid": mid, "late": late},
        }

    # ===== v5.3.5 新增：Agent 记忆主题聚类 + 行为洞察 =====

    def memory_cluster(self,
                       agent_id: str,
                       days: int = 30,
                       max_clusters: int = 10) -> Dict[str, Any]:
        """记忆主题聚类（v5.3.5 新增）

        基于关键词和标签相似度，将 Agent 记忆聚合成主题组，
        识别核心话题和知识结构。

        Args:
            agent_id: Agent ID
            days: 回溯天数（1-365）
            max_clusters: 最大聚类数（1-50）

        Returns:
            主题聚类结果：各主题簇列表、核心词、主题标签
        """
        conn = self._get_conn()
        aid = _filter_unicode_ctrl(agent_id[:128]) if isinstance(agent_id, str) else ""
        if not aid:
            return {"error": "Agent ID 不能为空"}

        days = max(1, min(365, int(days)))
        max_clusters = max(1, min(50, int(max_clusters)))
        now = time.time()
        since = now - days * 86400

        # v5.3.5 安全：参数化 SQL + 行数限制
        cur = conn.execute(
            "SELECT id, content, tags, category, importance, created_at "
            "FROM memories "
            "WHERE source_agent = ? AND category != 'trash' AND created_at >= ? "
            "ORDER BY importance DESC, created_at DESC",
            (aid, since)
        )
        rows = _limited_fetch(cur, limit=5000)

        if not rows:
            return {
                "agent_id": aid,
                "days": days,
                "total_memories": 0,
                "clusters": [],
                "unclustered": 0,
            }

        # 解析标签 + 抽取关键词（停用词过滤）
        stop_words = {
            "的", "了", "和", "是", "就", "都", "而", "及", "与", "着",
            "或", "一个", "没有", "我们", "你们", "他们", "这个", "那个",
            "the", "a", "an", "and", "or", "is", "are", "was", "were",
            "to", "of", "in", "for", "on", "with", "at", "by",
        }

        mem_vectors: List[Tuple[str, set, float]] = []
        for r in rows:
            mid = r[0]
            content = (r[1] or "").lower()
            tags = []
            try:
                if r[2]:
                    tags = json.loads(r[2]) if isinstance(r[2], str) else list(r[2])
            except Exception:
                tags = []
            imp = (r[3] or "MEDIUM").upper()
            imp_w = {"LOW": 0.5, "MEDIUM": 1.0, "HIGH": 1.5, "CRITICAL": 2.0}.get(imp, 1.0)
            # 分词关键词：取 >=2 字的非停用词 tokens
            words = set()
            # 英文单词
            import re as _re
            for w in _re.findall(r'[a-zA-Z]{2,}', content):
                if w not in stop_words:
                    words.add(w)
            # 中文字符：2-4 字片段（简单滑动窗口）
            for w in _re.findall(r'[\u4e00-\u9fff]{2,4}', content):
                if w not in stop_words:
                    words.add(w)
            # 叠加标签
            for t in tags:
                tl = str(t).lower()
                if tl and tl not in stop_words:
                    words.add(tl)
            if words:
                mem_vectors.append((mid, words, imp_w))

        # 贪心聚类：Jaccard 相似度阈值
        clusters: List[Dict[str, Any]] = []
        clustered = set()
        SIM_THRESHOLD = 0.2

        for (mid, words, w) in mem_vectors:
            if mid in clustered:
                continue
            # 寻找最匹配的既有簇
            best_idx = -1
            best_sim = 0.0
            for i, cl in enumerate(clusters):
                if len(cl["members"]) == 0:
                    continue
                # 与簇核心词的 Jaccard
                core = cl["core_words"]
                inter = len(words & core)
                union = len(words | core)
                sim = (inter / union) if union > 0 else 0.0
                if sim > best_sim:
                    best_sim = sim
                    best_idx = i
            if best_idx >= 0 and best_sim >= SIM_THRESHOLD:
                # 加入簇
                cl = clusters[best_idx]
                cl["members"].append({"id": mid, "weight": w})
                cl["core_words"] |= words
                cl["total_weight"] += w
            else:
                # 新簇
                if len(clusters) >= max_clusters:
                    continue
                clusters.append({
                    "cluster_id": len(clusters) + 1,
                    "members": [{"id": mid, "weight": w}],
                    "core_words": set(words),
                    "total_weight": w,
                })
            clustered.add(mid)

        # 计算簇标签（高频核心词 Top-5）
        result_clusters = []
        for cl in clusters:
            word_scores: Dict[str, float] = {}
            # 成员中的词频（加权）
            for (mid, words, w) in mem_vectors:
                if mid in {m["id"] for m in cl["members"]}:
                    for wd in words:
                        word_scores[wd] = word_scores.get(wd, 0.0) + w
            # 也加一次 core_words 的贡献
            for wd in cl["core_words"]:
                word_scores[wd] = word_scores.get(wd, 0.0) + 0.1
            top_words = sorted(word_scores.items(), key=lambda x: -x[1])[:8]
            label = " / ".join(w for w, _ in top_words[:5]) or "(通用)"
            result_clusters.append({
                "cluster_id": cl["cluster_id"],
                "label": label,
                "size": len(cl["members"]),
                "total_weight": round(cl["total_weight"], 2),
                "top_words": [w for w, _ in top_words],
                "sample_members": [m["id"] for m in cl["members"][:10]],
            })

        result_clusters.sort(key=lambda c: -c["size"])
        for i, cl in enumerate(result_clusters):
            cl["cluster_id"] = i + 1

        return {
            "agent_id": aid,
            "days": days,
            "total_memories": len(rows),
            "clustered_memories": len(clustered),
            "clusters": result_clusters,
            "unclustered": max(0, len(rows) - len(clustered)),
        }

    def agent_insight(self,
                      agent_id: str,
                      days: int = 30) -> Dict[str, Any]:
        """Agent 行为洞察（v5.3.5 新增）

        综合分析 Agent 记忆的活跃度趋势、标签偏好、
        记忆层分布变化、访问频率等行为模式。

        Args:
            agent_id: Agent ID
            days: 回溯天数（1-365）

        Returns:
            行为洞察报告
        """
        conn = self._get_conn()
        aid = _filter_unicode_ctrl(agent_id[:128]) if isinstance(agent_id, str) else ""
        if not aid:
            return {"error": "Agent ID 不能为空"}

        days = max(1, min(365, int(days)))
        now = time.time()
        since = now - days * 86400

        # v5.3.5 安全：参数化 SQL
        cur = conn.execute(
            "SELECT id, content, tags, category, importance, layer, "
            "created_at, access_count, privacy "
            "FROM memories "
            "WHERE source_agent = ? AND category != 'trash' AND created_at >= ?",
            (aid, since)
        )
        rows = _limited_fetch(cur, limit=10000)

        if not rows:
            return {
                "agent_id": aid,
                "days": days,
                "total_memories": 0,
                "activity_trend": {},
                "tag_preferences": [],
                "layer_distribution": {},
                "insights": [],
            }

        total = len(rows)
        layers: Dict[str, int] = {}
        importances: Dict[str, int] = {}
        categories: Dict[str, int] = {}
        privacies: Dict[str, int] = {}
        tag_counts: Dict[str, int] = {}
        # 每周活跃度（按周切片）
        week_buckets: Dict[str, int] = {}
        W = 7 * 86400
        total_access = 0
        creation_timestamps: List[float] = []

        for r in rows:
            layer = (r[5] or "SHORT_TERM").upper()
            imp = (r[3] or "MEDIUM").upper()
            cat = (r[3] or "general").lower()
            priv = (r[7] if len(r) > 7 else "INTERNAL") or "INTERNAL"
            priv = priv.upper()
            layers[layer] = layers.get(layer, 0) + 1
            importances[imp] = importances.get(imp, 0) + 1
            categories[cat] = categories.get(cat, 0) + 1
            privacies[priv] = privacies.get(priv, 0) + 1
            # 访问计数
            ac = r[6] if isinstance(r[6], int) else 0
            total_access += ac
            # 创建时间
            created = r[5] if isinstance(r[5], (int, float)) else now
            if isinstance(r[5], (int, float)) and r[5] > 1e9:
                created = r[5]
            else:
                # 回退：按列索引判断 created_at
                created = r[5] if isinstance(r[5], (int, float)) and r[5] > 1e8 else now
            # created_at 索引修正：SELECT 中是第 6 列（index=5）？重新按顺序
            # 顺序是: id[0],content[1],tags[2],category[3],importance[4],layer[5],created_at[6],access_count[7],privacy[8]
            # 所以应该修正：
            pass

        # 重新整理统计，修正列索引
        layers.clear()
        importances.clear()
        categories.clear()
        privacies.clear()
        tag_counts.clear()
        week_buckets.clear()
        total_access = 0
        creation_timestamps = []
        avg_len = 0.0

        for r in rows:
            mid, content, tags, category, importance, layer, created_at, access_count, privacy = r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8]
            layer = (layer or "SHORT_TERM").upper()
            imp = (importance or "MEDIUM").upper()
            cat = (category or "general").lower()
            priv = (privacy or "INTERNAL").upper()
            layers[layer] = layers.get(layer, 0) + 1
            importances[imp] = importances.get(imp, 0) + 1
            categories[cat] = categories.get(cat, 0) + 1
            privacies[priv] = privacies.get(priv, 0) + 1
            ac = access_count if isinstance(access_count, int) else 0
            total_access += ac
            created = created_at if isinstance(created_at, (int, float)) else now
            creation_timestamps.append(created)
            bucket_idx = int((now - created) / W) if created < now else 0
            bucket_label = f"W-{bucket_idx}" if bucket_idx > 0 else "本周"
            week_buckets[bucket_label] = week_buckets.get(bucket_label, 0) + 1
            # 标签统计
            try:
                tag_list = json.loads(tags) if isinstance(tags, str) else (tags or [])
                for t in tag_list:
                    ts = str(t).strip()
                    if ts:
                        tag_counts[ts] = tag_counts.get(ts, 0) + 1
            except Exception:
                pass
            # 内容长度
            avg_len += len(content or "") / max(1, total)

        # 活跃度趋势：比较最近一半 vs 前一半
        sorted_ts = sorted(creation_timestamps)
        half = len(sorted_ts) // 2
        insights: List[str] = []
        if half > 0:
            first_half_count = half
            second_half_count = total - half
            if second_half_count > first_half_count * 1.2:
                insights.append("近期活跃度上升 📈（后段记忆多于前段）")
            elif first_half_count > second_half_count * 1.2:
                insights.append("近期活跃度下降 📉（前段记忆多于后段）")
            else:
                insights.append("活跃度保持平稳 ➡️")

        # 记忆层洞察
        pct = lambda n: (n / total * 100) if total > 0 else 0.0
        long_pct = pct(layers.get("LONG_TERM", 0) + layers.get("PERMANENT", 0))
        if long_pct >= 60:
            insights.append(f"长期记忆占比偏高（{long_pct:.0f}%），知识沉淀良好")
        elif long_pct <= 15:
            insights.append(f"长期记忆占比偏低（{long_pct:.0f}%），建议强化记忆持久化")

        # 重要度洞察
        high_pct = pct(importances.get("HIGH", 0) + importances.get("CRITICAL", 0))
        if high_pct >= 40:
            insights.append(f"重要记忆占比较高（{high_pct:.0f}%），记忆质量优")
        elif high_pct <= 5:
            insights.append(f"重要记忆稀少（{high_pct:.0f}%），可考虑标注重要记忆")

        # 隐私洞察
        secret_pct = pct(privacies.get("SECRET", 0) + privacies.get("CONFIDENTIAL", 0))
        if secret_pct >= 30:
            insights.append(f"敏感记忆占比 {secret_pct:.0f}%，注意加密备份")

        # 平均访问频次
        avg_access = round(total_access / total, 2) if total > 0 else 0
        if avg_access >= 3.0:
            insights.append(f"记忆复用率高（平均访问 {avg_access} 次），价值密度好")

        # Top 标签偏好
        top_tags = sorted(tag_counts.items(), key=lambda x: -x[1])[:10]
        tag_preferences = [
            {"tag": t, "count": c, "ratio": round(c / total, 4)}
            for t, c in top_tags
        ]

        return {
            "agent_id": aid,
            "days": days,
            "total_memories": total,
            "activity": {
                "trend_by_week": week_buckets,
                "total_accesses": total_access,
                "avg_access_per_memory": avg_access,
                "avg_content_length": round(avg_len, 1),
            },
            "layer_distribution": layers,
            "importance_distribution": importances,
            "category_distribution": categories,
            "privacy_distribution": privacies,
            "tag_preferences": tag_preferences,
            "insights": insights,
        }

    # ===== v5.3.5 新增：AI 短剧剧情摘要 + 场景张力分析 =====

    def drama_summary(self,
                      drama_id: str,
                      max_length: int = 500) -> Dict[str, Any]:
        """短剧剧情摘要（v5.3.5 新增）

        基于场景描述和经典台词，生成短剧核心剧情摘要。

        Args:
            drama_id: 短剧 ID
            max_length: 摘要最大字符数（100-2000）

        Returns:
            剧情摘要、核心角色、关键场景索引
        """
        conn = self._get_conn()
        did = _filter_unicode_ctrl(drama_id[:64]) if isinstance(drama_id, str) and drama_id else ""
        if not did:
            return {"error": "短剧 ID 不能为空"}
        max_length = max(100, min(2000, int(max_length)))

        # 获取短剧元数据
        drow = conn.execute(
            "SELECT id, title, genre, total_episodes, current_episode, "
            "status, rating, description, metadata "
            "FROM drama_series WHERE id = ?",
            (did,)
        ).fetchone()
        if not drow:
            return {"error": "短剧不存在"}

        title = drow[1] or "未命名"
        genre = drow[2] or "未知"
        stored_summary = (drow[7] or "").strip()

        # 获取场景列表（按场景顺序）
        scene_rows = conn.execute(
            "SELECT id, episode, scene_number, title, content, metadata "
            "FROM drama_scenes WHERE drama_id = ? "
            "ORDER BY episode ASC, scene_number ASC",
            (did,)
        ).fetchall()

        # 获取经典台词
        classic_lines = conn.execute(
            "SELECT dl.line_text, dc.name, ds.episode "
            "FROM drama_lines dl "
            "LEFT JOIN drama_characters dc ON dl.character_id = dc.id "
            "LEFT JOIN drama_scenes ds ON dl.scene_id = ds.id "
            "WHERE dl.drama_id = ? AND dl.is_classic = 1 "
            "ORDER BY COALESCE(ds.episode, 0) ASC, COALESCE(ds.scene_number, 0) ASC, dl.created_at ASC",
            (did,)
        ).fetchall()

        # 获取主要角色（按台词量排序 Top-5）
        top_chars = conn.execute(
            "SELECT dc.name, COUNT(dl.id) as line_count "
            "FROM drama_characters dc "
            "LEFT JOIN drama_lines dl ON dc.id = dl.character_id "
            "WHERE dc.drama_id = ? "
            "GROUP BY dc.id ORDER BY line_count DESC LIMIT 5",
            (did,)
        ).fetchall()

        scenes = []
        key_scene_ids = []
        for sr in scene_rows:
            # v5.3.5: is_key_scene 存储在 metadata JSON 中
            meta_raw = sr[5] or "{}"
            try:
                meta = _safe_json_loads(meta_raw)
                is_key = bool(meta.get("is_key_scene", False))
            except Exception:
                is_key = False
            sc = {
                "id": sr[0],
                "episode": sr[1],
                "order": sr[2],
                "title": sr[3] or "",
                "description": (sr[4] or "")[:200],
                "is_key": is_key,
            }
            scenes.append(sc)
            if is_key:
                key_scene_ids.append(sc["id"])

        # 生成摘要（优先级：已有 summary > 关键场景描述拼接 > 普通场景拼接）
        parts = []
        if stored_summary:
            parts.append(stored_summary)
        else:
            # 优先关键场景
            key_scenes = [s for s in scenes if s["is_key"] and s["description"]]
            # 若关键场景太少，取首尾 + 中间分布
            if len(key_scenes) < 3:
                key_scenes = (
                    scenes[:1]
                    + scenes[len(scenes) // 3 : len(scenes) // 3 + 1]
                    + scenes[2 * len(scenes) // 3 : 2 * len(scenes) // 3 + 1]
                    + scenes[-1:]
                )
                key_scenes = [s for s in key_scenes if s.get("description")]
            for s in key_scenes:
                ep = f"第{s['episode']}集" if s.get("episode") else ""
                st = s.get("title") or f"场景{s.get('order') or ''}"
                desc = s["description"].strip().rstrip(".。")
                if desc:
                    parts.append(f"{ep}{st}：{desc}")

        # 添加经典台词（最多3句）
        quotes = []
        for cl in classic_lines[:3]:
            content, cname, ep = cl
            content = (content or "").strip()
            if len(content) > 80:
                content = content[:80] + "…"
            if content:
                who = cname or "角色"
                tag = f"（第{ep}集）" if ep else ""
                quotes.append(f"{who}{tag}：“{content}”")

        # 组装摘要，限制长度
        summary_text = "。".join(p for p in parts if p)
        if not summary_text:
            summary_text = f"{title}：暂无剧情描述数据"
        # 截断到 max_length（中文按字符截断更友好）
        if len(summary_text) > max_length:
            summary_text = summary_text[:max_length].rstrip() + "…"

        # 角色列表
        characters = [{"name": c[0], "lines": c[1] or 0} for c in top_chars if c[0]]

        return {
            "drama_id": did,
            "title": title,
            "genre": genre,
            "episodes": drow[3] or 0,
            "current_episode": drow[4] or 0,
            "status": drow[5] or "planned",
            "rating": drow[6] or 0.0,
            "summary": summary_text,
            "summary_source": "stored" if stored_summary else "derived",
            "characters": characters,
            "classic_quotes": quotes,
            "total_scenes": len(scenes),
            "key_scene_count": len(key_scene_ids),
        }

    def scene_tension(self,
                      drama_id: str,
                      top_k: int = 10) -> Dict[str, Any]:
        """场景张力分析（v5.3.5 新增）

        基于台词量、冲突词、角色数量等，识别高张力场景（冲突/高潮）。
        张力评分 = 台词量分 + 冲突关键词分 + 角色互动分 + 关键场景加成

        Args:
            drama_id: 短剧 ID
            top_k: 返回 Top-K 高张力场景（1-50）

        Returns:
            张力排行、各场景张力曲线、高潮场景索引
        """
        conn = self._get_conn()
        did = _filter_unicode_ctrl(drama_id[:64]) if isinstance(drama_id, str) and drama_id else ""
        if not did:
            return {"error": "短剧 ID 不能为空"}
        top_k = max(1, min(50, int(top_k)))

        # 获取短剧信息
        drow = conn.execute(
            "SELECT id, title, total_episodes FROM drama_series WHERE id = ?",
            (did,)
        ).fetchone()
        if not drow:
            return {"error": "短剧不存在"}
        title = drow[1] or "未命名"

        # 获取所有场景及其基础信息
        scene_rows = conn.execute(
            "SELECT id, episode, scene_number, title, content, metadata "
            "FROM drama_scenes WHERE drama_id = ? "
            "ORDER BY episode ASC, scene_number ASC",
            (did,)
        ).fetchall()

        if not scene_rows:
            return {
                "drama_id": did,
                "title": title,
                "total_scenes": 0,
                "tension_curve": [],
                "top_tension_scenes": [],
            }

        # 冲突关键词词典
        conflict_words = {
            "不", "别", "没", "错", "反对", "拒绝", "但是", "可是", "然而", "偏偏",
            "竟然", "居然", "为什么", "凭什么", "滚", "闭嘴", "打", "杀", "死",
            "no", "not", "never", "but", "however", "yet", "stop", "hate",
            "fight", "kill", "die", "attack", "angry", "rage", "damn",
            "冲突", "争吵", "吵架", "矛盾", "对抗", "崩溃", "爆炸", "危险",
            "陷阱", "阴谋", "背叛", "欺骗", "威胁", "警告", "逼迫", "绝境",
        }

        # 情感激烈词（提升张力）
        intensity_words = {
            "最", "非常", "极", "绝对", "立刻", "马上", "必须", "一定",
            "恨", "爱", "绝望", "希望", "真相", "秘密", "永远", "最后",
            "very", "most", "absolutely", "always", "forever", "final",
            "truth", "secret", "last", "never", "love", "hate",
            "惊喜", "震惊", "震撼", "奇迹", "命运", "抉择", "牺牲",
        }

        tension_curve = []
        scene_stats = []
        max_lines = 1

        # 对每个场景，统计台词、冲突词、角色数
        for sr in scene_rows:
            sid = sr[0]
            # 台词量
            line_data = conn.execute(
                "SELECT COUNT(*), COUNT(DISTINCT character_id) "
                "FROM drama_lines WHERE scene_id = ?",
                (sid,)
            ).fetchone()
            line_count = line_data[0] or 0
            char_count = line_data[1] or 0
            if line_count > max_lines:
                max_lines = line_count

            # 获取场景内台词文本 + 描述 合并做冲突检测
            texts = [sr[3] or "", sr[4] or ""]
            cur_l = conn.execute(
                "SELECT line_text FROM drama_lines WHERE scene_id = ? ORDER BY created_at ASC",
                (sid,)
            )
            for lr in _limited_fetch(cur_l, limit=500):
                texts.append(lr[0] or "")
            blob = " ".join(texts).lower()

            conflict_hits = sum(1 for w in conflict_words if w in blob)
            intensity_hits = sum(1 for w in intensity_words if w in blob)

            # 读取 is_key_scene 从 metadata
            meta_raw = sr[5] or "{}"
            try:
                meta_s = _safe_json_loads(meta_raw)
                is_key = 1 if meta_s.get("is_key_scene") else 0
            except Exception:
                is_key = 0

            # 张力分
            line_score = (line_count / max_lines) * 40.0
            char_score = min(char_count, 8) / 8.0 * 20.0
            conflict_score = min(conflict_hits, 10) / 10.0 * 25.0
            intensity_score = min(intensity_hits, 10) / 10.0 * 15.0
            key_bonus = 5.0 if is_key else 0.0

            tension = round(line_score + char_score + conflict_score + intensity_score + key_bonus, 2)
            entry = {
                "scene_id": sid,
                "episode": sr[1],
                "order": sr[2],
                "scene_title": sr[3] or f"场景{sr[2] or ''}",
                "is_key_scene": is_key,
                "line_count": line_count,
                "character_count": char_count,
                "conflict_hits": conflict_hits,
                "intensity_hits": intensity_hits,
                "tension": tension,
            }
            scene_stats.append(entry)
            tension_curve.append({
                "scene_id": sid,
                "episode": sr[1],
                "order": sr[2],
                "tension": tension,
            })

        # Top-K 高张力场景
        top_scenes = sorted(scene_stats, key=lambda s: -s["tension"])[:top_k]

        # 找高潮区间：连续 >=60 分的场景段
        threshold = 60.0 if len(scene_stats) >= 5 else 40.0
        climax_ranges = []
        cur_start = None
        for i, s in enumerate(scene_stats):
            if s["tension"] >= threshold:
                if cur_start is None:
                    cur_start = i
            else:
                if cur_start is not None:
                    climax_ranges.append((cur_start, i - 1))
                    cur_start = None
        if cur_start is not None:
            climax_ranges.append((cur_start, len(scene_stats) - 1))
        # 合并成摘要
        climax_summary = []
        for a, b in climax_ranges:
            ep_start = scene_stats[a].get("episode")
            ep_end = scene_stats[b].get("episode")
            scenes_str = f"{scene_stats[a]['scene_title']} → {scene_stats[b]['scene_title']}"
            climax_summary.append({
                "from_index": a,
                "to_index": b,
                "episodes": f"{ep_start}-{ep_end}" if ep_start != ep_end else str(ep_start),
                "description": scenes_str,
                "peak_tension": max(scene_stats[i]["tension"] for i in range(a, b + 1)),
            })
        # 取最长/峰值最高的一段为主高潮
        main_climax = None
        if climax_summary:
            main_climax = max(climax_summary, key=lambda c: c["peak_tension"])

        return {
            "drama_id": did,
            "title": title,
            "total_scenes": len(scene_stats),
            "avg_tension": round(sum(s["tension"] for s in scene_stats) / max(1, len(scene_stats)), 2),
            "tension_curve": tension_curve,
            "top_tension_scenes": top_scenes,
            "climax_segments": climax_summary,
            "main_climax": main_climax,
        }

    def analyze_similarity(self,
                           memory_id: str,
                           limit: int = 10,
                           min_similarity: float = 0.3) -> List[Dict[str, Any]]:
        """相似度分析（v5.2.2 新增）

        分析指定记忆与其他记忆的相似度。

        Args:
            memory_id: 目标记忆 ID
            limit: 返回数量
            min_similarity: 最低相似度阈值

        Returns:
            相似记忆列表，包含相似度分数
        """
        entry = self.get_memory(memory_id)
        if not entry:
            return []

        conn = self._get_conn()

        # 使用 FTS5 全文搜索找相似内容
        # v5.5.4 fix: 排除软删除记忆，防止回收站内容泄露
        try:
            rows = conn.execute(
                "SELECT m.id, m.content, m.category, m.layer, m.importance, m.starred, "
                "bm25(memory_fts) as relevance "
                "FROM memory_fts "
                "JOIN memories m ON memory_fts.rowid = m.rowid "
                "WHERE memory_fts MATCH ? AND m.id != ? AND m.category != 'trash' "
                "ORDER BY relevance "
                "LIMIT ?",
                (entry.content[:200], memory_id, limit * 2)
            ).fetchall()
        except sqlite3.OperationalError:
            # FTS 搜索失败，回退到 LIKE 搜索
            keywords = entry.content.split()[:5]
            if not keywords:
                return []

            query = "SELECT id, content, category, layer, importance, starred FROM memories WHERE id != ? AND category != 'trash' AND ("
            params = [memory_id]
            conditions = []
            for kw in keywords[:3]:
                # v5.3.3 安全加固：LIKE 通配符转义
                conditions.append("content LIKE ? ESCAPE '\\'")
                params.append(f"%{_escape_like(kw)}%")
            query += " OR ".join(conditions) + ") LIMIT ?"
            params.append(limit * 2)

            rows = conn.execute(query, params).fetchall()

        results = []
        for row in rows:
            other_id = row[0]
            other_content = row[1]

            # 计算相似度（基于内容重叠）
            similarity = self._calculate_similarity(entry.content, other_content)

            if similarity >= min_similarity:
                results.append({
                    "memory_id": other_id,
                    "similarity": round(similarity, 3),
                    "content_preview": other_content[:100] + "..." if len(other_content) > 100 else other_content,
                    "category": row[2],
                    "layer": row[3],
                    "importance": row[4],
                    "starred": bool(row[5]),
                })

        # 按相似度排序
        results.sort(key=lambda x: x["similarity"], reverse=True)
        return results[:limit]

    def _calculate_similarity(self, text1: str, text2: str) -> float:
        """计算文本相似度（简化版 Jaccard 相似度）"""
        if not text1 or not text2:
            return 0.0

        # 简单的词集合相似度
        words1 = set(text1.lower().split())
        words2 = set(text2.lower().split())

        if not words1 or not words2:
            return 0.0

        intersection = len(words1 & words2)
        union = len(words1 | words2)

        if union == 0:
            return 0.0

        return intersection / union

    def batch_quality_score(self,
                            category: Optional[str] = None,
                            limit: int = 100) -> Dict[str, Any]:
        """批量质量评分（v5.2.2 新增）

        对指定范围内的记忆进行批量质量评分。

        Args:
            category: 分类过滤
            limit: 数量限制

        Returns:
            批量评分结果，包含统计信息
        """
        entries = self.list_memories(category=category, limit=limit)
        scores = []
        grades = {"优秀": 0, "良好": 0, "中等": 0, "及格": 0, "需改进": 0}

        for entry in entries:
            score_data = self.quality_score(entry.id)
            if score_data:
                scores.append(score_data)
                grades[score_data["grade"]] = grades.get(score_data["grade"], 0) + 1

        if not scores:
            return {"total": 0, "average_score": 0, "grades": grades, "scores": []}

        avg_score = sum(s["total_score"] for s in scores) / len(scores)

        return {
            "total": len(scores),
            "average_score": round(avg_score, 1),
            "grades": grades,
            "top_scores": sorted(scores, key=lambda x: x["total_score"], reverse=True)[:10],
            "low_scores": sorted(scores, key=lambda x: x["total_score"])[:10],
        }

    def get_audit_log(self,
                      memory_id: Optional[str] = None,
                      actor: Optional[str] = None,
                      limit: int = 100) -> List[AuditRecord]:
        """获取审计日志"""
        conn = self._get_conn()
        query = "SELECT * FROM audit_log WHERE 1=1"
        params = []

        if memory_id:
            query += " AND memory_id = ?"
            params.append(memory_id)
        if actor:
            query += " AND actor = ?"
            params.append(actor)

        query += " ORDER BY timestamp DESC LIMIT ?"
        params.append(limit)

        rows = conn.execute(query, params).fetchall()
        return [AuditRecord(
            id=row["id"],
            action=row["action"],
            memory_id=row["memory_id"],
            actor=row["actor"],
            session_id=row["session_id"],
            privacy_level=row["privacy_level"],
            timestamp=row["timestamp"],
            details=self._safe_json_loads(row["details"], {}),
        ) for row in rows]

    def backup(self, backup_dir: str) -> Path:
        """备份数据库"""
        backup_path = Path(backup_dir)
        backup_path.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        dest = backup_path / f"memory_backup_{timestamp}.db"

        import shutil
        shutil.copy2(self.db_path, dest)
        return dest

    @staticmethod
    def _safe_json_loads(data: Optional[str], default: Any = None) -> Any:
        """安全解析 JSON，损坏时返回默认值（v5.2.3 安全加固）"""
        if not data:
            return default
        try:
            return json.loads(data)
        except (json.JSONDecodeError, TypeError):
            return default

    def _row_to_entry(self, row: sqlite3.Row) -> MemoryEntry:
        # v5.3.7 修复：tags/metadata 可能已经是 list/dict（非 JSON 字符串）
        raw_tags = row["tags"]
        if isinstance(raw_tags, (list, tuple)):
            tags_val = list(raw_tags)
        elif isinstance(raw_tags, str):
            try:
                tags_val = self._safe_json_loads(raw_tags, [])
            except (ValueError, json.JSONDecodeError):
                tags_val = []
        else:
            tags_val = []
        raw_meta = row["metadata"]
        if isinstance(raw_meta, dict):
            meta_val = raw_meta
        elif isinstance(raw_meta, str):
            try:
                meta_val = self._safe_json_loads(raw_meta, {})
            except (ValueError, json.JSONDecodeError):
                meta_val = {}
        else:
            meta_val = {}
        # v5.5.5 fix: 统一走 _downgrade_enum，兼容双重导入路径下的枚举类不一致
        return MemoryEntry(
            id=row["id"],
            content=row["content"] or "",
            category=row["category"],
            tags=tags_val,
            privacy=self._downgrade_enum(row["privacy"], PrivacyLevel, PrivacyLevel.INTERNAL),
            importance=self._downgrade_enum(row["importance"], Importance, Importance.MEDIUM),
            memory_type=self._downgrade_enum(row["memory_type"], MemoryType, MemoryType.TEXT),
            layer=self._downgrade_enum(row["layer"], MemoryLayer, MemoryLayer.SHORT_TERM),
            source_session=row["source_session"],
            source_agent=row["source_agent"],
            created_at=row["created_at"],
            updated_at=row["updated_at"],
            last_accessed_at=row["last_accessed_at"],
            access_count=row["access_count"],
            consolidation_count=row["consolidation_count"],
            forgetting_score=row["forgetting_score"],
            strength=row["strength"],
            starred=bool(row["starred"]) if "starred" in row.keys() else False,
            pinned=bool(row["pinned"]) if "pinned" in row.keys() else False,
            expires_at=float(row["expires_at"]) if "expires_at" in row.keys() and row["expires_at"] else 0.0,
            metadata=meta_val,
            encrypted=bool(row["encrypted"]),
            ciphertext=row["ciphertext"],
            nonce=row["nonce"],
            salt=row["salt"],
        )

    def _update_access(self, entry: MemoryEntry, actor: str, session_id: str):
        """更新访问计数（v5.1.2 优化：不再每次 get 都写审计日志，减轻低配电脑负担）"""
        conn = self._get_conn()
        now = time.time()
        conn.execute("""
            UPDATE memories SET access_count = access_count + 1, last_accessed_at = ?
            WHERE id = ?
        """, (now, entry.id))
        conn.commit()

    def _add_audit(self, action: str, memory_id: str, actor: str,
                   session_id: str, privacy_level: str, details: Optional[dict] = None):
        """添加审计记录

        v5.4.2 安全修复：审计写入失败不再静默吞没，改为 logging.error 记录。
        对 delete/purge/grant/revoke 等高敏操作，审计失败时抛出异常（fail-closed），
        防止在无审计记录下完成敏感操作。
        """
        # v5.4.0 安全加固：所有字段控制字符过滤 + 长度限制，防御审计日志污染
        ACTION_WHITELIST = {"add", "update", "delete", "restore", "purge", "export", "import",
                          "grant", "revoke", "merge", "access", "consolidate",
                          "forget", "share", "accept", "reject",
                          # v5.4.2 联邦 ACL + 共享冲突审计动作
                          "acl_deny", "acl_add_rule", "acl_remove_rule",
                          "conflict_detected", "conflict_resolved", "conflict_dismiss",
                          # v5.4.2 agent 高敏操作变体
                          "agent_purge", "agent_forget", "agent_merge", "agent_clean",
                          # v5.5.5 fix: agent_transfer 加入白名单，避免被降级为 other
                          "agent_transfer",
                          # v5.5.0 新增审计动作
                          "deduplicate_merge", "recalibrate", "reinforce",
                          # v5.5.5 fix: evolve 加入白名单
                          "evolve"}
        # v5.4.2：高敏感操作，审计失败时 fail-closed
        HIGH_SENSITIVE_ACTIONS = {"delete", "purge", "grant", "revoke", "forget",
                                  "agent_purge", "agent_forget"}
        if action not in ACTION_WHITELIST:
            action = "other"  # 非白名单降级为 other
        memory_id = self._strip_control(str(memory_id))[:64]
        actor = self._strip_control(str(actor))[:128]
        session_id = self._strip_control(str(session_id))[:128]
        privacy_level = self._strip_control(str(privacy_level))[:32]
        clean_details = self._sanitize_metadata(details or {}, max_depth=3, max_string_len=500)

        conn = self._get_conn()
        record_id = str(uuid.uuid4())
        now = time.time()
        try:
            conn.execute("""
                INSERT INTO audit_log (id, action, memory_id, actor, session_id, privacy_level, timestamp, details)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                record_id, action, memory_id, actor, session_id,
                privacy_level, now, json.dumps(clean_details, ensure_ascii=False)
            ))
            conn.commit()
        except sqlite3.Error as e:
            logger.error("审计日志写入失败 action=%s memory_id=%s error=%s", action, memory_id, e)
            if action in HIGH_SENSITIVE_ACTIONS:
                raise SecurityError(f"高敏感操作审计失败，拒绝执行: {action}") from e
        except Exception as e:
            logger.error("审计日志写入异常 action=%s memory_id=%s error=%s", action, memory_id, e)
            if action in HIGH_SENSITIVE_ACTIONS:
                raise SecurityError(f"高敏感操作审计失败，拒绝执行: {action}") from e

    def _close_conns(self):
        """关闭当前线程持有的 SQLite 连接（v5.4.6 线程安全改造）"""
        conn = getattr(self._conn_local, "conn", None)
        if conn is not None:
            try:
                conn.close()
            except sqlite3.Error:
                pass
            self._conn_local.conn = None

    def close(self):
        self._close_conns()

    def cleanup_expired(self, max_age_hours: int = 24, layer: str = "sensory") -> int:
        """清理过期记忆（v5.1.3 新增）

        Args:
            max_age_hours: 最大保留时长（小时），超过此时间的记忆将被软删除
            layer: 记忆层级（sensory/short_term/long_term/permanent）

        Returns:
            被清理的记忆数量
        """
        conn = self._get_conn()
        now = time.time()
        cutoff_time = now - (max_age_hours * 3600)

        rows = conn.execute("""
            SELECT id, category FROM memories
            WHERE layer = ? AND category != 'trash' AND created_at < ?
        """, (layer, cutoff_time)).fetchall()

        if not rows:
            return 0

        deleted_count = 0
        for row in rows:
            entry_id = row["id"]
            old_category = row["category"]

            try:
                conn.execute("""
                    UPDATE memories SET
                        category = 'trash',
                        metadata = JSON_SET(metadata, '$.original_category', ?),
                        updated_at = ?
                    WHERE id = ?
                """, (old_category, now, entry_id))

                if not self.encrypted:
                    conn.execute("""
                        INSERT INTO memory_fts(memory_fts, rowid, content, category, tags)
                        VALUES('delete', (SELECT rowid FROM memories WHERE id = ?), '', '', '')
                    """, (entry_id,))

                deleted_count += 1
            except (sqlite3.OperationalError, sqlite3.IntegrityError):
                pass

        conn.commit()
        return deleted_count

    def batch_add(self, entries: List[Dict[str, Any]]) -> int:
        """批量添加记忆（v5.1.3 新增，v5.5.6 增强：支持 expires_at/pinned/metadata）

        Args:
            entries: 记忆条目列表，每个条目包含 content、category、tags 等字段

        Returns:
            成功添加的记忆数量
        """
        conn = self._get_conn()
        now = time.time()
        added_count = 0

        for entry_data in entries:
            try:
                entry_id = str(uuid.uuid4())
                content = entry_data.get("content", "")
                # v5.4.1 修复：batch_add 此前绕过内容长度校验，可批量注入超长内容（DoS）
                _validate_content_len(content)
                category = entry_data.get("category", "general")
                tags = entry_data.get("tags", [])
                privacy_str = entry_data.get("privacy", "internal")
                importance_str = entry_data.get("importance", "medium")
                layer_str = entry_data.get("layer", "short_term")
                # v5.5.6 新增：支持 expires_at 和 pinned
                expires_at = float(entry_data.get("expires_at", 0.0) or 0.0)
                pinned = 1 if entry_data.get("pinned", False) else 0
                starred = 1 if entry_data.get("starred", False) else 0
                metadata = entry_data.get("metadata", {})
                if not isinstance(metadata, dict):
                    metadata = {}

                ciphertext = None
                nonce = None
                salt = None
                stored_content = content

                if self.encrypted and self.encryption:
                    blob = self.encryption.encrypt(content)
                    ciphertext = blob.ciphertext
                    nonce = blob.nonce
                    salt = blob.salt
                    stored_content = ""

                conn.execute("""
                    INSERT INTO memories (
                        id, content, ciphertext, nonce, salt, category, tags,
                        privacy, importance, memory_type, layer,
                        source_session, source_agent, created_at, updated_at,
                        last_accessed_at, access_count, consolidation_count,
                        forgetting_score, strength, starred, pinned, metadata, encrypted, expires_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    entry_id, stored_content, ciphertext, nonce, salt,
                    category, json.dumps(tags, ensure_ascii=False),
                    privacy_str, importance_str, "text",
                    layer_str, "", "",
                    now, now, now,
                    0, 0, 0.0, 1.0, starred, pinned,
                    json.dumps(metadata, ensure_ascii=False), int(self.encrypted),
                    expires_at
                ))

                if not self.encrypted:
                    conn.execute("""
                        INSERT INTO memory_fts (rowid, content, category, tags)
                        VALUES ((SELECT rowid FROM memories WHERE id = ?), ?, ?, ?)
                    """, (entry_id, content, category, json.dumps(tags, ensure_ascii=False)))

                added_count += 1
            except (sqlite3.OperationalError, sqlite3.IntegrityError, ValueError):
                # v5.4.1：ValueError（内容超长等校验失败）按单条失败处理，不中断批量
                pass

        conn.commit()
        return added_count

    def find_similar(self, content: str, limit: int = 5, threshold: float = 0.3) -> List[MemoryEntry]:
        """查找相似记忆（v5.1.3 新增）

        使用简单的 Jaccard 相似度匹配，在低配电脑上也能快速运行

        Args:
            content: 参考内容
            limit: 返回数量限制
            threshold: 相似度阈值（0-1）

        Returns:
            相似记忆列表（按相似度降序）
        """
        import re

        conn = self._get_conn()
        rows = conn.execute("""
            SELECT * FROM memories WHERE category != 'trash' AND encrypted = 0
        """).fetchall()

        def jaccard_similarity(s1: str, s2: str) -> float:
            words1 = set(re.findall(r'\w+', s1.lower()))
            words2 = set(re.findall(r'\w+', s2.lower()))
            if not words1 and not words2:
                return 0.0
            intersection = words1 & words2
            union = words1 | words2
            return len(intersection) / len(union)

        similarities = []
        for row in rows:
            entry_content = row["content"] or ""
            if entry_content:
                sim = jaccard_similarity(content, entry_content)
                if sim >= threshold:
                    similarities.append((sim, row))

        similarities.sort(key=lambda x: x[0], reverse=True)
        results = []
        for sim, row in similarities[:limit]:
            entry = self._row_to_entry(row)
            if self.encrypted and self.encryption:
                entry.content = self.encryption.decrypt(
                    EncryptedBlob(ciphertext=entry.ciphertext, nonce=entry.nonce, salt=entry.salt)
                )
            results.append(entry)

        return results

    def get_detailed_stats(self) -> Dict[str, Any]:
        """获取详细统计信息（v5.1.4 新增）"""
        conn = self._get_conn()

        stats = {}

        row = conn.execute("SELECT COUNT(*) FROM memories WHERE category != 'trash'").fetchone()
        stats["total"] = row[0] if row else 0

        row = conn.execute("SELECT COUNT(*) FROM memories WHERE category = 'trash'").fetchone()
        stats["trash"] = row[0] if row else 0

        row = conn.execute("SELECT COUNT(*) FROM memories WHERE starred = 1 AND category != 'trash'").fetchone()
        stats["starred"] = row[0] if row else 0

        row = conn.execute("SELECT COUNT(*) FROM memories WHERE encrypted = 1 AND category != 'trash'").fetchone()
        stats["encrypted"] = row[0] if row else 0

        row = conn.execute("SELECT MIN(created_at), MAX(created_at) FROM memories WHERE category != 'trash'").fetchone()
        if row and row[0]:
            stats["first_created"] = row[0]
            stats["last_created"] = row[1]
        else:
            stats["first_created"] = None
            stats["last_created"] = None

        row = conn.execute("SELECT AVG(access_count), AVG(strength), AVG(forgetting_score) FROM memories WHERE category != 'trash'").fetchone()
        if row:
            stats["avg_access_count"] = round(row[0], 2) if row[0] else 0
            stats["avg_strength"] = round(row[1], 2) if row[1] else 0
            stats["avg_forgetting_score"] = round(row[2], 2) if row[2] else 0

        row = conn.execute("SELECT MAX(access_count), MIN(strength) FROM memories WHERE category != 'trash'").fetchone()
        if row:
            stats["max_access_count"] = row[0] if row[0] else 0
            stats["min_strength"] = round(row[1], 2) if row[1] else 0

        stats["by_category"] = {}
        rows = conn.execute("""
            SELECT category, COUNT(*) FROM memories WHERE category != 'trash' GROUP BY category ORDER BY COUNT(*) DESC
        """).fetchall()
        for cat, cnt in rows:
            stats["by_category"][cat] = cnt

        stats["by_layer"] = {}
        rows = conn.execute("""
            SELECT layer, COUNT(*) FROM memories WHERE category != 'trash' GROUP BY layer ORDER BY COUNT(*) DESC
        """).fetchall()
        for lay, cnt in rows:
            stats["by_layer"][lay] = cnt

        stats["by_privacy"] = {}
        rows = conn.execute("""
            SELECT privacy, COUNT(*) FROM memories WHERE category != 'trash' GROUP BY privacy ORDER BY COUNT(*) DESC
        """).fetchall()
        for pri, cnt in rows:
            stats["by_privacy"][pri] = cnt

        stats["by_importance"] = {}
        rows = conn.execute("""
            SELECT importance, COUNT(*) FROM memories WHERE category != 'trash' GROUP BY importance ORDER BY COUNT(*) DESC
        """).fetchall()
        for imp, cnt in rows:
            stats["by_importance"][imp] = cnt

        row = conn.execute("SELECT COUNT(*) FROM audit_log").fetchone()
        stats["audit_records"] = row[0] if row else 0

        return stats

    def get_random_memories(self, count: int = 1,
                            category: Optional[str] = None,
                            layer: Optional[MemoryLayer] = None,
                            min_strength: Optional[float] = None) -> List[MemoryEntry]:
        """随机获取记忆（v5.1.7 新增）

        Args:
            count: 随机记忆数量
            category: 分类筛选
            layer: 记忆层级筛选
            min_strength: 最低记忆强度筛选

        Returns:
            随机记忆列表
        """
        conn = self._get_conn()
        query = "SELECT * FROM memories WHERE category != 'trash'"
        params = []

        if category:
            query += " AND category = ?"
            params.append(category)
        if layer:
            query += " AND layer = ?"
            params.append(layer.value)
        if min_strength is not None:
            query += " AND strength >= ?"
            params.append(min_strength)

        query += " ORDER BY RANDOM() LIMIT ?"
        params.append(count)

        rows = conn.execute(query, params).fetchall()
        results = [self._row_to_entry(row) for row in rows]

        if self.encrypted and self.encryption:
            for entry in results:
                try:
                    entry.content = self.encryption.decrypt(
                        EncryptedBlob(ciphertext=entry.ciphertext, nonce=entry.nonce, salt=entry.salt)
                    )
                except (ValueError, TypeError):
                    pass

        return results

    def rename_tag(self, old_tag: str, new_tag: str) -> int:
        """重命名标签（v5.1.7 新增，v5.5.6 修复：大小写不敏感 + 空值防御）

        Args:
            old_tag: 旧标签名
            new_tag: 新标签名

        Returns:
            受影响的记忆条数
        """
        # v5.5.6 fix: 空值防御
        if not old_tag or not new_tag or not isinstance(old_tag, str) or not isinstance(new_tag, str):
            return 0
        old_tag = old_tag.strip()
        new_tag = new_tag.strip()
        if not old_tag or not new_tag:
            return 0

        conn = self._get_conn()
        count = 0

        rows = conn.execute("SELECT id, tags FROM memories WHERE category != 'trash'").fetchall()
        for row in rows:
            tags = self._safe_json_loads(row["tags"], [])
            # v5.5.6 fix: 大小写不敏感匹配，但保留其他标签的原始大小写
            new_tags = []
            changed = False
            for t in tags:
                if isinstance(t, str) and t.lower() == old_tag.lower():
                    new_tags.append(new_tag)
                    changed = True
                else:
                    new_tags.append(t)
            if changed:
                # 去重（重命名后可能产生重复标签）
                seen = set()
                deduped = []
                for t in new_tags:
                    key = t.lower() if isinstance(t, str) else str(t)
                    if key not in seen:
                        seen.add(key)
                        deduped.append(t)
                conn.execute("UPDATE memories SET tags = ?, updated_at = ? WHERE id = ?",
                             (json.dumps(deduped, ensure_ascii=False), time.time(), row["id"]))
                count += 1

        conn.commit()
        return count

    def rename_category(self, old_cat: str, new_cat: str) -> int:
        """重命名分类（v5.1.7 新增）

        Args:
            old_cat: 旧分类名
            new_cat: 新分类名

        Returns:
            受影响的记忆条数
        """
        conn = self._get_conn()
        now = time.time()

        cursor = conn.execute("""
            UPDATE memories SET category = ?, updated_at = ?
            WHERE category = ? AND category != 'trash'
        """, (new_cat, now, old_cat))

        count = cursor.rowcount
        conn.commit()
        return count

    def get_config_summary(self) -> Dict[str, Any]:
        """获取配置摘要（v5.1.7 新增）"""
        return {
            "db_path": str(self.db_path),
            "encrypted": self.encrypted,
            "db_size_mb": round(Path(self.db_path).stat().st_size / (1024 * 1024), 2) if Path(self.db_path).exists() else 0,
        }

    # ===== 数据库迁移（v5.1.8 新增）=====

    _LATEST_DB_VERSION = 1

    def get_db_version(self) -> int:
        """获取当前数据库版本（v5.1.8 新增）"""
        conn = self._get_conn()
        try:
            row = conn.execute(
                "SELECT value FROM schema_version WHERE id = 1"
            ).fetchone()
            return int(row[0]) if row else 0
        except sqlite3.OperationalError:
            return 0

    def get_latest_db_version(self) -> int:
        """获取最新数据库版本（v5.1.8 新增）"""
        return self._LATEST_DB_VERSION

    def migrate_to_latest(self) -> Dict[str, Any]:
        """迁移数据库到最新版本（v5.1.8 新增）"""
        import time as _time
        start = _time.time()
        conn = self._get_conn()
        current = self.get_db_version()
        scripts_applied = 0

        # 确保 schema_version 表存在
        conn.execute("""
            CREATE TABLE IF NOT EXISTS schema_version (
                id INTEGER PRIMARY KEY,
                value INTEGER NOT NULL
            )
        """)

        # 迁移脚本（当前只有占位，未来版本可在此扩展）
        # if current < 2:
        #     ... 迁移逻辑 ...
        #     scripts_applied += 1

        # 更新版本号
        conn.execute(
            "INSERT OR REPLACE INTO schema_version (id, value) VALUES (1, ?)",
            (self._LATEST_DB_VERSION,)
        )
        conn.commit()

        return {
            "scripts_applied": scripts_applied,
            "duration_ms": round((_time.time() - start) * 1000, 2),
            "final_version": self._LATEST_DB_VERSION,
        }

    def export_as_excel(self,
                        output_path: str,
                        category: Optional[str] = None,
                        layer: Optional[MemoryLayer] = None,
                        starred_only: bool = False) -> Path:
        """导出记忆为 Excel 格式（v5.1.9 新增，v5.2.9 安全加固：路径校验 + CSV 公式注入防护）

        Args:
            output_path: 输出文件路径
            category: 限定分类
            layer: 限定层级
            starred_only: 仅导出收藏的记忆

        Returns:
            导出文件的 Path 对象
        """
        entries = self.list_memories(
            category=category,
            layer=layer,
            starred=starred_only if starred_only else None,
            limit=100000,
        )

        # v5.2.9 安全加固：路径校验
        out = _safe_path(output_path, allowed_exts={".xlsx", ".csv"})
        out.parent.mkdir(parents=True, exist_ok=True)

        # v5.2.9 安全加固：CSV/XLSX 公式注入防护
        def _cell_safe(v, max_len=5000):
            if v is None:
                return ""
            s = str(v)[:max_len]
            if s and s[0] in ("=", "+", "-", "@"):
                return "\t" + s
            return s

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "记忆数据"

            header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            headers = [
                "ID", "内容", "分类", "标签", "隐私等级", "重要性",
                "类型", "层级", "访问次数", "创建时间", "更新时间", "收藏"
            ]
            ws.append(headers)

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border

            def _fmt_time(ts: float) -> str:
                return datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%Y-%m-%d %H:%M")

            for row_num, entry in enumerate(entries, 2):
                ws.append([
                    _cell_safe(entry.id, 64),
                    _cell_safe(entry.content, 500),
                    _cell_safe(entry.category, 256),
                    _cell_safe(", ".join(entry.tags) if entry.tags else "", 1024),
                    _cell_safe(entry.privacy.value, 32),
                    _cell_safe(entry.importance.value, 32),
                    _cell_safe(entry.memory_type.value, 32),
                    _cell_safe(entry.layer.value, 32),
                    _cell_safe(entry.access_count, 32),
                    _cell_safe(_fmt_time(entry.created_at), 32),
                    _cell_safe(_fmt_time(entry.updated_at), 32),
                    "⭐" if entry.starred else "",
                ])

                for col_num in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.border = thin_border
                    cell.alignment = center_align

            ws.column_dimensions['A'].width = 22
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 14
            ws.column_dimensions['I'].width = 10
            ws.column_dimensions['J'].width = 18
            ws.column_dimensions['K'].width = 18
            ws.column_dimensions['L'].width = 6

            wb.save(str(out))
        except ImportError:
            import csv
            out_csv = Path(str(out).replace('.xlsx', '.csv'))
            with open(str(out_csv), 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(["ID", "内容", "分类", "标签", "隐私等级", "重要性", "类型", "层级", "访问次数", "创建时间", "更新时间", "收藏"])
                for entry in entries:
                    writer.writerow([
                        _cell_safe(entry.id, 64),
                        _cell_safe(entry.content, 500),
                        _cell_safe(entry.category, 256),
                        _cell_safe(", ".join(entry.tags) if entry.tags else "", 1024),
                        _cell_safe(entry.privacy.value, 32),
                        _cell_safe(entry.importance.value, 32),
                        _cell_safe(entry.memory_type.value, 32),
                        _cell_safe(entry.layer.value, 32),
                        _cell_safe(entry.access_count, 32),
                        _cell_safe(entry.created_at, 32),
                        _cell_safe(entry.updated_at, 32),
                        _cell_safe(entry.starred, 8),
                    ])
            out = out_csv

        # v5.2.9 安全加固：权限收紧
        try:
            import stat
            import os
            os.chmod(out, stat.S_IRUSR | stat.S_IWUSR | stat.S_IRGRP | stat.S_IROTH)
        except (OSError, ImportError):
            pass

        return out

    def import_from_excel(self,
                          input_path: str,
                          target_category: Optional[str] = None,
                          target_layer: Optional[MemoryLayer] = None) -> Dict[str, int]:
        """从 Excel 文件导入记忆（v5.1.9 新增，v5.2.9 安全加固：路径校验 + 长度限制 + 枚举白名单）

        Args:
            input_path: Excel 文件路径
            target_category: 目标分类（覆盖文件中的分类）
            target_layer: 目标记忆层级

        Returns:
            {imported, skipped, failed}
        """
        # v5.2.9 安全加固：路径校验 + 大小限制
        path = _safe_path(input_path, must_exist=True,
                          allowed_exts={".xlsx", ".xls", ".csv"},
                          max_size=500 * 1024 * 1024)

        entries = []
        _MAX_CONTENT = 1000000
        _MAX_CAT = 256
        _MAX_TAG = 128
        _MAX_TAGS = 64
        _MAX_ROWS = 100000

        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            ws = wb.active

            headers = {}
            header_row = next(ws.iter_rows(values_only=True))
            for col_num, cell in enumerate(header_row, 1):
                h = str(cell).strip().lower() if cell is not None else ""
                if h:
                    headers[h] = col_num

            content_col = headers.get("内容", 2)
            category_col = headers.get("分类", 3)
            tags_col = headers.get("标签", 4)
            privacy_col = headers.get("隐私等级", 5)
            importance_col = headers.get("重要性", 6)
            layer_col = headers.get("层级", 8)

            row_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_count += 1
                if row_count > _MAX_ROWS:
                    break
                cell_content = row[content_col - 1] if content_col - 1 < len(row) else None
                if cell_content and str(cell_content).strip():
                    content = str(cell_content).strip()[:_MAX_CONTENT]
                    cat_cell = row[category_col - 1] if category_col - 1 < len(row) else None
                    category = target_category or (
                        str(cat_cell).strip()[:_MAX_CAT] if cat_cell else "general"
                    )
                    tags_cell = row[tags_col - 1] if tags_col - 1 < len(row) else None
                    tags_str = str(tags_cell).strip() if tags_cell else ""
                    tags = [t.strip()[:_MAX_TAG] for t in tags_str.split(',') if t.strip()][:_MAX_TAGS] if tags_str else []
                    priv_cell = row[privacy_col - 1] if privacy_col - 1 < len(row) else None
                    privacy_str = str(priv_cell).strip() if priv_cell else "internal"
                    imp_cell = row[importance_col - 1] if importance_col - 1 < len(row) else None
                    importance_str = str(imp_cell).strip() if imp_cell else "medium"
                    lay_cell = row[layer_col - 1] if layer_col - 1 < len(row) else None
                    layer_str = str(lay_cell).strip() if lay_cell else "short_term"

                    entries.append({
                        "content": content,
                        "category": category,
                        "tags": tags,
                        "privacy": privacy_str,
                        "importance": importance_str,
                        "layer": layer_str,
                    })
        except ImportError:
            import csv
            with open(str(path), 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                row_count = 0
                for row in reader:
                    row_count += 1
                    if row_count > _MAX_ROWS:
                        break
                    if row.get("内容") and str(row["内容"]).strip():
                        content = str(row["内容"]).strip()[:_MAX_CONTENT]
                        cat_val = str(row.get("分类", "") or "general").strip()
                        category = target_category or (cat_val[:_MAX_CAT] if cat_val else "general")
                        tags_str = str(row.get("标签", ""))
                        tags = [t.strip()[:_MAX_TAG] for t in tags_str.split(',') if t.strip()][:_MAX_TAGS] if tags_str else []
                        privacy_str = str(row.get("隐私等级", "internal")).strip()
                        importance_str = str(row.get("重要性", "medium")).strip()
                        layer_str = str(row.get("层级", "short_term")).strip()

                        entries.append({
                            "content": content,
                            "category": category,
                            "tags": tags,
                            "privacy": privacy_str,
                            "importance": importance_str,
                            "layer": layer_str,
                        })

        imported = 0
        skipped = 0
        failed = 0

        # v5.2.9 安全加固：条数限制
        if len(entries) > _MAX_ROWS:
            entries = entries[:_MAX_ROWS]

        for entry_data in entries:
            try:
                existing = None
                rows = self._get_conn().execute(
                    "SELECT id FROM memories WHERE content = ? AND category = ?",
                    (entry_data["content"][:5000], entry_data["category"][:_MAX_CAT])
                ).fetchall()
                if rows:
                    existing = rows[0][0]

                if existing:
                    skipped += 1
                    continue

                self.add_memory(
                    content=entry_data["content"],
                    category=entry_data["category"],
                    tags=entry_data["tags"],
                    privacy=PrivacyLevel.from_string(entry_data["privacy"]),
                    importance=Importance.from_string(entry_data["importance"]),
                    layer=target_layer or MemoryLayer.from_string(entry_data["layer"]),
                )
                imported += 1
            except (ValueError, TypeError, KeyError):
                failed += 1

        return {"imported": imported, "skipped": skipped, "failed": failed}

    def copy_memory(self,
                    entry_id: str,
                    new_category: str,
                    actor: str = "",
                    session_id: str = "") -> bool:
        """复制记忆到新分类（v5.1.9 新增）

        创建一条新的记忆条目，内容与原条目相同，但分类为新分类。

        Args:
            entry_id: 原记忆 ID
            new_category: 新分类名
            actor: 操作者
            session_id: 会话 ID

        Returns:
            是否复制成功
        """
        conn = self._get_conn()
        row = conn.execute("SELECT * FROM memories WHERE id = ?", (entry_id,)).fetchone()
        if not row:
            return False

        original = self._row_to_entry(row)

        new_entry = self.add_memory(
            content=original.content,
            category=new_category,
            tags=original.tags,
            privacy=original.privacy,
            importance=original.importance,
            memory_type=original.memory_type,
            layer=original.layer,
            source_session=original.source_session,
            source_agent=original.source_agent,
            starred=original.starred,
            metadata={"_copied_from": original.id, **original.metadata},
        )

        self._add_audit(
            "copy", entry_id, actor, session_id, original.privacy.value,
            details={"message": f"复制到 {new_category}", "new_id": new_entry.id}
        )
        return True

    def move_memory(self,
                    entry_id: str,
                    new_category: str,
                    actor: str = "",
                    session_id: str = "") -> bool:
        """移动记忆到新分类（v5.1.9 新增）

        修改记忆的分类为新分类，同时同步更新 FTS 索引。

        Args:
            entry_id: 记忆 ID
            new_category: 新分类名
            actor: 操作者
            session_id: 会话 ID

        Returns:
            是否移动成功
        """
        conn = self._get_conn()
        # v5.5.5 fix: 同时获取 privacy 用于审计日志
        row = conn.execute("SELECT category, privacy FROM memories WHERE id = ?", (entry_id,)).fetchone()
        if not row or row["category"] == "trash":
            return False

        old_category = row["category"]
        now = time.time()

        self.update_memory(
            entry_id=entry_id,
            category=new_category,
            actor=actor,
            session_id=session_id,
        )

        self._add_audit(
            "move", entry_id, actor, session_id,
            row["privacy"] if row["privacy"] else "",
            details={"message": f"从 {old_category} 移动到 {new_category}"}
        )
        return True


    # ===== 嵌入向量（v5.4.5 新增）=====

    @property
    def embedding_engine(self):
        """懒加载 EmbeddingEngine 单例"""
        if not hasattr(self, '_embedding_eng'):
            try:
                from .embedding import EmbeddingEngine
                self._embedding_eng = EmbeddingEngine()
            except Exception:
                self._embedding_eng = None
        return self._embedding_eng

    def _store_embedding(self, memory_id: str, text_content: str):
        """为记忆生成并存储嵌入向量

        Args:
            memory_id: 记忆 ID
            text_content: 用于生成向量的文本（解密后的明文内容）
        """
        engine = self.embedding_engine
        if engine is None or not engine.is_available:
            return
        vec = engine.encode(text_content)
        if vec is None:
            return
        blob = engine.serialize(vec)
        now = time.time()
        conn = self._get_conn()
        conn.execute(
            "INSERT OR REPLACE INTO memory_embeddings"
            " (memory_id, embedding, model_name, dimension, created_at, updated_at)"
            " VALUES (?, ?, ?, ?, ?, ?)",
            (memory_id, blob, engine.model_name, engine.dimension, now, now)
        )
        conn.commit()

    def _delete_embedding(self, memory_id: str):
        """删除记忆的嵌入向量"""
        conn = self._get_conn()
        conn.execute(
            "DELETE FROM memory_embeddings WHERE memory_id = ?",
            (memory_id,)
        )
        conn.commit()

    def _get_all_embeddings(self, limit: int = 100000) -> list:
        """获取所有记忆的嵌入向量（用于向量召回）

        Returns:
            [(memory_id, embedding_blob), ...]
        """
        conn = self._get_conn()
        rows = conn.execute(
            "SELECT me.memory_id, me.embedding"
            " FROM memory_embeddings me"
            " INNER JOIN memories m ON m.id = me.memory_id"
            " WHERE m.category != 'trash'"
            " LIMIT ?",
            (limit,)
        ).fetchall()
        return [(row[0], row[1]) for row in rows]

    # ===== 归档机制（v5.4.6 新增）=====

    def auto_archive(self, max_age_hours: int = 24,
                     layer: str = "sensory",
                     actor: str = "system") -> Dict[str, Any]:
        """自动归档过期记忆（v5.4.6 新增）

        将到期的感官层/短期层记忆移到 archived_memories 表，
        而非直接删除。可配置保留天数，支持手动恢复。

        Args:
            max_age_hours: 最大保留时长（小时）
            layer: 记忆层级（sensory/short_term）
            actor: 操作者

        Returns:
            {archived, layer, max_age_hours}
        """
        conn = self._get_conn()
        now = time.time()
        cutoff_time = now - (max_age_hours * 3600)

        rows = conn.execute(
            "SELECT id, content, category, tags, privacy, importance,"
            " memory_type, layer, source_session, source_agent, metadata,"
            " created_at, updated_at"
            " FROM memories"
            " WHERE layer = ? AND category != 'trash' AND created_at < ?",
            (layer, cutoff_time)
        ).fetchall()

        if not rows:
            return {"archived": 0, "layer": layer, "max_age_hours": max_age_hours}

        archived_count = 0
        for row in rows:
            try:
                import uuid as _uuid
                archive_id = str(_uuid.uuid4())
                conn.execute(
                    "INSERT OR REPLACE INTO archived_memories"
                    " (id, original_id, content, category, tags, privacy,"
                    " importance, memory_type, layer, source_session,"
                    " source_agent, metadata, original_created_at,"
                    " original_updated_at, archived_at, archived_reason)"
                    " VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'expired')",
                    (archive_id, row["id"], row["content"], row["category"],
                     row["tags"], row["privacy"], row["importance"],
                     row["memory_type"], row["layer"], row["source_session"],
                     row["source_agent"], row["metadata"],
                     row["created_at"], row["updated_at"], now)
                )
                # 软删除原记忆
                conn.execute(
                    "UPDATE memories SET category = 'trash',"
                    " metadata = JSON_SET(metadata, '$.archived_id', ?),"
                    " updated_at = ? WHERE id = ?",
                    (archive_id, now, row["id"])
                )
                # 从 FTS 删除
                if not self.encrypted:
                    conn.execute(
                        "INSERT INTO memory_fts(memory_fts, rowid, content, category, tags)"
                        " VALUES('delete', (SELECT rowid FROM memories WHERE id = ?), '', '', '')",
                        (row["id"],)
                    )
                # 删除嵌入向量
                conn.execute(
                    "DELETE FROM memory_embeddings WHERE memory_id = ?",
                    (row["id"],)
                )
                archived_count += 1
            except Exception:
                continue

        conn.commit()

        # 审计日志
        if archived_count > 0:
            try:
                conn.execute(
                    "INSERT INTO audit_log (id, memory_id, action, actor,"
                    " session_id, details, timestamp)"
                    " VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (str(_uuid.uuid4()), "", "auto_archive", actor, "",
                     json.dumps({"count": archived_count, "layer": layer,
                                 "max_age_hours": max_age_hours}), now)
                )
                conn.commit()
            except Exception:
                pass

        return {"archived": archived_count, "layer": layer, "max_age_hours": max_age_hours}

    def list_archived(self, layer: Optional[str] = None,
                      category: Optional[str] = None,
                      limit: int = 50,
                      offset: int = 0) -> List[Dict[str, Any]]:
        """列出归档记忆（v5.4.6 新增）"""
        conn = self._get_conn()
        query = "SELECT * FROM archived_memories WHERE 1=1"
        params = []
        if layer:
            query += " AND layer = ?"
            params.append(layer)
        if category:
            query += " AND category = ?"
            params.append(category)
        query += " ORDER BY archived_at DESC LIMIT ? OFFSET ?"
        params.extend([limit, offset])
        rows = conn.execute(query, params).fetchall()
        return [dict(row) for row in rows]

    def restore_archived(self, archive_id: str,
                         actor: str = "system") -> Dict[str, Any]:
        """从归档恢复记忆（v5.4.6 新增）

        Args:
            archive_id: 归档记录 ID
            actor: 操作者

        Returns:
            {restored, memory_id, error}
        """
        conn = self._get_conn()
        row = conn.execute(
            "SELECT * FROM archived_memories WHERE id = ?",
            (archive_id,)
        ).fetchone()

        if not row:
            return {"restored": False, "error": "归档记录不存在"}

        now = time.time()
        original_id = row["original_id"]

        # 恢复到 memories 表（如果原记录还在 trash 中则恢复，否则新建）
        existing = conn.execute(
            "SELECT id FROM memories WHERE id = ?", (original_id,)
        ).fetchone()

        if existing:
            # 恢复原记录
            conn.execute(
                "UPDATE memories SET category = ?, updated_at = ?"
                " WHERE id = ?",
                (row["category"], now, original_id)
            )
            # 重新加入 FTS
            if not self.encrypted:
                conn.execute(
                    "INSERT INTO memory_fts(rowid, content, category, tags)"
                    " VALUES((SELECT rowid FROM memories WHERE id = ?), ?, ?, ?)",
                    (original_id, row["content"], row["category"], row["tags"])
                )
        else:
            # 原记录已硬删除，新建
            import uuid as _uuid
            new_id = str(_uuid.uuid4())
            conn.execute(
                "INSERT INTO memories (id, content, category, tags, privacy,"
                " importance, memory_type, layer, encrypted, access_count,"
                " strength, forgetting_score, source_session, source_agent,"
                " metadata, created_at, updated_at)"
                " VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0, 0, 1.0, 0.0, ?, ?, ?, ?, ?)",
                (new_id, row["content"], row["category"], row["tags"],
                 row["privacy"], row["importance"], row["memory_type"],
                 row["layer"], row["source_session"], row["source_agent"],
                 row["metadata"], row["original_created_at"], now)
            )
            if not self.encrypted:
                conn.execute(
                    "INSERT INTO memory_fts(rowid, content, category, tags)"
                    " VALUES((SELECT rowid FROM memories WHERE id = ?), ?, ?, ?)",
                    (new_id, row["content"], row["category"], row["tags"])
                )

        # 删除归档记录
        conn.execute("DELETE FROM archived_memories WHERE id = ?", (archive_id,))
        conn.commit()

        return {"restored": True, "memory_id": original_id}

    def purge_archived(self, older_than_days: int = 90,
                       actor: str = "system") -> int:
        """永久删除过期的归档记忆（v5.4.6 新增）

        Args:
            older_than_days: 归档超过 N 天的记忆将被永久删除
            actor: 操作者

        Returns:
            删除数量
        """
        conn = self._get_conn()
        now = time.time()
        cutoff = now - (older_than_days * 86400)
        cursor = conn.execute(
            "DELETE FROM archived_memories WHERE archived_at < ?",
            (cutoff,)
        )
        count = cursor.rowcount
        conn.commit()
        return count

    @staticmethod
    def _deserialize_vector_fallback(blob: bytes, expected_dim: int) -> Optional[List[float]]:
        """engine 不可用时的通用向量反序列化（v5.4.7 新增）

        格式与 EmbeddingEngine.serialize 一致：float32 小端序。
        """
        import struct
        if not blob:
            return None
        count = len(blob) // 4
        if count != expected_dim:
            return None
        try:
            return list(struct.unpack(f'<{count}f', blob))
        except struct.error:
            return None

    @staticmethod
    def _cosine_similarity_batch_fallback(
        query_vec: List[float],
        candidates: List[Tuple[str, List[float]]],
        top_k: int = 20,
    ) -> List[Tuple[str, float]]:
        """engine 不可用时的批量余弦相似度计算（v5.4.7 新增）"""
        if not query_vec or not candidates:
            return []
        scored = []
        for mem_id, vec in candidates:
            if not vec or len(vec) != len(query_vec):
                continue
            dot = sum(a * b for a, b in zip(query_vec, vec))
            norm1 = sum(a * a for a in query_vec) ** 0.5
            norm2 = sum(b * b for b in vec) ** 0.5
            if norm1 == 0.0 or norm2 == 0.0:
                continue
            score = max(-1.0, min(1.0, dot / (norm1 * norm2)))
            scored.append((mem_id, score))
        scored.sort(key=lambda x: x[1], reverse=True)
        return scored[:top_k]

    def vector_search(self, query: str = "", top_k: int = 20,
                      categories=None,
                      layers=None,
                      query_vector: Optional[List[float]] = None,
                      ) -> List[Dict[str, Any]]:
        """向量语义搜索（v5.4.5 新增，v5.4.7 修复）

        使用嵌入向量计算余弦相似度，召回语义相近但用词不同的记忆。

        v5.4.7 改进：支持传入预计算的 query_vector，当 embedding engine
        不可用时（如 sentence-transformers 被卸载），仍可利用已有向量检索。

        Args:
            query: 搜索查询（当 query_vector 未提供时，用于 encode）
            top_k: 返回前 k 个结果
            categories: 限定分类列表
            layers: 限定层级列表
            query_vector: 预计算的查询向量（可选）。提供后跳过 engine.encode()。

        Returns:
            [{entry, score, strategy: 'vector'}, ...]
        """
        engine = self.embedding_engine

        # 确定最终使用的查询向量
        vec = query_vector
        if vec is None:
            # 没有预计算向量，需要 engine 来 encode
            if engine is None or not engine.is_available:
                return []
            vec = engine.encode(query)
            if vec is None:
                return []

        # 获取所有嵌入向量
        all_embeddings = self._get_all_embeddings()
        if not all_embeddings:
            return []

        # 反序列化并计算相似度
        candidates = []
        for mem_id, blob in all_embeddings:
            if engine and engine.is_available:
                deserialized = engine.deserialize(blob)
                if deserialized and len(deserialized) == len(vec):
                    candidates.append((mem_id, deserialized))
            else:
                # engine 不可用时，使用通用反序列化（float32 小端序）
                deserialized = self._deserialize_vector_fallback(blob, len(vec))
                if deserialized:
                    candidates.append((mem_id, deserialized))

        if not candidates:
            return []

        # 批量计算余弦相似度，取 top_k
        if engine and engine.is_available:
            scored = engine.cosine_similarity_batch(vec, candidates, top_k=top_k * 2)
        else:
            scored = self._cosine_similarity_batch_fallback(vec, candidates, top_k=top_k * 2)

        results = []
        for mem_id, score in scored:
            entry = self.get_memory(mem_id)
            if not entry:
                continue
            # 分类过滤
            if categories and entry.category not in categories:
                continue
            # 层级过滤
            if layers and entry.layer not in layers:
                continue
            # 解密内容
            content_text = entry.content
            if entry.encrypted:
                content_text = self.decrypt_content(entry)
            results.append({
                "entry": entry,
                "score": float(score),
                "strategy": "vector",
            })
            if len(results) >= top_k:
                break

        return results

    def rebuild_embeddings(self, batch_size: int = 100,
                           incremental: bool = True) -> Dict[str, Any]:
        """重建/增量构建记忆的嵌入向量（v5.4.5 新增，v5.4.6 增量模式）

        v5.4.6 改进：默认 incremental=True，只处理缺失嵌入的记忆，
        5000+ 记忆时体感差异明显。设 incremental=False 则全量重建。

        Args:
            batch_size: 批量编码大小
            incremental: True=只处理缺失项（默认），False=全量重建

        Returns:
            {success, total, embedded, skipped, errors, incremental}
        """
        engine = self.embedding_engine
        if engine is None or not engine.is_available:
            return {"success": False, "error": "EmbeddingEngine 不可用（未安装 sentence-transformers 或未配置后端）"}

        conn = self._get_conn()

        if incremental:
            # v5.4.6 增量模式：只处理缺失嵌入的记忆
            rows = conn.execute(
                "SELECT m.id, m.content FROM memories m"
                " LEFT JOIN memory_embeddings me ON m.id = me.memory_id"
                " WHERE m.category != 'trash' AND m.encrypted = 0"
                " AND me.memory_id IS NULL"
            ).fetchall()
            mode_label = "incremental"
        else:
            # 全量重建模式
            rows = conn.execute(
                "SELECT id, content FROM memories"
                " WHERE category != 'trash' AND encrypted = 0"
            ).fetchall()
            mode_label = "full"

        total = len(rows)
        embedded = 0
        skipped = 0
        errors = 0
        now = time.time()

        for i in range(0, total, batch_size):
            batch = rows[i:i + batch_size]
            texts = [row[1] or "" for row in batch]
            # 过滤空文本
            valid = [(batch[j][0], texts[j]) for j in range(len(batch)) if texts[j].strip()]
            if not valid:
                skipped += len(batch)
                continue

            vecs = engine.encode_batch([t for _, t in valid])
            if vecs is None:
                errors += len(valid)
                continue

            for (mem_id, _), vec in zip(valid, vecs):
                if vec is None:
                    errors += 1
                    continue
                blob = engine.serialize(vec)
                conn.execute(
                    "INSERT OR REPLACE INTO memory_embeddings"
                    " (memory_id, embedding, model_name, dimension, created_at, updated_at)"
                    " VALUES (?, ?, ?, ?, ?, ?)",
                    (mem_id, blob, engine.model_name, engine.dimension, now, now)
                )
                embedded += 1

        conn.commit()
        return {
            "success": True,
            "total": total,
            "embedded": embedded,
            "skipped": skipped,
            "errors": errors,
            "mode": mode_label,
        }


    # ===== 搜索增强（v5.2.0 新增）=====

    def fuzzy_search(self,
                     query: str,
                     category: Optional[str] = None,
                     layer: Optional[MemoryLayer] = None,
                     limit: int = 20,
                     threshold: float = 0.3) -> List[Dict[str, Any]]:
        """模糊搜索记忆（v5.2.0 新增，v5.4.7 加入 SequenceMatcher 近似匹配）

        结合全文搜索、SequenceMatcher 近似匹配和词重叠计算，
        支持拼写纠错和近似匹配，无需额外依赖（使用标准库 difflib）。

        Args:
            query: 搜索关键词
            category: 限定分类
            layer: 限定层级
            limit: 返回结果数量
            threshold: 相似度阈值（0-1）

        Returns:
            带分数的搜索结果列表 [{entry, score, highlights}]
        """
        from difflib import SequenceMatcher

        # v5.5.6 fix: 防御 None / 非字符串 / 空查询
        if query is None:
            return []
        if not isinstance(query, str):
            query = str(query)
        if not query.strip():
            return []

        conn = self._get_conn()
        query_lower = query.lower()

        base_query = "SELECT * FROM memories WHERE category != 'trash'"
        params = []

        if category:
            base_query += " AND category = ?"
            params.append(category)
        if layer:
            base_query += " AND layer = ?"
            params.append(layer.value)

        rows = conn.execute(base_query, params).fetchall()
        entries = [self._row_to_entry(r) for r in rows]

        scored = []
        for entry in entries:
            content_lower = entry.content.lower()
            tags_lower = [t.lower() for t in entry.tags]
            cat_lower = entry.category.lower()

            score = 0.0
            highlights = []

            # 1. 精确子串匹配（最高权重）
            if query_lower in content_lower:
                score += 0.8
                pos = content_lower.find(query_lower)
                highlights.append({
                    "field": "content",
                    "start": pos,
                    "end": pos + len(query),
                    "text": entry.content[max(0, pos - 20):pos + len(query) + 20]
                })

            if query_lower in cat_lower:
                score += 0.5

            for tag in tags_lower:
                if query_lower in tag:
                    score += 0.4
                    break

            # 2. SequenceMatcher 近似匹配（支持拼写纠错 / 近似词）
            if score < 0.8:
                seq_ratio = SequenceMatcher(None, query_lower, content_lower).ratio()
                if seq_ratio > 0.4:
                    score += seq_ratio * 0.5
                    if not highlights:
                        pos = content_lower.find(query_lower[0])
                        if pos >= 0:
                            highlights.append({
                                "field": "content",
                                "start": max(0, pos - 20),
                                "end": min(len(entry.content), pos + len(query) + 20),
                                "text": entry.content[max(0, pos - 20):pos + len(query) + 20],
                                "match_type": "fuzzy"
                            })

                # 标签近似匹配
                for tag in tags_lower:
                    tag_ratio = SequenceMatcher(None, query_lower, tag).ratio()
                    if tag_ratio > 0.5:
                        score += tag_ratio * 0.3
                        break

            # 3. 词重叠兜底
            if score == 0:
                words = set(query_lower.split())
                content_words = set(content_lower.split())
                if words and content_words:
                    overlap = len(words & content_words)
                    score = overlap / max(len(words), 1) * 0.3

            if score >= threshold:
                scored.append({
                    "entry": entry,
                    "score": round(score, 4),
                    "highlights": highlights
                })

        scored.sort(key=lambda x: x["score"], reverse=True)
        return scored[:limit]

    def get_search_history(self, limit: int = 20) -> List[Dict[str, Any]]:
        """获取搜索历史（v5.2.0 新增）

        Args:
            limit: 返回条数

        Returns:
            搜索历史列表 [{query, count, last_used}]
        """
        conn = self._get_conn()
        try:
            rows = conn.execute("""
                SELECT query, MAX(timestamp) as last_used, COUNT(*) as cnt
                FROM audit_log
                WHERE action = 'search'
                GROUP BY query
                ORDER BY last_used DESC
                LIMIT ?
            """, (limit,)).fetchall()
            return [
                {"query": row["query"], "count": row["cnt"], "last_used": row["last_used"]}
                for row in rows if row["query"]
            ]
        except sqlite3.OperationalError:
            return []

    # ===== v5.5.6 新增方法 =====

    def batch_get_memories(self, memory_ids: List[str],
                           actor: str = "", session_id: str = "") -> List[MemoryEntry]:
        """批量获取记忆（v5.5.6 新增）

        单次 SQL 查询获取多条记忆，避免 N+1 查询。
        自动排除回收站和过期记忆。

        Args:
            memory_ids: 记忆 ID 列表
            actor: 操作者（审计用）
            session_id: 会话 ID

        Returns:
            按输入 ID 顺序排列的记忆列表（不存在的 ID 跳过）
        """
        if not memory_ids:
            return []
        # 去重并限制数量，防止 SQL 注入和性能问题
        unique_ids = []
        seen = set()
        for mid in memory_ids:
            if isinstance(mid, str) and mid and mid not in seen:
                seen.add(mid)
                unique_ids.append(mid)
        if not unique_ids:
            return []
        if len(unique_ids) > 10000:
            unique_ids = unique_ids[:10000]

        conn = self._get_conn()
        placeholders = ",".join(["?"] * len(unique_ids))
        rows = conn.execute(
            f"SELECT * FROM memories WHERE id IN ({placeholders}) AND category != 'trash'",
            unique_ids
        ).fetchall()

        # 按输入顺序排列
        id_to_row = {row["id"]: row for row in rows}
        now = time.time()
        result = []
        for mid in unique_ids:
            row = id_to_row.get(mid)
            if not row:
                continue
            entry = self._row_to_entry(row)
            # v5.5.2 兼容：过期记忆自动移入回收站
            if entry.expires_at and entry.expires_at > 0 and entry.expires_at <= now:
                try:
                    conn.execute(
                        "UPDATE memories SET category = 'trash', updated_at = ? WHERE id = ? AND category != 'trash'",
                        (now, mid)
                    )
                    conn.commit()
                except Exception:
                    pass
                continue
            self._update_access(entry, actor, session_id)
            result.append(entry)
        return result

    def timeline_view(self, category: Optional[str] = None,
                      layer: Optional[MemoryLayer] = None,
                      limit: int = 500) -> Dict[str, List[MemoryEntry]]:
        """记忆时间线视图（v5.5.6 新增）

        按创建时间分组为：今天、昨天、本周、本月、更早。

        Args:
            category: 限定分类
            layer: 限定层级
            limit: 最大返回条数

        Returns:
            {"today": [...], "yesterday": [...], "this_week": [...], "this_month": [...], "earlier": [...]}
        """
        import time as _time
        now = _time.time()
        # 计算各时间边界（本地日期）
        from datetime import datetime, timedelta
        now_dt = datetime.now()
        today_start = datetime(now_dt.year, now_dt.month, now_dt.day).timestamp()
        yesterday_start = today_start - 86400
        # 本周一零点
        week_start = today_start - (now_dt.weekday() * 86400)
        # 本月1号零点
        month_start = datetime(now_dt.year, now_dt.month, 1).timestamp()

        entries = self.list_memories(
            category=category, layer=layer, limit=limit,
            sort_by="created_at", sort_order="desc"
        )

        result = {"today": [], "yesterday": [], "this_week": [], "this_month": [], "earlier": []}
        for e in entries:
            ts = e.created_at
            if ts >= today_start:
                result["today"].append(e)
            elif ts >= yesterday_start:
                result["yesterday"].append(e)
            elif ts >= week_start:
                result["this_week"].append(e)
            elif ts >= month_start:
                result["this_month"].append(e)
            else:
                result["earlier"].append(e)
        return result

    def search_suggestions(self, prefix: str, limit: int = 10,
                           category: Optional[str] = None) -> Dict[str, List[str]]:
        """搜索建议（v5.5.6 新增）

        基于已有标签和分类，返回与前缀匹配的自动补全建议。

        Args:
            prefix: 输入前缀（大小写不敏感）
            limit: 每类返回数量上限
            category: 限定分类范围

        Returns:
            {"tags": [...], "categories": [...]}
        """
        if not prefix or not isinstance(prefix, str):
            return {"tags": [], "categories": []}
        prefix_lower = prefix.lower().strip()
        if not prefix_lower:
            return {"tags": [], "categories": []}

        conn = self._get_conn()
        query = "SELECT tags, category FROM memories WHERE category != 'trash'"
        params = []
        if category:
            query += " AND category = ?"
            params.append(category)
        rows = conn.execute(query, params).fetchall()

        tag_set = set()
        cat_set = set()
        for row in rows:
            cat = row["category"]
            if cat and cat.lower().startswith(prefix_lower):
                cat_set.add(cat)
            raw_tags = row["tags"]
            if raw_tags:
                try:
                    tags = json.loads(raw_tags) if isinstance(raw_tags, str) else raw_tags
                    for t in tags:
                        if isinstance(t, str) and t.lower().startswith(prefix_lower):
                            tag_set.add(t)
                except (json.JSONDecodeError, TypeError):
                    pass

        return {
            "tags": sorted(tag_set)[:limit],
            "categories": sorted(cat_set)[:limit],
        }

    def check_duplicates(self, content: str,
                         similarity_threshold: float = 0.85,
                         category: Optional[str] = None,
                         limit: int = 5) -> List[Dict[str, Any]]:
        """添加前检测近重复记忆（v5.5.6 新增）

        使用 Jaccard 词重叠 + SequenceMatcher 综合判断，
        返回与给定内容高度相似的已有记忆。

        Args:
            content: 待检测内容
            similarity_threshold: 相似度阈值（0-1），默认 0.85
            category: 限定分类范围
            limit: 最大返回数量

        Returns:
            [{"entry": MemoryEntry, "similarity": float, "match_type": str}, ...]
        """
        if not content or not isinstance(content, str) or not content.strip():
            return []

        # v5.5.6 fix: 与 add() 一致，先消毒再比较（否则 <b>bold</b> 无法匹配存储的 bold text）
        content = self._strip_control(content)
        content = _sanitize_html(content)

        from difflib import SequenceMatcher
        import re

        conn = self._get_conn()
        query = "SELECT * FROM memories WHERE category != 'trash' AND encrypted = 0"
        params = []
        if category:
            query += " AND category = ?"
            params.append(category)
        rows = conn.execute(query, params).fetchall()

        content_lower = content.lower().strip()
        content_words = set(re.findall(r'\w+', content_lower))

        results = []
        for row in rows:
            entry = self._row_to_entry(row)
            entry_content = (entry.content or "").lower().strip()
            if not entry_content:
                continue

            # 完全相同
            if entry_content == content_lower:
                results.append({"entry": entry, "similarity": 1.0, "match_type": "exact"})
                continue

            # Jaccard 词重叠
            entry_words = set(re.findall(r'\w+', entry_content))
            if content_words and entry_words:
                jaccard = len(content_words & entry_words) / len(content_words | entry_words)
            else:
                jaccard = 0.0

            # SequenceMatcher
            seq_ratio = SequenceMatcher(None, content_lower, entry_content).ratio()

            # 综合相似度（加权平均）
            similarity = max(jaccard * 0.6 + seq_ratio * 0.4, seq_ratio * 0.8)

            if similarity >= similarity_threshold:
                match_type = "high" if similarity >= 0.95 else "partial"
                results.append({"entry": entry, "similarity": round(similarity, 4), "match_type": match_type})

        results.sort(key=lambda x: x["similarity"], reverse=True)
        return results[:limit]

    def add_tags_to_ids(self, memory_ids: List[str], tags: List[str],
                        actor: str = "", session_id: str = "") -> int:
        """按 ID 列表批量添加标签（v5.5.6 新增）

        Args:
            memory_ids: 记忆 ID 列表
            tags: 要添加的标签列表
            actor: 操作者
            session_id: 会话 ID

        Returns:
            实际更新的记忆条数
        """
        if not memory_ids or not tags:
            return 0
        clean_tags = [t for t in tags if isinstance(t, str) and t.strip()]
        if not clean_tags:
            return 0

        conn = self._get_conn()
        now = time.time()
        updated = 0
        for mid in memory_ids:
            if not isinstance(mid, str) or not mid:
                continue
            row = conn.execute(
                "SELECT tags FROM memories WHERE id = ? AND category != 'trash'", (mid,)
            ).fetchone()
            if not row:
                continue
            existing = self._safe_json_loads(row["tags"], [])
            merged = list(existing)
            # v5.5.6 fix: 大小写不敏感去重，与 remove_tags_from_ids / rename_tag 一致
            existing_lower = {t.lower() for t in merged if isinstance(t, str)}
            for t in clean_tags:
                if t.lower() not in existing_lower:
                    merged.append(t)
                    existing_lower.add(t.lower())
            if len(merged) != len(existing):
                conn.execute(
                    "UPDATE memories SET tags = ?, updated_at = ? WHERE id = ?",
                    (json.dumps(merged, ensure_ascii=False), now, mid)
                )
                updated += 1
        conn.commit()
        return updated

    def remove_tags_from_ids(self, memory_ids: List[str], tags: List[str],
                             actor: str = "", session_id: str = "") -> int:
        """按 ID 列表批量移除标签（v5.5.6 新增）

        Args:
            memory_ids: 记忆 ID 列表
            tags: 要移除的标签列表
            actor: 操作者
            session_id: 会话 ID

        Returns:
            实际更新的记忆条数
        """
        if not memory_ids or not tags:
            return 0
        remove_lower = {t.lower() for t in tags if isinstance(t, str)}
        if not remove_lower:
            return 0

        conn = self._get_conn()
        now = time.time()
        updated = 0
        for mid in memory_ids:
            if not isinstance(mid, str) or not mid:
                continue
            row = conn.execute(
                "SELECT tags FROM memories WHERE id = ? AND category != 'trash'", (mid,)
            ).fetchone()
            if not row:
                continue
            existing = self._safe_json_loads(row["tags"], [])
            filtered = [t for t in existing if isinstance(t, str) and t.lower() not in remove_lower]
            if len(filtered) != len(existing):
                conn.execute(
                    "UPDATE memories SET tags = ?, updated_at = ? WHERE id = ?",
                    (json.dumps(filtered, ensure_ascii=False), now, mid)
                )
                updated += 1
        conn.commit()
        return updated

    def highlight_text(self, text: str, query: str,
                       before_tag: str = "<mark>",
                       after_tag: str = "</mark>") -> str:
        """高亮搜索关键词（v5.2.0 新增，v5.5.2 增强多关键词+中文支持）

        Args:
            text: 原始文本
            query: 搜索关键词（支持空格分隔的多关键词）
            before_tag: 高亮起始标签
            after_tag: 高亮结束标签

        Returns:
            带高亮标记的文本
        """
        if not query or not text:
            return text

        import re

        # 提取所有关键词：按空白分割，同时保留原始完整查询
        keywords = []
        for kw in re.split(r'\s+', query.strip()):
            kw = kw.strip()
            if kw and kw not in keywords:
                keywords.append(kw)
        # 完整查询也作为一个关键词（如果不在列表中）
        full_query = query.strip()
        if full_query and full_query not in keywords:
            keywords.insert(0, full_query)

        if not keywords:
            return text

        # 按长度降序排列，避免短关键词先替换破坏长关键词
        keywords.sort(key=len, reverse=True)

        result = text
        # 占位符策略：先用唯一占位符替换匹配项，最后统一替换为高亮标签，
        # 避免短关键词匹配到已插入的标签内部（如 <mark> 中的 "mark"）
        placeholder_map = {}  # placeholder -> original matched text
        for i, kw in enumerate(keywords):
            if not kw:
                continue
            pattern = re.compile(re.escape(kw), re.IGNORECASE)
            placeholder_prefix = f"\x00MFHL{i}_"

            def _make_replacer(prefix, pi_counter):
                def replacer(m):
                    matched = m.group()
                    ph = f"{prefix}{pi_counter[0]}\x00"
                    placeholder_map[ph] = matched
                    pi_counter[0] += 1
                    return ph
                return replacer

            result = pattern.sub(_make_replacer(placeholder_prefix, [0]), result)

        # 第二阶段：将占位符替换为带高亮标签的原文（保留原始大小写）
        for ph, matched_text in placeholder_map.items():
            result = result.replace(ph, before_tag + matched_text + after_tag)

        return result

    # ===== TTL / 过期管理（v5.5.2 新增）=====

    def set_ttl(self, memory_id: str, ttl_seconds: float,
                actor: str = "", session_id: str = "") -> bool:
        """为记忆设置 TTL（存活时间）

        v5.5.2 新增。

        Args:
            memory_id: 记忆 ID
            ttl_seconds: 存活秒数；<= 0 表示取消过期（永不过期）
            actor: 操作者
            session_id: 会话 ID

        Returns:
            是否成功
        """
        # v5.4.0 安全加固
        actor = self._strip_control(actor)[:128]
        session_id = self._strip_control(session_id)[:128]
        conn = self._get_conn()
        now = time.time()
        expires_at = (now + ttl_seconds) if ttl_seconds and ttl_seconds > 0 else 0.0
        # v5.5.5 fix: 获取 privacy 用于审计日志
        priv_row = conn.execute(
            "SELECT privacy FROM memories WHERE id = ? AND category != 'trash'",
            (memory_id,)
        ).fetchone()
        privacy_val = priv_row[0] if priv_row and priv_row[0] else ""
        cursor = conn.execute(
            "UPDATE memories SET expires_at = ?, updated_at = ? WHERE id = ? AND category != 'trash'",
            (expires_at, now, memory_id)
        )
        conn.commit()
        if cursor.rowcount > 0:
            self._add_audit("update", memory_id, actor, session_id, privacy_val,
                             details={"ttl_seconds": ttl_seconds, "expires_at": expires_at})
            return True
        return False

    def list_expired(self, limit: int = 1000) -> List[MemoryEntry]:
        """列出所有已过期但尚未清理的记忆

        v5.5.2 新增。

        Args:
            limit: 最大返回条数

        Returns:
            已过期记忆列表
        """
        conn = self._get_conn()
        now = time.time()
        rows = conn.execute(
            "SELECT * FROM memories WHERE expires_at > 0 AND expires_at <= ? AND category != 'trash' ORDER BY expires_at ASC LIMIT ?",
            (now, limit)
        ).fetchall()
        return [self._row_to_entry(row) for row in rows]

    def purge_expired(self, actor: str = "", session_id: str = "") -> int:
        """清理所有已过期记忆（移入回收站）

        v5.5.2 新增。

        Args:
            actor: 操作者
            session_id: 会话 ID

        Returns:
            清理的记忆条数
        """
        # v5.4.0 安全加固
        actor = self._strip_control(actor)[:128]
        session_id = self._strip_control(session_id)[:128]
        conn = self._get_conn()
        now = time.time()
        expired_rows = conn.execute(
            "SELECT id, privacy FROM memories WHERE expires_at > 0 AND expires_at <= ? AND category != 'trash'",
            (now,)
        ).fetchall()
        if not expired_rows:
            return 0
        count = 0
        for row in expired_rows:
            conn.execute(
                "UPDATE memories SET category = 'trash', updated_at = ? WHERE id = ?",
                (now, row["id"])
            )
            self._add_audit("forget", row["id"], actor, session_id,
                             row["privacy"] if row["privacy"] else "",
                             details={"reason": "ttl_purge", "expires_at": now})
            count += 1
        conn.commit()
        return count

    # ===== 按分类/标签批量删除（v5.5.2 新增）=====

    def batch_delete_by_category(self, category: str,
                                  actor: str = "", session_id: str = "",
                                  permanent: bool = False) -> int:
        """按分类批量删除记忆（移入回收站或永久删除）

        v5.5.2 新增。

        Args:
            category: 要删除的分类名称
            actor: 操作者
            session_id: 会话 ID
            permanent: True=永久删除，False=移入回收站

        Returns:
            删除的记忆条数
        """
        if not category:
            return 0
        # v5.4.0 安全加固
        category = self._strip_control(category)[:64]
        actor = self._strip_control(actor)[:128]
        session_id = self._strip_control(session_id)[:128]
        conn = self._get_conn()
        now = time.time()
        rows = conn.execute(
            "SELECT id, privacy FROM memories WHERE category = ?",
            (category,)
        ).fetchall()
        if not rows:
            return 0
        count = 0
        for row in rows:
            if permanent:
                self._delete_memory_cascade(conn, row["id"])
            else:
                conn.execute(
                    "UPDATE memories SET category = 'trash', updated_at = ? WHERE id = ?",
                    (now, row["id"])
                )
            self._add_audit("delete", row["id"], actor, session_id,
                             row["privacy"] if row["privacy"] else "",
                             details={"reason": "batch_by_category", "category": category, "permanent": permanent})
            count += 1
        conn.commit()
        return count

    def batch_delete_by_tag(self, tag: str,
                             actor: str = "", session_id: str = "",
                             permanent: bool = False) -> int:
        """按标签批量删除记忆（移入回收站或永久删除）

        v5.5.2 新增。

        Args:
            tag: 要删除的标签
            actor: 操作者
            session_id: 会话 ID
            permanent: True=永久删除，False=移入回收站

        Returns:
            删除的记忆条数
        """
        if not tag:
            return 0
        # v5.4.0 安全加固
        tag = self._strip_control(tag)[:64]
        actor = self._strip_control(actor)[:128]
        session_id = self._strip_control(session_id)[:128]
        conn = self._get_conn()
        now = time.time()
        rows = conn.execute(
            "SELECT id, privacy, tags FROM memories WHERE tags LIKE ? AND category != 'trash'",
            (f'%"{tag}"%',)
        ).fetchall()
        if not rows:
            return 0
        count = 0
        for row in rows:
            try:
                existing_tags = json.loads(row["tags"]) if isinstance(row["tags"], str) else row["tags"]
                if tag not in (existing_tags or []):
                    continue
            except (json.JSONDecodeError, TypeError):
                continue
            if permanent:
                self._delete_memory_cascade(conn, row["id"])
            else:
                conn.execute(
                    "UPDATE memories SET category = 'trash', updated_at = ? WHERE id = ?",
                    (now, row["id"])
                )
            self._add_audit("delete", row["id"], actor, session_id,
                             row["privacy"] if row["privacy"] else "",
                             details={"reason": "batch_by_tag", "tag": tag, "permanent": permanent})
            count += 1
        conn.commit()
        return count

    def get_ids_by_tag(self, tag: str) -> List[str]:
        """获取包含指定标签的记忆 ID 列表（SQL 级过滤，避免全表加载）

        v5.5.4 新增：供 mindforge 层清理索引用，替代 list_memories 全表加载方案。

        Args:
            tag: 标签名

        Returns:
            匹配的记忆 ID 列表
        """
        if not tag:
            return []
        tag = self._strip_control(tag)[:64]
        conn = self._get_conn()
        rows = conn.execute(
            "SELECT id, tags FROM memories WHERE tags LIKE ? AND category != 'trash'",
            (f'%"{tag}"%',)
        ).fetchall()
        result = []
        for row in rows:
            try:
                existing_tags = json.loads(row["tags"]) if isinstance(row["tags"], str) else row["tags"]
                if tag in (existing_tags or []):
                    result.append(row["id"])
            except (json.JSONDecodeError, TypeError):
                continue
        return result

    # ===== 记忆合并（v5.5.4 新增）=====

    def merge_memories(self, source_id: str, target_id: str,
                       actor: str = "", session_id: str = "") -> Optional[MemoryEntry]:
        """合并两条记忆

        v5.5.4 新增。

        合并规则：
        - 内容：按行去重合并（target 在前，source 在后）
        - 标签：取并集
        - 重要性：取较高值
        - 层级：取较深层级（数值更大）
        - 访问次数：相加
        - 元数据：合并（target 优先，source 补充不存在的 key）
        - source 记忆移入回收站（软删除）

        Args:
            source_id: 源记忆 ID（将被移入回收站）
            target_id: 目标记忆 ID（保留并吸收内容）
            actor: 操作者
            session_id: 会话 ID

        Returns:
            合并后的目标记忆条目，失败返回 None
        """
        if not source_id or not target_id or source_id == target_id:
            return None
        # v5.4.0 安全加固
        source_id = self._strip_control(source_id)[:64]
        target_id = self._strip_control(target_id)[:64]
        actor = self._strip_control(actor)[:128]
        session_id = self._strip_control(session_id)[:128]

        conn = self._get_conn()
        now = time.time()

        source_row = conn.execute(
            "SELECT * FROM memories WHERE id = ? AND category != 'trash'",
            (source_id,)
        ).fetchone()
        target_row = conn.execute(
            "SELECT * FROM memories WHERE id = ? AND category != 'trash'",
            (target_id,)
        ).fetchone()
        if not source_row or not target_row:
            return None

        source = self._row_to_entry(source_row)
        target = self._row_to_entry(target_row)

        # 内容按行去重合并
        target_lines = target.content.splitlines() if target.content else []
        source_lines = source.content.splitlines() if source.content else []
        seen = set(target_lines)
        merged_lines = list(target_lines)
        for line in source_lines:
            if line and line not in seen:
                merged_lines.append(line)
                seen.add(line)
        merged_content = "\n".join(merged_lines)

        # 标签取并集
        target_tags = set(target.tags or [])
        source_tags = set(source.tags or [])
        merged_tags = list(target_tags | source_tags)

        # 重要性取高
        # v5.5.5 fix: 使用 _downgrade_enum 兼容双重导入枚举类不一致
        source_imp = self._downgrade_enum(source.importance, Importance, Importance.MEDIUM)
        target_imp = self._downgrade_enum(target.importance, Importance, Importance.MEDIUM)
        merged_importance = target_imp if target_imp.to_int() >= source_imp.to_int() else source_imp
        merged_importance_val = merged_importance.value

        # 层级取深（数值更大的层级更深）
        # v5.5.5 fix: 使用 _downgrade_enum 兼容双重导入枚举类不一致
        source_layer = self._downgrade_enum(source.layer, MemoryLayer, MemoryLayer.SHORT_TERM)
        target_layer = self._downgrade_enum(target.layer, MemoryLayer, MemoryLayer.SHORT_TERM)
        merged_layer = target_layer if target_layer.to_int() >= source_layer.to_int() else source_layer

        # 访问次数相加
        merged_access_count = source.access_count + target.access_count

        # 元数据合并（target 优先）
        merged_metadata = dict(source.metadata or {})
        merged_metadata.update(target.metadata or {})

        # 平均 strength 和 forgetting_score
        merged_strength = (source.strength + target.strength) / 2
        merged_forgetting = (source.forgetting_score + target.forgetting_score) / 2

        # 更新目标记忆
        conn.execute(
            """UPDATE memories SET
                   content = ?, tags = ?, importance = ?, layer = ?,
                   access_count = ?, metadata = ?, strength = ?,
                   forgetting_score = ?, updated_at = ?
               WHERE id = ?""",
            (
                merged_content,
                json.dumps(merged_tags, ensure_ascii=False),
                merged_importance_val,
                merged_layer.value,
                merged_access_count,
                json.dumps(merged_metadata, ensure_ascii=False),
                merged_strength,
                merged_forgetting,
                now,
                target_id,
            )
        )

        # 源记忆移入回收站
        conn.execute(
            "UPDATE memories SET category = 'trash', updated_at = ? WHERE id = ?",
            (now, source_id)
        )

        # 审计日志
        self._add_audit("update", target_id, actor, session_id, target.privacy.value,
                         details={"action": "merge_from", "source_id": source_id})
        self._add_audit("delete", source_id, actor, session_id, source.privacy.value,
                         details={"reason": "merge_into", "target_id": target_id})

        # v5.5.5 fix: 合并后刷新 FTS 索引，避免全文搜索返回过期数据
        self._refresh_fts(conn, target_id)

        conn.commit()

        # 返回更新后的目标记忆
        updated_row = conn.execute(
            "SELECT * FROM memories WHERE id = ?", (target_id,)
        ).fetchone()
        return self._row_to_entry(updated_row) if updated_row else None

    # ===== 记忆分层与时间衰减（v5.5.5 新增）=====

    def memory_decay_weights(self, memory_id: str) -> Dict[str, float]:
        """计算单条记忆的综合权重（v5.5.5 新增）

        用于检索排序和自动分层，包含近期度、重要度、访问频率、强度四个维度。

        权重公式：
        - recency_weight = exp(-(now - last_accessed_at) / (7*86400))  # 7天半衰期
        - importance_weight = importance.to_int() / 3.0  # 0~1
        - access_weight = min(1.0, log(access_count + 1) / log(10))  # 访问次数对数缩放
        - strength_weight = min(1.0, max(0.0, strength))
        - final_score = 0.3*recency + 0.3*importance + 0.2*access + 0.2*strength

        Args:
            memory_id: 记忆 ID

        Returns:
            包含各权重分量和 final_score 的字典
        """
        from .types import Importance

        conn = self._get_conn()
        row = conn.execute(
            "SELECT * FROM memories WHERE id = ?",
            (memory_id,)
        ).fetchone()
        if not row:
            return {
                "recency_weight": 0.0,
                "importance_weight": 0.0,
                "access_weight": 0.0,
                "strength_weight": 0.0,
                "final_score": 0.0,
            }

        entry = self._row_to_entry(row)
        now = time.time()

        # 1. 近期度权重（7天半衰期的指数衰减）
        time_since_access = now - entry.last_accessed_at if entry.last_accessed_at > 0 else now
        recency_weight = math.exp(-time_since_access / (7 * 86400))

        # 2. 重要度权重（0~1 归一化）
        # v5.5.5 fix: 使用 _downgrade_enum 兼容双重导入
        imp = self._downgrade_enum(entry.importance, Importance, Importance.MEDIUM)
        importance_weight = imp.to_int() / 3.0

        # 3. 访问次数权重（对数缩放，防止高访问次数过度主导）
        access_weight = min(1.0, math.log(entry.access_count + 1) / math.log(10))

        # 4. 强度权重（直接使用，钳位到 0~1）
        strength_weight = min(1.0, max(0.0, entry.strength))

        # 5. 综合得分
        final_score = (
            0.3 * recency_weight
            + 0.3 * importance_weight
            + 0.2 * access_weight
            + 0.2 * strength_weight
        )

        return {
            "recency_weight": round(recency_weight, 4),
            "importance_weight": round(importance_weight, 4),
            "access_weight": round(access_weight, 4),
            "strength_weight": round(strength_weight, 4),
            "final_score": round(final_score, 4),
        }

    def auto_layer_consolidate(self, dry_run: bool = False) -> Dict[str, Any]:
        """自动记忆分层整合（v5.5.5 核心功能）

        遍历所有非 trash 记忆，根据综合权重（memory_decay_weights）自动调整记忆层级。

        升级规则（从低层到高层）：
        - SHORT_TERM → LONG_TERM: final_score >= 0.6 且 access_count >= 3
        - LONG_TERM → PERMANENT: final_score >= 0.8 且 access_count >= 10 且 consolidation_count >= 2

        降级规则（从高层到低层）：
        - PERMANENT → LONG_TERM: final_score < 0.4 且 30天未访问
        - LONG_TERM → SHORT_TERM: final_score < 0.2 且 14天未访问

        低优先级标记：
        - SHORT_TERM 且 final_score < 0.1 且 7天未访问 → metadata._low_priority = true

        Args:
            dry_run: 是否仅模拟不实际修改（True 时只返回统计结果）

        Returns:
            统计字典，包含：
            - promoted: 升级的记忆数量
            - demoted: 降级的记忆数量
            - low_priority: 标记为低优先级的记忆数量
            - dry_run: 是否为试运行
            - details: 详细变更列表 [{id, action, from_layer, to_layer, final_score, reason}]
        """
        from .types import MemoryLayer

        conn = self._get_conn()
        now = time.time()

        rows = conn.execute(
            "SELECT * FROM memories WHERE category != 'trash'"
        ).fetchall()

        promoted = 0
        demoted = 0
        low_priority_count = 0
        details = []

        for row in rows:
            entry = self._row_to_entry(row)
            weights = self.memory_decay_weights(entry.id)
            final_score = weights["final_score"]
            time_since_access = now - entry.last_accessed_at if entry.last_accessed_at > 0 else float('inf')

            # v5.5.5 fix: 使用 _downgrade_enum 兼容双重导入
            layer = self._downgrade_enum(entry.layer, MemoryLayer, MemoryLayer.SHORT_TERM)
            new_layer = layer
            action = None
            reason = ""

            # === 升级检查 ===
            if layer == MemoryLayer.SHORT_TERM:
                if final_score >= 0.6 and entry.access_count >= 3:
                    new_layer = MemoryLayer.LONG_TERM
                    action = "promote"
                    reason = (f"SHORT_TERM→LONG_TERM: score={final_score}>=0.6, "
                              f"access_count={entry.access_count}>=3")
                    promoted += 1
                # 低优先级标记（仅 SHORT_TERM 层）
                elif final_score < 0.1 and time_since_access > 7 * 86400:
                    action = "low_priority"
                    reason = (f"SHORT_TERM low_priority: score={final_score}<0.1, "
                              f"idle_days={time_since_access / 86400:.1f}>7")
                    low_priority_count += 1

            elif layer == MemoryLayer.LONG_TERM:
                if final_score >= 0.8 and entry.access_count >= 10 and entry.consolidation_count >= 2:
                    new_layer = MemoryLayer.PERMANENT
                    action = "promote"
                    reason = (f"LONG_TERM→PERMANENT: score={final_score}>=0.8, "
                              f"access_count={entry.access_count}>=10, "
                              f"consolidation={entry.consolidation_count}>=2")
                    promoted += 1
                elif final_score < 0.2 and time_since_access > 14 * 86400:
                    new_layer = MemoryLayer.SHORT_TERM
                    action = "demote"
                    reason = (f"LONG_TERM→SHORT_TERM: score={final_score}<0.2, "
                              f"idle_days={time_since_access / 86400:.1f}>14")
                    demoted += 1

            elif layer == MemoryLayer.PERMANENT:
                if final_score < 0.4 and time_since_access > 30 * 86400:
                    new_layer = MemoryLayer.LONG_TERM
                    action = "demote"
                    reason = (f"PERMANENT→LONG_TERM: score={final_score}<0.4, "
                              f"idle_days={time_since_access / 86400:.1f}>30")
                    demoted += 1

            if action:
                details.append({
                    "id": entry.id,
                    "action": action,
                    "from_layer": layer.value,
                    "to_layer": new_layer.value,
                    "final_score": final_score,
                    "reason": reason,
                })

                if not dry_run:
                    if action in ("promote", "demote"):
                        conn.execute(
                            """UPDATE memories SET
                                   layer = ?,
                                   consolidation_count = consolidation_count + 1,
                                   updated_at = ?
                               WHERE id = ?""",
                            (new_layer.value, now, entry.id)
                        )
                        priv_val = entry.privacy.value if hasattr(entry.privacy, 'value') else str(entry.privacy)
                        self._add_audit(
                            "consolidate", entry.id, "system", "", priv_val,
                            details={
                                "action": action,
                                "from_layer": layer.value,
                                "to_layer": new_layer.value,
                                "score": final_score,
                            }
                        )
                    elif action == "low_priority":
                        metadata = dict(entry.metadata or {})
                        metadata["_low_priority"] = True
                        conn.execute(
                            "UPDATE memories SET metadata = ?, updated_at = ? WHERE id = ?",
                            (json.dumps(metadata, ensure_ascii=False), now, entry.id)
                        )

        if not dry_run:
            conn.commit()

        return {
            "promoted": promoted,
            "demoted": demoted,
            "low_priority": low_priority_count,
            "dry_run": dry_run,
            "details": details,
        }

    def layered_retrieval(self, query: str, limit: int = 20,
                          min_score: float = 0.1) -> List[MemoryEntry]:
        """分层检索（v5.5.5 新检索入口）

        按记忆层级优先级检索：PERMANENT（高）→ LONG_TERM（中）→ SHORT_TERM（低），
        每层按比例分配检索名额，使用 fuzzy_search 作为基础匹配，
        过滤低权重记忆后按综合权重排序返回。

        分层配额：
        - PERMANENT: limit * 0.3（高优先级）
        - LONG_TERM: limit * 0.5（中优先级）
        - SHORT_TERM: limit * 0.2（低优先级）

        Args:
            query: 搜索关键词
            limit: 返回结果总数上限
            min_score: 最低综合权重阈值（final_score 低于此值的记忆被过滤）

        Returns:
            按 final_score 降序排列的记忆条目列表，最多 limit 条
        """
        from .types import MemoryLayer

        # 各层检索配置：(层级, 配额比例)
        layer_config = [
            (MemoryLayer.PERMANENT, 0.3),
            (MemoryLayer.LONG_TERM, 0.5),
            (MemoryLayer.SHORT_TERM, 0.2),
        ]

        collected = []  # List of (entry, final_score)
        seen_ids = set()

        for layer, ratio in layer_config:
            layer_quota = max(1, int(limit * ratio))
            # 多取一些，供 min_score 过滤后仍能凑够配额
            search_results = self.fuzzy_search(query, layer=layer, limit=layer_quota * 3)

            layer_count = 0
            for result in search_results:
                entry = result["entry"]
                if entry.id in seen_ids:
                    continue

                weights = self.memory_decay_weights(entry.id)
                final_score = weights["final_score"]

                if final_score >= min_score:
                    collected.append((entry, final_score))
                    seen_ids.add(entry.id)
                    layer_count += 1

                if layer_count >= layer_quota:
                    break

        # 按 final_score 降序排列，取前 limit 条
        collected.sort(key=lambda x: x[1], reverse=True)
        return [entry for entry, _ in collected[:limit]]

    def get_memory_layers_stats(self) -> Dict[str, Any]:
        """获取各层记忆统计（v5.5.5 新增）

        统计每层的数量、平均重要度、平均访问次数、平均遗忘分数。

        Returns:
            以层级 value 为 key 的统计字典，每层包含：
            - count: 记忆数量
            - avg_importance: 平均重要度（0~3 数值）
            - avg_access_count: 平均访问次数
            - avg_forgetting_score: 平均遗忘分数
        """
        from .types import MemoryLayer

        conn = self._get_conn()

        rows = conn.execute("""
            SELECT layer,
                   COUNT(*) as count,
                   AVG(CASE importance
                       WHEN 'LOW' THEN 0
                       WHEN 'MEDIUM' THEN 1
                       WHEN 'HIGH' THEN 2
                       WHEN 'CRITICAL' THEN 3
                       ELSE 1 END) as avg_importance,
                   AVG(access_count) as avg_access_count,
                   AVG(forgetting_score) as avg_forgetting_score
            FROM memories
            WHERE category != 'trash'
            GROUP BY layer
        """).fetchall()

        stats: Dict[str, Any] = {}
        for row in rows:
            layer_val = row["layer"]
            try:
                layer = MemoryLayer(layer_val)
            except ValueError:
                continue
            stats[layer.value] = {
                "count": row["count"],
                "avg_importance": round(row["avg_importance"] or 0.0, 4),
                "avg_access_count": round(row["avg_access_count"] or 0.0, 4),
                "avg_forgetting_score": round(row["avg_forgetting_score"] or 0.0, 4),
            }

        # 确保所有层都有记录（即使数量为 0）
        for layer in MemoryLayer:
            if layer.value not in stats:
                stats[layer.value] = {
                    "count": 0,
                    "avg_importance": 0.0,
                    "avg_access_count": 0.0,
                    "avg_forgetting_score": 0.0,
                }

        return stats

    # ===== 最常访问 / 最近访问（v5.5.4 新增）=====

    def most_accessed(self, limit: int = 10,
                      category: Optional[str] = None,
                      layer: Optional[MemoryLayer] = None) -> List[MemoryEntry]:
        """按访问次数降序返回最常访问的记忆

        v5.5.4 新增。

        Args:
            limit: 返回数量上限
            category: 按分类筛选（可选）
            layer: 按层级筛选（可选）

        Returns:
            记忆条目列表，按 access_count 降序
        """
        conn = self._get_conn()
        query = "SELECT * FROM memories WHERE category != 'trash'"
        params = []

        if category:
            query += " AND category = ?"
            params.append(category)
        if layer:
            query += " AND layer = ?"
            params.append(layer.value)

        query += " ORDER BY access_count DESC, last_accessed_at DESC LIMIT ?"
        params.append(max(1, min(limit, 1000)))

        rows = conn.execute(query, params).fetchall()
        return [self._row_to_entry(row) for row in rows]

    def recently_accessed(self, limit: int = 10,
                          category: Optional[str] = None,
                          layer: Optional[MemoryLayer] = None) -> List[MemoryEntry]:
        """按最近访问时间降序返回最近访问的记忆

        v5.5.4 新增。

        Args:
            limit: 返回数量上限
            category: 按分类筛选（可选）
            layer: 按层级筛选（可选）

        Returns:
            记忆条目列表，按 last_accessed_at 降序
        """
        conn = self._get_conn()
        query = "SELECT * FROM memories WHERE category != 'trash' AND last_accessed_at > 0"
        params = []

        if category:
            query += " AND category = ?"
            params.append(category)
        if layer:
            query += " AND layer = ?"
            params.append(layer.value)

        query += " ORDER BY last_accessed_at DESC LIMIT ?"
        params.append(max(1, min(limit, 1000)))

        rows = conn.execute(query, params).fetchall()
        return [self._row_to_entry(row) for row in rows]

    # ===== 按筛选批量更新（v5.5.4 新增）=====

    def bulk_update_by_filter(self,
                              category: Optional[str] = None,
                              tag: Optional[str] = None,
                              updates: Optional[Dict[str, Any]] = None,
                              actor: str = "",
                              session_id: str = "") -> int:
        """按分类或标签筛选后批量更新记忆

        v5.5.4 新增。

        支持的更新字段（updates dict 的 key）：
        - new_category: 新分类
        - new_layer: 新层级 (MemoryLayer)
        - new_importance: 新重要性 (0-10)
        - new_privacy: 新隐私级别 (PrivacyLevel)
        - add_tags: 要添加的标签列表
        - remove_tags: 要移除的标签列表
        - starred: 是否星标 (bool)
        - pinned: 是否置顶 (bool)

        Args:
            category: 按分类筛选（二选一或同时使用）
            tag: 按标签筛选（二选一或同时使用）
            updates: 要更新的字段字典
            actor: 操作者
            session_id: 会话 ID

        Returns:
            更新的记忆条数
        """
        if not updates or (not category and not tag):
            return 0
        # v5.4.0 安全加固
        if category:
            category = self._strip_control(category)[:64]
        if tag:
            tag = self._strip_control(tag)[:64]
        actor = self._strip_control(actor)[:128]
        session_id = self._strip_control(session_id)[:128]

        from .types import MemoryLayer, PrivacyLevel, Importance

        conn = self._get_conn()
        now = time.time()

        # 先筛选匹配的记忆
        query = "SELECT * FROM memories WHERE category != 'trash'"
        params = []
        if category:
            query += " AND category = ?"
            params.append(category)
        if tag:
            query += " AND tags LIKE ?"
            params.append(f'%"{tag}"%')

        rows = conn.execute(query, params).fetchall()
        if not rows:
            return 0

        count = 0
        for row in rows:
            entry = self._row_to_entry(row)

            # 如果按 tag 筛选，在 Python 端再精确验证
            if tag:
                if tag not in (entry.tags or []):
                    continue

            # 构建更新字段
            new_entry_tags = list(entry.tags or [])
            new_category = entry.category
            new_layer = entry.layer
            new_importance = entry.importance
            new_privacy = entry.privacy
            new_starred = entry.starred
            new_pinned = entry.pinned
            new_metadata = dict(entry.metadata or {})

            if "new_category" in updates and updates["new_category"]:
                new_category = self._strip_control(updates["new_category"])[:64]

            if "new_layer" in updates and updates["new_layer"] is not None:
                new_layer = updates["new_layer"]

            if "new_importance" in updates and updates["new_importance"] is not None:
                new_imp_val = updates["new_importance"]
                # v5.5.5 fix: 使用 _downgrade_enum 兼容双重导入
                coerced = self._downgrade_enum(new_imp_val, Importance, None)
                if coerced is not None:
                    new_importance = coerced
                else:
                    # 如果是数字，尝试映射到最近的级别
                    try:
                        level = int(float(new_imp_val))
                        if level <= 0:
                            new_importance = Importance.LOW
                        elif level == 1:
                            new_importance = Importance.MEDIUM
                        elif level == 2:
                            new_importance = Importance.HIGH
                        else:
                            new_importance = Importance.CRITICAL
                    except (ValueError, TypeError):
                        pass  # 保持 entry.importance 不变

            if "new_privacy" in updates and updates["new_privacy"] is not None:
                new_privacy = updates["new_privacy"]

            if "add_tags" in updates and updates["add_tags"]:
                # v5.5.6 fix: 大小写不敏感去重，与 add_tags_to_ids / batch_add_tags 一致
                existing_lower = {t.lower() for t in new_entry_tags if isinstance(t, str)}
                for t in updates["add_tags"]:
                    t_clean = self._strip_control(t)[:64]
                    if t_clean and t_clean.lower() not in existing_lower:
                        new_entry_tags.append(t_clean)
                        existing_lower.add(t_clean.lower())

            if "remove_tags" in updates and updates["remove_tags"]:
                for t in updates["remove_tags"]:
                    if t in new_entry_tags:
                        new_entry_tags.remove(t)

            if "starred" in updates and updates["starred"] is not None:
                new_starred = bool(updates["starred"])

            if "pinned" in updates and updates["pinned"] is not None:
                new_pinned = bool(updates["pinned"])

            # 执行更新
            # v5.5.5 fix: 使用 _downgrade_enum 兼容双重导入枚举类不一致
            layer_val = self._downgrade_enum(new_layer, MemoryLayer, MemoryLayer.SHORT_TERM).value
            privacy_val = self._downgrade_enum(new_privacy, PrivacyLevel, PrivacyLevel.INTERNAL).value
            imp_val = self._downgrade_enum(new_importance, Importance, Importance.MEDIUM).value

            conn.execute(
                """UPDATE memories SET
                       category = ?, layer = ?, importance = ?, privacy = ?,
                       tags = ?, starred = ?, pinned = ?, metadata = ?, updated_at = ?
                   WHERE id = ?""",
                (
                    new_category,
                    layer_val,
                    imp_val,
                    privacy_val,
                    json.dumps(new_entry_tags, ensure_ascii=False),
                    1 if new_starred else 0,
                    1 if new_pinned else 0,
                    json.dumps(new_metadata, ensure_ascii=False),
                    now,
                    entry.id,
                )
            )

            self._add_audit("update", entry.id, actor, session_id,
                             entry.privacy.value,
                             details={"action": "bulk_update", "changes": list(updates.keys())})
            # v5.5.5 fix: 批量更新后刷新 FTS 索引，避免全文搜索返回过期数据
            self._refresh_fts(conn, entry.id)
            count += 1

        conn.commit()
        return count

    # ===== 标签统计（v5.5.4 新增）=====

    def tag_stats(self, limit: int = 20) -> Dict[str, Any]:
        """标签统计信息

        v5.5.4 新增。

        Args:
            limit: Top 标签数量

        Returns:
            统计字典：total_tags, memories_with_tags, memories_without_tags,
            avg_tags_per_memory, top_tags (list of {tag, count, categories})
        """
        conn = self._get_conn()
        limit = max(1, min(limit, 100))

        rows = conn.execute(
            "SELECT tags, category FROM memories WHERE category != 'trash'"
        ).fetchall()

        total_memories = len(rows)
        memories_with_tags = 0
        total_tag_count = 0
        tag_counter: Dict[str, Dict[str, Any]] = {}

        for row in rows:
            if not row["tags"]:
                continue
            try:
                tags = json.loads(row["tags"]) if isinstance(row["tags"], str) else row["tags"]
            except (json.JSONDecodeError, TypeError):
                continue
            if not tags:
                continue
            memories_with_tags += 1
            total_tag_count += len(tags)
            for tag in tags:
                if tag not in tag_counter:
                    tag_counter[tag] = {"count": 0, "categories": set()}
                tag_counter[tag]["count"] += 1
                tag_counter[tag]["categories"].add(row["category"])

        # 排序取 top
        sorted_tags = sorted(tag_counter.items(), key=lambda x: x[1]["count"], reverse=True)[:limit]
        top_tags = [
            {
                "tag": tag,
                "count": info["count"],
                "categories": sorted(list(info["categories"])),
            }
            for tag, info in sorted_tags
        ]

        avg_tags = round(total_tag_count / memories_with_tags, 2) if memories_with_tags > 0 else 0

        return {
            "total_tags": len(tag_counter),
            "memories_with_tags": memories_with_tags,
            "memories_without_tags": total_memories - memories_with_tags,
            "avg_tags_per_memory": avg_tags,
            "top_tags": top_tags,
        }

    # ===== 索引一致性检查（v5.5.4 新增）=====

    def check_index_consistency(self) -> Dict[str, Any]:
        """检查 FTS5 索引与主表的一致性

        v5.5.4 新增。

        检测两类问题（contentless FTS5 无法对比内容）：
        - missing_in_fts: 主表有记录，但 FTS5 索引中缺失
        - stale_in_fts: FTS5 索引中有记录，但主表中已删除或已进入回收站

        Returns:
            检查结果字典
        """
        conn = self._get_conn()
        issues = {
            "missing_in_fts": [],     # 主表有，FTS 没有
            "stale_in_fts": [],       # FTS 有，主表没有或已删除
        }

        try:
            # 获取主表中所有非删除记忆的 rowid
            main_rows = conn.execute(
                "SELECT m.rowid, m.id FROM memories m WHERE m.category != 'trash'"
            ).fetchall()

            # 获取 FTS 表中所有 rowid（contentless FTS5 只能查 rowid）
            fts_rows = conn.execute(
                "SELECT rowid FROM memory_fts"
            ).fetchall()

        except sqlite3.OperationalError:
            # FTS 表可能不存在
            return {
                "status": "error",
                "error": "memory_fts table not found or inaccessible",
                "total_main": 0,
                "total_fts": 0,
                "issues_found": 0,
                "issues": issues,
            }

        main_ids = {row["rowid"]: row["id"] for row in main_rows}
        fts_ids = {row["rowid"] for row in fts_rows}

        # 检查缺失（主表有，FTS 没有）
        for rowid, memory_id in main_ids.items():
            if rowid not in fts_ids:
                issues["missing_in_fts"].append({
                    "rowid": rowid,
                    "memory_id": memory_id,
                })

        # 检查陈旧（FTS 有，主表没有或已删除）
        for rowid in fts_ids:
            if rowid not in main_ids:
                issues["stale_in_fts"].append({
                    "rowid": rowid,
                })

        total_issues = len(issues["missing_in_fts"]) + len(issues["stale_in_fts"])

        return {
            "status": "ok" if total_issues == 0 else "issues_found",
            "total_main": len(main_ids),
            "total_fts": len(fts_ids),
            "issues_found": total_issues,
            "issues": issues,
        }

    def _delete_memory_cascade(self, conn: sqlite3.Connection, memory_id: str):
        """级联删除记忆及其关联数据（v5.5.2 新增内部辅助方法）

        v5.5.2 修复：补充 review_schedules、kg_edges 清理，并修正 FTS5 删除方式
        （contentless FTS5 表必须用 'delete' 特殊命令，不能用普通 DELETE）。
        """
        # 先取 rowid 和 FTS 字段用于清理 FTS 索引
        row = conn.execute(
            "SELECT rowid, content, category, tags FROM memories WHERE id = ?",
            (memory_id,)
        ).fetchone()
        # 清理带外键的从表
        conn.execute("DELETE FROM memory_versions WHERE memory_id = ?", (memory_id,))
        conn.execute("DELETE FROM memory_links WHERE source_id = ? OR target_id = ?", (memory_id, memory_id))
        conn.execute("DELETE FROM review_schedules WHERE memory_id = ?", (memory_id,))
        try:
            conn.execute("DELETE FROM memory_notes WHERE memory_id = ?", (memory_id,))
        except sqlite3.OperationalError:
            pass
        try:
            conn.execute("DELETE FROM kg_edges WHERE source = ? OR target = ?", (memory_id, memory_id))
        except sqlite3.OperationalError:
            pass
        conn.execute("DELETE FROM memory_embeddings WHERE memory_id = ?", (memory_id,))
        try:
            conn.execute("DELETE FROM memory_templates WHERE memory_id = ?", (memory_id,))
        except sqlite3.OperationalError:
            pass
        if row:
            try:
                # contentless FTS5 必须用 'delete' 特殊命令
                conn.execute(
                    "INSERT INTO memory_fts(memory_fts, rowid, content, category, tags) "
                    "VALUES('delete', ?, ?, ?, ?)",
                    (row[0], row[1] or "", row[2] or "", row[3] or "[]")
                )
            except sqlite3.OperationalError:
                pass
        conn.execute("DELETE FROM memories WHERE id = ?", (memory_id,))

    # ===== 标签批量管理（v5.2.0 新增）=====

    def batch_add_tags(self,
                       entry_ids: List[str],
                       tags: List[str],
                       actor: str = "",
                       session_id: str = "") -> int:
        """批量添加标签（v5.2.0 新增）

        Args:
            entry_ids: 记忆 ID 列表
            tags: 要添加的标签列表
            actor: 操作者
            session_id: 会话 ID

        Returns:
            受影响的记忆条数
        """
        conn = self._get_conn()
        count = 0
        now = time.time()

        for entry_id in entry_ids:
            row = conn.execute(
                "SELECT tags, privacy FROM memories WHERE id = ? AND category != 'trash'",
                (entry_id,)
            ).fetchone()
            if not row:
                continue

            existing_tags = self._safe_json_loads(row["tags"], [])
            # v5.5.6 fix: 保序去重（原 list(set()) 会打乱顺序），与 add_tags_to_ids 一致
            new_tags = list(existing_tags)
            seen_lower = {t.lower() for t in new_tags if isinstance(t, str)}
            for t in tags:
                if isinstance(t, str) and t.strip() and t.lower() not in seen_lower:
                    new_tags.append(t)
                    seen_lower.add(t.lower())
            if new_tags != existing_tags:
                conn.execute(
                    "UPDATE memories SET tags = ?, updated_at = ? WHERE id = ?",
                    (json.dumps(new_tags, ensure_ascii=False), now, entry_id)
                )
                # v5.5.5 fix: 更新标签后刷新 FTS 索引
                self._refresh_fts(conn, entry_id)
                self._add_audit(
                    "batch_add_tags", entry_id, actor, session_id,
                    row["privacy"] if row["privacy"] else "",
                    details={"added_tags": tags, "result_tags": new_tags}
                )
                count += 1

        conn.commit()
        return count

    def batch_remove_tags(self,
                          entry_ids: List[str],
                          tags: List[str],
                          actor: str = "",
                          session_id: str = "") -> int:
        """批量移除标签（v5.2.0 新增）

        Args:
            entry_ids: 记忆 ID 列表
            tags: 要移除的标签列表
            actor: 操作者
            session_id: 会话 ID

        Returns:
            受影响的记忆条数
        """
        conn = self._get_conn()
        count = 0
        now = time.time()

        for entry_id in entry_ids:
            row = conn.execute(
                "SELECT tags, privacy FROM memories WHERE id = ? AND category != 'trash'",
                (entry_id,)
            ).fetchone()
            if not row:
                continue

            existing_tags = self._safe_json_loads(row["tags"], [])
            new_tags = [t for t in existing_tags if t not in tags]
            if new_tags != existing_tags:
                conn.execute(
                    "UPDATE memories SET tags = ?, updated_at = ? WHERE id = ?",
                    (json.dumps(new_tags, ensure_ascii=False), now, entry_id)
                )
                # v5.5.5 fix: 更新标签后刷新 FTS 索引
                self._refresh_fts(conn, entry_id)
                self._add_audit(
                    "batch_remove_tags", entry_id, actor, session_id,
                    row["privacy"] if row["privacy"] else "",
                    details={"removed_tags": tags, "result_tags": new_tags}
                )
                count += 1

        conn.commit()
        return count

    def merge_tags(self,
                   source_tags: List[str],
                   target_tag: str,
                   actor: str = "",
                   session_id: str = "") -> int:
        """合并多个标签为一个标签（v5.2.0 新增）

        Args:
            source_tags: 要合并的源标签列表
            target_tag: 目标标签名
            actor: 操作者
            session_id: 会话 ID

        Returns:
            受影响的记忆条数
        """
        conn = self._get_conn()
        count = 0
        now = time.time()

        rows = conn.execute(
            "SELECT id, tags, privacy FROM memories WHERE category != 'trash'"
        ).fetchall()

        for row in rows:
            tags = self._safe_json_loads(row["tags"], [])
            has_source = any(t in source_tags for t in tags)
            if not has_source:
                continue

            new_tags = [t for t in tags if t not in source_tags]
            if target_tag not in new_tags:
                new_tags.append(target_tag)

            if new_tags != tags:
                conn.execute(
                    "UPDATE memories SET tags = ?, updated_at = ? WHERE id = ?",
                    (json.dumps(new_tags, ensure_ascii=False), now, row["id"])
                )
                # v5.5.5 fix: 更新标签后刷新 FTS 索引
                self._refresh_fts(conn, row["id"])
                self._add_audit(
                    "merge_tags", row["id"], actor, session_id,
                    row["privacy"] if row["privacy"] else "",
                    details={"source_tags": source_tags, "target_tag": target_tag}
                )
                count += 1

        conn.commit()
        return count

    def add_tags_by_category(self,
                             category: str,
                             tags: List[str],
                             actor: str = "",
                             session_id: str = "") -> int:
        """按分类批量添加标签（v5.2.0 新增）

        Args:
            category: 分类名
            tags: 要添加的标签列表
            actor: 操作者
            session_id: 会话 ID

        Returns:
            受影响的记忆条数
        """
        conn = self._get_conn()
        rows = conn.execute(
            "SELECT id FROM memories WHERE category = ? AND category != 'trash'",
            (category,)
        ).fetchall()
        entry_ids = [r["id"] for r in rows]
        return self.batch_add_tags(entry_ids, tags, actor, session_id)

    # ===== 数据备份与恢复（v5.2.0 新增）=====

    def create_backup(self, backup_dir: str = "./data/backups") -> Dict[str, Any]:
        """创建数据库备份（v5.2.0 新增）

        Args:
            backup_dir: 备份目录

        Returns:
            {path, size_mb, timestamp, success}
        """
        import shutil
        # v5.4.8 安全修复：路径遍历防护
        backup_dir = _safe_path(backup_dir, allowed_exts=None)
        backup_path = Path(backup_dir)
        backup_path.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file = backup_path / f"memory_backup_{timestamp}.db"

        try:
            shutil.copy2(str(self.db_path), str(backup_file))
            size_mb = round(backup_file.stat().st_size / (1024 * 1024), 2)
            return {
                "success": True,
                "path": str(backup_file),
                "size_mb": size_mb,
                "timestamp": timestamp,
                "filename": backup_file.name,
            }
        except (OSError, IOError) as e:
            return {
                "success": False,
                "error": str(e),
                "path": str(backup_file),
            }

    def list_backups(self, backup_dir: str = "./data/backups") -> List[Dict[str, Any]]:
        """列出所有备份（v5.2.0 新增）

        Args:
            backup_dir: 备份目录

        Returns:
            备份列表 [{filename, path, size_mb, created_at}]
        """
        # v5.4.8 安全修复：路径遍历防护
        backup_dir = _safe_path(backup_dir, allowed_exts=None)
        backup_path = Path(backup_dir)
        if not backup_path.exists():
            return []

        backups = []
        for f in sorted(backup_path.glob("memory_backup_*.db"), reverse=True):
            stat = f.stat()
            backups.append({
                "filename": f.name,
                "path": str(f),
                "size_mb": round(stat.st_size / (1024 * 1024), 2),
                "created_at": stat.st_ctime,
            })
        return backups

    def restore_backup(self, backup_path: str,
                       create_backup_before: bool = True) -> Dict[str, Any]:
        """从备份恢复数据库（v5.2.0 新增）

        Args:
            backup_path: 备份文件路径
            create_backup_before: 恢复前是否先备份当前数据库

        Returns:
            {success, restored_from, backup_created, error}
        """
        import shutil
        # v5.4.8 安全修复：路径遍历防护
        backup_path = _safe_path(backup_path, allowed_exts={".db"})
        backup_file = Path(backup_path)
        if not backup_file.exists():
            return {"success": False, "error": "备份文件不存在"}

        result = {
            "success": False,
            "restored_from": str(backup_file),
            "backup_created": None,
        }

        pre_backup_path = None
        try:
            if create_backup_before:
                pre_backup = self.create_backup()
                if pre_backup["success"]:
                    result["backup_created"] = pre_backup["path"]
                    pre_backup_path = pre_backup["path"]

            self._close_conns()

            shutil.copy2(str(backup_file), str(self.db_path))

            self._init_db()

            # PRAGMA integrity_check 校验（v5.2.7 新增：确保恢复的数据库结构完整）
            conn_now = self._get_conn()
            try:
                cur = conn_now.execute("PRAGMA integrity_check")
                integrity_row = cur.fetchone()
                cur.close()
            except sqlite3.DatabaseError as ie:
                result["error"] = f"数据库完整性校验异常: {ie}"
                # 校验异常时尝试回滚到恢复前的备份
                if pre_backup_path:
                    try:
                        self._close_conns()
                        shutil.copy2(pre_backup_path, str(self.db_path))
                        self._init_db()
                    except (OSError, IOError):
                        pass
                return result

            if integrity_row is None or str(integrity_row[0]).lower() != "ok":
                integrity_msg = integrity_row[0] if integrity_row else "无结果"
                result["error"] = f"数据库完整性校验失败: {integrity_msg}"
                # 完整性校验失败时尝试回滚到恢复前的备份
                if pre_backup_path:
                    try:
                        self._close_conns()
                        shutil.copy2(pre_backup_path, str(self.db_path))
                        self._init_db()
                    except (OSError, IOError):
                        pass
                return result

            result["success"] = True
        except (OSError, IOError) as e:
            result["error"] = str(e)

        return result

    def delete_old_backups(self, backup_dir: str = "./data/backups",
                           keep_count: int = 10) -> int:
        """删除旧备份，保留最新的 N 个（v5.2.0 新增）

        Args:
            backup_dir: 备份目录
            keep_count: 保留数量

        Returns:
            删除的备份数量
        """
        backups = self.list_backups(backup_dir)
        if len(backups) <= keep_count:
            return 0

        deleted = 0
        for backup in backups[keep_count:]:
            try:
                Path(backup["path"]).unlink()
                deleted += 1
            except (OSError, IOError):
                pass
        return deleted

    # ===== AI 短剧记忆模块（v5.2.1 新增）=====

    # --- 安全验证辅助方法（v5.2.1 新增）---

    @staticmethod
    def _validate_str(value: str, name: str, max_len: int = 1000) -> str:
        """字符串输入验证（v5.2.1 新增）"""
        if value is None:
            return ""
        value = str(value)
        if len(value) > max_len:
            value = value[:max_len]
        return value

    @staticmethod
    def _validate_int(value: int, name: str, min_val: int = 0, max_val: int = 100000) -> int:
        """整数输入验证（v5.2.1 新增）"""
        try:
            value = int(value)
        except (TypeError, ValueError):
            return min_val
        if value < min_val:
            return min_val
        if value > max_val:
            return max_val
        return value

    @staticmethod
    def _validate_float(value: float, name: str, min_val: float = 0.0, max_val: float = 100.0) -> float:
        """浮点数输入验证（v5.2.1 新增）"""
        try:
            value = float(value)
        except (TypeError, ValueError):
            return min_val
        if value < min_val:
            return min_val
        if value > max_val:
            return max_val
        return value

    # --- 短剧系列 ---

    def add_drama(self,
                  title: str,
                  genre: DramaGenre = DramaGenre.OTHER,
                  total_episodes: int = 0,
                  status: DramaStatus = DramaStatus.PLANNED,
                  platform: str = "",
                  rating: float = 0.0,
                  description: str = "",
                  tags: Optional[List[str]] = None,
                  cover_url: str = "",
                  metadata: Optional[Dict[str, Any]] = None) -> DramaSeries:
        """添加短剧（v5.2.1 新增，v5.3.3 安全加固：XSS 消毒）"""
        now = time.time()
        drama_id = str(uuid.uuid4())

        # v5.3.3 安全加固：XSS 消毒
        title = _sanitize_html(self._validate_str(title, "title", max_len=200))
        platform = _sanitize_html(self._validate_str(platform, "platform", max_len=100))
        rating = self._validate_float(rating, "rating", min_val=0.0, max_val=10.0)
        description = _sanitize_html(self._validate_str(description, "description", max_len=5000))
        cover_url = self._validate_str(cover_url, "cover_url", max_len=500)
        total_episodes = self._validate_int(total_episodes, "total_episodes", min_val=0, max_val=10000)

        if not isinstance(genre, DramaGenre):
            genre = DramaGenre.OTHER
        if not isinstance(status, DramaStatus):
            status = DramaStatus.PLANNED

        drama = DramaSeries(
            id=drama_id,
            title=title,
            genre=genre,
            total_episodes=total_episodes,
            current_episode=0,
            status=status,
            platform=platform,
            rating=rating,
            description=description,
            tags=tags or [],
            cover_url=cover_url,
            metadata=metadata or {},
            created_at=now,
            updated_at=now,
            last_watched_at=0.0,
        )

        conn = self._get_conn()
        conn.execute("""
            INSERT INTO drama_series (
                id, title, genre, total_episodes, current_episode,
                status, platform, rating, description, tags,
                cover_url, metadata, created_at, updated_at, last_watched_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            drama.id, drama.title, drama.genre.value,
            drama.total_episodes, drama.current_episode,
            drama.status.value, drama.platform, drama.rating,
            drama.description, json.dumps(drama.tags, ensure_ascii=False),
            drama.cover_url, json.dumps(drama.metadata, ensure_ascii=False),
            drama.created_at, drama.updated_at, drama.last_watched_at,
        ))
        conn.commit()
        return drama

    def get_drama(self, drama_id: str) -> Optional[DramaSeries]:
        """获取短剧详情（v5.2.1 新增）"""
        conn = self._get_conn()
        row = conn.execute(
            "SELECT * FROM drama_series WHERE id = ?",
            (drama_id,)
        ).fetchone()
        if not row:
            return None
        return self._row_to_drama(row)

    def list_dramas(self,
                    genre: Optional[DramaGenre] = None,
                    status: Optional[DramaStatus] = None,
                    platform: Optional[str] = None,
                    min_rating: float = 0.0,
                    limit: int = 50,
                    offset: int = 0,
                    sort_by: str = "updated_at",
                    sort_order: str = "desc") -> List[DramaSeries]:
        """列出短剧（v5.2.1 新增）"""
        conn = self._get_conn()
        query = "SELECT * FROM drama_series WHERE 1=1"
        params = []

        if genre:
            query += " AND genre = ?"
            params.append(genre.value)
        if status:
            query += " AND status = ?"
            params.append(status.value)
        if platform:
            query += " AND platform = ?"
            params.append(platform)
        if min_rating > 0:
            query += " AND rating >= ?"
            params.append(min_rating)

        valid_sorts = ["created_at", "updated_at", "rating", "last_watched_at", "title"]
        if sort_by not in valid_sorts:
            sort_by = "updated_at"
        sort_order = "DESC" if sort_order.lower() == "desc" else "ASC"

        query += f" ORDER BY {sort_by} {sort_order} LIMIT ? OFFSET ?"
        params.extend([limit, offset])

        rows = conn.execute(query, params).fetchall()
        return [self._row_to_drama(r) for r in rows]

    def update_drama(self,
                     drama_id: str,
                     title: Optional[str] = None,
                     genre: Optional[DramaGenre] = None,
                     total_episodes: Optional[int] = None,
                     current_episode: Optional[int] = None,
                     status: Optional[DramaStatus] = None,
                     platform: Optional[str] = None,
                     rating: Optional[float] = None,
                     description: Optional[str] = None,
                     tags: Optional[List[str]] = None,
                     cover_url: Optional[str] = None,
                     metadata: Optional[Dict[str, Any]] = None,
                     mark_watched: bool = False) -> bool:
        """更新短剧信息（v5.2.1 新增）"""
        conn = self._get_conn()
        now = time.time()

        updates = []
        params = []

        if title is not None:
            updates.append("title = ?")
            params.append(self._validate_str(title, "title", max_len=200))
        if genre is not None:
            if not isinstance(genre, DramaGenre):
                genre = DramaGenre.OTHER
            updates.append("genre = ?")
            params.append(genre.value)
        if total_episodes is not None:
            updates.append("total_episodes = ?")
            params.append(self._validate_int(total_episodes, "total_episodes", min_val=0, max_val=10000))
        if current_episode is not None:
            updates.append("current_episode = ?")
            params.append(self._validate_int(current_episode, "current_episode", min_val=0, max_val=10000))
        if status is not None:
            if not isinstance(status, DramaStatus):
                status = DramaStatus.PLANNED
            updates.append("status = ?")
            params.append(status.value)
        if platform is not None:
            updates.append("platform = ?")
            params.append(self._validate_str(platform, "platform", max_len=100))
        if rating is not None:
            updates.append("rating = ?")
            params.append(self._validate_float(rating, "rating", min_val=0.0, max_val=10.0))
        if description is not None:
            updates.append("description = ?")
            params.append(self._validate_str(description, "description", max_len=5000))
        if tags is not None:
            updates.append("tags = ?")
            params.append(json.dumps(tags, ensure_ascii=False))
        if cover_url is not None:
            updates.append("cover_url = ?")
            params.append(self._validate_str(cover_url, "cover_url", max_len=500))
        if metadata is not None:
            updates.append("metadata = ?")
            params.append(json.dumps(metadata, ensure_ascii=False))

        if not updates and not mark_watched:
            return False

        updates.append("updated_at = ?")
        params.append(now)

        if mark_watched:
            updates.append("last_watched_at = ?")
            params.append(now)

        params.append(drama_id)

        cursor = conn.execute(
            f"UPDATE drama_series SET {', '.join(updates)} WHERE id = ?",
            params
        )
        conn.commit()
        return cursor.rowcount > 0

    def delete_drama(self, drama_id: str) -> bool:
        """删除短剧（v5.2.1 新增）
        同时删除关联的场次、角色、台词
        """
        conn = self._get_conn()
        conn.execute("DELETE FROM drama_lines WHERE drama_id = ?", (drama_id,))
        conn.execute("DELETE FROM drama_characters WHERE drama_id = ?", (drama_id,))
        conn.execute("DELETE FROM drama_scenes WHERE drama_id = ?", (drama_id,))
        cursor = conn.execute("DELETE FROM drama_series WHERE id = ?", (drama_id,))
        conn.commit()
        return cursor.rowcount > 0

    def drama_stats(self) -> Dict[str, Any]:
        """短剧统计（v5.2.1 新增）"""
        conn = self._get_conn()
        total = conn.execute("SELECT COUNT(*) as cnt FROM drama_series").fetchone()["cnt"]
        by_genre = {}
        by_status = {}

        for row in conn.execute("SELECT genre, COUNT(*) as cnt FROM drama_series GROUP BY genre"):
            by_genre[row["genre"]] = row["cnt"]
        for row in conn.execute("SELECT status, COUNT(*) as cnt FROM drama_series GROUP BY status"):
            by_status[row["status"]] = row["cnt"]

        watching = conn.execute(
            "SELECT COUNT(*) as cnt FROM drama_series WHERE status = 'watching'"
        ).fetchone()["cnt"]
        completed = conn.execute(
            "SELECT COUNT(*) as cnt FROM drama_series WHERE status = 'completed'"
        ).fetchone()["cnt"]
        total_lines = conn.execute("SELECT COUNT(*) as cnt FROM drama_lines").fetchone()["cnt"]
        classic_lines = conn.execute(
            "SELECT COUNT(*) as cnt FROM drama_lines WHERE is_classic = 1"
        ).fetchone()["cnt"]

        return {
            "total": total,
            "by_genre": by_genre,
            "by_status": by_status,
            "watching": watching,
            "completed": completed,
            "total_lines": total_lines,
            "classic_lines": classic_lines,
        }

    def _row_to_drama(self, row) -> DramaSeries:
        return DramaSeries(
            id=row["id"],
            title=row["title"],
            genre=DramaGenre(row["genre"]) if row["genre"] else DramaGenre.OTHER,
            total_episodes=row["total_episodes"] or 0,
            current_episode=row["current_episode"] or 0,
            status=DramaStatus(row["status"]) if row["status"] else DramaStatus.PLANNED,
            platform=row["platform"] or "",
            rating=row["rating"] or 0.0,
            description=row["description"] or "",
            tags=self._safe_json_loads(row["tags"], []),
            cover_url=row["cover_url"] or "",
            metadata=self._safe_json_loads(row["metadata"], {}),
            created_at=row["created_at"] or 0.0,
            updated_at=row["updated_at"] or 0.0,
            last_watched_at=row["last_watched_at"] or 0.0,
        )

    # --- 短剧场次 ---

    def add_scene(self,
                  drama_id: str,
                  episode: int,
                  scene_number: int,
                  title: str,
                  content: str = "",
                  location: str = "",
                  time_of_day: str = "",
                  tags: Optional[List[str]] = None,
                  metadata: Optional[Dict[str, Any]] = None) -> DramaScene:
        """添加短剧场次（v5.2.1 新增）"""
        now = time.time()
        scene_id = str(uuid.uuid4())

        title = self._validate_str(title, "title", max_len=200)
        content = self._validate_str(content, "content", max_len=10000)
        location = self._validate_str(location, "location", max_len=200)
        time_of_day = self._validate_str(time_of_day, "time_of_day", max_len=50)
        episode = self._validate_int(episode, "episode", min_val=0, max_val=10000)
        scene_number = self._validate_int(scene_number, "scene_number", min_val=0, max_val=10000)

        scene = DramaScene(
            id=scene_id,
            drama_id=drama_id,
            episode=episode,
            scene_number=scene_number,
            title=title,
            content=content,
            location=location,
            time_of_day=time_of_day,
            tags=tags or [],
            metadata=metadata or {},
            created_at=now,
        )

        conn = self._get_conn()
        conn.execute("""
            INSERT INTO drama_scenes (
                id, drama_id, episode, scene_number, title,
                content, location, time_of_day, tags, metadata, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            scene.id, scene.drama_id, scene.episode, scene.scene_number,
            scene.title, scene.content, scene.location, scene.time_of_day,
            json.dumps(scene.tags, ensure_ascii=False),
            json.dumps(scene.metadata, ensure_ascii=False),
            scene.created_at,
        ))
        conn.commit()
        return scene

    def get_scene(self, scene_id: str) -> Optional[DramaScene]:
        """获取场次详情（v5.2.1 新增）"""
        conn = self._get_conn()
        row = conn.execute(
            "SELECT * FROM drama_scenes WHERE id = ?",
            (scene_id,)
        ).fetchone()
        if not row:
            return None
        return self._row_to_scene(row)

    def list_scenes(self,
                    drama_id: Optional[str] = None,
                    episode: Optional[int] = None,
                    limit: int = 100,
                    offset: int = 0) -> List[DramaScene]:
        """列出短剧场次（v5.2.1 新增）"""
        conn = self._get_conn()
        query = "SELECT * FROM drama_scenes WHERE 1=1"
        params = []

        if drama_id:
            query += " AND drama_id = ?"
            params.append(drama_id)
        if episode is not None:
            query += " AND episode = ?"
            params.append(episode)

        query += " ORDER BY episode ASC, scene_number ASC LIMIT ? OFFSET ?"
        params.extend([limit, offset])

        rows = conn.execute(query, params).fetchall()
        return [self._row_to_scene(r) for r in rows]

    def update_scene(self,
                     scene_id: str,
                     title: Optional[str] = None,
                     content: Optional[str] = None,
                     location: Optional[str] = None,
                     time_of_day: Optional[str] = None,
                     tags: Optional[List[str]] = None,
                     metadata: Optional[Dict[str, Any]] = None) -> bool:
        """更新场次（v5.2.1 新增）"""
        conn = self._get_conn()
        updates = []
        params = []

        if title is not None:
            updates.append("title = ?")
            params.append(self._validate_str(title, "title", max_len=200))
        if content is not None:
            updates.append("content = ?")
            params.append(self._validate_str(content, "content", max_len=10000))
        if location is not None:
            updates.append("location = ?")
            params.append(self._validate_str(location, "location", max_len=200))
        if time_of_day is not None:
            updates.append("time_of_day = ?")
            params.append(self._validate_str(time_of_day, "time_of_day", max_len=50))
        if tags is not None:
            updates.append("tags = ?")
            params.append(json.dumps(tags, ensure_ascii=False))
        if metadata is not None:
            updates.append("metadata = ?")
            params.append(json.dumps(metadata, ensure_ascii=False))

        if not updates:
            return False

        params.append(scene_id)
        cursor = conn.execute(
            f"UPDATE drama_scenes SET {', '.join(updates)} WHERE id = ?",
            params
        )
        conn.commit()
        return cursor.rowcount > 0

    def delete_scene(self, scene_id: str) -> bool:
        """删除场次（v5.2.1 新增）"""
        conn = self._get_conn()
        conn.execute("DELETE FROM drama_lines WHERE scene_id = ?", (scene_id,))
        cursor = conn.execute("DELETE FROM drama_scenes WHERE id = ?", (scene_id,))
        conn.commit()
        return cursor.rowcount > 0

    def _row_to_scene(self, row) -> DramaScene:
        return DramaScene(
            id=row["id"],
            drama_id=row["drama_id"],
            episode=row["episode"] or 0,
            scene_number=row["scene_number"] or 0,
            title=row["title"] or "",
            content=row["content"] or "",
            location=row["location"] or "",
            time_of_day=row["time_of_day"] or "",
            tags=self._safe_json_loads(row["tags"], []),
            metadata=self._safe_json_loads(row["metadata"], {}),
            created_at=row["created_at"] or 0.0,
        )

    # --- 短剧角色 ---

    def add_character(self,
                      drama_id: str,
                      name: str,
                      role: str = "supporting",
                      actor: str = "",
                      description: str = "",
                      personality: str = "",
                      avatar_url: str = "",
                      tags: Optional[List[str]] = None,
                      metadata: Optional[Dict[str, Any]] = None) -> DramaCharacter:
        """添加短剧角色（v5.2.1 新增）"""
        now = time.time()
        char_id = str(uuid.uuid4())

        name = self._validate_str(name, "name", max_len=100)
        role = self._validate_str(role, "role", max_len=50)
        actor = self._validate_str(actor, "actor", max_len=100)
        description = self._validate_str(description, "description", max_len=2000)
        personality = self._validate_str(personality, "personality", max_len=1000)
        avatar_url = self._validate_str(avatar_url, "avatar_url", max_len=500)

        character = DramaCharacter(
            id=char_id,
            drama_id=drama_id,
            name=name,
            role=role,
            actor=actor,
            description=description,
            personality=personality,
            avatar_url=avatar_url,
            tags=tags or [],
            metadata=metadata or {},
            created_at=now,
        )

        conn = self._get_conn()
        conn.execute("""
            INSERT INTO drama_characters (
                id, drama_id, name, role, actor, description,
                personality, avatar_url, tags, metadata, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            character.id, character.drama_id, character.name,
            character.role, character.actor, character.description,
            character.personality, character.avatar_url,
            json.dumps(character.tags, ensure_ascii=False),
            json.dumps(character.metadata, ensure_ascii=False),
            character.created_at,
        ))
        conn.commit()
        return character

    def get_character(self, char_id: str) -> Optional[DramaCharacter]:
        """获取角色详情（v5.2.1 新增）"""
        conn = self._get_conn()
        row = conn.execute(
            "SELECT * FROM drama_characters WHERE id = ?",
            (char_id,)
        ).fetchone()
        if not row:
            return None
        return self._row_to_character(row)

    def list_characters(self,
                        drama_id: Optional[str] = None,
                        role: Optional[str] = None,
                        limit: int = 100,
                        offset: int = 0) -> List[DramaCharacter]:
        """列出短剧角色（v5.2.1 新增）"""
        conn = self._get_conn()
        query = "SELECT * FROM drama_characters WHERE 1=1"
        params = []

        if drama_id:
            query += " AND drama_id = ?"
            params.append(drama_id)
        if role:
            query += " AND role = ?"
            params.append(role)

        query += " ORDER BY created_at ASC LIMIT ? OFFSET ?"
        params.extend([limit, offset])

        rows = conn.execute(query, params).fetchall()
        return [self._row_to_character(r) for r in rows]

    def update_character(self,
                         char_id: str,
                         name: Optional[str] = None,
                         role: Optional[str] = None,
                         actor: Optional[str] = None,
                         description: Optional[str] = None,
                         personality: Optional[str] = None,
                         avatar_url: Optional[str] = None,
                         tags: Optional[List[str]] = None,
                         metadata: Optional[Dict[str, Any]] = None) -> bool:
        """更新角色信息（v5.2.1 新增）"""
        conn = self._get_conn()
        updates = []
        params = []

        if name is not None:
            updates.append("name = ?")
            params.append(self._validate_str(name, "name", max_len=100))
        if role is not None:
            updates.append("role = ?")
            params.append(self._validate_str(role, "role", max_len=50))
        if actor is not None:
            updates.append("actor = ?")
            params.append(self._validate_str(actor, "actor", max_len=100))
        if description is not None:
            updates.append("description = ?")
            params.append(self._validate_str(description, "description", max_len=2000))
        if personality is not None:
            updates.append("personality = ?")
            params.append(self._validate_str(personality, "personality", max_len=1000))
        if avatar_url is not None:
            updates.append("avatar_url = ?")
            params.append(self._validate_str(avatar_url, "avatar_url", max_len=500))
        if tags is not None:
            updates.append("tags = ?")
            params.append(json.dumps(tags, ensure_ascii=False))
        if metadata is not None:
            updates.append("metadata = ?")
            params.append(json.dumps(metadata, ensure_ascii=False))

        if not updates:
            return False

        params.append(char_id)
        cursor = conn.execute(
            f"UPDATE drama_characters SET {', '.join(updates)} WHERE id = ?",
            params
        )
        conn.commit()
        return cursor.rowcount > 0

    def delete_character(self, char_id: str) -> bool:
        """删除角色（v5.2.1 新增）"""
        conn = self._get_conn()
        cursor = conn.execute("DELETE FROM drama_characters WHERE id = ?", (char_id,))
        conn.commit()
        return cursor.rowcount > 0

    def _row_to_character(self, row) -> DramaCharacter:
        return DramaCharacter(
            id=row["id"],
            drama_id=row["drama_id"],
            name=row["name"],
            role=row["role"] or "supporting",
            actor=row["actor"] or "",
            description=row["description"] or "",
            personality=row["personality"] or "",
            avatar_url=row["avatar_url"] or "",
            tags=self._safe_json_loads(row["tags"], []),
            metadata=self._safe_json_loads(row["metadata"], {}),
            created_at=row["created_at"] or 0.0,
        )

    # --- 短剧台词 ---

    def add_line(self,
                 drama_id: str,
                 line_text: str,
                 scene_id: str = "",
                 character_id: str = "",
                 character_name: str = "",
                 context: str = "",
                 episode: int = 0,
                 timestamp: str = "",
                 is_classic: bool = False,
                 tags: Optional[List[str]] = None,
                 metadata: Optional[Dict[str, Any]] = None) -> DramaLine:
        """添加短剧台词（v5.2.1 新增）"""
        now = time.time()
        line_id = str(uuid.uuid4())

        line_text = self._validate_str(line_text, "line_text", max_len=2000)
        character_name = self._validate_str(character_name, "character_name", max_len=100)
        context = self._validate_str(context, "context", max_len=2000)
        episode = self._validate_int(episode, "episode", min_val=0, max_val=10000)
        timestamp_float = self._validate_float(timestamp if timestamp else 0.0, "timestamp", min_val=0.0, max_val=100000.0)
        timestamp = str(timestamp_float)

        line = DramaLine(
            id=line_id,
            drama_id=drama_id,
            scene_id=scene_id,
            character_id=character_id,
            character_name=character_name,
            line_text=line_text,
            context=context,
            episode=episode,
            timestamp=timestamp,
            is_classic=is_classic,
            memory_id="",
            tags=tags or [],
            metadata=metadata or {},
            created_at=now,
        )

        conn = self._get_conn()
        conn.execute("""
            INSERT INTO drama_lines (
                id, drama_id, scene_id, character_id, character_name,
                line_text, context, episode, timestamp, is_classic,
                memory_id, tags, metadata, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            line.id, line.drama_id, line.scene_id, line.character_id,
            line.character_name, line.line_text, line.context,
            line.episode, line.timestamp, int(line.is_classic),
            line.memory_id, json.dumps(line.tags, ensure_ascii=False),
            json.dumps(line.metadata, ensure_ascii=False), line.created_at,
        ))
        conn.commit()
        return line

    def get_line(self, line_id: str) -> Optional[DramaLine]:
        """获取台词详情（v5.2.1 新增）"""
        conn = self._get_conn()
        row = conn.execute(
            "SELECT * FROM drama_lines WHERE id = ?",
            (line_id,)
        ).fetchone()
        if not row:
            return None
        return self._row_to_line(row)

    def list_lines(self,
                   drama_id: Optional[str] = None,
                   scene_id: Optional[str] = None,
                   character_id: Optional[str] = None,
                   is_classic: Optional[bool] = None,
                   episode: Optional[int] = None,
                   limit: int = 100,
                   offset: int = 0) -> List[DramaLine]:
        """列出台词（v5.2.1 新增）"""
        conn = self._get_conn()
        query = "SELECT * FROM drama_lines WHERE 1=1"
        params = []

        if drama_id:
            query += " AND drama_id = ?"
            params.append(drama_id)
        if scene_id:
            query += " AND scene_id = ?"
            params.append(scene_id)
        if character_id:
            query += " AND character_id = ?"
            params.append(character_id)
        if is_classic is not None:
            query += " AND is_classic = ?"
            params.append(1 if is_classic else 0)
        if episode is not None:
            query += " AND episode = ?"
            params.append(episode)

        query += " ORDER BY episode ASC, created_at ASC LIMIT ? OFFSET ?"
        params.extend([limit, offset])

        rows = conn.execute(query, params).fetchall()
        return [self._row_to_line(r) for r in rows]

    def update_line(self,
                    line_id: str,
                    line_text: Optional[str] = None,
                    character_name: Optional[str] = None,
                    context: Optional[str] = None,
                    is_classic: Optional[bool] = None,
                    tags: Optional[List[str]] = None,
                    metadata: Optional[Dict[str, Any]] = None) -> bool:
        """更新台词（v5.2.1 新增）"""
        conn = self._get_conn()
        updates = []
        params = []

        if line_text is not None:
            updates.append("line_text = ?")
            params.append(self._validate_str(line_text, "line_text", max_len=2000))
        if character_name is not None:
            updates.append("character_name = ?")
            params.append(self._validate_str(character_name, "character_name", max_len=100))
        if context is not None:
            updates.append("context = ?")
            params.append(self._validate_str(context, "context", max_len=2000))
        if is_classic is not None:
            updates.append("is_classic = ?")
            params.append(1 if is_classic else 0)
        if tags is not None:
            updates.append("tags = ?")
            params.append(json.dumps(tags, ensure_ascii=False))
        if metadata is not None:
            updates.append("metadata = ?")
            params.append(json.dumps(metadata, ensure_ascii=False))

        if not updates:
            return False

        params.append(line_id)
        cursor = conn.execute(
            f"UPDATE drama_lines SET {', '.join(updates)} WHERE id = ?",
            params
        )
        conn.commit()
        return cursor.rowcount > 0

    def delete_line(self, line_id: str) -> bool:
        """删除台词（v5.2.1 新增）"""
        conn = self._get_conn()
        cursor = conn.execute("DELETE FROM drama_lines WHERE id = ?", (line_id,))
        conn.commit()
        return cursor.rowcount > 0

    def search_lines(self,
                     query: str,
                     drama_id: Optional[str] = None,
                     is_classic_only: bool = False,
                     limit: int = 20) -> List[DramaLine]:
        """搜索台词（v5.2.1 新增）"""
        conn = self._get_conn()
        # v5.3.3 安全加固：LIKE 通配符转义
        safe_query = query[:200] if isinstance(query, str) else ""
        sql = "SELECT * FROM drama_lines WHERE line_text LIKE ? ESCAPE '\\'"
        params = [f"%{_escape_like(safe_query)}%"]

        if drama_id:
            sql += " AND drama_id = ?"
            params.append(drama_id)
        if is_classic_only:
            sql += " AND is_classic = 1"

        sql += " LIMIT ?"
        params.append(limit)

        rows = conn.execute(sql, params).fetchall()
        return [self._row_to_line(r) for r in rows]

    def classic_lines(self,
                      drama_id: Optional[str] = None,
                      limit: int = 20) -> List[DramaLine]:
        """获取经典台词（v5.2.1 新增）"""
        return self.list_lines(
            drama_id=drama_id,
            is_classic=True,
            limit=limit,
        )

    def _row_to_line(self, row) -> DramaLine:
        return DramaLine(
            id=row["id"],
            drama_id=row["drama_id"],
            scene_id=row["scene_id"] or "",
            character_id=row["character_id"] or "",
            character_name=row["character_name"] or "",
            line_text=row["line_text"],
            context=row["context"] or "",
            episode=row["episode"] or 0,
            timestamp=row["timestamp"] or "",
            is_classic=bool(row["is_classic"]),
            memory_id=row["memory_id"] or "",
            tags=self._safe_json_loads(row["tags"], []),
            metadata=self._safe_json_loads(row["metadata"], {}),
            created_at=row["created_at"] or 0.0,
        )

    # ===== v5.2.4 新增方法 =====

    def add_note(self, memory_id: str, content: str, author: str = "",
                 tags: Optional[List[str]] = None) -> Dict[str, Any]:
        """添加记忆笔记/批注（v5.2.4 新增）"""
        conn = self._get_conn()
        # 验证记忆存在
        row = conn.execute("SELECT id FROM memories WHERE id = ?", (memory_id,)).fetchone()
        if not row:
            return {"success": False, "error": f"记忆不存在: {memory_id}"}

        now = time.time()
        note_id = str(uuid.uuid4())
        conn.execute("""
            INSERT INTO memory_notes (id, memory_id, content, author, tags, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (note_id, memory_id, content, author,
              json.dumps(tags or [], ensure_ascii=False), now, now))
        conn.commit()
        self._add_audit("add_note", memory_id, author, "", "INTERNAL")
        return {"success": True, "note_id": note_id, "memory_id": memory_id}

    def list_notes(self, memory_id: str, limit: int = 50, offset: int = 0) -> List[Dict[str, Any]]:
        """列出记忆的笔记（v5.2.4 新增）"""
        conn = self._get_conn()
        rows = conn.execute("""
            SELECT * FROM memory_notes WHERE memory_id = ?
            ORDER BY created_at DESC LIMIT ? OFFSET ?
        """, (memory_id, limit, offset)).fetchall()
        results = []
        for row in rows:
            results.append({
                "id": row["id"],
                "memory_id": row["memory_id"],
                "content": row["content"],
                "author": row["author"],
                "tags": self._safe_json_loads(row["tags"], []),
                "created_at": row["created_at"],
                "updated_at": row["updated_at"],
            })
        return results

    def delete_note(self, note_id: str) -> Dict[str, Any]:
        """删除笔记（v5.2.4 新增）"""
        conn = self._get_conn()
        row = conn.execute("SELECT * FROM memory_notes WHERE id = ?", (note_id,)).fetchone()
        if not row:
            return {"success": False, "error": f"笔记不存在: {note_id}"}
        conn.execute("DELETE FROM memory_notes WHERE id = ?", (note_id,))
        conn.commit()
        self._add_audit("delete_note", row["memory_id"], "", "", "INTERNAL")
        return {"success": True, "note_id": note_id}

    def add_template(self, name: str, content_template: str, category: str = "general",
                     tags: Optional[List[str]] = None, importance: str = "MEDIUM",
                     layer: str = "short_term", description: str = "") -> Dict[str, Any]:
        """添加记忆模板（v5.2.4 新增）"""
        conn = self._get_conn()
        # 检查名称重复
        existing = conn.execute("SELECT id FROM memory_templates WHERE name = ?", (name,)).fetchone()
        if existing:
            return {"success": False, "error": f"模板名称已存在: {name}"}

        now = time.time()
        template_id = str(uuid.uuid4())
        conn.execute("""
            INSERT INTO memory_templates (id, name, content_template, category, tags,
                                          importance, layer, description, use_count, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?)
        """, (template_id, name, content_template, category,
              json.dumps(tags or [], ensure_ascii=False), importance, layer, description, now, now))
        conn.commit()
        return {"success": True, "template_id": template_id, "name": name}

    def list_templates(self, category: Optional[str] = None,
                       limit: int = 50, offset: int = 0) -> List[Dict[str, Any]]:
        """列出记忆模板（v5.2.4 新增）"""
        conn = self._get_conn()
        query = "SELECT * FROM memory_templates WHERE 1=1"
        params = []
        if category:
            query += " AND category = ?"
            params.append(category)
        query += " ORDER BY use_count DESC, created_at DESC LIMIT ? OFFSET ?"
        params.extend([limit, offset])
        rows = conn.execute(query, params).fetchall()
        results = []
        for row in rows:
            results.append({
                "id": row["id"],
                "name": row["name"],
                "content_template": row["content_template"],
                "category": row["category"],
                "tags": self._safe_json_loads(row["tags"], []),
                "importance": row["importance"],
                "layer": row["layer"],
                "description": row["description"],
                "use_count": row["use_count"],
                "created_at": row["created_at"],
            })
        return results

    def use_template(self, template_id: str, variables: Optional[Dict[str, str]] = None,
                     actor: str = "", session_id: str = "") -> Dict[str, Any]:
        """使用模板创建记忆（v5.2.4 新增）"""
        conn = self._get_conn()
        row = conn.execute("SELECT * FROM memory_templates WHERE id = ?", (template_id,)).fetchone()
        if not row:
            return {"success": False, "error": f"模板不存在: {template_id}"}

        # 替换模板变量 {var_name}
        content = row["content_template"]
        if variables:
            for key, value in variables.items():
                content = content.replace("{" + key + "}", value)

        # 创建记忆
        entry = self.add_memory(
            content=content,
            category=row["category"],
            tags=self._safe_json_loads(row["tags"], []),
            importance=Importance.from_string(row["importance"]) if hasattr(Importance, 'from_string') else Importance.MEDIUM,
            layer=MemoryLayer.from_string(row["layer"]) if hasattr(MemoryLayer, 'from_string') else MemoryLayer.SHORT_TERM,
            source_session=session_id,
            source_agent=actor,
        )

        # 更新使用次数
        conn.execute("UPDATE memory_templates SET use_count = use_count + 1, updated_at = ? WHERE id = ?",
                     (time.time(), template_id))
        conn.commit()

        return {"success": True, "memory_id": entry.id, "template_name": row["name"], "content": content}

    def delete_template(self, template_id: str) -> Dict[str, Any]:
        """删除模板（v5.2.4 新增）"""
        conn = self._get_conn()
        row = conn.execute("SELECT id, name FROM memory_templates WHERE id = ?", (template_id,)).fetchone()
        if not row:
            return {"success": False, "error": f"模板不存在: {template_id}"}
        conn.execute("DELETE FROM memory_templates WHERE id = ?", (template_id,))
        conn.commit()
        return {"success": True, "template_id": template_id, "name": row["name"]}

    def batch_update(self, memory_ids: Optional[List[str]] = None,
                     category: Optional[str] = None,
                     tags: Optional[List[str]] = None,
                     importance: Optional[str] = None,
                     layer: Optional[str] = None,
                     starred: Optional[bool] = None,
                     actor: str = "", session_id: str = "") -> Dict[str, Any]:
        """批量更新记忆（v5.2.4 新增）

        Args:
            memory_ids: 要更新的记忆 ID 列表
            category: 新分类（None 表示不修改）
            tags: 新标签（None 表示不修改）
            importance: 新重要性（None 表示不修改）
            layer: 新层级（None 表示不修改）
            starred: 新收藏状态（None 表示不修改）
        """
        if not memory_ids:
            return {"success": False, "error": "未指定记忆 ID", "updated": 0}

        conn = self._get_conn()
        now = time.time()
        updated = 0
        errors = []

        for mid in memory_ids:
            row = conn.execute("SELECT id FROM memories WHERE id = ?", (mid,)).fetchone()
            if not row:
                errors.append(mid)
                continue

            updates = ["updated_at = ?"]
            params = [now]

            if category is not None:
                updates.append("category = ?")
                params.append(category)
            if tags is not None:
                updates.append("tags = ?")
                params.append(json.dumps(tags, ensure_ascii=False))
            if importance is not None:
                updates.append("importance = ?")
                params.append(importance)
            if layer is not None:
                updates.append("layer = ?")
                params.append(layer)
            if starred is not None:
                updates.append("starred = ?")
                params.append(1 if starred else 0)

            params.append(mid)
            conn.execute(f"UPDATE memories SET {', '.join(updates)} WHERE id = ?", params)
            updated += 1

        conn.commit()
        self._add_audit("batch_update", ",".join(memory_ids[:5]), actor, session_id, "INTERNAL",
                        {"updated": updated, "category": category, "importance": importance})
        return {"success": True, "updated": updated, "errors": errors, "total": len(memory_ids)}

    def create_review_schedule(self, memory_id: str, interval_days: float = 1.0,
                               actor: str = "") -> Dict[str, Any]:
        """创建复习计划（v5.2.4 新增）"""
        conn = self._get_conn()
        row = conn.execute("SELECT id FROM memories WHERE id = ?", (memory_id,)).fetchone()
        if not row:
            return {"success": False, "error": f"记忆不存在: {memory_id}"}

        now = time.time()
        schedule_id = str(uuid.uuid4())
        scheduled_at = now + interval_days * 86400

        conn.execute("""
            INSERT INTO review_schedules (id, memory_id, scheduled_at, interval_days,
                                          review_count, last_reviewed_at, status, created_at)
            VALUES (?, ?, ?, ?, 0, 0.0, 'pending', ?)
        """, (schedule_id, memory_id, scheduled_at, interval_days, now))
        conn.commit()
        return {"success": True, "schedule_id": schedule_id, "memory_id": memory_id,
                "scheduled_at": scheduled_at, "interval_days": interval_days}

    def list_due_reviews(self, limit: int = 20) -> List[Dict[str, Any]]:
        """列出到期复习（v5.2.4 新增）"""
        conn = self._get_conn()
        now = time.time()
        rows = conn.execute("""
            SELECT rs.*, m.content, m.category, m.importance, m.layer
            FROM review_schedules rs
            JOIN memories m ON rs.memory_id = m.id
            WHERE rs.status = 'pending' AND rs.scheduled_at <= ?
            ORDER BY rs.scheduled_at ASC
            LIMIT ?
        """, (now, limit)).fetchall()
        results = []
        for row in rows:
            results.append({
                "schedule_id": row["id"],
                "memory_id": row["memory_id"],
                "content": row["content"][:100] if row["content"] else "[已加密]",
                "category": row["category"],
                "importance": row["importance"],
                "layer": row["layer"],
                "scheduled_at": row["scheduled_at"],
                "interval_days": row["interval_days"],
                "review_count": row["review_count"],
            })
        return results

    def complete_review(self, schedule_id: str) -> Dict[str, Any]:
        """完成复习，自动安排下次（v5.2.4 新增）

        使用间隔重复算法：每次复习后间隔翻倍（1天→2天→4天→7天→15天→30天）
        """
        conn = self._get_conn()
        row = conn.execute("SELECT * FROM review_schedules WHERE id = ?", (schedule_id,)).fetchone()
        if not row:
            return {"success": False, "error": f"复习计划不存在: {schedule_id}"}

        now = time.time()
        new_count = row["review_count"] + 1
        # 间隔重复：1, 2, 4, 7, 15, 30 天
        intervals = [1, 2, 4, 7, 15, 30]
        new_interval = intervals[min(new_count, len(intervals) - 1)]
        next_scheduled = now + new_interval * 86400

        conn.execute("""
            UPDATE review_schedules
            SET review_count = ?, last_reviewed_at = ?, interval_days = ?,
                scheduled_at = ?, status = 'pending'
            WHERE id = ?
        """, (new_count, now, new_interval, next_scheduled, schedule_id))

        # 更新记忆的巩固次数和强度
        conn.execute("""
            UPDATE memories SET consolidation_count = consolidation_count + 1,
                   strength = MIN(strength + 0.1, 2.0), last_accessed_at = ?
            WHERE id = ?
        """, (now, row["memory_id"]))

        conn.commit()
        self._add_audit("complete_review", row["memory_id"], "", "", "INTERNAL")
        return {"success": True, "schedule_id": schedule_id, "review_count": new_count,
                "next_interval_days": new_interval, "next_scheduled_at": next_scheduled}

    def get_review_stats(self) -> Dict[str, Any]:
        """复习计划统计（v5.2.4 新增）"""
        conn = self._get_conn()
        now = time.time()
        total = conn.execute("SELECT COUNT(*) as cnt FROM review_schedules").fetchone()["cnt"]
        pending = conn.execute("SELECT COUNT(*) as cnt FROM review_schedules WHERE status = 'pending'").fetchone()["cnt"]
        due = conn.execute("SELECT COUNT(*) as cnt FROM review_schedules WHERE status = 'pending' AND scheduled_at <= ?",
                           (now,)).fetchone()["cnt"]
        total_reviews = conn.execute("SELECT COALESCE(SUM(review_count), 0) as cnt FROM review_schedules").fetchone()["cnt"]
        return {
            "total_schedules": total,
            "pending": pending,
            "due_now": due,
            "total_reviews_completed": total_reviews,
        }

    # ===== 记忆关联（v5.2.5 新增）=====

    def link_memories(self, source_id: str, target_id: str,
                      link_type: str = "related", note: str = "") -> Dict[str, Any]:
        """创建记忆关联（双向）"""
        if source_id == target_id:
            return {"success": False, "error": "不能关联自己"}

        # 输入校验
        link_type = link_type.strip()[:50] if link_type else "related"
        note = note.strip()[:500] if note else ""

        conn = self._get_conn()

        # 检查两条记忆是否存在
        for mid in (source_id, target_id):
            row = conn.execute("SELECT id FROM memories WHERE id = ? AND category != 'trash'", (mid,)).fetchone()
            if not row:
                return {"success": False, "error": f"记忆不存在: {mid}"}

        # 检查是否已存在关联（任一方向）
        existing = conn.execute(
            "SELECT id FROM memory_links WHERE (source_id = ? AND target_id = ?) OR (source_id = ? AND target_id = ?)",
            (source_id, target_id, target_id, source_id)
        ).fetchone()
        if existing:
            return {"success": False, "error": "关联已存在"}

        link_id = f"link_{uuid.uuid4().hex[:12]}"
        now = time.time()
        conn.execute(
            "INSERT INTO memory_links (id, source_id, target_id, link_type, note, created_at) VALUES (?, ?, ?, ?, ?, ?)",
            (link_id, source_id, target_id, link_type, note, now)
        )
        conn.commit()
        return {"success": True, "link_id": link_id, "source_id": source_id, "target_id": target_id,
                "link_type": link_type, "note": note}

    def list_links(self, memory_id: str) -> List[Dict[str, Any]]:
        """列出记忆的所有关联（双向）"""
        conn = self._get_conn()
        rows = conn.execute(
            """SELECT ml.*, m.content as target_content, m.category as target_category
               FROM memory_links ml
               JOIN memories m ON (ml.target_id = m.id AND ml.source_id = ?)
                                  OR (ml.source_id = m.id AND ml.target_id = ?)
               WHERE ml.source_id = ? OR ml.target_id = ?
               ORDER BY ml.created_at DESC""",
            (memory_id, memory_id, memory_id, memory_id)
        ).fetchall()
        results = []
        for row in rows:
            # 确定关联的另一端
            linked_id = row["target_id"] if row["source_id"] == memory_id else row["source_id"]
            results.append({
                "link_id": row["id"],
                "linked_id": linked_id,
                "linked_content": row["target_content"],
                "linked_category": row["target_category"],
                "link_type": row["link_type"],
                "note": row["note"],
                "created_at": row["created_at"],
            })
        return results

    def unlink_memories(self, link_id: str) -> bool:
        """删除记忆关联"""
        conn = self._get_conn()
        cursor = conn.execute("DELETE FROM memory_links WHERE id = ?", (link_id,))
        conn.commit()
        return cursor.rowcount > 0

    # ===== 记忆版本历史（v5.2.7 新增）=====

    def save_version(self, memory_id: str, content: str, category: str,
                     tags, importance, actor: str = "") -> Dict[str, Any]:
        """保存记忆的历史版本（v5.2.7 新增）"""
        # 输入校验
        if not memory_id or len(memory_id) > 128:
            return {"success": False, "error": "记忆 ID 无效"}
        if not content or len(content) > 100000:
            return {"success": False, "error": "内容无效或过长"}

        conn = self._get_conn()
        try:
            # 检查记忆是否存在
            row = conn.execute("SELECT id FROM memories WHERE id = ?", (memory_id,)).fetchone()
            if not row:
                return {"success": False, "error": "记忆不存在"}

            # 获取当前版本号
            row = conn.execute(
                "SELECT MAX(version_number) FROM memory_versions WHERE memory_id = ?",
                (memory_id,)
            ).fetchone()
            next_version = (row[0] or 0) + 1

            version_id = f"ver_{uuid.uuid4().hex[:12]}"
            now = time.time()
            tags_str = json.dumps(tags, ensure_ascii=False) if isinstance(tags, list) else str(tags or "")

            conn.execute(
                """INSERT INTO memory_versions
                   (id, memory_id, version_number, content, category, tags, importance, actor, changed_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (version_id, memory_id, next_version, content, category or "",
                 tags_str, str(importance or ""), actor or "", now)
            )
            conn.commit()
            return {
                "success": True,
                "version_id": version_id,
                "version_number": next_version,
                "memory_id": memory_id,
            }
        except Exception as e:
            conn.rollback()
            return {"success": False, "error": str(e)}

    def list_versions(self, memory_id: str, limit: int = 50) -> List[Dict[str, Any]]:
        """列出记忆的所有历史版本（v5.2.7 新增）"""
        if not memory_id or len(memory_id) > 128:
            return []
        # 限制 limit 范围
        limit = max(1, min(limit, 500))

        conn = self._get_conn()
        rows = conn.execute(
            """SELECT id, version_number, content, category, tags, importance, actor, changed_at
               FROM memory_versions WHERE memory_id = ?
               ORDER BY version_number DESC LIMIT ?""",
            (memory_id, limit)
        ).fetchall()

        versions = []
        for r in rows:
            versions.append({
                "version_id": r[0],
                "version_number": r[1],
                "content": r[2],
                "content_preview": r[2][:80] + ("..." if len(r[2]) > 80 else ""),
                "category": r[3],
                "tags": json.loads(r[4]) if r[4] and r[4].startswith("[") else r[4],
                "importance": r[5],
                "actor": r[6],
                "changed_at": r[7],
            })
        return versions

    def get_version(self, version_id: str) -> Optional[Dict[str, Any]]:
        """获取指定版本详情（v5.2.7 新增）"""
        if not version_id or len(version_id) > 128:
            return None
        conn = self._get_conn()
        row = conn.execute(
            """SELECT id, memory_id, version_number, content, category, tags, importance, actor, changed_at
               FROM memory_versions WHERE id = ?""",
            (version_id,)
        ).fetchone()
        if not row:
            return None
        return {
            "version_id": row[0],
            "memory_id": row[1],
            "version_number": row[2],
            "content": row[3],
            "category": row[4],
            "tags": json.loads(row[5]) if row[5] and row[5].startswith("[") else row[5],
            "importance": row[6],
            "actor": row[7],
            "changed_at": row[8],
        }

    def rollback_to_version(self, version_id: str, actor: str = "") -> Dict[str, Any]:
        """回滚记忆到指定历史版本（v5.2.7 新增）

        会先保存当前内容为新版本，再回滚到目标版本。
        """
        if not version_id or len(version_id) > 128:
            return {"success": False, "error": "版本 ID 无效"}

        conn = self._get_conn()
        try:
            # 获取目标版本
            target = conn.execute(
                "SELECT memory_id, content, category, tags, importance FROM memory_versions WHERE id = ?",
                (version_id,)
            ).fetchone()
            if not target:
                return {"success": False, "error": "版本不存在"}

            memory_id = target[0]
            content = target[1]
            category = target[2]
            tags_json = target[3]
            importance = target[4]

            # 获取当前记忆内容（保存为新版本）
            current = conn.execute(
                "SELECT content, category, tags, importance FROM memories WHERE id = ?",
                (memory_id,)
            ).fetchone()
            if not current:
                return {"success": False, "error": "记忆不存在"}

            # 保存当前状态为新版本
            save_result = self.save_version(
                memory_id, current[0], current[1],
                json.loads(current[2]) if current[2] and current[2].startswith("[") else current[2],
                current[3], actor
            )

            # 回滚：更新记忆为目标版本的内容
            conn.execute(
                """UPDATE memories SET content = ?, category = ?, tags = ?, importance = ?,
                   updated_at = ? WHERE id = ?""",
                (content, category, tags_json, importance, time.time(), memory_id)
            )
            # v5.5.5 fix: 回滚后刷新 FTS 索引，避免全文搜索返回过期数据
            self._refresh_fts(conn, memory_id)
            conn.commit()

            return {
                "success": True,
                "memory_id": memory_id,
                "rolled_back_to_version": target[1] if len(target) > 1 else None,
                "saved_current_as_version": save_result.get("version_number"),
            }
        except Exception as e:
            conn.rollback()
            return {"success": False, "error": str(e)}

    # ===== 置顶功能（v5.2.5 新增）=====

    def pin_memory(self, memory_id: str) -> bool:
        """置顶记忆"""
        return self.update_memory(entry_id=memory_id, pinned=True)

    def unpin_memory(self, memory_id: str) -> bool:
        """取消置顶"""
        return self.update_memory(entry_id=memory_id, pinned=False)

    def list_pinned(self, limit: int = 50) -> List[MemoryEntry]:
        """列出所有置顶记忆"""
        limit = max(1, min(500, int(limit)))
        conn = self._get_conn()
        rows = conn.execute(
            "SELECT * FROM memories WHERE pinned = 1 AND category != 'trash' ORDER BY updated_at DESC LIMIT ?",
            (limit,)
        ).fetchall()
        return [self._row_to_entry(row) for row in rows]

    # ===== v5.3.6 新增：Agent 记忆关联推理 + 智能召回 + 短剧节奏/互动分析 =====

    def memory_link(self,
                    agent_id: str,
                    memory_id: str,
                    top_k: int = 10,
                    days: int = 90) -> Dict[str, Any]:
        """记忆关联推理（v5.3.6 新增）

        基于关键词重叠、标签共享、时间邻近度，自动发现指定记忆
        与同 Agent 其他记忆之间的隐式关联，构建记忆关联网络。

        Args:
            agent_id: Agent ID
            memory_id: 目标记忆 ID
            top_k: 返回 Top-K 关联记忆（1-50）
            days: 回溯窗口天数（1-365）

        Returns:
            关联记忆列表（含关联类型与关联强度）、关联图谱摘要
        """
        conn = self._get_conn()
        aid = _filter_unicode_ctrl(agent_id[:128]) if isinstance(agent_id, str) else ""
        mid = _filter_unicode_ctrl(memory_id[:64]) if isinstance(memory_id, str) else ""
        if not aid or not mid:
            return {"error": "Agent ID 和记忆 ID 不能为空"}
        top_k = max(1, min(50, int(top_k)))
        days = max(1, min(365, int(days)))
        now = time.time()
        since = now - days * 86400

        # 获取目标记忆
        target = conn.execute(
            "SELECT id, content, tags, category, importance, created_at "
            "FROM memories WHERE id = ? AND source_agent = ?",
            (mid, aid)
        ).fetchone()
        if not target:
            return {"error": "目标记忆不存在或不属于该 Agent"}

        t_content = (target[1] or "").lower()
        t_tags_raw = target[2]
        try:
            t_tags = set(json.loads(t_tags_raw)) if isinstance(t_tags_raw, str) and t_tags_raw else set()
        except Exception:
            t_tags = set()
        t_created = target[5] if isinstance(target[5], (int, float)) else now

        import re as _re
        # 提取目标关键词
        t_words = set()
        for w in _re.findall(r'[a-zA-Z]{2,}', t_content):
            t_words.add(w)
        for w in _re.findall(r'[\u4e00-\u9fff]{2,4}', t_content):
            t_words.add(w)
        for tg in t_tags:
            t_words.add(str(tg).lower())

        if not t_words:
            return {
                "agent_id": aid,
                "memory_id": mid,
                "links": [],
                "total_candidates": 0,
            }

        # 获取同 Agent 候选记忆（v5.3.6 安全：参数化 + 行数限制）
        cur = conn.execute(
            "SELECT id, content, tags, category, importance, created_at "
            "FROM memories "
            "WHERE source_agent = ? AND id != ? AND category != 'trash' AND created_at >= ? "
            "ORDER BY created_at DESC",
            (aid, mid, since)
        )
        rows = _limited_fetch(cur, limit=5000)

        links: List[Dict[str, Any]] = []
        for r in rows:
            cid = r[0]
            c_content = (r[1] or "").lower()
            c_tags_raw = r[2]
            try:
                c_tags = set(json.loads(c_tags_raw)) if isinstance(c_tags_raw, str) and c_tags_raw else set()
            except Exception:
                c_tags = set()
            c_created = r[5] if isinstance(r[5], (int, float)) else now

            c_words = set()
            for w in _re.findall(r'[a-zA-Z]{2,}', c_content):
                c_words.add(w)
            for w in _re.findall(r'[\u4e00-\u9fff]{2,4}', c_content):
                c_words.add(w)
            for tg in c_tags:
                c_words.add(str(tg).lower())

            # 关键词重叠分（Jaccard）
            inter = len(t_words & c_words)
            union = len(t_words | c_words) or 1
            kw_score = inter / union

            # 标签共享分
            shared_tags = t_tags & c_tags
            tag_score = min(len(shared_tags) / 3.0, 1.0) if t_tags else 0.0

            # 时间邻近度分（7 天内满分，30 天内线性衰减）
            time_diff_days = abs(t_created - c_created) / 86400.0
            if time_diff_days <= 7:
                time_score = 1.0
            elif time_diff_days <= 30:
                time_score = 1.0 - (time_diff_days - 7) / 23.0
            else:
                time_score = 0.0

            # 综合关联强度
            strength = round(kw_score * 0.5 + tag_score * 0.3 + time_score * 0.2, 4)

            if strength < 0.05:
                continue

            # 关联类型判定
            link_types = []
            if kw_score >= 0.25:
                link_types.append("keyword")
            if tag_score >= 0.34:
                link_types.append("tag")
            if time_score >= 0.7:
                link_types.append("temporal")
            if not link_types:
                link_types.append("weak")

            links.append({
                "memory_id": cid,
                "strength": strength,
                "link_types": link_types,
                "shared_keywords": sorted(list(t_words & c_words))[:10],
                "shared_tags": sorted(list(shared_tags))[:8],
                "time_diff_days": round(time_diff_days, 1),
                "content_preview": (r[1] or "")[:80],
            })

        links.sort(key=lambda x: -x["strength"])
        top_links = links[:top_k]

        # 图谱摘要
        type_counts: Dict[str, int] = {}
        for lk in top_links:
            for lt in lk["link_types"]:
                type_counts[lt] = type_counts.get(lt, 0) + 1

        return {
            "agent_id": aid,
            "memory_id": mid,
            "total_candidates": len(rows),
            "total_links": len(links),
            "returned": len(top_links),
            "links": top_links,
            "link_type_distribution": type_counts,
            "strongest_link": top_links[0] if top_links else None,
        }

    def memory_recall(self,
                      agent_id: str,
                      query: str,
                      top_k: int = 10,
                      days: int = 180) -> Dict[str, Any]:
        """智能记忆召回（v5.3.6 新增）

        基于查询关键词的语义召回，按重要度、访问频次、
        时间衰减综合评分返回最相关记忆。

        Args:
            agent_id: Agent ID
            query: 查询文本
            top_k: 返回 Top-K 召回记忆（1-50）
            days: 回溯窗口天数（1-365）

        Returns:
            召回记忆列表（含召回分、匹配关键词）、召回统计
        """
        conn = self._get_conn()
        aid = _filter_unicode_ctrl(agent_id[:128]) if isinstance(agent_id, str) else ""
        q = _filter_unicode_ctrl(query[:500]) if isinstance(query, str) else ""
        if not aid or not q:
            return {"error": "Agent ID 和查询文本不能为空"}
        top_k = max(1, min(50, int(top_k)))
        days = max(1, min(365, int(days)))
        now = time.time()
        since = now - days * 86400

        import re as _re
        q_lower = q.lower()
        # 提取查询关键词
        q_words = set()
        for w in _re.findall(r'[a-zA-Z]{2,}', q_lower):
            q_words.add(w)
        for w in _re.findall(r'[\u4e00-\u9fff]{2,4}', q_lower):
            q_words.add(w)

        if not q_words:
            return {
                "agent_id": aid,
                "query": q,
                "recalled": [],
                "total_scanned": 0,
            }

        # 获取候选记忆（v5.3.6 安全：参数化 + 行数限制）
        cur = conn.execute(
            "SELECT id, content, tags, category, importance, layer, "
            "created_at, access_count, starred "
            "FROM memories "
            "WHERE source_agent = ? AND category != 'trash' AND created_at >= ?",
            (aid, since)
        )
        rows = _limited_fetch(cur, limit=10000)

        recalled: List[Dict[str, Any]] = []
        for r in rows:
            mid, content, tags_raw, category, importance, layer, created_at, access_count, starred = (
                r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8]
            )
            c_lower = (content or "").lower()
            # 匹配关键词
            matched = sorted([w for w in q_words if w in c_lower])
            if not matched:
                # 检查标签
                try:
                    c_tags = json.loads(tags_raw) if isinstance(tags_raw, str) and tags_raw else []
                except Exception:
                    c_tags = []
                tag_set = {str(t).lower() for t in c_tags}
                matched = sorted([w for w in q_words if w in tag_set])
                if not matched:
                    continue

            # 召回评分组件
            # 1) 匹配覆盖率（匹配词 / 查询词）
            coverage = len(matched) / len(q_words)
            # 2) 重要度加权
            imp = (importance or "MEDIUM").upper()
            imp_w = {"LOW": 0.6, "MEDIUM": 1.0, "HIGH": 1.4, "CRITICAL": 1.8}.get(imp, 1.0)
            # 3) 访问频次（对数压缩）
            ac = access_count if isinstance(access_count, int) and access_count > 0 else 0
            access_w = 1.0 + min(1.0, (ac / 10.0) * 0.5)
            # 4) 时间衰减（30 天内满分，180 天内线性衰减到 0.2）
            created = created_at if isinstance(created_at, (int, float)) else now
            age_days = max(0.0, (now - created) / 86400.0)
            if age_days <= 30:
                recency = 1.0
            elif age_days <= 180:
                recency = 1.0 - (age_days - 30) / 150.0 * 0.8
            else:
                recency = 0.2
            # 5) 置顶加成
            star_bonus = 1.1 if starred else 1.0

            # 综合召回分
            score = round(coverage * 40.0 + imp_w * 20.0 + access_w * 15.0
                          + recency * 20.0, 2)
            score = round(score * star_bonus, 2)

            recalled.append({
                "memory_id": mid,
                "score": score,
                "coverage": round(coverage, 3),
                "matched_keywords": matched[:12],
                "importance": imp,
                "access_count": ac,
                "starred": bool(starred),
                "age_days": round(age_days, 1),
                "category": category or "general",
                "layer": (layer or "SHORT_TERM").upper(),
                "content_preview": (content or "")[:100],
            })

        recalled.sort(key=lambda x: -x["score"])
        top_recalled = recalled[:top_k]

        return {
            "agent_id": aid,
            "query": q,
            "query_keywords": sorted(list(q_words)),
            "total_scanned": len(rows),
            "total_matched": len(recalled),
            "returned": len(top_recalled),
            "recalled": top_recalled,
            "avg_score": round(sum(r["score"] for r in top_recalled) / max(1, len(top_recalled)), 2),
        }

    def drama_pacing(self,
                     drama_id: str,
                     window: int = 3) -> Dict[str, Any]:
        """剧集节奏分析（v5.3.6 新增）

        按集/场景分析节奏分布（快/中/慢），识别拖沓段和密集段，
        给出节奏健康度评分。

        Args:
            drama_id: 短剧 ID
            window: 滑动窗口大小（场景数，1-10）

        Returns:
            节奏分布、拖沓/密集段、节奏健康度
        """
        conn = self._get_conn()
        did = _filter_unicode_ctrl(drama_id[:64]) if isinstance(drama_id, str) and drama_id else ""
        if not did:
            return {"error": "短剧 ID 不能为空"}
        window = max(1, min(10, int(window)))

        drow = conn.execute(
            "SELECT id, title, total_episodes FROM drama_series WHERE id = ?",
            (did,)
        ).fetchone()
        if not drow:
            return {"error": "短剧不存在"}
        title = drow[1] or "未命名"

        scene_rows = conn.execute(
            "SELECT id, episode, scene_number "
            "FROM drama_scenes WHERE drama_id = ? "
            "ORDER BY episode ASC, scene_number ASC",
            (did,)
        ).fetchall()

        if not scene_rows:
            return {
                "drama_id": did,
                "title": title,
                "total_scenes": 0,
                "pacing_curve": [],
                "pacing_distribution": {"fast": 0, "medium": 0, "slow": 0},
                "health_score": 0.0,
            }

        # 统计每个场景的台词量与角色数
        scene_density: List[Dict[str, Any]] = []
        max_lines = 1
        for sr in scene_rows:
            sid = sr[0]
            ld = conn.execute(
                "SELECT COUNT(*), COUNT(DISTINCT character_id) "
                "FROM drama_lines WHERE scene_id = ?",
                (sid,)
            ).fetchone()
            lc = ld[0] or 0
            cc = ld[1] or 0
            if lc > max_lines:
                max_lines = lc
            scene_density.append({
                "scene_id": sid,
                "episode": sr[1],
                "order": sr[2],
                "line_count": lc,
                "character_count": cc,
            })

        # 归一化密度 = 台词量占比
        for sd in scene_density:
            sd["density"] = round(sd["line_count"] / max_lines, 3)

        # 滑动窗口平均密度
        n = len(scene_density)
        pacing_curve = []
        for i, sd in enumerate(scene_density):
            lo = max(0, i - window + 1)
            seg = scene_density[lo:i + 1]
            avg_d = sum(s["density"] for s in seg) / len(seg)
            # 节奏分类
            if avg_d >= 0.6:
                pace = "fast"
            elif avg_d >= 0.3:
                pace = "medium"
            else:
                pace = "slow"
            pacing_curve.append({
                "scene_id": sd["scene_id"],
                "episode": sd["episode"],
                "order": sd["order"],
                "density": sd["density"],
                "avg_density": round(avg_d, 3),
                "pace": pace,
            })

        # 节奏分布
        dist = {"fast": 0, "medium": 0, "slow": 0}
        for p in pacing_curve:
            dist[p["pace"]] = dist.get(p["pace"], 0) + 1

        # 识别拖沓段（连续 >=3 个 slow）和密集段（连续 >=3 个 fast）
        slow_segments = []
        fast_segments = []
        cur_pace = None
        seg_start = 0
        for i, p in enumerate(pacing_curve):
            if p["pace"] != cur_pace:
                if cur_pace == "slow" and i - seg_start >= 3:
                    slow_segments.append((seg_start, i - 1))
                elif cur_pace == "fast" and i - seg_start >= 3:
                    fast_segments.append((seg_start, i - 1))
                cur_pace = p["pace"]
                seg_start = i
        # 收尾
        if cur_pace == "slow" and n - seg_start >= 3:
            slow_segments.append((seg_start, n - 1))
        elif cur_pace == "fast" and n - seg_start >= 3:
            fast_segments.append((seg_start, n - 1))

        def _seg_summary(segs):
            out = []
            for a, b in segs:
                ep_s = pacing_curve[a].get("episode")
                ep_e = pacing_curve[b].get("episode")
                out.append({
                    "from_index": a,
                    "to_index": b,
                    "length": b - a + 1,
                    "episodes": f"{ep_s}-{ep_e}" if ep_s != ep_e else str(ep_s),
                    "avg_density": round(sum(pacing_curve[i]["avg_density"] for i in range(a, b + 1)) / (b - a + 1), 3),
                })
            return out

        # 节奏健康度评分：medium 占比越高越健康
        total = n or 1
        medium_ratio = dist["medium"] / total
        fast_ratio = dist["fast"] / total
        slow_ratio = dist["slow"] / total
        # 惩罚拖沓段
        slow_penalty = min(0.3, len(slow_segments) * 0.1)
        health = round(max(0.0, min(100.0, medium_ratio * 100.0 * 1.2 + fast_ratio * 30.0 - slow_penalty * 100.0)), 1)

        # 建议洞察
        insights: List[str] = []
        if slow_ratio > 0.4:
            insights.append(f"拖沓场景占比 {slow_ratio:.0%}，建议精简低密度场景")
        if fast_ratio > 0.5:
            insights.append(f"密集场景占比 {fast_ratio:.0%}，节奏紧凑但可能疲劳")
        if medium_ratio >= 0.5:
            insights.append(f"中等节奏占比 {medium_ratio:.0%}，整体节奏稳定")
        if not insights:
            insights.append("节奏分布均衡")

        return {
            "drama_id": did,
            "title": title,
            "total_scenes": n,
            "window": window,
            "pacing_curve": pacing_curve,
            "pacing_distribution": dist,
            "slow_segments": _seg_summary(slow_segments),
            "fast_segments": _seg_summary(fast_segments),
            "health_score": health,
            "insights": insights,
        }

    def char_interaction(self,
                         drama_id: str,
                         top_k: int = 15) -> Dict[str, Any]:
        """角色互动分析（v5.3.6 新增）

        分析角色两两之间的台词互动频率、冲突度，
        构建角色互动矩阵，识别核心关系。

        Args:
            drama_id: 短剧 ID
            top_k: 返回 Top-K 互动关系（1-50）

        Returns:
            互动矩阵、Top 互动关系、核心角色识别
        """
        conn = self._get_conn()
        did = _filter_unicode_ctrl(drama_id[:64]) if isinstance(drama_id, str) and drama_id else ""
        if not did:
            return {"error": "短剧 ID 不能为空"}
        top_k = max(1, min(50, int(top_k)))

        drow = conn.execute(
            "SELECT id, title FROM drama_series WHERE id = ?",
            (did,)
        ).fetchone()
        if not drow:
            return {"error": "短剧不存在"}
        title = drow[1] or "未命名"

        # 获取所有场景（按顺序）
        scene_rows = conn.execute(
            "SELECT id, episode, scene_number "
            "FROM drama_scenes WHERE drama_id = ? "
            "ORDER BY episode ASC, scene_number ASC",
            (did,)
        ).fetchall()

        if not scene_rows:
            return {
                "drama_id": did,
                "title": title,
                "total_characters": 0,
                "interactions": [],
                "matrix": {},
            }

        # 获取角色名称映射
        char_rows = conn.execute(
            "SELECT id, name FROM drama_characters WHERE drama_id = ?",
            (did,)
        ).fetchall()
        char_names = {r[0]: r[1] or "未命名" for r in char_rows}

        # 冲突词
        conflict_words = {
            "不", "别", "错", "滚", "闭嘴", "打", "杀", "死",
            "no", "not", "stop", "hate", "fight", "kill",
            "冲突", "争吵", "背叛", "欺骗", "威胁", "逼迫",
        }

        # 统计每对角色在同一场景的共现 + 台词交替
        pair_stats: Dict[Tuple[str, str], Dict[str, Any]] = {}

        for sr in scene_rows:
            sid = sr[0]
            # 获取该场景内按顺序的角色台词
            cur_l = conn.execute(
                "SELECT character_id, line_text FROM drama_lines "
                "WHERE scene_id = ? ORDER BY created_at ASC",
                (sid,)
            )
            lines = _limited_fetch(cur_l, limit=500)
            if len(lines) < 2:
                continue

            # 场景内出现的角色集合
            chars_in_scene = list(dict.fromkeys(l[0] for l in lines if l[0]))
            # 两两共现计数
            for i in range(len(chars_in_scene)):
                for j in range(i + 1, len(chars_in_scene)):
                    first_char, second_char = chars_in_scene[i], chars_in_scene[j]
                    key = (first_char, second_char) if first_char <= second_char else (second_char, first_char)
                    if key not in pair_stats:
                        pair_stats[key] = {
                            "co_scenes": 0,
                            "alternations": 0,
                            "conflict_hits": 0,
                            "total_lines": 0,
                        }
                    pair_stats[key]["co_scenes"] += 1

            # 台词交替（相邻不同角色）+ 冲突词
            prev_char = None
            for l in lines:
                cid, text = l[0], (l[1] or "").lower()
                if cid and prev_char and cid != prev_char:
                    key = (cid, prev_char) if cid <= prev_char else (prev_char, cid)
                    if key in pair_stats:
                        pair_stats[key]["alternations"] += 1
                # 冲突词命中（归属到该角色在场景内所有 pair）
                if cid:
                    hits = sum(1 for w in conflict_words if w in text)
                    if hits > 0:
                        for other in chars_in_scene:
                            if other == cid:
                                continue
                            key = (cid, other) if cid <= other else (other, cid)
                            if key in pair_stats:
                                pair_stats[key]["conflict_hits"] += hits
                prev_char = cid

        # 修正：重新统计 total_lines（用 pair 内 alternations 近似）
        # 并构建互动关系列表
        interactions: List[Dict[str, Any]] = []
        for (a, b), st in pair_stats.items():
            co = st["co_scenes"]
            alt = st["alternations"]
            conf = st["conflict_hits"]
            if co == 0 and alt == 0:
                continue
            # 互动强度 = 共现场景 * 1.0 + 交替 * 0.5 + 冲突 * 0.3
            strength = round(co * 1.0 + alt * 0.5 + conf * 0.3, 2)
            # 关系类型
            if conf >= 5 and conf / max(1, alt) >= 0.3:
                rel_type = "antagonist"
            elif alt >= 10:
                rel_type = "close"
            elif co >= 3:
                rel_type = "frequent"
            else:
                rel_type = "casual"
            interactions.append({
                "char_a": a,
                "char_b": b,
                "name_a": char_names.get(a, a[:8]),
                "name_b": char_names.get(b, b[:8]),
                "co_scenes": co,
                "alternations": alt,
                "conflict_hits": conf,
                "strength": strength,
                "relation_type": rel_type,
            })

        interactions.sort(key=lambda x: -x["strength"])
        top_inter = interactions[:top_k]

        # 矩阵（角色 ID -> 角色 ID -> 强度）
        matrix: Dict[str, Dict[str, float]] = {}
        for it in interactions:
            a, b, s = it["char_a"], it["char_b"], it["strength"]
            matrix.setdefault(a, {})[b] = s
            matrix.setdefault(b, {})[a] = s

        # 核心角色识别（互动强度总和 Top-3）
        char_total: Dict[str, float] = {}
        for it in interactions:
            char_total[it["char_a"]] = char_total.get(it["char_a"], 0.0) + it["strength"]
            char_total[it["char_b"]] = char_total.get(it["char_b"], 0.0) + it["strength"]
        core_chars = sorted(char_total.items(), key=lambda x: -x[1])[:3]
        core_chars_out = [
            {"character_id": cid, "name": char_names.get(cid, cid[:8]), "total_strength": round(s, 2)}
            for cid, s in core_chars
        ]

        return {
            "drama_id": did,
            "title": title,
            "total_characters": len(char_names),
            "total_pairs": len(interactions),
            "returned": len(top_inter),
            "interactions": top_inter,
            "core_characters": core_chars_out,
            "matrix_size": len(matrix),
        }

    # ===== v5.3.7 新增：Agent 记忆重要度/上下文注入/情感追踪 + 短剧类型趋势/追剧粘性/角色关系 =====

    def memory_importance(self,
                          agent_id: str,
                          days: int = 30) -> Dict[str, Any]:
        """记忆重要度分析（v5.3.7 新增）

        分析 Agent 记忆的重要度分布趋势、重要度漂移、
        低估/高估记忆识别，并给出动态重评估建议。
        参考 Mem0 的动态记忆评分机制。
        """
        conn = self._get_conn()
        aid = _filter_unicode_ctrl(agent_id[:128]) if isinstance(agent_id, str) else ""
        if not aid:
            return {"error": "Agent ID 不能为空"}
        days = max(1, min(365, int(days)))
        now = time.time()
        since = now - days * 86400

        # v5.3.7 安全：参数化 SQL + 行数限制
        cur = conn.execute(
            "SELECT id, content, importance, access_count, created_at, category "
            "FROM memories "
            "WHERE source_agent = ? AND category != 'trash' AND created_at >= ? "
            "ORDER BY created_at ASC",
            (aid, since)
        )
        rows = _limited_fetch(cur, limit=10000)

        if not rows:
            return {
                "agent_id": aid,
                "days": days,
                "total_memories": 0,
                "importance_distribution": {},
                "drift_analysis": {},
                "underrated": [],
                "overrated": [],
                "re_evaluation_suggestions": [],
            }

        # 重要度分布
        imp_dist: Dict[str, int] = {"LOW": 0, "MEDIUM": 0, "HIGH": 0, "CRITICAL": 0}
        # 按时间段分（前半段 vs 后半段）分析漂移
        midpoint = len(rows) // 2
        first_half_imp: Dict[str, int] = {"LOW": 0, "MEDIUM": 0, "HIGH": 0, "CRITICAL": 0}
        second_half_imp: Dict[str, int] = {"LOW": 0, "MEDIUM": 0, "HIGH": 0, "CRITICAL": 0}

        underrated: List[Dict[str, Any]] = []
        overrated: List[Dict[str, Any]] = []

        for i, r in enumerate(rows):
            mid, content, importance, access_count, created_at, category = (
                r[0], r[1], r[2], r[3], r[4], r[5]
            )
            imp = (importance or "MEDIUM").upper()
            if imp not in imp_dist:
                imp = "MEDIUM"
            imp_dist[imp] += 1

            # 漂移分析
            half = first_half_imp if i < midpoint else second_half_imp
            half[imp] += 1

            ac = access_count if isinstance(access_count, int) and access_count > 0 else 0

            # 低估记忆：高访问量但低重要度
            if ac >= 5 and imp in ("LOW", "MEDIUM"):
                underrated.append({
                    "memory_id": mid,
                    "importance": imp,
                    "access_count": ac,
                    "suggested_importance": "HIGH" if ac >= 10 else "MEDIUM",
                    "content_preview": (content or "")[:80],
                })

            # 高估记忆：高重要度但低访问量
            if imp in ("HIGH", "CRITICAL") and ac <= 1:
                overrated.append({
                    "memory_id": mid,
                    "importance": imp,
                    "access_count": ac,
                    "suggested_importance": "MEDIUM",
                    "content_preview": (content or "")[:80],
                })

        # 漂移趋势
        total_first = max(1, sum(first_half_imp.values()))
        total_second = max(1, sum(second_half_imp.values()))
        drift: Dict[str, Dict[str, Any]] = {}
        for imp_level in ("LOW", "MEDIUM", "HIGH", "CRITICAL"):
            first_ratio = first_half_imp[imp_level] / total_first
            second_ratio = second_half_imp[imp_level] / total_second
            direction = "stable"
            if second_ratio > first_ratio + 0.05:
                direction = "increasing"
            elif second_ratio < first_ratio - 0.05:
                direction = "decreasing"
            drift[imp_level] = {
                "first_half_ratio": round(first_ratio, 4),
                "second_half_ratio": round(second_ratio, 4),
                "direction": direction,
            }

        # 重评估建议
        suggestions: List[str] = []
        if len(underrated) > len(overrated):
            suggestions.append(f"发现 {len(underrated)} 条被低估的记忆（高访问低重要度），建议提升其重要度")
        if len(overrated) > 5:
            suggestions.append(f"发现 {len(overrated)} 条被高估的记忆（高重要度低访问），建议降低其重要度")
        if not suggestions:
            suggestions.append("重要度分配合理，无需大规模调整")

        underrated.sort(key=lambda x: -x["access_count"])
        overrated.sort(key=lambda x: x["access_count"])

        return {
            "agent_id": aid,
            "days": days,
            "total_memories": len(rows),
            "importance_distribution": imp_dist,
            "drift_analysis": drift,
            "underrated": underrated[:20],
            "overrated": overrated[:20],
            "re_evaluation_suggestions": suggestions,
        }

    def memory_context(self,
                       agent_id: str,
                       query: str,
                       max_tokens: int = 4000) -> Dict[str, Any]:
        """上下文记忆注入（v5.3.7 新增）

        给定查询，选择并格式化最相关的记忆以适配 token 预算。
        参考 Letta 的上下文窗口管理。
        """
        conn = self._get_conn()
        aid = _filter_unicode_ctrl(agent_id[:128]) if isinstance(agent_id, str) else ""
        q = _filter_unicode_ctrl(query[:500]) if isinstance(query, str) else ""
        if not aid or not q:
            return {"error": "Agent ID 和查询文本不能为空"}
        max_tokens = max(500, min(32000, int(max_tokens)))
        now = time.time()
        since = now - 180 * 86400  # 默认回溯 180 天

        import re as _re
        q_lower = q.lower()
        q_words = set()
        for w in _re.findall(r'[a-zA-Z]{2,}', q_lower):
            q_words.add(w)
        for w in _re.findall(r'[\u4e00-\u9fff]{2,4}', q_lower):
            q_words.add(w)

        if not q_words:
            return {
                "agent_id": aid,
                "query": q,
                "context": "",
                "included_count": 0,
                "excluded_count": 0,
                "token_estimate": 0,
            }

        # v5.3.7 安全：参数化 SQL + 行数限制
        cur = conn.execute(
            "SELECT id, content, tags, category, importance, layer, "
            "created_at, access_count, starred "
            "FROM memories "
            "WHERE source_agent = ? AND category != 'trash' AND created_at >= ?",
            (aid, since)
        )
        rows = _limited_fetch(cur, limit=10000)

        # 复用 memory_recall 评分逻辑
        scored: List[Dict[str, Any]] = []
        for r in rows:
            mid, content, tags_raw, category, importance, layer, created_at, access_count, starred = (
                r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8]
            )
            c_lower = (content or "").lower()
            matched = sorted([w for w in q_words if w in c_lower])
            if not matched:
                try:
                    c_tags = json.loads(tags_raw) if isinstance(tags_raw, str) and tags_raw else []
                except Exception:
                    c_tags = []
                tag_set = {str(t).lower() for t in c_tags}
                matched = sorted([w for w in q_words if w in tag_set])
                if not matched:
                    continue

            coverage = len(matched) / len(q_words)
            imp = (importance or "MEDIUM").upper()
            imp_w = {"LOW": 0.6, "MEDIUM": 1.0, "HIGH": 1.4, "CRITICAL": 1.8}.get(imp, 1.0)
            ac = access_count if isinstance(access_count, int) and access_count > 0 else 0
            access_w = 1.0 + min(1.0, (ac / 10.0) * 0.5)
            created = created_at if isinstance(created_at, (int, float)) else now
            age_days = max(0.0, (now - created) / 86400.0)
            if age_days <= 30:
                recency = 1.0
            elif age_days <= 180:
                recency = 1.0 - (age_days - 30) / 150.0 * 0.8
            else:
                recency = 0.2
            star_bonus = 1.1 if starred else 1.0
            score = round(coverage * 40.0 + imp_w * 20.0 + access_w * 15.0
                          + recency * 20.0, 2)
            score = round(score * star_bonus, 2)

            scored.append({
                "memory_id": mid,
                "score": score,
                "content": content or "",
                "importance": imp,
                "category": category or "general",
                "layer": (layer or "SHORT_TERM").upper(),
                "matched_keywords": matched[:12],
            })

        scored.sort(key=lambda x: -x["score"])

        # token 估算：每 4 字符 ≈ 1 token（粗略）
        chars_per_token = 4
        budget_chars = max_tokens * chars_per_token
        used_chars = 0
        included: List[Dict[str, Any]] = []
        excluded_count = 0

        # 预留 header 和 footer 空间
        header = f"# Agent Memory Context\nQuery: {q}\n\n"
        footer = f"\n---\n{len(scored)} memories scanned."
        budget_chars -= len(header) + len(footer)

        context_parts: List[str] = [header]
        for s in scored:
            entry_text = f"## Memory: {s['memory_id'][:8]}...\n" \
                        f"Importance: {s['importance']} | Category: {s['category']}\n" \
                        f"Content: {s['content']}\n\n"
            entry_len = len(entry_text)
            if used_chars + entry_len > budget_chars:
                excluded_count += 1
                continue
            used_chars += entry_len
            context_parts.append(entry_text)
            included.append(s)

        context_parts.append(footer)
        context_str = "".join(context_parts)
        token_estimate = len(context_str) // chars_per_token

        return {
            "agent_id": aid,
            "query": q,
            "context": context_str,
            "included_count": len(included),
            "excluded_count": excluded_count,
            "token_estimate": token_estimate,
            "max_tokens": max_tokens,
        }

    def agent_emotion(self,
                     agent_id: str,
                     days: int = 30) -> Dict[str, Any]:
        """Agent 情感追踪（v5.3.7 新增）

        基于记忆情感的时间追踪，构建情感时间线、转换序列、
        主导情感与波动性评分。参考 Zep 的情感记忆功能。
        """
        conn = self._get_conn()
        aid = _filter_unicode_ctrl(agent_id[:128]) if isinstance(agent_id, str) else ""
        if not aid:
            return {"error": "Agent ID 不能为空"}
        days = max(1, min(365, int(days)))
        now = time.time()
        since = now - days * 86400

        # v5.3.7 安全：参数化 SQL + 行数限制
        cur = conn.execute(
            "SELECT content, importance, created_at "
            "FROM memories "
            "WHERE source_agent = ? AND category != 'trash' AND created_at >= ? "
            "ORDER BY created_at ASC",
            (aid, since)
        )
        rows = _limited_fetch(cur, limit=10000)

        if not rows:
            return {
                "agent_id": aid,
                "days": days,
                "total_memories": 0,
                "emotion_distribution": {},
                "dominant_emotion": "no_data",
                "volatility_score": 0.0,
            }

        # 情感关键词词典（扩展版）
        positive_words = {
            "好", "棒", "优秀", "成功", "完成", "解决", "开心", "满意", "喜欢",
            "good", "great", "excellent", "success", "happy", "love", "perfect",
            "突破", "提升", "优化", "改进", "有效", "正确", "赞同",
        }
        negative_words = {
            "坏", "差", "失败", "错误", "问题", "bug", "崩溃", "讨厌", "不满",
            "bad", "fail", "error", "broken", "crash", "hate", "wrong", "issue",
            "缺失", "丢失", "异常", "警告", "危险", "漏洞", "冲突", "阻塞",
        }
        # 情感分类
        emotion_map = {"positive": "joy", "negative": "frustration", "neutral": "calm"}

        # 按天分组
        day_emotions: Dict[str, Dict[str, int]] = {}
        emotion_dist: Dict[str, int] = {"joy": 0, "frustration": 0, "calm": 0}
        transitions: List[str] = []
        prev_emotion: Optional[str] = None

        for r in rows:
            content = (r[0] or "").lower()
            created_at = r[2] if isinstance(r[2], (int, float)) else now
            day_key = time.strftime("%Y-%m-%d", time.localtime(created_at))

            pos_hits = sum(1 for w in positive_words if w in content)
            neg_hits = sum(1 for w in negative_words if w in content)

            if pos_hits > neg_hits:
                emotion = "joy"
            elif neg_hits > pos_hits:
                emotion = "frustration"
            else:
                emotion = "calm"

            emotion_dist[emotion] += 1

            if day_key not in day_emotions:
                day_emotions[day_key] = {"joy": 0, "frustration": 0, "calm": 0}
            day_emotions[day_key][emotion] += 1

            # 转换追踪
            cur_e = emotion_map.get(emotion, emotion)
            if prev_emotion and prev_emotion != cur_e:
                transitions.append(f"{prev_emotion}→{cur_e}")
            prev_emotion = cur_e

        # 情感时间线（按天）
        timeline: List[Dict[str, Any]] = []
        for day, counts in sorted(day_emotions.items()):
            total = max(1, sum(counts.values()))
            dom = max(counts, key=counts.get)
            timeline.append({
                "date": day,
                "dominant": dom,
                "distribution": counts,
                "total": sum(counts.values()),
            })

        # 主导情感
        total_mems = len(rows)
        dominant = max(emotion_dist, key=emotion_dist.get)
        if emotion_dist[dominant] == 0:
            dominant = "no_data"

        # 情感波动性评分（转换频率 / 总记忆数）
        volatility = round(min(100.0, len(transitions) / max(1, total_mems) * 100.0), 1)

        # 情感分布百分比
        emotion_pct = {
            e: round(c / total_mems, 4) if total_mems else 0
            for e, c in emotion_dist.items()
        }

        return {
            "agent_id": aid,
            "days": days,
            "total_memories": total_mems,
            "emotion_distribution": emotion_dist,
            "emotion_percentages": emotion_pct,
            "dominant_emotion": dominant,
            "timeline": timeline,
            "transitions": transitions[:50],
            "transition_count": len(transitions),
            "volatility_score": volatility,
        }

    def drama_genre_trend(self,
                          days: int = 90) -> Dict[str, Any]:
        """短剧类型趋势分析（v5.3.7 新增）

        分析所有短剧的类型分布与流行度趋势。
        """
        conn = self._get_conn()
        days = max(1, min(365, int(days)))
        now = time.time()
        since = now - days * 86400

        # v5.3.7 安全：参数化 SQL
        cur = conn.execute(
            "SELECT id, title, genre, rating, total_episodes, created_at "
            "FROM drama_series WHERE created_at >= ? "
            "ORDER BY created_at ASC",
            (since,)
        )
        rows = _limited_fetch(cur, limit=10000)

        if not rows:
            return {
                "days": days,
                "total_dramas": 0,
                "genre_distribution": {},
                "trends": {},
                "top_genre": None,
            }

        # 类型分布
        genre_counts: Dict[str, int] = {}
        genre_ratings: Dict[str, List[float]] = {}
        genre_first_half: Dict[str, int] = {}
        genre_second_half: Dict[str, int] = {}

        midpoint = len(rows) // 2

        for i, r in enumerate(rows):
            title = r[1] or "未命名"
            genre = (r[2] or "未分类").strip()
            rating = r[3] if isinstance(r[3], (int, float)) else 0

            genre_counts[genre] = genre_counts.get(genre, 0) + 1
            genre_ratings.setdefault(genre, []).append(rating)

            if i < midpoint:
                genre_first_half[genre] = genre_first_half.get(genre, 0) + 1
            else:
                genre_second_half[genre] = genre_second_half.get(genre, 0) + 1

        # 趋势方向
        total_first = max(1, sum(genre_first_half.values()))
        total_second = max(1, sum(genre_second_half.values()))
        trends: Dict[str, Dict[str, Any]] = {}
        for genre, count in genre_counts.items():
            first_ratio = genre_first_half.get(genre, 0) / total_first
            second_ratio = genre_second_half.get(genre, 0) / total_second
            if second_ratio > first_ratio + 0.05:
                direction = "rising"
            elif second_ratio < first_ratio - 0.05:
                direction = "declining"
            else:
                direction = "stable"
            ratings = genre_ratings.get(genre, [])
            avg_rating = round(sum(ratings) / len(ratings), 2) if ratings else 0
            trends[genre] = {
                "count": count,
                "share": round(count / len(rows), 4),
                "trend": direction,
                "avg_rating": avg_rating,
                "first_half_ratio": round(first_ratio, 4),
                "second_half_ratio": round(second_ratio, 4),
            }

        # 热门类型（按数量）
        top_genre = max(genre_counts, key=genre_counts.get) if genre_counts else None

        return {
            "days": days,
            "total_dramas": len(rows),
            "genre_distribution": genre_counts,
            "trends": trends,
            "top_genre": top_genre,
            "top_genre_count": genre_counts.get(top_genre, 0) if top_genre else 0,
        }

    def drama_binge_score(self,
                          drama_id: str) -> Dict[str, Any]:
        """追剧粘性评分（v5.3.7 新增）

        多因子加权评分：节奏健康度 25% + 平均张力 25% +
        互动密度 20% + 经典台词比 15% + 完成率 15%。
        """
        conn = self._get_conn()
        did = _filter_unicode_ctrl(drama_id[:64]) if isinstance(drama_id, str) and drama_id else ""
        if not did:
            return {"error": "短剧 ID 不能为空"}

        drow = conn.execute(
            "SELECT id, title, total_episodes, rating "
            "FROM drama_series WHERE id = ?",
            (did,)
        ).fetchone()
        if not drow:
            return {"error": "短剧不存在"}
        title = drow[1] or "未命名"
        total_eps = drow[2] if isinstance(drow[2], int) and drow[2] > 0 else 1

        # 因子 1: 节奏健康度（复用 drama_pacing 逻辑）
        pacing_result = self.drama_pacing(did, 3)
        pacing_health = pacing_result.get("health_score", 50.0)
        pacing_health = max(0.0, min(100.0, pacing_health))

        # 因子 2: 平均场景张力（复用 scene_tension 逻辑）
        scene_rows = conn.execute(
            "SELECT id FROM drama_scenes WHERE drama_id = ?",
            (did,)
        ).fetchall()
        total_scenes = len(scene_rows)

        tension_scores: List[float] = []
        conflict_words = {
            "不", "别", "错", "滚", "闭嘴", "打", "杀", "死",
            "no", "not", "stop", "hate", "fight", "kill",
            "冲突", "争吵", "背叛", "欺骗", "威胁", "逼迫",
        }
        intensity_words = {
            "必须", "马上", "立刻", "快", "紧急", "危险",
            "终于", "竟然", "居然", "到底", "不可能",
        }

        for sr in scene_rows:
            sid = sr[0]
            lrow = conn.execute(
                "SELECT COUNT(*), COUNT(DISTINCT character_id) "
                "FROM drama_lines WHERE scene_id = ?",
                (sid,)
            ).fetchone()
            lc = lrow[0] or 0
            cc = lrow[1] or 0

            lines_cur = conn.execute(
                "SELECT line_text FROM drama_lines WHERE scene_id = ?",
                (sid,)
            )
            lines = _limited_fetch(lines_cur, limit=500)

            conflict_hits = 0
            intensity_hits = 0
            for l in lines:
                text = (l[0] or "").lower()
                conflict_hits += sum(1 for w in conflict_words if w in text)
                intensity_hits += sum(1 for w in intensity_words if w in text)

            tension = min(100.0, lc * 5 + conflict_hits * 10 + intensity_hits * 8 + cc * 3)
            tension_scores.append(tension)

        tension_avg = round(sum(tension_scores) / len(tension_scores), 1) if tension_scores else 0.0
        tension_avg = max(0.0, min(100.0, tension_avg))

        # 因子 3: 互动密度（复用 char_interaction 逻辑简化版）
        char_rows = conn.execute(
            "SELECT id, name FROM drama_characters WHERE drama_id = ?",
            (did,)
        ).fetchall()
        total_chars = len(char_rows)

        interaction_count = 0
        if total_chars > 0 and total_scenes > 0:
            for sr in scene_rows:
                sid = sr[0]
                lcount = conn.execute(
                    "SELECT COUNT(DISTINCT character_id) FROM drama_lines WHERE scene_id = ?",
                    (sid,)
                ).fetchone()
                cc = lcount[0] if lcount and lcount[0] else 0
                if cc >= 2:
                    interaction_count += cc * (cc - 1) // 2  # C(n,2)
            interaction_density = min(100.0, (interaction_count / max(1, total_scenes)) * 20.0)
        else:
            interaction_density = 0.0

        # 因子 4: 经典台词比
        classic_count_row = conn.execute(
            "SELECT COUNT(*) FROM drama_lines WHERE drama_id = ? AND is_classic = 1",
            (did,)
        ).fetchone()
        classic_count = classic_count_row[0] if classic_count_row else 0

        total_lines_row = conn.execute(
            "SELECT COUNT(*) FROM drama_lines WHERE drama_id = ?",
            (did,)
        ).fetchone()
        total_lines = total_lines_row[0] if total_lines_row else 0

        classic_ratio = round(classic_count / total_lines * 100, 2) if total_lines > 0 else 0.0
        classic_ratio = min(100.0, classic_ratio)

        # 因子 5: 完成率（有台词的场景数 / 总场景数）
        scenes_with_lines = 0
        for sr in scene_rows:
            sid = sr[0]
            lc = conn.execute(
                "SELECT COUNT(*) FROM drama_lines WHERE scene_id = ?",
                (sid,)
            ).fetchone()
            if lc and lc[0] > 0:
                scenes_with_lines += 1
        completion_rate = round(scenes_with_lines / max(1, total_scenes) * 100, 2) if total_scenes > 0 else 0.0

        # 加权总分
        binge_score = round(
            pacing_health * 0.25 +
            tension_avg * 0.25 +
            interaction_density * 0.20 +
            classic_ratio * 0.15 +
            completion_rate * 0.15,
            1
        )
        binge_score = max(0.0, min(100.0, binge_score))

        # 评级
        if binge_score >= 80:
            rating_label = "extreme"
            recommendation = "极度推荐：追剧粘性极高，大概率一口气看完"
        elif binge_score >= 60:
            rating_label = "high"
            recommendation = "高度推荐：节奏紧凑，角色互动丰富，值得追看"
        elif binge_score >= 40:
            rating_label = "medium"
            recommendation = "中等推荐：有一定吸引力，但存在拖沓或互动不足"
        else:
            rating_label = "low"
            recommendation = "低度推荐：追剧粘性较低，可能需要优化节奏和角色互动"

        return {
            "drama_id": did,
            "title": title,
            "binge_score": binge_score,
            "rating": rating_label,
            "recommendation": recommendation,
            "factors": {
                "pacing_health": {"score": round(pacing_health, 1), "weight": 0.25, "contribution": round(pacing_health * 0.25, 1)},
                "tension_avg": {"score": round(tension_avg, 1), "weight": 0.25, "contribution": round(tension_avg * 0.25, 1)},
                "interaction_density": {"score": round(interaction_density, 1), "weight": 0.20, "contribution": round(interaction_density * 0.20, 1)},
                "classic_ratio": {"score": round(classic_ratio, 2), "weight": 0.15, "contribution": round(classic_ratio * 0.15, 2)},
                "completion_rate": {"score": round(completion_rate, 2), "weight": 0.15, "contribution": round(completion_rate * 0.15, 2)},
            },
            "total_scenes": total_scenes,
            "total_characters": total_chars,
            "total_lines": total_lines,
            "classic_lines": classic_count,
        }

    def char_relationship(self,
                           drama_id: str,
                           char1_id: str,
                           char2_id: str) -> Dict[str, Any]:
        """角色关系深度分析（v5.3.7 新增）

        分析两个特定角色之间的关系。
        """
        conn = self._get_conn()
        did = _filter_unicode_ctrl(drama_id[:64]) if isinstance(drama_id, str) and drama_id else ""
        c1 = _filter_unicode_ctrl(char1_id[:64]) if isinstance(char1_id, str) and char1_id else ""
        c2 = _filter_unicode_ctrl(char2_id[:64]) if isinstance(char2_id, str) and char2_id else ""
        if not did:
            return {"error": "短剧 ID 不能为空"}
        if not c1 or not c2:
            return {"error": "角色 ID 不能为空"}
        if c1 == c2:
            return {"error": "两个角色 ID 不能相同"}

        drow = conn.execute(
            "SELECT id, title FROM drama_series WHERE id = ?",
            (did,)
        ).fetchone()
        if not drow:
            return {"error": "短剧不存在"}
        title = drow[1] or "未命名"

        # 获取角色名称
        ch1_row = conn.execute(
            "SELECT name FROM drama_characters WHERE id = ? AND drama_id = ?",
            (c1, did)
        ).fetchone()
        ch2_row = conn.execute(
            "SELECT name FROM drama_characters WHERE id = ? AND drama_id = ?",
            (c2, did)
        ).fetchone()
        name1 = (ch1_row[0] if ch1_row else "角色1") or "角色1"
        name2 = (ch2_row[0] if ch2_row else "角色2") or "角色2"

        # 获取所有场景
        scene_rows = conn.execute(
            "SELECT id, episode, scene_number FROM drama_scenes "
            "WHERE drama_id = ? ORDER BY episode ASC, scene_number ASC",
            (did,)
        ).fetchall()

        if not scene_rows:
            return {
                "drama_id": did,
                "title": title,
                "char1_id": c1,
                "char2_id": c2,
                "name1": name1,
                "name2": name2,
                "relationship_type": "stranger",
                "interaction_count": 0,
                "relationship_strength": 0.0,
            }

        # 冲突词
        conflict_words = {
            "不", "别", "错", "滚", "闭嘴", "打", "杀", "死",
            "no", "not", "stop", "hate", "fight", "kill",
            "冲突", "争吵", "背叛", "欺骗", "威胁", "逼迫",
        }
        # 情感词
        positive_words = {
            "好", "棒", "喜欢", "爱", "开心", "满意", "谢谢", "感谢",
            "good", "love", "happy", "thanks", "great", "perfect",
        }
        negative_words = {
            "坏", "差", "恨", "讨厌", "滚", "走开", "不",
            "bad", "hate", "go away", "wrong", "stupid",
        }

        co_scenes = 0
        total_alternations = 0
        total_conflict = 0
        total_positive = 0
        total_negative = 0
        key_scenes: List[Dict[str, Any]] = []
        emotion_progression: List[Dict[str, Any]] = []

        for sr in scene_rows:
            sid = sr[0]
            episode = sr[1]
            scene_number = sr[2]

            # 获取该场景中两个角色的台词
            cur_l = conn.execute(
                "SELECT character_id, line_text FROM drama_lines "
                "WHERE scene_id = ? ORDER BY created_at ASC",
                (sid,)
            )
            lines = _limited_fetch(cur_l, limit=500)

            # 检查是否两个角色都在该场景中
            char1_lines = [l for l in lines if l[0] == c1]
            char2_lines = [l for l in lines if l[0] == c2]

            if not char1_lines and not char2_lines:
                continue

            co_scenes += 1

            # 统计交替次数
            alternations = 0
            prev_char = None
            for l in lines:
                cid = l[0]
                if cid in (c1, c2):
                    if prev_char and prev_char != cid:
                        alternations += 1
                    prev_char = cid
            total_alternations += alternations

            # 统计冲突/情感词
            scene_conflict = 0
            scene_positive = 0
            scene_negative = 0
            for l in char1_lines + char2_lines:
                text = (l[1] or "").lower()
                scene_conflict += sum(1 for w in conflict_words if w in text)
                scene_positive += sum(1 for w in positive_words if w in text)
                scene_negative += sum(1 for w in negative_words if w in text)
            total_conflict += scene_conflict
            total_positive += scene_positive
            total_negative += scene_negative

            # 情感方向
            if scene_positive > scene_negative:
                emotion = "positive"
            elif scene_negative > scene_positive:
                emotion = "negative"
            else:
                emotion = "neutral"
            emotion_progression.append({
                "episode": episode,
                "scene": scene_number,
                "emotion": emotion,
                "alternations": alternations,
            })

            # 关键场景（高交替或高冲突）
            if alternations >= 5 or scene_conflict >= 3:
                key_scenes.append({
                    "episode": episode,
                    "scene_number": scene_number,
                    "alternations": alternations,
                    "conflict_hits": scene_conflict,
                    "emotion": emotion,
                })

        # 关系类型判定
        if co_scenes == 0:
            rel_type = "stranger"
            rel_strength = 0.0
        else:
            conflict_ratio = total_conflict / max(1, total_alternations)
            positive_ratio = total_positive / max(1, total_alternations)

            if conflict_ratio >= 0.3:
                rel_type = "rival"
            elif positive_ratio >= 0.25 and total_positive > total_negative:
                rel_type = "romance"
            elif total_alternations >= 10 and conflict_ratio < 0.15:
                rel_type = "ally"
            elif co_scenes >= 3 and total_alternations < 3:
                rel_type = "family"
            elif total_alternations >= 5 and positive_ratio >= 0.1:
                rel_type = "mentor"
            else:
                rel_type = "stranger"

            # 关系强度
            rel_strength = round(
                co_scenes * 1.0 +
                total_alternations * 0.5 +
                total_conflict * 0.3 +
                (total_positive + total_negative) * 0.2,
                2
            )

        # 冲突水平
        if total_conflict == 0:
            conflict_level = "none"
        elif total_conflict < 5:
            conflict_level = "low"
        elif total_conflict < 15:
            conflict_level = "moderate"
        else:
            conflict_level = "high"

        # 情感弧线
        emotion_arc = "stable"
        if len(emotion_progression) >= 2:
            first_emotions = [e["emotion"] for e in emotion_progression[:len(emotion_progression)//2]]
            last_emotions = [e["emotion"] for e in emotion_progression[len(emotion_progression)//2:]]
            first_dom = max(set(first_emotions), key=first_emotions.count) if first_emotions else "neutral"
            last_dom = max(set(last_emotions), key=last_emotions.count) if last_emotions else "neutral"
            if first_dom != last_dom:
                emotion_arc = f"{first_dom}→{last_dom}"
            else:
                emotion_arc = f"stable ({first_dom})"

        return {
            "drama_id": did,
            "title": title,
            "char1_id": c1,
            "char2_id": c2,
            "name1": name1,
            "name2": name2,
            "relationship_type": rel_type,
            "interaction_count": co_scenes,
            "total_alternations": total_alternations,
            "conflict_level": conflict_level,
            "conflict_hits": total_conflict,
            "positive_hits": total_positive,
            "negative_hits": total_negative,
            "emotion_arc": emotion_arc,
            "key_scenes": key_scenes[:20],
            "relationship_strength": rel_strength,
            "emotion_progression": emotion_progression[:30],
        }

    # ===== v5.4.1 新增：Agent 记忆三大能力 =====

    def memory_reflection(self,
                          agent_id: str,
                          days: int = 30) -> Dict[str, Any]:
        """记忆反思（v5.4.1 新增）

        对时间窗口内的记忆做元认知反思：聚合主题分布、情感基调、
        关键经验教训与注意力焦点漂移，生成结构化反思报告与建议。
        参考 Generative Agents 的 reflection 机制。
        """
        conn = self._get_conn()
        aid = _filter_unicode_ctrl(agent_id[:128]) if isinstance(agent_id, str) else ""
        if not aid:
            return {"error": "Agent ID 不能为空"}
        days = max(1, min(365, int(days)))
        now = time.time()
        since = now - days * 86400

        # v5.4.1 安全：参数化 SQL + 行数限制
        cur = conn.execute(
            "SELECT id, content, category, tags, importance, access_count, "
            "starred, created_at "
            "FROM memories "
            "WHERE source_agent = ? AND category != 'trash' AND created_at >= ? "
            "ORDER BY created_at ASC",
            (aid, since)
        )
        rows = _limited_fetch(cur, limit=10000)

        if not rows:
            return {
                "agent_id": aid,
                "days": days,
                "total_memories": 0,
                "top_categories": [],
                "top_tags": [],
                "importance_distribution": {},
                "emotional_tone": {"dominant": "no_data"},
                "key_lessons": [],
                "focus_shift": {"changed": False},
                "recurring_themes": [],
                "reflection_summary": "窗口内无记忆可供反思。",
                "suggestions": [],
            }

        # 分布统计
        cat_counts: Dict[str, int] = {}
        tag_counts: Dict[str, int] = {}
        imp_dist: Dict[str, int] = {"LOW": 0, "MEDIUM": 0, "HIGH": 0, "CRITICAL": 0}
        first_half_cats: Dict[str, int] = {}
        second_half_cats: Dict[str, int] = {}
        midpoint = len(rows) // 2

        positive_words = {
            "好", "棒", "优秀", "成功", "完成", "解决", "开心", "满意", "喜欢",
            "good", "great", "excellent", "success", "happy", "love", "perfect",
            "突破", "提升", "优化", "改进", "有效", "正确", "赞同",
        }
        negative_words = {
            "坏", "差", "失败", "错误", "问题", "bug", "崩溃", "讨厌", "不满",
            "bad", "fail", "error", "broken", "crash", "hate", "wrong", "issue",
            "缺失", "丢失", "异常", "警告", "危险", "漏洞", "冲突", "阻塞",
        }
        tone_counts = {"positive": 0, "negative": 0, "neutral": 0}

        lessons: List[Dict[str, Any]] = []

        for i, r in enumerate(rows):
            mid, content, category, tags_raw, importance, access_count, starred, created_at = (
                r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7]
            )
            cat = (category or "general").strip() or "general"
            cat_counts[cat] = cat_counts.get(cat, 0) + 1
            half = first_half_cats if i < midpoint else second_half_cats
            half[cat] = half.get(cat, 0) + 1

            try:
                c_tags = json.loads(tags_raw) if isinstance(tags_raw, str) and tags_raw else []
            except Exception:
                c_tags = []
            for t in c_tags:
                tk = str(t).strip().lower()
                if tk:
                    tag_counts[tk] = tag_counts.get(tk, 0) + 1

            imp = (importance or "MEDIUM").upper()
            if imp not in imp_dist:
                imp = "MEDIUM"
            imp_dist[imp] += 1

            # 情感基调
            c_lower = (content or "").lower()
            pos_hits = sum(1 for w in positive_words if w in c_lower)
            neg_hits = sum(1 for w in negative_words if w in c_lower)
            if pos_hits > neg_hits:
                tone_counts["positive"] += 1
            elif neg_hits > pos_hits:
                tone_counts["negative"] += 1
            else:
                tone_counts["neutral"] += 1

            # 关键经验候选：重要度权重 * 访问活跃度 + 星标加成
            imp_w = {"LOW": 0.5, "MEDIUM": 1.0, "HIGH": 1.6, "CRITICAL": 2.2}.get(imp, 1.0)
            ac = access_count if isinstance(access_count, int) and access_count > 0 else 0
            lesson_score = round(imp_w * 10.0 + min(ac, 10) * 1.5 + (5.0 if starred else 0.0), 2)
            lessons.append({
                "memory_id": mid,
                "score": lesson_score,
                "importance": imp,
                "access_count": ac,
                "content_preview": (content or "")[:100],
            })

        total = len(rows)
        top_categories = [
            {"category": k, "count": v, "share": round(v / total, 4)}
            for k, v in sorted(cat_counts.items(), key=lambda x: -x[1])[:5]
        ]
        top_tags = [
            {"tag": k, "count": v}
            for k, v in sorted(tag_counts.items(), key=lambda x: -x[1])[:10]
        ]
        recurring_themes = [k for k, v in tag_counts.items() if v >= 3][:10]

        dominant_tone = max(tone_counts, key=tone_counts.get)
        if tone_counts[dominant_tone] == 0:
            dominant_tone = "no_data"

        lessons.sort(key=lambda x: -x["score"])
        key_lessons = lessons[:5]

        # 焦点漂移：前半段 vs 后半段的主导分类
        def _dominant(d: Dict[str, int]) -> Optional[str]:
            if not d:
                return None
            return max(d.items(), key=lambda x: x[1])[0]

        first_dom = _dominant(first_half_cats)
        second_dom = _dominant(second_half_cats)
        focus_changed = bool(first_dom and second_dom and first_dom != second_dom)

        # 反思摘要
        cat_brief = "、".join([t["category"] for t in top_categories[:3]]) or "无"
        tag_brief = "、".join(recurring_themes[:3]) or "无明显主题"
        summary_parts = [
            f"过去 {days} 天共沉淀 {total} 条记忆，主要集中在：{cat_brief}。",
            f"反复出现的主题：{tag_brief}。",
            f"整体情感基调为 {dominant_tone}。",
        ]
        if focus_changed:
            summary_parts.append(f"注意力焦点已从「{first_dom}」转移到「{second_dom}」。")
        else:
            summary_parts.append("注意力焦点保持稳定。")
        reflection_summary = " ".join(summary_parts)

        suggestions: List[str] = []
        if focus_changed:
            suggestions.append(f"焦点从「{first_dom}」转向「{second_dom}」，建议回顾旧焦点下是否有未完结事项")
        if tone_counts["negative"] > total * 0.4:
            suggestions.append("负面记忆占比偏高，建议对高频负面对主题做一次根因反思")
        if key_lessons:
            suggestions.append("建议将 Top 关键经验固化为长期记忆或技能模板，避免随时间衰减")
        if not suggestions:
            suggestions.append("记忆结构健康，保持当前节奏即可")

        return {
            "agent_id": aid,
            "days": days,
            "total_memories": total,
            "top_categories": top_categories,
            "top_tags": top_tags,
            "importance_distribution": imp_dist,
            "emotional_tone": {
                **tone_counts,
                "dominant": dominant_tone,
            },
            "key_lessons": key_lessons,
            "focus_shift": {
                "changed": focus_changed,
                "from": first_dom,
                "to": second_dom,
            },
            "recurring_themes": recurring_themes,
            "reflection_summary": reflection_summary,
            "suggestions": suggestions,
        }

    def memory_lineage(self, memory_id: str) -> Dict[str, Any]:
        """记忆血缘/溯源追踪（v5.4.1 新增）

        追踪单条记忆的完整来源脉络：基础快照、版本历史、
        关联链接（出/入）、审计事件与生命周期时间线。
        """
        conn = self._get_conn()
        mid = _filter_unicode_ctrl(memory_id[:64]) if isinstance(memory_id, str) and memory_id else ""
        if not mid:
            return {"error": "记忆 ID 不能为空"}

        row = conn.execute(
            "SELECT id, content, category, importance, layer, created_at, "
            "updated_at, last_accessed_at, access_count, starred, pinned, "
            "source_agent, source_session "
            "FROM memories WHERE id = ?",
            (mid,)
        ).fetchone()
        if not row:
            return {"error": "记忆不存在", "memory_id": mid}

        (rid, content, category, importance, layer, created_at,
         updated_at, last_accessed_at, access_count, starred, pinned,
         source_agent, source_session) = row

        now = time.time()
        created = created_at if isinstance(created_at, (int, float)) else now
        age_days = round(max(0.0, (now - created) / 86400.0), 2)

        # 版本历史
        vrows = conn.execute(
            "SELECT version_number, actor, changed_at, content "
            "FROM memory_versions WHERE memory_id = ? "
            "ORDER BY version_number ASC",
            (mid,)
        )
        versions = [
            {
                "version_number": v[0],
                "actor": v[1] or "",
                "changed_at": v[2],
                "content_preview": (v[3] or "")[:100],
            }
            for v in _limited_fetch(vrows, limit=200)
        ]

        # 关联链接（出/入）
        orows = conn.execute(
            "SELECT id, target_id, link_type, note, created_at "
            "FROM memory_links WHERE source_id = ? ORDER BY created_at ASC",
            (mid,)
        )
        links_out = [
            {"link_id": o[0], "target_id": o[1], "link_type": o[2] or "related",
             "note": o[3] or "", "created_at": o[4]}
            for o in _limited_fetch(orows, limit=200)
        ]
        irows = conn.execute(
            "SELECT id, source_id, link_type, note, created_at "
            "FROM memory_links WHERE target_id = ? ORDER BY created_at ASC",
            (mid,)
        )
        links_in = [
            {"link_id": i[0], "source_id": i[1], "link_type": i[2] or "related",
             "note": i[3] or "", "created_at": i[4]}
            for i in _limited_fetch(irows, limit=200)
        ]

        # 审计事件（最近 50 条）
        arows = conn.execute(
            "SELECT action, actor, timestamp, details "
            "FROM audit_log WHERE memory_id = ? "
            "ORDER BY timestamp DESC LIMIT 50",
            (mid,)
        )
        audit_events = [
            {"action": a[0] or "", "actor": a[1] or "",
             "timestamp": a[2], "details": a[3] or "{}"}
            for a in arows.fetchall()
        ]

        # 生命周期时间线
        timeline: List[Dict[str, Any]] = [
            {"event": "created", "timestamp": created,
             "description": "记忆创建"}
        ]
        for v in versions:
            timeline.append({
                "event": "version",
                "timestamp": v["changed_at"],
                "description": f"更新到 v{v['version_number']}"
                               + (f"（by {v['actor']}）" if v["actor"] else ""),
            })
        if isinstance(last_accessed_at, (int, float)) and last_accessed_at > created:
            timeline.append({
                "event": "last_accessed",
                "timestamp": last_accessed_at,
                "description": "最近一次访问",
            })
        timeline.sort(key=lambda x: x["timestamp"] if isinstance(x["timestamp"], (int, float)) else 0)

        return {
            "memory_id": rid,
            "basic": {
                "content_preview": (content or "")[:120],
                "category": category or "general",
                "importance": importance or "MEDIUM",
                "layer": layer or "short_term",
                "source_agent": source_agent or "",
                "source_session": source_session or "",
                "starred": bool(starred),
                "pinned": bool(pinned),
                "access_count": access_count or 0,
            },
            "stats": {
                "age_days": age_days,
                "version_count": len(versions),
                "link_count_out": len(links_out),
                "link_count_in": len(links_in),
                "audit_event_count": len(audit_events),
            },
            "versions": versions,
            "links_out": links_out,
            "links_in": links_in,
            "audit_events": audit_events,
            "lifecycle_timeline": timeline,
        }

    def memory_reinforce(self,
                         agent_id: str,
                         days: int = 90,
                         limit: int = 10) -> Dict[str, Any]:
        """记忆强化候选（v5.4.1 新增）

        前瞻性识别「高价值但正在衰减」的记忆：综合重要度、星标、
        访问活跃度、记忆强度与遗忘分数，输出强化候选排序、
        每条候选的原因与推荐动作（复习/提权/合并观察）。
        """
        conn = self._get_conn()
        aid = _filter_unicode_ctrl(agent_id[:128]) if isinstance(agent_id, str) else ""
        if not aid:
            return {"error": "Agent ID 不能为空"}
        days = max(1, min(365, int(days)))
        limit = max(1, min(50, int(limit)))
        now = time.time()
        since = now - days * 86400

        # v5.4.1 安全：参数化 SQL + 行数限制
        cur = conn.execute(
            "SELECT id, content, importance, starred, strength, "
            "forgetting_score, last_accessed_at, access_count, category "
            "FROM memories "
            "WHERE source_agent = ? AND category != 'trash' AND created_at >= ?",
            (aid, since)
        )
        rows = _limited_fetch(cur, limit=10000)

        if not rows:
            return {
                "agent_id": aid,
                "days": days,
                "total_scanned": 0,
                "candidates": [],
                "summary": "窗口内无记忆，无需强化。",
            }

        candidates: List[Dict[str, Any]] = []
        for r in rows:
            (mid, content, importance, starred, strength,
             forgetting_score, last_accessed_at, access_count, category) = r

            imp = (importance or "MEDIUM").upper()
            if imp not in ("LOW", "MEDIUM", "HIGH", "CRITICAL"):
                imp = "MEDIUM"
            imp_w = {"LOW": 0.5, "MEDIUM": 1.0, "HIGH": 1.6, "CRITICAL": 2.2}[imp]
            ac = access_count if isinstance(access_count, int) and access_count > 0 else 0
            is_starred = bool(starred)

            # 价值分：重要度 + 星标 + 历史活跃度
            value_score = min(100.0, imp_w * 30.0 + (15.0 if is_starred else 0.0)
                              + min(ac, 10) * 2.0)

            # 风险分：闲置时长 + 遗忘分数 + 强度衰减
            last_acc = last_accessed_at if isinstance(last_accessed_at, (int, float)) else now
            days_idle = max(0.0, (now - last_acc) / 86400.0)
            fs = forgetting_score if isinstance(forgetting_score, (int, float)) else 0.0
            fs = max(0.0, min(100.0, float(fs)))
            st = strength if isinstance(strength, (int, float)) else 1.0
            weakness = max(0.0, 1.0 - min(1.0, float(st)))
            risk_score = min(100.0, (days_idle / float(days)) * 50.0
                             + fs * 0.3 + weakness * 20.0)

            reinforce_score = round(value_score * 0.5 + risk_score * 0.5, 1)

            reasons: List[str] = []
            if imp in ("HIGH", "CRITICAL"):
                reasons.append(f"重要度为 {imp}，属于高价值记忆")
            if is_starred:
                reasons.append("已被星标标记")
            if days_idle >= days * 0.5:
                reasons.append(f"已闲置 {round(days_idle, 1)} 天，超过窗口一半")
            if fs >= 40.0:
                reasons.append(f"遗忘分数偏高（{round(fs, 1)}）")
            if weakness > 0.2:
                reasons.append(f"记忆强度衰减至 {round(st, 2)}")
            if ac >= 5:
                reasons.append(f"历史访问活跃（{ac} 次），遗忘损失大")
            if not reasons:
                reasons.append("综合评分进入候选区间")

            # 推荐动作
            if is_starred or imp == "CRITICAL":
                action = "priority_review"
            elif imp in ("LOW", "MEDIUM") and ac >= 5:
                action = "promote_importance"
            elif days_idle >= days * 0.6:
                action = "schedule_review"
            else:
                action = "keep_monitoring"

            candidates.append({
                "memory_id": mid,
                "content_preview": (content or "")[:80],
                "category": category or "general",
                "importance": imp,
                "starred": is_starred,
                "days_idle": round(days_idle, 1),
                "value_score": round(value_score, 1),
                "risk_score": round(risk_score, 1),
                "reinforce_score": reinforce_score,
                "reasons": reasons,
                "recommended_action": action,
            })

        candidates.sort(key=lambda x: -x["reinforce_score"])
        top = candidates[:limit]

        action_counts: Dict[str, int] = {}
        for t in top:
            action_counts[t["recommended_action"]] = action_counts.get(t["recommended_action"], 0) + 1
        summary_parts = [f"扫描 {len(rows)} 条记忆，识别出 {len(top)} 条强化候选。"]
        if action_counts.get("priority_review"):
            summary_parts.append(f"{action_counts['priority_review']} 条建议优先复习。")
        if action_counts.get("schedule_review"):
            summary_parts.append(f"{action_counts['schedule_review']} 条建议创建复习计划。")
        if action_counts.get("promote_importance"):
            summary_parts.append(f"{action_counts['promote_importance']} 条被低估，建议提升重要度。")

        return {
            "agent_id": aid,
            "days": days,
            "total_scanned": len(rows),
            "candidates": top,
            "action_distribution": action_counts,
            "summary": " ".join(summary_parts),
        }

    # ===== v5.4.1 新增：AI 短剧三大能力 =====

    def drama_plot_thread(self, drama_id: str) -> Dict[str, Any]:
        """剧情线索/伏笔追踪（v5.4.1 新增）

        从场景与台词中识别「埋设伏笔（setup）」与「揭示回收（payoff）」
        标记，按时间顺序贪心匹配，输出全部线索、未回收的开放式
        线索与回收率，辅助编剧检查伏笔闭环。
        """
        conn = self._get_conn()
        did = _filter_unicode_ctrl(drama_id[:64]) if isinstance(drama_id, str) and drama_id else ""
        if not did:
            return {"error": "短剧 ID 不能为空"}

        drow = conn.execute(
            "SELECT id, title FROM drama_series WHERE id = ?",
            (did,)
        ).fetchone()
        if not drow:
            return {"error": "短剧不存在"}
        title = drow[1] or "未命名"

        srows = conn.execute(
            "SELECT id, episode, scene_number, title, content, tags "
            "FROM drama_scenes WHERE drama_id = ? "
            "ORDER BY episode ASC, scene_number ASC",
            (did,)
        )
        scenes = _limited_fetch(srows, limit=5000)

        if not scenes:
            return {
                "drama_id": did,
                "title": title,
                "total_scenes": 0,
                "threads": [],
                "open_count": 0,
                "resolved_count": 0,
                "resolution_rate": 0.0,
                "suggestions": ["暂无场景数据，无法进行线索追踪。"],
            }

        setup_words = {
            "伏笔", "暗示", "预言", "约定", "秘密", "谜", "埋下", "线索",
            "信物", "悬念", "setup", "foreshadow", "promise", "secret",
            "clue", "mystery",
        }
        payoff_words = {
            "真相", "揭晓", "实现", "应验", "解开", "发现", "原来", "竟然",
            "终于", "回收", "payoff", "reveal", "truth", "finally",
        }

        # 预取每个场景的台词文本（一次查询，按 scene_id 分组）
        lrows = conn.execute(
            "SELECT scene_id, line_text, tags FROM drama_lines WHERE drama_id = ?",
            (did,)
        )
        scene_lines: Dict[str, List[str]] = {}
        for l in _limited_fetch(lrows, limit=10000):
            sid = l[0] or ""
            scene_lines.setdefault(sid, []).append(l[1] or "")
            if l[2]:
                scene_lines[sid].append(l[2])

        def _match_words(text: str, words) -> List[str]:
            t = text.lower()
            return sorted({w for w in words if w in t})

        open_threads: List[Dict[str, Any]] = []
        threads: List[Dict[str, Any]] = []
        thread_seq = 0

        for s in scenes:
            sid, episode, scene_number, s_title, s_content, s_tags = (
                s[0], s[1] or 0, s[2] or 0, s[3] or "", s[4] or "", s[5] or ""
            )
            text_blob = " ".join(
                [s_title, s_content, s_tags] + scene_lines.get(sid, [])
            )
            setup_hits = _match_words(text_blob, setup_words)
            payoff_hits = _match_words(text_blob, payoff_words)

            # 先回收（payoff 只回收更早埋设的线索）
            if payoff_hits and open_threads:
                th = open_threads.pop(0)
                th["status"] = "resolved"
                th["payoff_episode"] = episode
                th["payoff_scene_id"] = sid
                th["payoff_markers"] = payoff_hits

            # 再埋设
            if setup_hits:
                thread_seq += 1
                thread = {
                    "thread_id": f"T{thread_seq}",
                    "name": s_title or setup_hits[0],
                    "setup_episode": episode,
                    "setup_scene_id": sid,
                    "markers": setup_hits,
                    "status": "open",
                    "payoff_episode": None,
                    "payoff_scene_id": None,
                }
                threads.append(thread)
                open_threads.append(thread)

        resolved_count = sum(1 for t in threads if t["status"] == "resolved")
        open_count = len(threads) - resolved_count
        resolution_rate = round(resolved_count / len(threads) * 100, 2) if threads else 0.0

        suggestions: List[str] = []
        if open_count > 0:
            open_eps = [t["setup_episode"] for t in threads if t["status"] == "open"]
            suggestions.append(
                f"有 {open_count} 条伏笔尚未回收（最早埋设于第 {min(open_eps)} 集），建议尽快安排回收"
            )
        if threads and resolution_rate < 50.0:
            suggestions.append("伏笔回收率低于 50%，观众可能产生'挖坑不填'的负面体验")
        if not threads:
            suggestions.append("未检测到明显伏笔标记，可增加悬念/伏笔以提升追剧粘性")
        elif resolution_rate >= 80.0:
            suggestions.append("伏笔闭环良好，剧情完整度高")

        return {
            "drama_id": did,
            "title": title,
            "total_scenes": len(scenes),
            "threads": threads[:50],
            "open_count": open_count,
            "resolved_count": resolved_count,
            "resolution_rate": resolution_rate,
            "suggestions": suggestions,
        }

    def drama_episode_curve(self, drama_id: str) -> Dict[str, Any]:
        """分集张力曲线（v5.4.1 新增）

        按集聚合台词量、冲突词与强度词，生成全剧张力曲线，
        识别高潮集、波动率，并对曲线形态分类（上升/下降/
        中段高峰/平稳），辅助节奏诊断。
        """
        conn = self._get_conn()
        did = _filter_unicode_ctrl(drama_id[:64]) if isinstance(drama_id, str) and drama_id else ""
        if not did:
            return {"error": "短剧 ID 不能为空"}

        drow = conn.execute(
            "SELECT id, title, total_episodes FROM drama_series WHERE id = ?",
            (did,)
        ).fetchone()
        if not drow:
            return {"error": "短剧不存在"}
        title = drow[1] or "未命名"
        declared_eps = drow[2] if isinstance(drow[2], int) and drow[2] > 0 else 0

        conflict_words = {
            "不", "别", "错", "滚", "闭嘴", "打", "杀", "死",
            "no", "not", "stop", "hate", "fight", "kill",
            "冲突", "争吵", "背叛", "欺骗", "威胁", "逼迫",
        }
        intensity_words = {
            "必须", "马上", "立刻", "快", "紧急", "危险",
            "终于", "竟然", "居然", "到底", "不可能",
        }

        # 场景按集聚合
        srows = conn.execute(
            "SELECT episode, id FROM drama_scenes WHERE drama_id = ?",
            (did,)
        )
        ep_scenes: Dict[int, int] = {}
        scene_ids: List[str] = []
        for s in _limited_fetch(srows, limit=5000):
            ep = s[0] if isinstance(s[0], int) and s[0] > 0 else 0
            ep_scenes[ep] = ep_scenes.get(ep, 0) + 1
            scene_ids.append(s[1])

        # 台词按集聚合
        ep_lines: Dict[int, int] = {}
        ep_conflict: Dict[int, int] = {}
        ep_intensity: Dict[int, int] = {}
        ep_chars: Dict[int, set] = {}
        lrows = conn.execute(
            "SELECT episode, line_text, character_id FROM drama_lines WHERE drama_id = ?",
            (did,)
        )
        for l in _limited_fetch(lrows, limit=10000):
            ep = l[0] if isinstance(l[0], int) and l[0] > 0 else 0
            text = (l[1] or "").lower()
            ep_lines[ep] = ep_lines.get(ep, 0) + 1
            ep_conflict[ep] = ep_conflict.get(ep, 0) + sum(1 for w in conflict_words if w in text)
            ep_intensity[ep] = ep_intensity.get(ep, 0) + sum(1 for w in intensity_words if w in text)
            if l[2]:
                ep_chars.setdefault(ep, set()).add(l[2])

        all_eps = sorted(set(ep_scenes.keys()) | set(ep_lines.keys()))
        if not all_eps:
            return {
                "drama_id": did,
                "title": title,
                "total_episodes": declared_eps,
                "curve": [],
                "climax_episode": None,
                "avg_tension": 0.0,
                "volatility": 0.0,
                "shape": "no_data",
                "suggestions": ["暂无场景/台词数据，无法生成张力曲线。"],
            }

        raw_scores: Dict[int, float] = {}
        for ep in all_eps:
            lc = ep_lines.get(ep, 0)
            cc = len(ep_chars.get(ep, set()))
            raw = (lc * 2 + ep_conflict.get(ep, 0) * 10
                   + ep_intensity.get(ep, 0) * 8 + cc * 3)
            raw_scores[ep] = float(raw)

        max_raw = max(raw_scores.values()) if raw_scores else 0.0
        curve: List[Dict[str, Any]] = []
        for ep in all_eps:
            tension = round(raw_scores[ep] / max_raw * 100.0, 1) if max_raw > 0 else 0.0
            curve.append({
                "episode": ep,
                "tension": tension,
                "scenes": ep_scenes.get(ep, 0),
                "lines": ep_lines.get(ep, 0),
            })

        climax = max(curve, key=lambda x: x["tension"])
        tensions = [p["tension"] for p in curve]
        avg_tension = round(sum(tensions) / len(tensions), 1)

        # 波动率：标准差 / 均值
        if avg_tension > 0:
            mean = sum(tensions) / len(tensions)
            var = sum((t - mean) ** 2 for t in tensions) / len(tensions)
            volatility = round((var ** 0.5) / mean, 3)
        else:
            volatility = 0.0

        # 形态分类：按顺序三段比较
        shape = "steady"
        if len(tensions) >= 3:
            third = max(1, len(tensions) // 3)
            seg = lambda part: sum(part) / len(part) if part else 0.0
            s1 = seg(tensions[:third])
            s2 = seg(tensions[third:2 * third])
            s3 = seg(tensions[2 * third:])
            if s3 > s1 * 1.2 and s3 >= s2:
                shape = "rising"
            elif s1 > s3 * 1.2 and s1 >= s2:
                shape = "falling"
            elif s2 > s1 and s2 > s3:
                shape = "mid_peak"
            else:
                shape = "steady"

        suggestions: List[str] = []
        total_eps = declared_eps if declared_eps else len(all_eps)
        if climax["episode"] > 0 and total_eps > 0:
            pos = climax["episode"] / max(1, total_eps)
            if pos < 0.3:
                suggestions.append(f"高潮出现在第 {climax['episode']} 集（前段），后段可能缺乏张力，建议增设二次高潮")
            elif pos > 0.9:
                suggestions.append(f"高潮贴近结局（第 {climax['episode']} 集），整体铺垫充分")
        if volatility < 0.15 and len(tensions) >= 3:
            suggestions.append("各集张力差异较小，节奏偏平，建议制造更明显的起伏")
        if shape == "falling":
            suggestions.append("曲线整体走低，建议在中后段加强冲突或悬念")
        if not suggestions:
            suggestions.append("张力曲线健康，节奏把控良好")

        return {
            "drama_id": did,
            "title": title,
            "total_episodes": total_eps,
            "curve": curve,
            "climax_episode": climax["episode"],
            "climax_tension": climax["tension"],
            "avg_tension": avg_tension,
            "volatility": volatility,
            "shape": shape,
            "suggestions": suggestions,
        }

    def drama_screen_time(self, drama_id: str) -> Dict[str, Any]:
        """角色戏份平衡分析（v5.4.1 新增）

        统计每个角色的台词量、字数、出场场景与集数占比，
        计算群像平衡度（Top 占比 + 基尼系数），识别
        独角戏/双核/群像结构并给出建议。
        """
        conn = self._get_conn()
        did = _filter_unicode_ctrl(drama_id[:64]) if isinstance(drama_id, str) and drama_id else ""
        if not did:
            return {"error": "短剧 ID 不能为空"}

        drow = conn.execute(
            "SELECT id, title FROM drama_series WHERE id = ?",
            (did,)
        ).fetchone()
        if not drow:
            return {"error": "短剧不存在"}
        title = drow[1] or "未命名"

        crows = conn.execute(
            "SELECT id, name, role FROM drama_characters WHERE drama_id = ?",
            (did,)
        )
        chars = _limited_fetch(crows, limit=1000)
        if not chars:
            return {
                "drama_id": did,
                "title": title,
                "total_lines": 0,
                "characters": [],
                "balance": {},
                "suggestions": ["暂无角色数据。"],
            }

        total_row = conn.execute(
            "SELECT COUNT(*) FROM drama_lines WHERE drama_id = ?",
            (did,)
        ).fetchone()
        total_lines = total_row[0] if total_row else 0

        char_stats: List[Dict[str, Any]] = []
        for c in chars:
            cid, cname, crole = c[0], c[1] or "未命名角色", c[2] or "supporting"
            lrow = conn.execute(
                "SELECT COUNT(*), COUNT(DISTINCT scene_id), COUNT(DISTINCT episode) "
                "FROM drama_lines WHERE drama_id = ? AND character_id = ?",
                (did, cid)
            ).fetchone()
            line_count = lrow[0] or 0
            scene_count = lrow[1] or 0
            episode_count = lrow[2] or 0

            wrow = conn.execute(
                "SELECT COALESCE(SUM(LENGTH(line_text)), 0) "
                "FROM drama_lines WHERE drama_id = ? AND character_id = ?",
                (did, cid)
            ).fetchone()
            word_count = wrow[0] or 0

            share_pct = round(line_count / total_lines * 100, 2) if total_lines > 0 else 0.0
            char_stats.append({
                "char_id": cid,
                "name": cname,
                "role": crole,
                "line_count": line_count,
                "word_count": word_count,
                "scene_count": scene_count,
                "episode_count": episode_count,
                "share_pct": share_pct,
            })

        char_stats.sort(key=lambda x: -x["line_count"])
        for rank, cs in enumerate(char_stats, start=1):
            cs["rank"] = rank

        # 基尼系数（基于台词量）
        values = sorted([cs["line_count"] for cs in char_stats])
        n = len(values)
        total_val = sum(values)
        if n > 0 and total_val > 0:
            cumulative = 0.0
            weighted_sum = 0.0
            for i, v in enumerate(values, start=1):
                weighted_sum += i * v
            gini = round((2.0 * weighted_sum) / (n * total_val) - (n + 1.0) / n, 3)
            gini = max(0.0, min(1.0, gini))
        else:
            gini = 0.0

        top_share = char_stats[0]["share_pct"] if char_stats else 0.0
        top2_share = round(sum(cs["share_pct"] for cs in char_stats[:2]), 2) if len(char_stats) >= 2 else top_share
        if top_share >= 50.0:
            structure = "one_lead"
            structure_label = "独角戏结构"
        elif top2_share >= 70.0:
            structure = "dual_lead"
            structure_label = "双核结构"
        else:
            structure = "ensemble"
            structure_label = "群像结构"

        suggestions: List[str] = []
        zero_chars = [cs["name"] for cs in char_stats if cs["line_count"] == 0]
        if zero_chars:
            suggestions.append(f"角色 {('、'.join(zero_chars[:5]))} 没有台词，考虑删减或增加戏份")
        if structure == "one_lead" and n >= 4:
            suggestions.append("主角戏份占比过高，配角空间不足，可能影响群像丰满度")
        if structure == "ensemble" and gini < 0.15 and n >= 4:
            suggestions.append("群像戏份非常均衡，注意保持主线焦点，避免观众注意力分散")
        if not suggestions:
            suggestions.append("角色戏份分布合理")

        return {
            "drama_id": did,
            "title": title,
            "total_lines": total_lines,
            "characters": char_stats,
            "balance": {
                "top_character": char_stats[0]["name"] if char_stats else None,
                "top_share_pct": top_share,
                "top2_share_pct": top2_share,
                "gini_coefficient": gini,
                "structure": structure,
                "structure_label": structure_label,
            },
            "suggestions": suggestions,
        }

    # ===== v5.4.3 新增：Agent 记忆影响力图谱 + 记忆重叠 + 冲突检测 =====

    def agent_influence_map(self,
                            agent_id: str,
                            days: int = 30) -> Dict[str, Any]:
        """Agent 记忆影响力图谱（v5.4.3 新增）

        分析指定 Agent 的记忆如何被其他 Agent 引用/关联，以及该 Agent
        引用了哪些其他 Agent 的记忆，构建双向影响力网络。

        影响力来源：
        - memory_links 表中的跨 Agent 关联
        - memories.metadata 中记录的 referenced_agents 字段
        - 共享标签/分类的间接关联

        Args:
            agent_id: Agent ID
            days: 回溯天数（1-365）

        Returns:
            影响力图谱：节点列表、边列表、入度/出度排行、核心影响力 Agent
        """
        conn = self._get_conn()
        aid = _filter_unicode_ctrl(agent_id[:128]) if isinstance(agent_id, str) else ""
        if not aid:
            return {"error": "Agent ID 不能为空"}

        days = max(1, min(365, int(days)))
        now = time.time()
        since = now - days * 86400

        # v5.4.3 安全：参数化 SQL + 行数限制
        cur = conn.execute(
            "SELECT ml.source_id, ml.target_id, ml.link_type, "
            "m_src.source_agent AS src_agent, m_tgt.source_agent AS tgt_agent "
            "FROM memory_links ml "
            "JOIN memories m_src ON ml.source_id = m_src.id "
            "JOIN memories m_tgt ON ml.target_id = m_tgt.id "
            "WHERE (m_src.source_agent = ? OR m_tgt.source_agent = ?) "
            "AND m_src.created_at >= ? AND m_tgt.created_at >= ?",
            (aid, aid, since, since)
        )
        link_rows = _limited_fetch(cur, limit=5000)

        cur2 = conn.execute(
            "SELECT id, source_agent, metadata, category, tags, created_at "
            "FROM memories "
            "WHERE category != 'trash' AND created_at >= ? "
            "AND source_agent != ?",
            (since, aid)
        )
        meta_rows = _limited_fetch(cur2, limit=5000)

        nodes: Dict[str, Dict[str, Any]] = {}
        edges: List[Dict[str, Any]] = []

        def _ensure_node(nid: str):
            if nid and nid not in nodes:
                nodes[nid] = {
                    "agent_id": nid,
                    "influence_in": 0,
                    "influence_out": 0,
                    "shared_tags": 0,
                }

        _ensure_node(aid)

        for lr in link_rows:
            src_agent = lr[3] or ""
            tgt_agent = lr[4] or ""
            if not src_agent or not tgt_agent or src_agent == tgt_agent:
                continue
            _ensure_node(src_agent)
            _ensure_node(tgt_agent)

            if src_agent == aid:
                nodes[aid]["influence_in"] += 1
                nodes[tgt_agent]["influence_out"] += 1
                edges.append({
                    "from": tgt_agent,
                    "to": aid,
                    "type": lr[2] or "related",
                    "direction": "influences",
                })
            elif tgt_agent == aid:
                nodes[aid]["influence_out"] += 1
                nodes[tgt_agent]["influence_in"] += 1
                edges.append({
                    "from": aid,
                    "to": tgt_agent,
                    "type": lr[2] or "related",
                    "direction": "influences",
                })

        for mr in meta_rows:
            other_agent = mr[1] or ""
            if not other_agent:
                continue
            try:
                meta = json.loads(mr[2]) if mr[2] else {}
            except (json.JSONDecodeError, TypeError):
                meta = {}
            ref_agents = meta.get("referenced_agents", [])
            if not isinstance(ref_agents, list):
                continue
            if aid in ref_agents:
                _ensure_node(other_agent)
                nodes[aid]["influence_out"] += 1
                nodes[other_agent]["influence_in"] += 1
                edges.append({
                    "from": aid,
                    "to": other_agent,
                    "type": "metadata_ref",
                    "direction": "influences",
                })

        # 共享标签的间接关联
        aid_tags_cur = conn.execute(
            "SELECT tags FROM memories "
            "WHERE source_agent = ? AND category != 'trash' AND created_at >= ?",
            (aid, since)
        )
        aid_tag_rows = _limited_fetch(aid_tags_cur, limit=2000)
        aid_tags: set = set()
        for tr in aid_tag_rows:
            try:
                tags = json.loads(tr[0]) if tr[0] else []
            except (json.JSONDecodeError, TypeError):
                tags = []
            if isinstance(tags, list):
                aid_tags.update(t[:64] for t in tags if isinstance(t, str) and t)

        if aid_tags:
            for mr in meta_rows:
                other_agent = mr[1] or ""
                if not other_agent or other_agent == aid:
                    continue
                try:
                    tags = json.loads(mr[3]) if mr[3] else []
                except (json.JSONDecodeError, TypeError):
                    tags = []
                if not isinstance(tags, list):
                    continue
                other_tags = set(t[:64] for t in tags if isinstance(t, str) and t)
                shared = aid_tags & other_tags
                if shared:
                    _ensure_node(other_agent)
                    nodes[other_agent]["shared_tags"] = len(shared)

        node_list = sorted(nodes.values(),
                           key=lambda x: x["influence_in"] + x["influence_out"],
                           reverse=True)
        top_influencers = [n for n in node_list if n["agent_id"] != aid][:10]

        total_edges = len(edges)
        return {
            "agent_id": aid,
            "days": days,
            "total_nodes": len(nodes),
            "total_edges": total_edges,
            "nodes": node_list[:50],
            "edges": edges[:200],
            "top_influencers": top_influencers,
            "self_influence_in": nodes[aid]["influence_in"],
            "self_influence_out": nodes[aid]["influence_out"],
            "influence_score": round(
                (nodes[aid]["influence_out"] * 2 + nodes[aid]["influence_in"]) /
                max(1, total_edges), 4
            ) if total_edges else 0.0,
        }

    def memory_overlap(self,
                       agent_id_a: str,
                       agent_id_b: str,
                       days: int = 30) -> Dict[str, Any]:
        """记忆重叠分析（v5.4.3 新增）

        分析两个 Agent 的记忆在标签、分类、关键词层面的重叠度，
        识别共同知识领域和各自独有的知识领域。

        Args:
            agent_id_a: Agent A ID
            agent_id_b: Agent B ID
            days: 回溯天数（1-365）

        Returns:
            重叠分析：标签重叠率、分类重叠率、共享标签、独有标签、
            共享分类、Jaccard 相似度
        """
        conn = self._get_conn()
        aid_a = _filter_unicode_ctrl(agent_id_a[:128]) if isinstance(agent_id_a, str) else ""
        aid_b = _filter_unicode_ctrl(agent_id_b[:128]) if isinstance(agent_id_b, str) else ""
        if not aid_a or not aid_b:
            return {"error": "两个 Agent ID 都不能为空"}
        if aid_a == aid_b:
            return {"error": "两个 Agent ID 不能相同"}

        days = max(1, min(365, int(days)))
        now = time.time()
        since = now - days * 86400

        def _get_agent_tags_cats(aid: str) -> Tuple[set, set, List[str]]:
            cur = conn.execute(
                "SELECT tags, category, content FROM memories "
                "WHERE source_agent = ? AND category != 'trash' AND created_at >= ?",
                (aid, since)
            )
            rows = _limited_fetch(cur, limit=5000)
            tags_set: set = set()
            cats_set: set = set()
            contents: List[str] = []
            for r in rows:
                try:
                    tags = json.loads(r[0]) if r[0] else []
                except (json.JSONDecodeError, TypeError):
                    tags = []
                if isinstance(tags, list):
                    for t in tags:
                        if isinstance(t, str) and t:
                            tags_set.add(t[:64])
                if r[1]:
                    cats_set.add(r[1][:64])
                if r[2]:
                    contents.append(r[2][:200])
            return tags_set, cats_set, contents

        tags_a, cats_a, contents_a = _get_agent_tags_cats(aid_a)
        tags_b, cats_b, contents_b = _get_agent_tags_cats(aid_b)

        shared_tags = tags_a & tags_b
        unique_tags_a = tags_a - tags_b
        unique_tags_b = tags_b - tags_a
        tag_union = tags_a | tags_b
        tag_jaccard = round(len(shared_tags) / len(tag_union), 4) if tag_union else 0.0
        tag_overlap_pct = round(len(shared_tags) / len(tags_a) * 100, 2) if tags_a else 0.0

        shared_cats = cats_a & cats_b
        unique_cats_a = cats_a - cats_b
        unique_cats_b = cats_b - cats_a
        cat_union = cats_a | cats_b
        cat_jaccard = round(len(shared_cats) / len(cat_union), 4) if cat_union else 0.0

        stop_words = {
            "的", "了", "和", "是", "就", "都", "而", "及", "与", "在",
            "the", "a", "an", "and", "or", "is", "are", "was", "were",
            "to", "of", "in", "for", "on", "at", "by", "it", "as",
        }

        def _extract_keywords(texts: List[str], top_n: int = 50) -> set:
            kw: Dict[str, int] = {}
            for text in texts:
                words = text.lower().split()
                for w in words:
                    w = w.strip(".,!?;:\"'()[]{}").strip()
                    if len(w) >= 2 and w not in stop_words:
                        kw[w] = kw.get(w, 0) + 1
                clean = "".join(c for c in text if c.isalpha())
                for i in range(len(clean) - 1):
                    gram = clean[i:i + 2]
                    if gram not in stop_words:
                        kw[gram] = kw.get(gram, 0) + 1
            return set(sorted(kw, key=kw.get, reverse=True)[:top_n])

        kw_a = _extract_keywords(contents_a)
        kw_b = _extract_keywords(contents_b)
        shared_kw = kw_a & kw_b
        kw_jaccard = round(len(shared_kw) / len(kw_a | kw_b), 4) if (kw_a | kw_b) else 0.0

        overall_sim = round(
            tag_jaccard * 0.4 + cat_jaccard * 0.3 + kw_jaccard * 0.3, 4
        )

        return {
            "agent_a": aid_a,
            "agent_b": aid_b,
            "days": days,
            "tags": {
                "shared": sorted(shared_tags),
                "unique_a": sorted(unique_tags_a),
                "unique_b": sorted(unique_tags_b),
                "overlap_pct": tag_overlap_pct,
                "jaccard": tag_jaccard,
            },
            "categories": {
                "shared": sorted(shared_cats),
                "unique_a": sorted(unique_cats_a),
                "unique_b": sorted(unique_cats_b),
                "jaccard": cat_jaccard,
            },
            "keywords": {
                "shared_count": len(shared_kw),
                "shared": sorted(shared_kw)[:20],
                "jaccard": kw_jaccard,
            },
            "overall_similarity": overall_sim,
            "similarity_level": (
                "high" if overall_sim >= 0.5 else
                "medium" if overall_sim >= 0.2 else
                "low"
            ),
        }

    def conflict_graph(self,
                       agent_id: str,
                       days: int = 30) -> Dict[str, Any]:
        """记忆冲突检测图（v5.4.3 新增）

        检测同一 Agent 记忆中潜在的知识冲突：相同标签/分类下
        重要性差异显著的记忆对，或内容关键词高度重叠但重要性
        矛盾的记忆对。

        Args:
            agent_id: Agent ID
            days: 回溯天数（1-365）

        Returns:
            冲突图：冲突节点对、冲突类型、严重度分布、冲突密度
        """
        conn = self._get_conn()
        aid = _filter_unicode_ctrl(agent_id[:128]) if isinstance(agent_id, str) else ""
        if not aid:
            return {"error": "Agent ID 不能为空"}

        days = max(1, min(365, int(days)))
        now = time.time()
        since = now - days * 86400

        cur = conn.execute(
            "SELECT id, content, category, tags, importance, created_at "
            "FROM memories "
            "WHERE source_agent = ? AND category != 'trash' AND created_at >= ? "
            "ORDER BY created_at DESC",
            (aid, since)
        )
        rows = _limited_fetch(cur, limit=3000)

        if len(rows) < 2:
            return {
                "agent_id": aid,
                "days": days,
                "total_memories": len(rows),
                "conflicts": [],
                "conflict_count": 0,
                "severity_distribution": {"high": 0, "medium": 0, "low": 0},
                "conflict_density": 0.0,
            }

        memories: List[Dict[str, Any]] = []
        for r in rows:
            try:
                tags = json.loads(r[3]) if r[3] else []
            except (json.JSONDecodeError, TypeError):
                tags = []
            if not isinstance(tags, list):
                tags = []
            memories.append({
                "id": r[0],
                "content": (r[1] or "")[:200],
                "category": r[2] or "general",
                "tags": set(t[:64] for t in tags if isinstance(t, str) and t),
                "importance": (r[4] or "MEDIUM").upper(),
                "created_at": r[5] or now,
            })

        imp_order = {"LOW": 0, "MEDIUM": 1, "HIGH": 2, "CRITICAL": 3}

        conflicts: List[Dict[str, Any]] = []
        severity_dist = {"high": 0, "medium": 0, "low": 0}

        n = len(memories)
        max_pairs = min(n * (n - 1) // 2, 50000)

        checked = 0
        for i in range(n):
            if checked >= max_pairs:
                break
            for j in range(i + 1, n):
                if checked >= max_pairs:
                    break
                checked += 1

                mi = memories[i]
                mj = memories[j]

                same_cat = mi["category"] == mj["category"]
                tag_overlap = mi["tags"] & mj["tags"]
                tag_shared = len(tag_overlap) > 0

                if not (same_cat or tag_shared):
                    continue

                imp_diff = abs(
                    imp_order.get(mi["importance"], 1) -
                    imp_order.get(mj["importance"], 1)
                )

                words_i = set(mi["content"].lower().split())
                words_j = set(mj["content"].lower().split())
                content_overlap = words_i & words_j
                content_overlap_count = len(content_overlap)

                if content_overlap_count < 2 and not tag_shared:
                    continue

                if imp_diff >= 2 and content_overlap_count >= 3:
                    severity = "high"
                elif imp_diff >= 1 and content_overlap_count >= 2:
                    severity = "medium"
                else:
                    severity = "low"

                severity_dist[severity] += 1

                conflicts.append({
                    "memory_a": {
                        "id": mi["id"],
                        "content_preview": mi["content"][:80],
                        "importance": mi["importance"],
                        "category": mi["category"],
                    },
                    "memory_b": {
                        "id": mj["id"],
                        "content_preview": mj["content"][:80],
                        "importance": mj["importance"],
                        "category": mj["category"],
                    },
                    "shared_tags": sorted(tag_overlap)[:10],
                    "content_overlap_words": sorted(content_overlap)[:10],
                    "importance_diff": imp_diff,
                    "severity": severity,
                    "conflict_type": (
                        "importance_contradiction" if imp_diff >= 2 else
                        "tag_content_overlap" if tag_shared else
                        "content_similarity"
                    ),
                })

        severity_order = {"high": 0, "medium": 1, "low": 2}
        conflicts.sort(key=lambda x: severity_order.get(x["severity"], 3))

        total_pairs = n * (n - 1) // 2
        conflict_density = round(len(conflicts) / max(1, total_pairs), 4)

        return {
            "agent_id": aid,
            "days": days,
            "total_memories": n,
            "conflicts": conflicts[:100],
            "conflict_count": len(conflicts),
            "severity_distribution": severity_dist,
            "conflict_density": conflict_density,
            "top_conflict_tags": sorted(
                set(t for c in conflicts[:50] for t in c["shared_tags"])
            )[:20],
        }

    # ===== v5.4.3 新增：AI 短剧经典台词地图 + 角色成长 + 场景节奏 =====

    def drama_quote_map(self,
                        drama_id: str) -> Dict[str, Any]:
        """经典台词地图（v5.4.3 新增）

        将短剧中的经典台词映射到集/场景/角色维度，分析经典台词的
        分布密度、角色贡献度和集数集中度。

        Args:
            drama_id: 短剧 ID

        Returns:
            台词地图：按集分布、按角色分布、按场景分布、台词时间线、
            经典密度评级
        """
        conn = self._get_conn()
        did = _filter_unicode_ctrl(drama_id[:64]) if isinstance(drama_id, str) else ""
        if not did:
            return {"error": "短剧 ID 不能为空"}

        drama_row = conn.execute(
            "SELECT title, total_episodes FROM drama_series WHERE id = ?",
            (did,)
        ).fetchone()
        if not drama_row:
            return {"error": "短剧不存在"}
        title = drama_row[0] or "未命名"
        total_eps = drama_row[1] or 0

        cur = conn.execute(
            "SELECT id, scene_id, character_id, character_name, line_text, "
            "episode, is_classic, created_at "
            "FROM drama_lines WHERE drama_id = ? "
            "ORDER BY episode ASC, created_at ASC",
            (did,)
        )
        rows = _limited_fetch(cur, limit=10000)

        if not rows:
            return {
                "drama_id": did,
                "title": title,
                "total_lines": 0,
                "classic_count": 0,
                "by_episode": {},
                "by_character": {},
                "by_scene": {},
                "timeline": [],
                "density_rating": "no_data",
            }

        total_lines = len(rows)
        classic_lines = [r for r in rows if r[6]]
        classic_count = len(classic_lines)

        by_episode: Dict[int, Dict[str, Any]] = {}
        by_character: Dict[str, Dict[str, Any]] = {}
        by_scene: Dict[str, Dict[str, Any]] = {}
        timeline: List[Dict[str, Any]] = []

        for r in rows:
            ep = r[5] or 0
            char_id = r[2] or ""
            char_name = r[3] or "未知角色"
            scene_id = r[1] or ""
            is_classic = bool(r[6])
            line_text = (r[4] or "")[:100]

            if ep not in by_episode:
                by_episode[ep] = {"total": 0, "classic": 0, "classic_lines": []}
            by_episode[ep]["total"] += 1
            if is_classic:
                by_episode[ep]["classic"] += 1
                if len(by_episode[ep]["classic_lines"]) < 5:
                    by_episode[ep]["classic_lines"].append({
                        "text": line_text,
                        "character": char_name,
                    })

            char_key = char_id or char_name
            if char_key not in by_character:
                by_character[char_key] = {
                    "name": char_name,
                    "total": 0,
                    "classic": 0,
                    "classic_lines": [],
                }
            by_character[char_key]["total"] += 1
            if is_classic:
                by_character[char_key]["classic"] += 1
                if len(by_character[char_key]["classic_lines"]) < 3:
                    by_character[char_key]["classic_lines"].append(line_text)

            if scene_id:
                if scene_id not in by_scene:
                    by_scene[scene_id] = {"total": 0, "classic": 0}
                by_scene[scene_id]["total"] += 1
                if is_classic:
                    by_scene[scene_id]["classic"] += 1

            if is_classic:
                timeline.append({
                    "episode": ep,
                    "scene_id": scene_id,
                    "character": char_name,
                    "text": line_text,
                })

        classic_ratio = classic_count / total_lines if total_lines else 0
        if classic_ratio >= 0.15:
            density_rating = "rich"
        elif classic_ratio >= 0.08:
            density_rating = "moderate"
        elif classic_ratio >= 0.03:
            density_rating = "sparse"
        else:
            density_rating = "minimal"

        char_ranking = sorted(
            by_character.values(),
            key=lambda x: x["classic"],
            reverse=True
        )[:10]

        ep_classic_counts = [(ep, d["classic"]) for ep, d in by_episode.items()]
        ep_classic_counts.sort(key=lambda x: x[1], reverse=True)
        top_episodes = ep_classic_counts[:5]

        return {
            "drama_id": did,
            "title": title,
            "total_episodes": total_eps,
            "total_lines": total_lines,
            "classic_count": classic_count,
            "classic_ratio": round(classic_ratio, 4),
            "density_rating": density_rating,
            "by_episode": {str(k): v for k, v in sorted(by_episode.items())},
            "by_character": char_ranking,
            "by_scene": {k: v for k, v in sorted(by_scene.items())[:20]},
            "timeline": timeline[:50],
            "top_episodes": [{"episode": e, "classic_count": c} for e, c in top_episodes],
        }

    def character_growth(self,
                         drama_id: str,
                         character_id: str) -> Dict[str, Any]:
        """角色成长深度分析（v5.4.3 新增）

        在 character_arc 基础上深化分析：情感成长轨迹（台词情感变化）、
        对话复杂度演变（台词长度/词汇丰富度）、角色活跃度阶段划分。

        Args:
            drama_id: 短剧 ID
            character_id: 角色 ID

        Returns:
            成长分析：情感弧线、复杂度曲线、活跃度阶段、成长评分、
            成长总结
        """
        conn = self._get_conn()
        did = _filter_unicode_ctrl(drama_id[:64]) if isinstance(drama_id, str) else ""
        cid = character_id[:64] if isinstance(character_id, str) else ""
        if not did or not cid:
            return {"error": "短剧 ID 和角色 ID 不能为空"}

        char_row = conn.execute(
            "SELECT name, role, description FROM drama_characters WHERE id = ? AND drama_id = ?",
            (cid, did)
        ).fetchone()
        if not char_row:
            return {"error": "角色不存在"}
        char_name = char_row[0] or "未知"
        char_role = char_row[1] or "supporting"

        cur = conn.execute(
            "SELECT id, line_text, episode, scene_id, created_at "
            "FROM drama_lines WHERE drama_id = ? AND character_id = ? "
            "ORDER BY episode ASC, created_at ASC",
            (did, cid)
        )
        rows = _limited_fetch(cur, limit=5000)

        if not rows:
            return {
                "drama_id": did,
                "character_id": cid,
                "character_name": char_name,
                "total_lines": 0,
                "emotion_arc": [],
                "complexity_curve": [],
                "activity_stages": [],
                "growth_score": 0,
                "growth_summary": "无台词数据",
            }

        positive_words = {"好", "爱", "开心", "高兴", "希望", "幸福", "感谢",
                          "美", "喜欢", "成功", "赢", "胜利", "笑", "温暖",
                          "good", "love", "happy", "hope", "great", "win"}
        negative_words = {"坏", "恨", "伤心", "难过", "绝望", "愤怒", "怕",
                          "失败", "输", "哭", "痛", "苦", "死", "离开",
                          "bad", "hate", "sad", "angry", "fear", "lose", "fail"}

        total_lines = len(rows)

        ep_data: Dict[int, List[Dict[str, Any]]] = {}
        for r in rows:
            ep = r[2] or 0
            if ep not in ep_data:
                ep_data[ep] = []
            ep_data[ep].append({
                "id": r[0],
                "text": r[1] or "",
                "scene_id": r[3] or "",
            })

        emotion_arc: List[Dict[str, Any]] = []
        complexity_curve: List[Dict[str, Any]] = []

        for ep in sorted(ep_data.keys()):
            lines = ep_data[ep]
            texts = [l["text"] for l in lines]

            all_text = " ".join(texts).lower()
            pos_hits = sum(1 for w in positive_words if w in all_text)
            neg_hits = sum(1 for w in negative_words if w in all_text)
            if pos_hits > neg_hits:
                emotion = "positive"
            elif neg_hits > pos_hits:
                emotion = "negative"
            else:
                emotion = "neutral"
            emotion_score = round((pos_hits - neg_hits) / max(1, len(lines)), 4)

            emotion_arc.append({
                "episode": ep,
                "line_count": len(lines),
                "emotion": emotion,
                "emotion_score": emotion_score,
            })

            avg_len = round(sum(len(t) for t in texts) / len(texts), 1) if texts else 0
            words = set()
            for t in texts:
                words.update(t.lower().split())
                words.update(c for c in t if c.isalpha())
            vocab_size = len(words)
            complexity = round(vocab_size / max(1, len(texts)), 2)

            complexity_curve.append({
                "episode": ep,
                "avg_line_length": avg_len,
                "vocabulary_size": vocab_size,
                "complexity_score": complexity,
            })

        episodes = sorted(ep_data.keys())
        total_eps_count = len(episodes)
        if total_eps_count >= 3:
            third = max(1, total_eps_count // 3)
            early_eps = episodes[:third]
            mid_eps = episodes[third:third * 2]
            late_eps = episodes[third * 2:]

            early_lines = sum(len(ep_data[e]) for e in early_eps)
            mid_lines = sum(len(ep_data[e]) for e in mid_eps)
            late_lines = sum(len(ep_data[e]) for e in late_eps)

            activity_stages = [
                {"stage": "early", "episodes": early_eps, "line_count": early_lines},
                {"stage": "middle", "episodes": mid_eps, "line_count": mid_lines},
                {"stage": "late", "episodes": late_eps, "line_count": late_lines},
            ]

            if late_lines > early_lines * 1.3:
                activity_trend = "rising"
            elif early_lines > late_lines * 1.3:
                activity_trend = "declining"
            elif mid_lines > early_lines * 1.2 and mid_lines > late_lines * 1.2:
                activity_trend = "peak_middle"
            else:
                activity_trend = "stable"
        else:
            activity_stages = [{"stage": "all", "episodes": episodes,
                                "line_count": total_lines}]
            activity_trend = "insufficient_data"

        if len(emotion_arc) >= 2:
            first_emotion = emotion_arc[0]["emotion_score"]
            last_emotion = emotion_arc[-1]["emotion_score"]
            emotion_delta = last_emotion - first_emotion

            first_complexity = complexity_curve[0]["complexity_score"]
            last_complexity = complexity_curve[-1]["complexity_score"]
            complexity_delta = last_complexity - first_complexity

            growth_score = min(100, max(0, round(
                (emotion_delta * 30 + complexity_delta * 20 +
                 (1 if activity_trend == "rising" else 0) * 25 +
                 min(100, total_lines * 2) / 100 * 25)
            )))
        else:
            emotion_delta = 0
            complexity_delta = 0
            growth_score = min(100, round(total_lines * 2))

        if growth_score >= 70:
            growth_summary = f"角色「{char_name}」展现出显著的成长轨迹，" \
                             f"情感{'正向转变' if emotion_delta > 0 else '深度增加'}，" \
                             f"对话复杂度提升，活跃度{activity_trend}"
        elif growth_score >= 40:
            growth_summary = f"角色「{char_name}」有一定成长，" \
                             f"活跃度{activity_trend}，" \
                             f"情感和复杂度有所变化"
        else:
            growth_summary = f"角色「{char_name}」成长幅度有限，" \
                             f"活跃度{activity_trend}，" \
                             f"建议增加戏份或情感冲突"

        return {
            "drama_id": did,
            "character_id": cid,
            "character_name": char_name,
            "character_role": char_role,
            "total_lines": total_lines,
            "total_episodes_active": total_eps_count,
            "emotion_arc": emotion_arc,
            "complexity_curve": complexity_curve,
            "activity_stages": activity_stages,
            "activity_trend": activity_trend,
            "emotion_delta": round(emotion_delta, 4),
            "complexity_delta": round(complexity_delta, 4),
            "growth_score": growth_score,
            "growth_summary": growth_summary,
        }

    def scene_rhythm(self,
                     drama_id: str) -> Dict[str, Any]:
        """场景节奏分析（v5.4.3 新增）

        分析短剧各场景的台词密度、对话节奏和场景长度分布，
        识别快节奏/慢节奏场景，评估整体节奏健康度。

        Args:
            drama_id: 短剧 ID

        Returns:
            节奏分析：各场景节奏数据、节奏曲线、节奏分类、
            整体节奏评估、节奏建议
        """
        conn = self._get_conn()
        did = _filter_unicode_ctrl(drama_id[:64]) if isinstance(drama_id, str) else ""
        if not did:
            return {"error": "短剧 ID 不能为空"}

        drama_row = conn.execute(
            "SELECT title FROM drama_series WHERE id = ?",
            (did,)
        ).fetchone()
        if not drama_row:
            return {"error": "短剧不存在"}
        title = drama_row[0] or "未命名"

        cur_scenes = conn.execute(
            "SELECT id, episode, scene_number, title, content "
            "FROM drama_scenes WHERE drama_id = ? "
            "ORDER BY episode ASC, scene_number ASC",
            (did,)
        )
        scene_rows = _limited_fetch(cur_scenes, limit=2000)

        if not scene_rows:
            return {
                "drama_id": did,
                "title": title,
                "total_scenes": 0,
                "scene_rhythms": [],
                "rhythm_curve": [],
                "overall_pace": "no_data",
                "suggestions": [],
            }

        scene_data: Dict[str, Dict[str, Any]] = {}
        for sr in scene_rows:
            scene_data[sr[0]] = {
                "scene_id": sr[0],
                "episode": sr[1] or 0,
                "scene_number": sr[2] or 0,
                "title": sr[3] or "",
                "content_length": len(sr[4] or ""),
                "line_count": 0,
                "total_chars": 0,
                "speaker_count": 0,
                "speakers": set(),
            }

        cur_lines = conn.execute(
            "SELECT scene_id, character_id, line_text "
            "FROM drama_lines WHERE drama_id = ? AND scene_id != ''",
            (did,)
        )
        line_rows = _limited_fetch(cur_lines, limit=10000)

        for lr in line_rows:
            sid = lr[0] or ""
            if sid not in scene_data:
                continue
            scene_data[sid]["line_count"] += 1
            scene_data[sid]["total_chars"] += len(lr[2] or "")
            if lr[1]:
                scene_data[sid]["speakers"].add(lr[1])

        scene_rhythms: List[Dict[str, Any]] = []
        for sd in scene_data.values():
            sd["speaker_count"] = len(sd["speakers"])
            del sd["speakers"]

            line_count = sd["line_count"]
            total_chars = sd["total_chars"]
            avg_line_len = round(total_chars / line_count, 1) if line_count else 0

            content_len = max(1, sd["content_length"])
            density = round(line_count / content_len * 100, 2)

            if line_count == 0:
                pace = "silent"
            elif density >= 5.0 or line_count >= 20:
                pace = "fast"
            elif density >= 2.0 or line_count >= 8:
                pace = "moderate"
            else:
                pace = "slow"

            sd["avg_line_length"] = avg_line_len
            sd["density"] = density
            sd["pace"] = pace
            scene_rhythms.append(sd)

        scene_rhythms.sort(key=lambda x: (x["episode"], x["scene_number"]))

        rhythm_curve = [
            {
                "scene_id": sr["scene_id"],
                "episode": sr["episode"],
                "pace": sr["pace"],
                "line_count": sr["line_count"],
                "density": sr["density"],
            }
            for sr in scene_rhythms
        ]

        total_scenes = len(scene_rhythms)
        pace_counts = {"fast": 0, "moderate": 0, "slow": 0, "silent": 0}
        for sr in scene_rhythms:
            pace_counts[sr["pace"]] += 1

        active_scenes = total_scenes - pace_counts["silent"]
        if active_scenes == 0:
            overall_pace = "no_data"
        elif pace_counts["fast"] >= active_scenes * 0.5:
            overall_pace = "fast_paced"
        elif pace_counts["slow"] >= active_scenes * 0.5:
            overall_pace = "slow_paced"
        else:
            overall_pace = "balanced"

        pace_transitions = 0
        prev_pace = None
        for rc in rhythm_curve:
            if rc["pace"] in ("silent",):
                continue
            if prev_pace and prev_pace != rc["pace"]:
                pace_transitions += 1
            prev_pace = rc["pace"]

        rhythm_variability = round(
            pace_transitions / max(1, active_scenes), 4
        ) if active_scenes else 0

        suggestions: List[str] = []
        if pace_counts["silent"] > total_scenes * 0.3:
            suggestions.append(f"{pace_counts['silent']} 个场景无台词，考虑增加对话或删减空场景")
        if overall_pace == "fast_paced":
            suggestions.append("整体节奏偏快，建议穿插慢节奏场景以提供情感沉淀空间")
        if overall_pace == "slow_paced":
            suggestions.append("整体节奏偏慢，建议增加快节奏场景以提升观众注意力")
        if rhythm_variability < 0.1 and active_scenes >= 5:
            suggestions.append("节奏变化度低，建议增加快慢交替以增强观感层次")
        if not suggestions:
            suggestions.append("场景节奏分布合理，快慢搭配得当")

        return {
            "drama_id": did,
            "title": title,
            "total_scenes": total_scenes,
            "scene_rhythms": scene_rhythms[:100],
            "rhythm_curve": rhythm_curve[:200],
            "pace_distribution": pace_counts,
            "overall_pace": overall_pace,
            "rhythm_variability": rhythm_variability,
            "suggestions": suggestions,
        }

    # ===== v5.4.8 新增：Agent 记忆强化 + 跨 Agent 共享 + AI 短剧增强 =====

    def agent_memory_reinforce(self,
                               agent_id: str,
                               min_access_count: int = 3,
                               boost_importance: bool = True,
                               dry_run: bool = False) -> Dict[str, Any]:
        """Agent 记忆强化（v5.4.8 新增）

        基于访问频率自动提升高频记忆的重要性等级。
        频繁被检索的记忆说明价值更高，应自动强化。

        Args:
            agent_id: Agent ID
            min_access_count: 最低访问次数阈值
            boost_importance: 是否自动提升重要性
            dry_run: 仅预览不执行

        Returns:
            {evaluated, reinforced, details}
        """
        conn = self._get_conn()
        aid = _filter_unicode_ctrl(agent_id[:128]) if isinstance(agent_id, str) else ""
        if not aid:
            return {"error": "Agent ID 不能为空"}

        # 查找高频记忆
        # v5.5.5 fix: 添加 privacy 用于审计日志
        rows = conn.execute(
            "SELECT id, content, importance, access_count, privacy FROM memories "
            "WHERE source_agent = ? AND category != 'trash' AND access_count >= ?",
            (aid, min_access_count)
        ).fetchall()

        if not rows:
            return {
                "agent_id": aid,
                "evaluated": 0,
                "reinforced": 0,
                "details": [],
            }

        # 重要性升级路径
        imp_upgrade = {"LOW": "MEDIUM", "MEDIUM": "HIGH"}
        reinforced = 0
        details = []

        for r in rows:
            mem_id, content, cur_imp, access_cnt, privacy = r[0], r[1], (r[2] or "MEDIUM").upper(), r[3], r[4]
            new_imp = imp_upgrade.get(cur_imp)

            if new_imp and boost_importance:
                if not dry_run:
                    conn.execute(
                        "UPDATE memories SET importance = ?, updated_at = ? WHERE id = ?",
                        (new_imp, time.time(), mem_id)
                    )
                    # v5.5.5 fix: 审计日志记录真实隐私级别
                    self._add_audit("reinforce", mem_id, agent_id, "",
                                     privacy if privacy else "",
                                     details={"from": cur_imp, "to": new_imp})
                reinforced += 1
                details.append({
                    "id": mem_id,
                    "content": (content or "")[:80],
                    "access_count": access_cnt,
                    "from": cur_imp,
                    "to": new_imp,
                })

        if not dry_run and reinforced > 0:
            conn.commit()

        return {
            "agent_id": aid,
            "evaluated": len(rows),
            "reinforced": reinforced,
            "dry_run": dry_run,
            "details": details[:50],
        }

    def agent_shared_memories(self,
                              from_agent: str,
                              to_agent: str,
                              categories: Optional[List[str]] = None,
                              max_count: int = 50,
                              dry_run: bool = False) -> Dict[str, Any]:
        """跨 Agent 记忆共享（v5.4.8 新增）

        将一个 Agent 的记忆共享给另一个 Agent（复制，非移动）。
        共享后目标 Agent 可以检索到源 Agent 的知识。

        Args:
            from_agent: 源 Agent ID
            to_agent: 目标 Agent ID
            categories: 只共享指定分类（None 表示全部）
            max_count: 最大共享数量
            dry_run: 仅预览

        Returns:
            {shared_count, details}
        """
        conn = self._get_conn()
        from_aid = _filter_unicode_ctrl(from_agent[:128]) if isinstance(from_agent, str) else ""
        to_aid = _filter_unicode_ctrl(to_agent[:128]) if isinstance(to_agent, str) else ""

        if not from_aid or not to_aid:
            return {"error": "Agent ID 不能为空"}
        if from_aid == to_aid:
            return {"error": "不能共享给自身"}

        # 查询源 Agent 的记忆
        query = "SELECT id, content, category, tags, importance, privacy, layer FROM memories WHERE source_agent = ? AND category != 'trash'"
        params: list = [from_aid]

        if categories:
            placeholders = ",".join(["?"] * len(categories))
            query += f" AND category IN ({placeholders})"
            params.extend(categories[:10])

        query += " LIMIT ?"
        params.append(max(1, min(1000, max_count)))

        rows = conn.execute(query, params).fetchall()

        if not rows:
            return {"from_agent": from_aid, "to_agent": to_aid, "shared_count": 0, "details": []}

        # 批量去重：一次性获取目标 Agent 已有内容（避免 N+1 查询）
        existing_contents = set(
            r[0] for r in conn.execute(
                "SELECT content FROM memories WHERE source_agent = ?", (to_aid,)
            ).fetchall()
        )

        shared = 0
        details = []
        now = time.time()

        for r in rows:
            mem_id, content, category, tags, importance, privacy, layer = r
            # 检查目标 Agent 是否已有相同内容（O(1) 查找）
            if content in existing_contents:
                continue

            if not dry_run:
                new_id = str(uuid.uuid4())
                conn.execute("""
                    INSERT INTO memories (
                        id, content, category, tags, importance, privacy, memory_type, layer,
                        source_agent, created_at, updated_at, last_accessed_at,
                        access_count, consolidation_count, forgetting_score, strength,
                        starred, pinned, metadata, encrypted
                    ) VALUES (?, ?, ?, ?, ?, ?, 'text', ?, ?, ?, ?, ?, 0, 0, 0, 1.0, 0, 0, '{}', 0)
                """, (
                    new_id, content, category or "general", tags or "[]",
                    importance or "MEDIUM", privacy or "INTERNAL", layer or "short_term",
                    to_aid, now, now, now
                ))
                # 同步 FTS（失败时记录日志）
                try:
                    conn.execute(
                        "INSERT INTO memory_fts (rowid, content, category, tags) "
                        "VALUES ((SELECT rowid FROM memories WHERE id = ?), ?, ?, ?)",
                        (new_id, content, category or "general", tags or "[]")
                    )
                except Exception as e:
                    logger.warning("FTS sync failed for shared memory %s: %s", new_id, e)
                # 审计记录
                # v5.5.5 fix: 审计日志记录真实隐私级别
                self._add_audit("share", new_id, to_aid, "",
                                 privacy if privacy else "")

            shared += 1
            existing_contents.add(content)  # 更新去重集合
            details.append({
                "source_id": mem_id,
                "content_preview": (content or "")[:60],
                "category": category,
            })

        if not dry_run and shared > 0:
            conn.commit()

        return {
            "from_agent": from_aid,
            "to_agent": to_aid,
            "shared_count": shared,
            "dry_run": dry_run,
            "details": details[:50],
        }

    def agent_knowledge_domains(self,
                                agent_id: str,
                                top_n: int = 10) -> Dict[str, Any]:
        """Agent 知识领域分析（v5.4.8 新增）

        基于记忆分类和标签分析 Agent 的知识分布。

        Args:
            agent_id: Agent ID
            top_n: 返回前 N 个领域

        Returns:
            {domains: [{name, count, tags}], total_memories}
        """
        conn = self._get_conn()
        aid = _filter_unicode_ctrl(agent_id[:128]) if isinstance(agent_id, str) else ""
        if not aid:
            return {"error": "Agent ID 不能为空"}

        # 按分类统计
        cat_rows = conn.execute(
            "SELECT category, COUNT(*) as cnt FROM memories "
            "WHERE source_agent = ? AND category != 'trash' "
            "GROUP BY category ORDER BY cnt DESC LIMIT ?",
            (aid, top_n)
        ).fetchall()

        domains = []
        for cat, cnt in cat_rows:
            # 获取该分类下最常见的标签
            tag_rows = conn.execute(
                "SELECT tags FROM memories WHERE source_agent = ? AND category = ? LIMIT 100",
                (aid, cat)
            ).fetchall()

            tag_counts: Dict[str, int] = {}
            for tr in tag_rows:
                try:
                    tags = json.loads(tr[0]) if tr[0] and tr[0].strip().startswith('[') else []
                    for t in tags:
                        if isinstance(t, str) and t:
                            tag_counts[t] = tag_counts.get(t, 0) + 1
                except (json.JSONDecodeError, TypeError):
                    pass

            top_tags = sorted(tag_counts.items(), key=lambda x: -x[1])[:5]

            domains.append({
                "name": cat or "general",
                "count": cnt,
                "top_tags": [t[0] for t in top_tags],
            })

        total = conn.execute(
            "SELECT COUNT(*) FROM memories WHERE source_agent = ? AND category != 'trash'",
            (aid,)
        ).fetchone()[0]

        return {
            "agent_id": aid,
            "total_memories": total,
            "domains": domains,
        }

    def drama_generate_scene(self,
                             drama_id: str,
                             scene_title: str,
                             characters: Optional[List[str]] = None,
                             mood: str = "neutral",
                             setting: str = "") -> Dict[str, Any]:
        """AI 短剧场景生成（v5.4.8 新增）

        基于剧本上下文生成新场景的框架结构。
        生成内容包括：场景描述、角色对话提示、情感基调建议。

        Args:
            drama_id: 短剧 ID
            scene_title: 场景标题
            characters: 参与角色列表
            mood: 情感基调 (neutral/happy/sad/tense/romantic)
            setting: 场景设置描述

        Returns:
            {scene_id, title, structure, suggestions}
        """
        conn = self._get_conn()
        did = _filter_unicode_ctrl(drama_id[:128]) if isinstance(drama_id, str) else ""
        if not did:
            return {"error": "短剧 ID 不能为空"}

        # 获取短剧信息
        drama = conn.execute(
            "SELECT id, title, genre FROM drama_series WHERE id = ?", (did,)
        ).fetchone()

        if not drama:
            return {"error": "短剧不存在"}

        drama_title = drama[1] or "未知短剧"
        genre = drama[2] or "general"

        # 获取已有场景数量
        scene_count = conn.execute(
            "SELECT COUNT(*) FROM drama_scenes WHERE drama_id = ?", (did,)
        ).fetchone()[0]

        # 获取角色信息
        char_list = []
        if characters:
            for cname in characters[:10]:
                char = conn.execute(
                    "SELECT id, name, role FROM drama_characters WHERE drama_id = ? AND name = ?",
                    (did, cname[:128])
                ).fetchone()
                if char:
                    char_list.append({"id": char[0], "name": char[1], "role": char[2]})

        # 生成场景结构
        scene_id = str(uuid.uuid4())
        now = time.time()

        # 情感基调映射
        mood_templates = {
            "happy": "轻松愉快的氛围，角色之间互动积极正面",
            "sad": "悲伤沉重的氛围，可能有离别或失落的情节",
            "tense": "紧张悬疑的氛围，冲突即将爆发或正在进行",
            "romantic": "温馨浪漫的氛围，角色之间情感升温",
            "neutral": "平稳叙事的氛围，推进剧情发展",
        }
        mood_desc = mood_templates.get(mood, mood_templates["neutral"])

        # 生成场景建议
        suggestions = []
        if genre == "romance":
            suggestions.append("可以加入角色之间的微妙互动或误会")
            suggestions.append("考虑设置一个促进感情发展的契机")
        elif genre == "suspense":
            suggestions.append("埋下一个伏笔或线索供后续揭示")
            suggestions.append("制造一个出乎意料的转折")
        elif genre == "comedy":
            suggestions.append("加入一个幽默的误会或巧合")
            suggestions.append("角色可以有夸张但可爱的反应")
        else:
            suggestions.append("确保场景推进了主线剧情")
            suggestions.append("考虑角色在这个场景中的成长或变化")

        if scene_count == 0:
            suggestions.insert(0, "这是开场场景，需要建立世界观和主要角色")
        suggestions.append(f"情感基调：{mood_desc}")

        # 存储场景
        scene_data = {
            "title": scene_title[:256],
            "description": f"{setting or scene_title}\n\n[{mood_desc}]",
            "mood": mood,
            "characters": [c["name"] for c in char_list],
            "generated": True,
            "suggestions": suggestions,
        }

        conn.execute("""
            INSERT INTO drama_scenes (
                id, drama_id, episode, scene_number, title, content,
                tags, metadata, created_at
            ) VALUES (?, ?, 0, ?, ?, ?, ?, ?, ?)
        """, (
            scene_id, did, scene_count + 1, scene_title[:256],
            scene_data["description"],
            json.dumps([mood], ensure_ascii=False),
            json.dumps(scene_data, ensure_ascii=False), now
        ))
        conn.commit()

        return {
            "scene_id": scene_id,
            "drama_id": did,
            "drama_title": drama_title,
            "title": scene_title,
            "scene_order": scene_count + 1,
            "mood": mood,
            "mood_description": mood_desc,
            "characters": char_list,
            "suggestions": suggestions,
        }

    def drama_emotion_timeline(self,
                               drama_id: str) -> Dict[str, Any]:
        """短剧情感时间线（v5.4.8 新增）

        分析短剧各场景的情感走向，生成情感曲线。

        Args:
            drama_id: 短剧 ID

        Returns:
            {drama_id, title, emotion_points, trend, summary}
        """
        conn = self._get_conn()
        did = _filter_unicode_ctrl(drama_id[:128]) if isinstance(drama_id, str) else ""
        if not did:
            return {"error": "短剧 ID 不能为空"}

        drama = conn.execute(
            "SELECT id, title FROM drama_series WHERE id = ?", (did,)
        ).fetchone()

        if not drama:
            return {"error": "短剧不存在"}

        title = drama[1] or "未知短剧"

        # 获取所有场景（按集数和场景号排序）
        scenes = conn.execute(
            "SELECT id, title, content, scene_number FROM drama_scenes "
            "WHERE drama_id = ? ORDER BY episode, scene_number",
            (did,)
        ).fetchall()

        if not scenes:
            return {
                "drama_id": did,
                "title": title,
                "emotion_points": [],
                "trend": "no_data",
                "summary": "暂无场景数据",
            }

        # 情感关键词
        emotion_lexicon = {
            "positive": {"好", "开心", "成功", "爱", "喜", "乐", "幸福", "happy", "love", "joy", "success"},
            "negative": {"坏", "悲伤", "失败", "恨", "悲", "痛", "绝望", "sad", "hate", "pain", "fail"},
            "tense": {"紧张", "危险", "冲突", "对抗", "危机", "tense", "danger", "conflict", "crisis"},
            "calm": {"平静", "安宁", "日常", "悠闲", "calm", "peace", "quiet", "daily"},
        }

        emotion_points = []
        emotion_values = []

        for scene in scenes:
            scene_id, scene_title, description, order = scene
            text = ((scene_title or "") + " " + (description or "")).lower()

            # 计算各情感得分
            scores = {}
            for emotion, keywords in emotion_lexicon.items():
                score = sum(1 for kw in keywords if kw in text)
                scores[emotion] = score

            # 主导情感
            dominant = max(scores, key=scores.get) if any(scores.values()) else "neutral"
            intensity = sum(scores.values())

            # 情感值 (-2 到 +2)
            value = 0
            if scores["positive"] > scores["negative"]:
                value = min(2, scores["positive"] - scores["negative"])
            elif scores["negative"] > scores["positive"]:
                value = max(-2, -(scores["negative"] - scores["positive"]))

            if scores["tense"] > 2:
                # 紧张场景增加波动
                value = value * 0.5 if value > 0 else value * 1.5

            emotion_values.append(value)
            emotion_points.append({
                "scene_id": scene_id,
                "scene_order": order,
                "title": (scene_title or "")[:50],
                "dominant_emotion": dominant,
                "intensity": intensity,
                "emotion_value": round(value, 2),
            })

        # 分析趋势
        if len(emotion_values) < 2:
            trend = "insufficient_data"
        else:
            # 简单线性趋势
            n = len(emotion_values)
            x_mean = (n - 1) / 2
            y_mean = sum(emotion_values) / n
            numerator = sum((i - x_mean) * (v - y_mean) for i, v in enumerate(emotion_values))
            denominator = sum((i - x_mean) ** 2 for i in range(n))

            slope = numerator / denominator if denominator > 0 else 0

            if slope > 0.3:
                trend = "ascending"  # 情感走向积极
            elif slope < -0.3:
                trend = "descending"  # 情感走向消极
            elif abs(slope) <= 0.1:
                trend = "stable"  # 情感平稳
            else:
                # 检查波动
                diffs = [abs(emotion_values[i] - emotion_values[i-1]) for i in range(1, n)]
                avg_diff = sum(diffs) / len(diffs)
                if avg_diff > 1.0:
                    trend = "volatile"  # 情感波动大
                else:
                    trend = "moderate"  # 适度变化

        # 生成摘要
        summary_parts = []
        if trend == "ascending":
            summary_parts.append("剧情情感走向逐渐积极，可能是从困境走向圆满")
        elif trend == "descending":
            summary_parts.append("剧情情感走向逐渐消极，可能是悲剧或困境加深")
        elif trend == "volatile":
            summary_parts.append("剧情情感波动较大，充满戏剧性转折")
        elif trend == "stable":
            summary_parts.append("剧情情感较为平稳，叙事节奏均匀")
        else:
            summary_parts.append("剧情情感变化适度，节奏把控良好")

        pos_count = sum(1 for p in emotion_points if p["dominant_emotion"] == "positive")
        neg_count = sum(1 for p in emotion_points if p["dominant_emotion"] == "negative")
        tense_count = sum(1 for p in emotion_points if p["dominant_emotion"] == "tense")

        if pos_count > len(emotion_points) * 0.4:
            summary_parts.append("整体基调偏积极")
        elif neg_count > len(emotion_points) * 0.4:
            summary_parts.append("整体基调偏沉重")
        if tense_count > 0:
            summary_parts.append(f"有 {tense_count} 个紧张场景")

        return {
            "drama_id": did,
            "title": title,
            "total_scenes": len(emotion_points),
            "emotion_points": emotion_points,
            "trend": trend,
            "summary": "。".join(summary_parts),
        }

    # ===== v5.5.0 新增：Agent 记忆终极增强实现 =====

    def _ensure_snapshots_table(self):
        """确保记忆快照表存在（v5.5.0 新增）"""
        conn = self._get_conn()
        conn.execute("""
            CREATE TABLE IF NOT EXISTS memory_snapshots (
                id TEXT PRIMARY KEY,
                agent_id TEXT NOT NULL,
                label TEXT DEFAULT '',
                memory_count INTEGER DEFAULT 0,
                snapshot_data TEXT DEFAULT '{}',
                categories TEXT DEFAULT '{}',
                created_at REAL NOT NULL
            )
        """)
        conn.commit()

    def agent_memory_snapshot(self, agent_id: str, label: str = "") -> Dict[str, Any]:
        """Agent 记忆快照（v5.5.0 新增）"""
        conn = self._get_conn()
        aid = _filter_unicode_ctrl(agent_id[:128]) if isinstance(agent_id, str) else ""
        if not aid:
            return {"error": "Agent ID 不能为空"}

        self._ensure_snapshots_table()

        # 获取该 Agent 的所有记忆
        rows = conn.execute(
            "SELECT id, content, category, tags, importance, privacy, layer, "
            "access_count, created_at, updated_at FROM memories "
            "WHERE source_agent = ? AND category != 'trash'",
            (aid,)
        ).fetchall()

        # 按分类统计
        categories: Dict[str, int] = {}
        snapshot_entries = []
        for r in rows:
            cat = r[2] or "general"
            categories[cat] = categories.get(cat, 0) + 1
            snapshot_entries.append({
                "id": r[0],
                "content_preview": (r[1] or "")[:100],
                "category": cat,
                "importance": r[4] or "MEDIUM",
                "layer": r[6] or "short_term",
                "access_count": r[7] or 0,
            })

        snapshot_id = str(uuid.uuid4())
        now = time.time()

        conn.execute(
            "INSERT INTO memory_snapshots (id, agent_id, label, memory_count, snapshot_data, categories, created_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?)",
            (snapshot_id, aid, label or "", len(rows),
             json.dumps(snapshot_entries[:500], ensure_ascii=False),
             json.dumps(categories, ensure_ascii=False), now)
        )
        conn.commit()

        return {
            "snapshot_id": snapshot_id,
            "agent_id": aid,
            "label": label or "",
            "memory_count": len(rows),
            "created_at": now,
            "categories": categories,
            "sample_entries": snapshot_entries[:10],
        }

    def agent_memory_deduplicate(self, agent_id: str,
                                   similarity_threshold: float = 0.85,
                                   dry_run: bool = False) -> Dict[str, Any]:
        """Agent 记忆去重（v5.5.0 新增）"""
        conn = self._get_conn()
        aid = _filter_unicode_ctrl(agent_id[:128]) if isinstance(agent_id, str) else ""
        if not aid:
            return {"error": "Agent ID 不能为空"}

        rows = conn.execute(
            "SELECT id, content, category, tags, importance, access_count FROM memories "
            "WHERE source_agent = ? AND category != 'trash' ORDER BY access_count DESC",
            (aid,)
        ).fetchall()

        if len(rows) < 2:
            return {
                "agent_id": aid,
                "evaluated": len(rows),
                "duplicates_found": 0,
                "merged": 0,
                "groups": [],
                "dry_run": dry_run,
            }

        # 基于内容相似度检测重复（使用简单的字符级 Jaccard + 标签匹配）
        def content_similarity(a: str, b: str) -> float:
            if not a or not b:
                return 0.0
            # 简单的字符集合 Jaccard 相似度
            set_a = set(a.lower())
            set_b = set(b.lower())
            if not set_a or not set_b:
                return 0.0
            intersection = len(set_a & set_b)
            union = len(set_a | set_b)
            return intersection / union if union > 0 else 0.0

        merged_ids = set()
        groups = []
        duplicates_found = 0
        merged = 0
        pending_audits = []

        try:
            for i, r1 in enumerate(rows):
                if r1[0] in merged_ids:
                    continue
                group = {"keep": r1[0], "keep_preview": (r1[1] or "")[:60], "duplicates": []}
                for j, r2 in enumerate(rows):
                    if i >= j or r2[0] in merged_ids:
                        continue
                    # 同分类才考虑去重
                    if (r1[2] or "") != (r2[2] or ""):
                        continue
                    sim = content_similarity(r1[1] or "", r2[1] or "")
                    if sim >= similarity_threshold:
                        group["duplicates"].append({
                            "id": r2[0],
                            "preview": (r2[1] or "")[:60],
                            "similarity": round(sim, 3),
                        })
                        duplicates_found += 1
                        if not dry_run:
                            existing_meta = conn.execute(
                                "SELECT metadata FROM memories WHERE id = ?", (r2[0],)
                            ).fetchone()
                            try:
                                meta = json.loads(existing_meta["metadata"]) if existing_meta and existing_meta["metadata"] else {}
                            except (json.JSONDecodeError, TypeError, KeyError):
                                meta = {}
                            meta["merged_into"] = r1[0]
                            meta["dedup_similarity"] = sim
                            meta["_original_category"] = r2[2]
                            conn.execute(
                                "UPDATE memories SET category = 'trash', updated_at = ?, metadata = ? WHERE id = ?",
                                (time.time(), json.dumps(meta, ensure_ascii=False), r2[0])
                            )
                            pending_audits.append(("deduplicate_merge", r2[0], aid, "", "",
                                {"merged_into": r1[0], "dedup_similarity": round(sim, 3)}))
                            merged += 1
                        merged_ids.add(r2[0])
                if group["duplicates"]:
                    groups.append(group)

            if not dry_run and merged > 0:
                conn.commit()
                for audit_args in pending_audits:
                    self._add_audit(*audit_args)
        except sqlite3.Error as e:
            conn.rollback()
            logger.error("agent_memory_deduplicate failed: %s", e)
            return {"error": f"数据库操作失败: {e}", "agent_id": aid}

        return {
            "agent_id": aid,
            "evaluated": len(rows),
            "duplicates_found": duplicates_found,
            "merged": merged,
            "groups": groups[:20],
            "dry_run": dry_run,
        }

    def agent_memory_health_check(self, agent_id: str) -> Dict[str, Any]:
        """Agent 记忆健康检查（v5.5.0 新增）"""
        conn = self._get_conn()
        aid = _filter_unicode_ctrl(agent_id[:128]) if isinstance(agent_id, str) else ""
        if not aid:
            return {"error": "Agent ID 不能为空"}

        rows = conn.execute(
            "SELECT id, importance, layer, access_count, created_at, updated_at, "
            "last_accessed_at, category, privacy, encrypted FROM memories "
            "WHERE source_agent = ? AND category != 'trash'",
            (aid,)
        ).fetchall()

        total = len(rows)
        if total == 0:
            return {
                "agent_id": aid,
                "overall_score": 0,
                "total_memories": 0,
                "dimensions": {},
                "issues": ["该 Agent 暂无记忆数据"],
                "recommendations": ["建议开始添加记忆以建立知识库"],
            }

        now = time.time()
        issues = []
        recommendations = []

        # 1. 层级分布
        layer_dist: Dict[str, int] = {}
        for r in rows:
            layer = r[2] or "short_term"
            layer_dist[layer] = layer_dist.get(layer, 0) + 1

        short_pct = layer_dist.get("short_term", 0) / total * 100
        if short_pct > 80:
            issues.append(f"短期记忆占比过高（{short_pct:.1f}%），建议执行记忆巩固")
            recommendations.append("运行 consolidate() 将高频短期记忆提升为长期记忆")

        # 2. 重要性分布
        imp_dist: Dict[str, int] = {}
        for r in rows:
            imp = r[1] or "MEDIUM"
            imp_dist[imp] = imp_dist.get(imp, 0) + 1

        # 3. 访问活跃度
        never_accessed = sum(1 for r in rows if (r[3] or 0) == 0)
        never_pct = never_accessed / total * 100
        if never_pct > 50:
            issues.append(f"{never_pct:.1f}% 的记忆从未被访问，可能存在冗余")
            recommendations.append("运行 agent_memory_deduplicate() 清理重复和低价值记忆")

        # 4. 遗忘风险（超过90天未访问的非永久记忆）
        at_risk = 0
        for r in rows:
            last_acc = r[6] or r[5] or 0
            layer = r[2] or "short_term"
            if layer != "permanent" and (now - last_acc) > 90 * 86400:
                at_risk += 1
        at_risk_pct = at_risk / total * 100
        if at_risk_pct > 30:
            issues.append(f"{at_risk_pct:.1f}% 的记忆超过90天未访问，存在遗忘风险")
            recommendations.append("考虑对重要记忆执行记忆强化或提升为永久记忆")

        # 5. 分类均衡度
        cat_dist: Dict[str, int] = {}
        for r in rows:
            cat = r[7] or "general"
            cat_dist[cat] = cat_dist.get(cat, 0) + 1
        dominant_cat = max(cat_dist, key=cat_dist.get)
        dominant_pct = cat_dist[dominant_cat] / total * 100
        if dominant_pct > 70 and len(cat_dist) > 1:
            issues.append(f"分类不均衡：'{dominant_cat}' 占比 {dominant_pct:.1f}%")

        # 6. 加密状态
        unencrypted = sum(1 for r in rows if not r[9])
        if unencrypted > 0:
            issues.append(f"{unencrypted} 条记忆未加密")
            recommendations.append("建议启用加密以保护敏感记忆")

        # 计算综合健康分（0-100）
        score = 100
        score -= min(30, short_pct * 0.2)  # 短期占比扣分
        score -= min(20, never_pct * 0.2)  # 未访问扣分
        score -= min(20, at_risk_pct * 0.15)  # 遗忘风险扣分
        score -= min(10, len(issues) * 3)  # 问题数量扣分
        score = max(0, min(100, round(score)))

        dimensions = {
            "total_memories": total,
            "layer_distribution": layer_dist,
            "importance_distribution": imp_dist,
            "category_distribution": cat_dist,
            "never_accessed": never_accessed,
            "never_accessed_pct": round(never_pct, 1),
            "at_risk_forget": at_risk,
            "at_risk_pct": round(at_risk_pct, 1),
            "unencrypted": unencrypted,
        }

        if not issues:
            issues.append("记忆系统状态良好，未发现明显问题")

        return {
            "agent_id": aid,
            "overall_score": score,
            "total_memories": total,
            "dimensions": dimensions,
            "issues": issues,
            "recommendations": recommendations[:5],
        }

    def agent_memory_importance_recalibrate(self, agent_id: str,
                                              dry_run: bool = False) -> Dict[str, Any]:
        """Agent 记忆重要度重校准（v5.5.0 新增）"""
        conn = self._get_conn()
        aid = _filter_unicode_ctrl(agent_id[:128]) if isinstance(agent_id, str) else ""
        if not aid:
            return {"error": "Agent ID 不能为空"}

        rows = conn.execute(
            "SELECT id, content, importance, access_count, created_at, updated_at, "
            "last_accessed_at, layer FROM memories "
            "WHERE source_agent = ? AND category != 'trash'",
            (aid,)
        ).fetchall()

        if not rows:
            return {
                "agent_id": aid,
                "evaluated": 0,
                "upgraded": 0,
                "downgraded": 0,
                "unchanged": 0,
                "details": [],
                "dry_run": dry_run,
            }

        now = time.time()
        imp_levels = ["LOW", "MEDIUM", "HIGH", "CRITICAL"]
        upgraded = 0
        downgraded = 0
        unchanged = 0
        details = []
        pending_audits = []

        try:
            for r in rows:
                mem_id, content, cur_imp, access_cnt, created, updated, last_acc, layer = r
                cur_imp = (cur_imp or "MEDIUM").upper()
                access_cnt = access_cnt or 0
                last_acc = last_acc or updated or created or now

                # 计算综合使用分（0-100）
                recency_days = (now - last_acc) / 86400
                recency_score = max(0, 100 - recency_days * 2)  # 每天扣2分
                frequency_score = min(100, access_cnt * 10)  # 每次访问10分，上限100
                age_days = (now - (created or now)) / 86400
                age_score = max(0, 100 - age_days * 0.5)  # 越新分越高

                usage_score = (recency_score * 0.4 + frequency_score * 0.4 + age_score * 0.2)

                # 确定目标重要性
                if usage_score >= 80:
                    target_imp = "CRITICAL"
                elif usage_score >= 55:
                    target_imp = "HIGH"
                elif usage_score >= 25:
                    target_imp = "MEDIUM"
                else:
                    target_imp = "LOW"

                cur_idx = imp_levels.index(cur_imp) if cur_imp in imp_levels else 1
                tgt_idx = imp_levels.index(target_imp)

                if tgt_idx > cur_idx:
                    upgraded += 1
                    action = "upgrade"
                elif tgt_idx < cur_idx:
                    downgraded += 1
                    action = "downgrade"
                else:
                    unchanged += 1
                    action = "unchanged"
                    continue  # 不变的不记录详情

                if not dry_run and action != "unchanged":
                    conn.execute(
                        "UPDATE memories SET importance = ?, updated_at = ? WHERE id = ?",
                        (target_imp, now, mem_id)
                    )
                    pending_audits.append(("recalibrate", mem_id, aid, "", "",
                        {"from": cur_imp, "to": target_imp, "usage_score": round(usage_score, 1)}))

                details.append({
                    "id": mem_id,
                    "content_preview": (content or "")[:60],
                    "from": cur_imp,
                    "to": target_imp,
                    "usage_score": round(usage_score, 1),
                    "access_count": access_cnt,
                    "action": action,
                })

            if not dry_run and (upgraded + downgraded) > 0:
                conn.commit()
                for audit_args in pending_audits:
                    self._add_audit(*audit_args)
        except sqlite3.Error as e:
            conn.rollback()
            logger.error("agent_memory_importance_recalibrate failed: %s", e)
            return {"error": f"数据库操作失败: {e}", "agent_id": aid}

        return {
            "agent_id": aid,
            "evaluated": len(rows),
            "upgraded": upgraded,
            "downgraded": downgraded,
            "unchanged": unchanged,
            "details": details[:50],
            "dry_run": dry_run,
        }

    # ===== v5.5.0 新增：AI 短剧终极增强实现 =====

    def drama_generate_episode(self, drama_id: str,
                                episode_number: int,
                                theme: str = "",
                                num_scenes: int = 3,
                                mood: str = "neutral") -> Dict[str, Any]:
        """AI 短剧分集生成（v5.5.0 新增）"""
        conn = self._get_conn()
        did = _filter_unicode_ctrl(drama_id[:64]) if isinstance(drama_id, str) else ""
        if not did:
            return {"error": "短剧 ID 不能为空"}

        drama = conn.execute(
            "SELECT id, title, genre, description FROM drama_series WHERE id = ?", (did,)
        ).fetchone()
        if not drama:
            return {"error": "短剧不存在"}

        drama_title = drama[1] or "未知短剧"
        genre = drama[2] or "drama"

        # 获取已有场景数
        existing_scenes = conn.execute(
            "SELECT COUNT(*) FROM drama_scenes WHERE drama_id = ? AND episode = ?",
            (did, episode_number)
        ).fetchone()[0]

        # 获取角色列表
        chars = conn.execute(
            "SELECT id, name, role, personality FROM drama_characters WHERE drama_id = ? LIMIT 10",
            (did,)
        ).fetchall()
        char_names = [c[1] for c in chars]

        # 场景模板（基于类型和集数）
        scene_templates = {
            "opening": ["建立场景", "引入冲突", "角色出场"],
            "development": ["冲突升级", "角色互动", "线索揭示"],
            "climax": ["高潮对峙", "真相揭露", "情感爆发"],
            "resolution": ["冲突解决", "情感收束", "伏笔回收"],
        }

        # 根据集数确定阶段
        if episode_number <= 1:
            phase = "opening"
        elif episode_number <= 3:
            phase = "development"
        elif episode_number <= 5:
            phase = "climax"
        else:
            phase = "resolution"

        base_templates = scene_templates.get(phase, scene_templates["development"])

        # 生成场景
        scenes = []
        now = time.time()
        for i in range(num_scenes):
            scene_idx = existing_scenes + i + 1
            title_base = base_templates[i % len(base_templates)]
            scene_title = f"第{episode_number}集 - {title_base}" if not theme else f"{theme} - {title_base}"

            # 场景内容建议
            content_parts = [f"【{title_base}】"]
            if i == 0:
                content_parts.append(f"本集主题：{theme or '推进剧情发展'}")
                content_parts.append(f"情感基调：{mood}")
            if char_names and i < len(char_names):
                content_parts.append(f"重点角色：{char_names[i % len(char_names)]}")
            content_parts.append("场景目的：推进剧情，展现角色性格变化")

            scene_id = str(uuid.uuid4())
            scene_data = {
                "scene_id": scene_id,
                "episode": episode_number,
                "scene_number": scene_idx,
                "title": scene_title,
                "content": "\n".join(content_parts),
                "mood": mood,
                "phase": phase,
                "generated": True,
            }
            scenes.append(scene_data)

            # 存储场景
            conn.execute("""
                INSERT INTO drama_scenes (
                    id, drama_id, episode, scene_number, title, content,
                    tags, metadata, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                scene_id, did, episode_number, scene_idx, scene_title,
                scene_data["content"], json.dumps([mood, phase], ensure_ascii=False),
                json.dumps(scene_data, ensure_ascii=False), now
            ))

        conn.commit()

        # 生成建议
        suggestions = [
            f"本集属于{phase}阶段，重点是{base_templates[0]}",
            f"建议围绕'{theme or '主线剧情'}'展开，保持{mood}的情感基调",
        ]
        if char_names:
            suggestions.append(f"可重点刻画角色：{', '.join(char_names[:3])}")
        if genre == "romance":
            suggestions.append("加入情感互动和微妙关系变化")
        elif genre == "suspense":
            suggestions.append("埋设线索，保持悬念感")

        return {
            "episode_id": f"ep_{episode_number}",
            "drama_id": did,
            "drama_title": drama_title,
            "episode_number": episode_number,
            "theme": theme or "",
            "phase": phase,
            "mood": mood,
            "scenes": scenes,
            "suggestions": suggestions,
        }

    def drama_character_dialogue(self, drama_id: str,
                                   character_name: str,
                                   context: str = "",
                                   emotion: str = "neutral",
                                   num_lines: int = 3) -> Dict[str, Any]:
        """AI 短剧角色台词生成（v5.5.0 新增）"""
        conn = self._get_conn()
        did = _filter_unicode_ctrl(drama_id[:64]) if isinstance(drama_id, str) else ""
        if not did:
            return {"error": "短剧 ID 不能为空"}

        # 查找角色
        char = conn.execute(
            "SELECT id, name, role, personality, description FROM drama_characters "
            "WHERE drama_id = ? AND name = ?",
            (did, character_name)
        ).fetchone()
        if not char:
            return {"error": f"角色 '{character_name}' 不存在"}

        char_id, name, role, personality, description = char

        # 获取角色历史台词风格
        past_lines = conn.execute(
            "SELECT line_text, context FROM drama_lines "
            "WHERE drama_id = ? AND character_name = ? ORDER BY created_at DESC LIMIT 20",
            (did, name)
        ).fetchall()

        # 情感模板
        emotion_styles = {
            "happy": ["太好了！", "真让人开心", "我就知道会这样"],
            "sad": ["怎么会这样...", "我很难过", "这太令人失望了"],
            "angry": ["岂有此理！", "我不能接受", "这太过分了"],
            "surprised": ["什么？！", "真的假的？", "这怎么可能！"],
            "tense": ["我们得小心", "情况不太对", "做好准备"],
            "neutral": ["原来如此", "我明白了", "这样啊"],
        }

        style_lines = emotion_styles.get(emotion, emotion_styles["neutral"])

        # 生成台词（基于角色性格和情感）
        generated_lines = []
        personality_traits = (personality or "").split("、") if personality else []

        for i in range(num_lines):
            base = style_lines[i % len(style_lines)]
            line = base
            # 根据角色角色调整
            if role == "protagonist":
                line = f"{line} 我们必须继续前进。"
            elif role == "antagonist":
                line = f"{line} 这还没完。"
            elif role == "mentor":
                line = f"{line} 记住这个教训。"
            # 根据性格调整
            if personality_traits:
                trait = personality_traits[i % len(personality_traits)]
                if "冷静" in trait or "理性" in trait:
                    line = f"（冷静地）{line}"
                elif "热情" in trait or "活泼" in trait:
                    line = f"（热情地）{line}"

            generated_lines.append({
                "line_id": f"gen_{i}_{int(time.time())}",
                "character": name,
                "emotion": emotion,
                "text": line,
                "context_note": context[:100] if context else "",
            })

        # 上下文分析
        context_analysis = {
            "character": name,
            "role": role or "unknown",
            "personality": personality or "",
            "past_lines_count": len(past_lines),
            "requested_emotion": emotion,
            "context": context[:200] if context else "",
        }

        return {
            "character": name,
            "character_id": char_id,
            "drama_id": did,
            "emotion": emotion,
            "lines": generated_lines,
            "context_analysis": context_analysis,
        }

    def drama_plot_twist_suggest(self, drama_id: str,
                                   num_suggestions: int = 3) -> Dict[str, Any]:
        """AI 短剧剧情反转建议（v5.5.0 新增）"""
        conn = self._get_conn()
        did = _filter_unicode_ctrl(drama_id[:64]) if isinstance(drama_id, str) else ""
        if not did:
            return {"error": "短剧 ID 不能为空"}

        drama = conn.execute(
            "SELECT id, title, genre, description FROM drama_series WHERE id = ?", (did,)
        ).fetchone()
        if not drama:
            return {"error": "短剧不存在"}

        drama_title = drama[1] or "未知短剧"
        genre = drama[2] or "drama"

        # 获取已有场景和角色
        scenes = conn.execute(
            "SELECT title, content FROM drama_scenes WHERE drama_id = ? ORDER BY episode, scene_number LIMIT 50",
            (did,)
        ).fetchall()
        chars = conn.execute(
            "SELECT name, role, description FROM drama_characters WHERE drama_id = ? LIMIT 10",
            (did,)
        ).fetchall()

        # 当前剧情分析
        all_content = " ".join((s[0] or "") + " " + (s[1] or "") for s in scenes)
        current_analysis = {
            "total_scenes": len(scenes),
            "total_characters": len(chars),
            "genre": genre,
            "key_elements": [],
        }

        # 检测关键元素
        key_words = {"秘密": "秘密", "复仇": "复仇", "爱情": "感情线", "背叛": "背叛",
                     "真相": "真相", "身份": "身份谜团", "死亡": "生死", "失踪": "失踪"}
        for kw, label in key_words.items():
            if kw in all_content:
                current_analysis["key_elements"].append(label)

        # 反转类型库
        twist_templates = [
            {
                "type": "身份反转",
                "description": "某个角色的真实身份与表面完全不同",
                "setup": "在前期场景中暗示角色行为的异常之处",
                "impact": "颠覆观众对角色关系的认知，重新解读之前的剧情",
                "suitable_genres": ["suspense", "scifi", "fantasy", "drama"],
            },
            {
                "type": "背叛反转",
                "description": "最信任的角色原来是幕后黑手",
                "setup": "通过细节暗示该角色与反派有隐秘联系",
                "impact": "制造强烈的情感冲击，推动主角成长",
                "suitable_genres": ["suspense", "action", "drama", "romance"],
            },
            {
                "type": "时间线反转",
                "description": "故事的时间顺序被重新排列，过去与现在交织",
                "setup": "在叙事中留下时间线索的矛盾点",
                "impact": "让观众重新理解整个故事的因果关系",
                "suitable_genres": ["suspense", "scifi", "drama"],
            },
            {
                "type": "死亡反转",
                "description": "看似死亡的角色其实还活着，或以另一种形式存在",
                "setup": "不直接展示死亡过程，留下模糊的结局",
                "impact": "在关键时刻回归，制造高潮",
                "suitable_genres": ["suspense", "fantasy", "scifi", "action"],
            },
            {
                "type": "动机反转",
                "description": "反派的行为动机被揭示为正义或无奈",
                "setup": "展示反派行为的背景和苦衷",
                "impact": "模糊善恶边界，增加故事深度",
                "suitable_genres": ["drama", "suspense", "action"],
            },
            {
                "type": "关系反转",
                "description": "两个角色的真实关系与表面呈现的完全不同",
                "setup": "通过对话和互动留下关系的疑点",
                "impact": "改变角色互动的意义，增加戏剧张力",
                "suitable_genres": ["romance", "drama", "suspense"],
            },
        ]

        # 筛选适合当前类型的反转，不足时用通用反转补充
        suitable_twists = [t for t in twist_templates if genre in t["suitable_genres"]]
        # 补充通用反转（drama 类型适用的），确保数量足够
        if len(suitable_twists) < num_suggestions:
            general_twists = [t for t in twist_templates if t not in suitable_twists]
            suitable_twists.extend(general_twists)
        if not suitable_twists:
            suitable_twists = twist_templates

        # 生成建议
        twists = []
        for i in range(min(num_suggestions, len(suitable_twists))):
            twist = suitable_twists[i]
            # 结合当前角色个性化
            char_suggestion = ""
            if chars:
                char = chars[i % len(chars)]
                char_suggestion = f"可围绕角色'{char[0]}'展开此反转"

            twists.append({
                "twist_id": f"twist_{i}",
                "type": twist["type"],
                "description": twist["description"],
                "setup_method": twist["setup"],
                "story_impact": twist["impact"],
                "character_suggestion": char_suggestion,
                "dramatic_level": min(10, 6 + i),
            })

        recommendations = [
            "建议在剧情中段（约60%进度）设置主要反转",
            "反转前至少铺垫3个以上的伏笔线索",
            "反转后需要给观众足够的信息来重新理解剧情",
        ]

        return {
            "drama_id": did,
            "drama_title": drama_title,
            "current_analysis": current_analysis,
            "twists": twists,
            "recommendations": recommendations,
        }

    def drama_script_export(self, drama_id: str,
                             format: str = "standard") -> Dict[str, Any]:
        """AI 短剧剧本导出（v5.5.0 新增）"""
        conn = self._get_conn()
        did = _filter_unicode_ctrl(drama_id[:64]) if isinstance(drama_id, str) else ""
        if not did:
            return {"error": "短剧 ID 不能为空"}

        drama = conn.execute(
            "SELECT id, title, genre, description, total_episodes FROM drama_series WHERE id = ?",
            (did,)
        ).fetchone()
        if not drama:
            return {"error": "短剧不存在"}

        drama_title = drama[1] or "未知短剧"
        genre = drama[2] or "drama"
        description = drama[3] or ""

        # 获取所有场景（按集数和场景号排序）
        scenes = conn.execute(
            "SELECT id, episode, scene_number, title, content, location, time_of_day "
            "FROM drama_scenes WHERE drama_id = ? ORDER BY episode, scene_number",
            (did,)
        ).fetchall()

        # 获取所有台词
        lines = conn.execute(
            "SELECT scene_id, character_name, line_text, episode, timestamp, is_classic "
            "FROM drama_lines WHERE drama_id = ? ORDER BY episode, id",
            (did,)
        ).fetchall()

        # 按场景分组台词
        lines_by_scene: Dict[str, list] = {}
        for ln in lines:
            sid = ln[0] or ""
            if sid not in lines_by_scene:
                lines_by_scene[sid] = []
            lines_by_scene[sid].append(ln)

        # 生成剧本
        script_parts = []
        script_parts.append("=" * 60)
        script_parts.append(f"《{drama_title}》")
        script_parts.append(f"类型：{genre}")
        if description:
            script_parts.append(f"简介：{description}")
        script_parts.append("=" * 60)
        script_parts.append("")

        current_episode = -1
        total_scenes = len(scenes)
        total_lines = len(lines)

        for sc in scenes:
            sc_id, episode, scene_num, title, content, location, time_of_day = sc

            # 集数标题
            if episode != current_episode:
                current_episode = episode
                script_parts.append("")
                script_parts.append(f"{'─' * 40}")
                script_parts.append(f"第 {episode} 集")
                script_parts.append(f"{'─' * 40}")
                script_parts.append("")

            # 场景标题
            script_parts.append(f"【场景 {scene_num}】{title or ''}")
            if location:
                script_parts.append(f"地点：{location}")
            if time_of_day:
                script_parts.append(f"时间：{time_of_day}")
            script_parts.append("")

            # 场景描述
            if content:
                if format == "detailed":
                    script_parts.append(f"〔场景描述〕{content}")
                elif format == "condensed":
                    script_parts.append(f"〔{content[:100]}〕")
                else:
                    script_parts.append(f"〔{content}〕")
                script_parts.append("")

            # 台词
            scene_lines = lines_by_scene.get(sc_id, [])
            for ln in scene_lines:
                _, char_name, line_text, _, timestamp, is_classic = ln
                classic_mark = " ★" if is_classic else ""
                if format == "detailed" and timestamp:
                    script_parts.append(f"  {char_name or '旁白'}（{timestamp}）{classic_mark}：{line_text}")
                else:
                    script_parts.append(f"  {char_name or '旁白'}{classic_mark}：{line_text}")

            if scene_lines:
                script_parts.append("")

        script_content = "\n".join(script_parts)

        return {
            "drama_id": did,
            "title": drama_title,
            "format": format,
            "script_content": script_content,
            "total_scenes": total_scenes,
            "total_lines": total_lines,
            "total_episodes": len(set(sc[1] for sc in scenes)),
        }

    # ===== v5.5.5 硬件自适应与缓存优化 =====

    def get_hardware_profile(self) -> Dict[str, Any]:
        """获取硬件性能画像（v5.5.5 新增）

        返回当前设备的 CPU、内存、磁盘性能检测结果及推荐配置。
        """
        return self._hardware_profile

    def get_cache_stats(self) -> Dict[str, Any]:
        """获取缓存统计（v5.5.5 新增）

        返回记忆缓存的命中率、当前大小、最大容量等统计信息。
        """
        return self._memory_cache.stats()

    def adapt_retrieval_params(self, base_limit: int = 20) -> Dict[str, Any]:
        """根据硬件性能自适应调整检索参数（v5.5.5 新增）

        低配设备自动关闭向量检索，仅使用 FTS + 模糊匹配，
        同时降低返回数量和检索深度，减少资源占用。

        Args:
            base_limit: 基础检索数量上限

        Returns:
            调整后的参数字典：{limit, depth, use_vector, use_fts}
        """
        profile = self._hardware_profile
        level = profile["performance_level"]

        if level == "low":
            # 低配：关闭向量检索，降低数量和深度
            return {
                "limit": min(base_limit, profile["recommended_limit"]),
                "depth": profile["recommended_depth"],
                "use_vector": False,
                "use_fts": True,
            }
        elif level == "medium":
            # 中配：启用 FTS + 向量检索，适度限制
            return {
                "limit": min(base_limit, profile["recommended_limit"]),
                "depth": profile["recommended_depth"],
                "use_vector": True,
                "use_fts": True,
            }
        else:
            # 高配：完整启用，较高上限
            return {
                "limit": max(base_limit, profile["recommended_limit"]),
                "depth": profile["recommended_depth"],
                "use_vector": True,
                "use_fts": True,
            }

    # ===== 前置过滤与冲突检测（v5.5.5 新增）=====

    def prefilter_memories(self, memory_ids: List[str],
                           query: str = "") -> Dict[str, Any]:
        """前置过滤：在将记忆送入大模型前，筛掉重复、低关联、过期的记忆。

        v5.5.5 新增。

        过滤规则（按顺序）：
        1. 去重：内容相似度 >= 0.85 的记忆只保留最新的一条（Jaccard 相似度）。
        2. 低关联过滤：如果提供了 query，计算记忆内容与 query 的关键词重叠率，得分 < 0.2 的移除。
        3. 过期过滤：对于有 expires_at 字段且已过期的记忆直接移除。
        4. 低质量过滤：内容长度 < 10 字符或纯空白的移除。

        Args:
            memory_ids: 候选记忆 ID 列表
            query: 可选查询文本，用于相关性计算

        Returns:
            字典结构：
            {
                "kept": [...],          # 保留的记忆 ID
                "removed": [...],       # 被过滤掉的记忆 ID
                "reasons": {id: reason},  # 每条被移除的原因
                "stats": {
                    "total": int,
                    "kept_count": int,
                    "removed_count": int,
                    "dedup": int,
                    "low_relevance": int,
                    "expired": int,
                    "low_quality": int,
                }
            }
        """
        import re as _re

        if not memory_ids:
            return {
                "kept": [],
                "removed": [],
                "reasons": {},
                "stats": {
                    "total": 0, "kept_count": 0, "removed_count": 0,
                    "dedup": 0, "low_relevance": 0, "expired": 0, "low_quality": 0,
                },
            }

        # v5.4.0 安全加固：ID 过滤
        cleaned_ids = []
        for mid in memory_ids:
            if not mid or not isinstance(mid, str):
                continue
            cid = self._strip_control(mid)[:64]
            if cid and cid not in cleaned_ids:
                cleaned_ids.append(cid)

        if not cleaned_ids:
            return {
                "kept": [],
                "removed": [],
                "reasons": {},
                "stats": {
                    "total": 0, "kept_count": 0, "removed_count": 0,
                    "dedup": 0, "low_relevance": 0, "expired": 0, "low_quality": 0,
                },
            }

        conn = self._get_conn()
        total = len(cleaned_ids)
        removed: List[str] = []
        reasons: Dict[str, str] = {}

        # 批量获取记忆
        placeholders = ",".join("?" * len(cleaned_ids))
        rows = conn.execute(
            f"SELECT id, content, created_at, expires_at, tags, importance "
            f"FROM memories WHERE id IN ({placeholders}) AND category != 'trash'",
            cleaned_ids,
        ).fetchall()

        # 构建记忆字典
        mem_map: Dict[str, Dict[str, Any]] = {}
        for row in rows:
            content = row["content"] or ""
            mem_map[row["id"]] = {
                "id": row["id"],
                "content": content,
                "created_at": row["created_at"] or 0.0,
                "expires_at": float(row["expires_at"]) if row["expires_at"] else 0.0,
                "tags": self._safe_json_loads(row["tags"], []),
                "importance": row["importance"] or "MEDIUM",
            }

        # 当前存活的记忆 ID 列表（按顺序处理）
        current_ids = [mid for mid in cleaned_ids if mid in mem_map]

        # ---- 辅助函数：Jaccard 相似度 ----
        def _jaccard_sim(text_a: str, text_b: str) -> float:
            """基于字符级 bigram 的 Jaccard 相似度"""
            if not text_a or not text_b:
                return 0.0
            # 生成字符 bigram 集合
            def _bigrams(t: str) -> set:
                t = t.strip().lower()
                if len(t) < 2:
                    return {t} if t else set()
                return {t[i:i + 2] for i in range(len(t) - 1)}
            set_a = _bigrams(text_a)
            set_b = _bigrams(text_b)
            if not set_a or not set_b:
                return 0.0
            inter = len(set_a & set_b)
            union = len(set_a | set_b)
            return inter / union if union > 0 else 0.0

        # ---- 辅助函数：关键词重叠率 ----
        def _keyword_overlap(text: str, q: str) -> float:
            """基于关键词（中文按字、英文按词）的重叠率"""
            if not text or not q:
                return 0.0
            # 简单分词：中文单字 + 英文单词
            def _tokens(t: str) -> set:
                t = t.lower().strip()
                tokens = set()
                # 提取英文单词
                for m in _re.findall(r'[a-zA-Z]{2,}', t):
                    tokens.add(m)
                # 提取中文字符
                for ch in t:
                    if '\u4e00' <= ch <= '\u9fff':
                        tokens.add(ch)
                # 提取数字
                for m in _re.findall(r'\d+', t):
                    tokens.add(m)
                return tokens
            t_tokens = _tokens(text)
            q_tokens = _tokens(q)
            if not q_tokens:
                return 0.0
            overlap = len(t_tokens & q_tokens)
            return overlap / len(q_tokens) if len(q_tokens) > 0 else 0.0

        # ---- 第 1 步：去重 ----
        dedup_count = 0
        deduped_ids: List[str] = []
        # 按创建时间降序排列（最新在前），遍历时保留第一个，移除后续相似的
        sorted_by_time = sorted(
            current_ids,
            key=lambda mid: mem_map[mid]["created_at"],
            reverse=True,
        )
        seen_contents: List[str] = []
        for mid in sorted_by_time:
            content = mem_map[mid]["content"]
            is_dup = False
            for seen in seen_contents:
                if _jaccard_sim(content, seen) >= 0.85:
                    is_dup = True
                    break
            if is_dup:
                removed.append(mid)
                reasons[mid] = "dedup"
                dedup_count += 1
            else:
                seen_contents.append(content)
                deduped_ids.append(mid)
        # 保持原有顺序
        current_ids = [mid for mid in current_ids if mid in deduped_ids]

        # ---- 第 2 步：低关联过滤 ----
        low_rel_count = 0
        if query:
            query_clean = self._strip_control(query)
            relevant_ids = []
            for mid in current_ids:
                content = mem_map[mid]["content"]
                score = _keyword_overlap(content, query_clean)
                if score < 0.2:
                    removed.append(mid)
                    reasons[mid] = "low_relevance"
                    low_rel_count += 1
                else:
                    relevant_ids.append(mid)
            current_ids = relevant_ids

        # ---- 第 3 步：过期过滤 ----
        expired_count = 0
        now = time.time()
        non_expired = []
        for mid in current_ids:
            exp = mem_map[mid]["expires_at"]
            if exp > 0 and exp <= now:
                removed.append(mid)
                reasons[mid] = "expired"
                expired_count += 1
            else:
                non_expired.append(mid)
        current_ids = non_expired

        # ---- 第 4 步：低质量过滤 ----
        low_quality_count = 0
        high_quality = []
        for mid in current_ids:
            content = mem_map[mid]["content"]
            if len(content.strip()) < 10:
                removed.append(mid)
                reasons[mid] = "low_quality"
                low_quality_count += 1
            else:
                high_quality.append(mid)
        current_ids = high_quality

        kept = current_ids
        removed_count = len(removed)

        return {
            "kept": kept,
            "removed": removed,
            "reasons": reasons,
            "stats": {
                "total": total,
                "kept_count": len(kept),
                "removed_count": removed_count,
                "dedup": dedup_count,
                "low_relevance": low_rel_count,
                "expired": expired_count,
                "low_quality": low_quality_count,
            },
        }

    def detect_conflicts(self, memory_ids: Optional[List[str]] = None
                         ) -> List[Dict[str, Any]]:
        """事实冲突检测：识别内容相互矛盾的记忆对。

        v5.5.5 新增。

        检测策略：
        1. 数值/时间冲突：提取日期、数字，对比是否矛盾。
        2. 时间线冲突：时间表述顺序不一致。
        3. 事实陈述冲突：高度相似内容但一条含否定词。
        4. 标签冲突：相同实体但标签完全相反。

        Args:
            memory_ids: 可选记忆 ID 列表，不传则检测所有非 trash 记忆

        Returns:
            冲突列表，每个元素：
            {
                "conflict_id": str,
                "memory_a_id": str,
                "memory_b_id": str,
                "conflict_type": str,  # factual_contradiction / value_update / timeline_inconsistency
                "confidence": float,   # 0~1
                "description": str,
                "suggested_action": str,  # merge / keep_newer / keep_higher_importance / review_needed
            }
        """
        import re as _re

        conn = self._get_conn()

        # 获取候选记忆
        if memory_ids:
            cleaned_ids = []
            for mid in memory_ids:
                if not mid or not isinstance(mid, str):
                    continue
                cid = self._strip_control(mid)[:64]
                if cid and cid not in cleaned_ids:
                    cleaned_ids.append(cid)
            if not cleaned_ids:
                return []
            placeholders = ",".join("?" * len(cleaned_ids))
            rows = conn.execute(
                f"SELECT id, content, created_at, updated_at, tags, importance, "
                f"access_count FROM memories WHERE id IN ({placeholders}) "
                f"AND category != 'trash' AND encrypted = 0",
                cleaned_ids,
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT id, content, created_at, updated_at, tags, importance, "
                "access_count FROM memories WHERE category != 'trash' AND encrypted = 0 "
                "ORDER BY created_at DESC LIMIT 500"
            ).fetchall()

        if len(rows) < 2:
            return []

        # 构建记忆数据
        memories: List[Dict[str, Any]] = []
        for row in rows:
            content = row["content"] or ""
            if len(content.strip()) < 3:
                continue
            tags_val = self._safe_json_loads(row["tags"], [])
            memories.append({
                "id": row["id"],
                "content": content,
                "created_at": row["created_at"] or 0.0,
                "updated_at": row["updated_at"] or 0.0,
                "tags": list(tags_val) if isinstance(tags_val, (list, tuple)) else [],
                "importance": row["importance"] or "MEDIUM",
                "access_count": row["access_count"] or 0,
            })

        if len(memories) < 2:
            return []

        conflicts: List[Dict[str, Any]] = []
        conflict_id_set = set()

        # ---- 辅助函数 ----
        def _make_cid(a_id: str, b_id: str) -> str:
            """生成稳定的冲突 ID（按字典序排列避免重复）"""
            if a_id < b_id:
                return f"conflict_{a_id}_{b_id}"
            return f"conflict_{b_id}_{a_id}"

        def _add_conflict(a: Dict, b: Dict, ctype: str,
                          confidence: float, desc: str, action: str):
            cid = _make_cid(a["id"], b["id"])
            if cid in conflict_id_set:
                return  # 避免重复
            conflict_id_set.add(cid)
            conflicts.append({
                "conflict_id": cid,
                "memory_a_id": a["id"],
                "memory_b_id": b["id"],
                "conflict_type": ctype,
                "confidence": round(confidence, 2),
                "description": desc,
                "suggested_action": action,
            })

        def _suggest_action(a: Dict, b: Dict) -> str:
            """根据重要度和时间建议处理方式"""
            imp_map = {"LOW": 0, "MEDIUM": 1, "HIGH": 2, "CRITICAL": 3}
            a_imp = imp_map.get(a["importance"], 1)
            b_imp = imp_map.get(b["importance"], 1)
            imp_diff = abs(a_imp - b_imp)
            time_diff = abs(a["created_at"] - b["created_at"])

            if imp_diff >= 2:
                return "keep_higher_importance"
            if time_diff > 86400 * 7:  # 相差超过 7 天
                return "keep_newer"
            if imp_diff >= 1 and time_diff > 86400:
                return "keep_newer"
            return "review_needed"

        def _extract_numbers(text: str) -> List[Tuple[str, float]]:
            """提取文本中的数值及其上下文（前 2~4 个字符的关键词）"""
            results = []
            # 匹配 "关键词是/为/有/= 数字" 模式
            pattern = _re.compile(
                r'([\u4e00-\u9fa5A-Za-z]{1,6})(?:是|为|有|等于|=|：|:)\s*(\d+(?:\.\d+)?)'
            )
            for m in pattern.finditer(text):
                keyword = m.group(1)
                value = float(m.group(2))
                results.append((keyword, value))
            # 匹配 "数字+单位" 模式
            unit_pattern = _re.compile(
                r'(\d+(?:\.\d+)?)\s*(元|块|个|岁|年|月|天|小时|分钟|次|斤|公斤|克|千克|米|厘米|毫米|公里|km|kg|g|m|cm|mm)'
            )
            for m in unit_pattern.finditer(text):
                value = float(m.group(1))
                unit = m.group(2)
                results.append((unit, value))
            return results

        def _has_negation(text: str) -> bool:
            """检测文本中是否包含否定词"""
            neg_words = ["不是", "没有", "不会", "不能", "不可", "并非", "非也",
                         "没", "不", "否", "非", "无", "未", "莫", "勿", "别"]
            # 优先检测长词
            for w in sorted(neg_words, key=len, reverse=True):
                if w in text:
                    return True
            return False

        def _jaccard_sim(text_a: str, text_b: str) -> float:
            """字符 bigram Jaccard 相似度"""
            if not text_a or not text_b:
                return 0.0
            def _bigrams(t: str) -> set:
                t = t.strip().lower()
                if len(t) < 2:
                    return {t} if t else set()
                return {t[i:i + 2] for i in range(len(t) - 1)}
            sa = _bigrams(text_a)
            sb = _bigrams(text_b)
            if not sa or not sb:
                return 0.0
            inter = len(sa & sb)
            union = len(sa | sb)
            return inter / union if union > 0 else 0.0

        def _extract_dates(text: str) -> List[str]:
            """提取文本中的日期表述"""
            dates = []
            # YYYY-MM-DD / YYYY/MM/DD
            for m in _re.finditer(r'\d{4}[-/]\d{1,2}[-/]\d{1,2}', text):
                dates.append(m.group())
            # MM月DD日
            for m in _re.finditer(r'\d{1,2}月\d{1,2}日', text):
                dates.append(m.group())
            # 相对时间
            rel_patterns = ["今天", "昨天", "前天", "上周", "本周", "下周",
                            "上个月", "这个月", "下个月", "去年", "今年", "明年"]
            for rp in rel_patterns:
                if rp in text:
                    dates.append(rp)
            return dates

        # 相反标签对（用于标签冲突检测）
        opposite_tag_pairs = [
            {"happy", "sad"},
            {"good", "bad"},
            {"positive", "negative"},
            {"true", "false"},
            {"success", "failure"},
            {"like", "dislike"},
            {"开心", "难过"},
            {"好", "坏"},
            {"对", "错"},
            {"是", "否"},
            {"真", "假"},
            {"成功", "失败"},
            {"喜欢", "讨厌"},
            {"记得", "忘记"},
        ]

        # ---- 两两比较 ----
        n = len(memories)
        for i in range(n):
            for j in range(i + 1, n):
                mem_a = memories[i]
                mem_b = memories[j]
                content_a = mem_a["content"]
                content_b = mem_b["content"]

                # 跳过完全相同的（已由去重处理）
                if content_a == content_b:
                    continue

                sim = _jaccard_sim(content_a, content_b)

                # ---- 1. 数值/时间冲突检测 ----
                nums_a = _extract_numbers(content_a)
                nums_b = _extract_numbers(content_b)
                value_conflicts = []
                for kw_a, val_a in nums_a:
                    for kw_b, val_b in nums_b:
                        if kw_a == kw_b and val_a != val_b:
                            # 数值差异比例
                            if val_a > 0 and val_b > 0:
                                ratio = min(val_a, val_b) / max(val_a, val_b)
                                if ratio < 0.9:  # 差异超过 10%
                                    value_conflicts.append((kw_a, val_a, val_b))

                if value_conflicts and sim >= 0.3:
                    kw, va, vb = value_conflicts[0]
                    confidence = min(0.95, 0.5 + sim * 0.3 + 0.15 * len(value_conflicts))
                    desc = f"数值冲突：关键词「{kw}」在两条记忆中分别为 {va} 和 {vb}"
                    action = _suggest_action(mem_a, mem_b)
                    _add_conflict(mem_a, mem_b, "value_update", confidence, desc, action)
                    continue  # 已检测到冲突，跳过后续检测避免重复

                # ---- 2. 时间线冲突检测 ----
                dates_a = _extract_dates(content_a)
                dates_b = _extract_dates(content_b)
                if dates_a and dates_b and sim >= 0.25:
                    # 检测是否存在互斥的相对时间表述
                    rel_exclusive_pairs = [
                        ("上周", "本周"), ("上周", "下周"),
                        ("上个月", "这个月"), ("上个月", "下个月"),
                        ("去年", "今年"), ("去年", "明年"),
                        ("昨天", "今天"), ("前天", "昨天"),
                    ]
                    has_timeline_conflict = False
                    for ra, rb in rel_exclusive_pairs:
                        if (ra in content_a and rb in content_b) or \
                           (ra in content_b and rb in content_a):
                            has_timeline_conflict = True
                            break
                    if has_timeline_conflict:
                        confidence = min(0.85, 0.4 + sim * 0.4)
                        desc = "时间线冲突：两条记忆中的时间表述不一致"
                        action = _suggest_action(mem_a, mem_b)
                        _add_conflict(mem_a, mem_b, "timeline_inconsistency",
                                      confidence, desc, action)
                        continue

                # ---- 3. 事实陈述冲突（否定词检测）----
                if sim >= 0.6:
                    neg_a = _has_negation(content_a)
                    neg_b = _has_negation(content_b)
                    if neg_a != neg_b:  # 一条有否定词，一条没有
                        confidence = min(0.9, 0.5 + sim * 0.4)
                        desc = "事实陈述冲突：两条记忆内容高度相似但一条含否定表述"
                        action = _suggest_action(mem_a, mem_b)
                        _add_conflict(mem_a, mem_b, "factual_contradiction",
                                      confidence, desc, action)
                        continue

                # ---- 4. 标签冲突检测 ----
                tags_a = set(t.lower() for t in mem_a["tags"] if isinstance(t, str))
                tags_b = set(t.lower() for t in mem_b["tags"] if isinstance(t, str))
                if tags_a and tags_b:
                    has_tag_conflict = False
                    for pair in opposite_tag_pairs:
                        pair_lower = {p.lower() for p in pair}
                        if pair_lower.issubset(tags_a | tags_b) and \
                           not pair_lower.issubset(tags_a) and \
                           not pair_lower.issubset(tags_b):
                            # 一对相反标签分别在两条记忆中
                            if (pair_lower & tags_a) and (pair_lower & tags_b):
                                has_tag_conflict = True
                                break
                    if has_tag_conflict and sim >= 0.2:
                        confidence = min(0.75, 0.35 + sim * 0.3)
                        desc = "标签冲突：两条记忆包含相反的情感/状态标签"
                        action = _suggest_action(mem_a, mem_b)
                        _add_conflict(mem_a, mem_b, "factual_contradiction",
                                      confidence, desc, action)

        # 按置信度降序排列
        conflicts.sort(key=lambda c: c["confidence"], reverse=True)
        return conflicts

    def resolve_conflict(self, conflict: Dict[str, Any],
                         strategy: str = "auto") -> Optional[str]:
        """冲突解决：根据策略处理冲突。

        v5.5.5 新增。

        strategy 可选：
        - "auto": 自动选择（优先更新的 > 重要度高的 > 访问次数多的）
        - "keep_a": 保留 memory_a，将 memory_b 移入 trash
        - "keep_b": 保留 memory_b，将 memory_a 移入 trash
        - "merge": 合并两条（复用 merge_memories 逻辑，较新的作为 target）

        Args:
            conflict: 冲突字典，需包含 memory_a_id 和 memory_b_id
            strategy: 解决策略

        Returns:
            保留的记忆 ID，失败返回 None
        """
        if not conflict or not isinstance(conflict, dict):
            return None

        a_id = conflict.get("memory_a_id", "")
        b_id = conflict.get("memory_b_id", "")

        if not a_id or not b_id:
            return None

        # v5.4.0 安全加固
        a_id = self._strip_control(a_id)[:64]
        b_id = self._strip_control(b_id)[:64]
        strategy = self._strip_control(strategy)[:32].lower()

        if a_id == b_id:
            return None

        conn = self._get_conn()
        now = time.time()

        # 获取两条记忆
        row_a = conn.execute(
            "SELECT * FROM memories WHERE id = ? AND category != 'trash'",
            (a_id,)
        ).fetchone()
        row_b = conn.execute(
            "SELECT * FROM memories WHERE id = ? AND category != 'trash'",
            (b_id,)
        ).fetchone()

        if not row_a and not row_b:
            return None
        if not row_a:
            return b_id
        if not row_b:
            return a_id

        entry_a = self._row_to_entry(row_a)
        entry_b = self._row_to_entry(row_b)

        if strategy == "keep_a":
            # 保留 A，删除 B
            conn.execute(
                "UPDATE memories SET category = 'trash', updated_at = ? WHERE id = ?",
                (now, b_id)
            )
            # 审计
            self._add_audit("delete", b_id, "", "", entry_b.privacy.value,
                            details={"reason": "conflict_resolved",
                                     "strategy": "keep_a",
                                     "kept_id": a_id})
            conn.commit()
            return a_id

        elif strategy == "keep_b":
            # 保留 B，删除 A
            conn.execute(
                "UPDATE memories SET category = 'trash', updated_at = ? WHERE id = ?",
                (now, a_id)
            )
            self._add_audit("delete", a_id, "", "", entry_a.privacy.value,
                            details={"reason": "conflict_resolved",
                                     "strategy": "keep_b",
                                     "kept_id": b_id})
            conn.commit()
            return b_id

        elif strategy == "merge":
            # 合并：较新的作为 target，较旧的作为 source
            if entry_a.created_at >= entry_b.created_at:
                target_id, source_id = a_id, b_id
            else:
                target_id, source_id = b_id, a_id
            result = self.merge_memories(source_id, target_id)
            return target_id if result else None

        elif strategy == "auto":
            # 自动选择：更新的 > 重要度高的 > 访问次数多的
            imp_map = {"LOW": 0, "MEDIUM": 1, "HIGH": 2, "CRITICAL": 3}

            a_newer = entry_a.updated_at >= entry_b.updated_at
            a_higher_imp = imp_map.get(str(entry_a.importance), 1) > \
                           imp_map.get(str(entry_b.importance), 1)
            a_more_access = entry_a.access_count >= entry_b.access_count

            # 评分：更新时间权重最高，其次重要度，最后访问次数
            score_a = (3 if a_newer else 0) + (2 if a_higher_imp else 0) + (1 if a_more_access else 0)
            score_b = (3 if not a_newer else 0) + (2 if not a_higher_imp else 0) + (1 if not a_more_access else 0)

            if score_a >= score_b:
                keep_id, remove_id = a_id, b_id
                keep_entry, remove_entry = entry_a, entry_b
            else:
                keep_id, remove_id = b_id, a_id
                keep_entry, remove_entry = entry_b, entry_a

            conn.execute(
                "UPDATE memories SET category = 'trash', updated_at = ? WHERE id = ?",
                (now, remove_id)
            )
            self._add_audit("delete", remove_id, "", "", remove_entry.privacy.value,
                            details={"reason": "conflict_resolved",
                                     "strategy": "auto",
                                     "kept_id": keep_id,
                                     "score_kept": max(score_a, score_b),
                                     "score_removed": min(score_a, score_b)})
            conn.commit()
            return keep_id

        else:
            logger.warning("Unknown conflict resolution strategy: %s", strategy)
            return None

    def get_conflict_stats(self) -> Dict[str, Any]:
        """获取冲突统计：总冲突数、各类型数量、待处理数。

        v5.5.5 新增。

        Returns:
            冲突统计字典：
            {
                "total_conflicts": int,
                "by_type": {
                    "factual_contradiction": int,
                    "value_update": int,
                    "timeline_inconsistency": int,
                },
                "high_confidence": int,    # confidence >= 0.8
                "medium_confidence": int,  # 0.5 <= confidence < 0.8
                "low_confidence": int,     # confidence < 0.5
                "pending_review": int,     # 建议人工审核的数量
                "total_memories_checked": int,
            }
        """
        conn = self._get_conn()

        # 统计记忆总数（非 trash、非加密）
        total_mem = conn.execute(
            "SELECT COUNT(*) FROM memories WHERE category != 'trash' AND encrypted = 0"
        ).fetchone()[0]

        # 检测所有冲突
        conflicts = self.detect_conflicts()

        total = len(conflicts)
        by_type = {
            "factual_contradiction": 0,
            "value_update": 0,
            "timeline_inconsistency": 0,
        }
        high_conf = 0
        medium_conf = 0
        low_conf = 0
        pending_review = 0

        for c in conflicts:
            ctype = c.get("conflict_type", "")
            if ctype in by_type:
                by_type[ctype] += 1
            conf = c.get("confidence", 0.0)
            if conf >= 0.8:
                high_conf += 1
            elif conf >= 0.5:
                medium_conf += 1
            else:
                low_conf += 1
            if c.get("suggested_action") == "review_needed":
                pending_review += 1

        return {
            "total_conflicts": total,
            "by_type": by_type,
            "high_confidence": high_conf,
            "medium_confidence": medium_conf,
            "low_confidence": low_conf,
            "pending_review": pending_review,
            "total_memories_checked": total_mem,
        }

